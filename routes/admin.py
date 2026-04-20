"""
绠＄悊鍛樿矾鐢?
"""

import os
import re
import uuid
from datetime import datetime, timedelta
from urllib.parse import unquote

from flask import Blueprint, render_template, request, jsonify, redirect, url_for, send_file, current_app
from flask_login import login_required, current_user
from sqlalchemy import case, func, or_

from extensions import db
from models.user import User, UserSession
from models.device import DeviceRegistry
from models.file import UploadedFile
from models.ticket import LotteryTicket
from models.winning import WinningRecord
from models.result import MatchResult, ResultFile
from models.settings import SystemSettings
from models.audit import AuditLog
from services.file_parser import process_uploaded_file, revoke_file
from services.session_service import force_logout_user
from services.ticket_pool import get_pool_status
from services.notify_service import notify_admins, notify_all
from utils.decorators import admin_required, get_client_ip, login_required_json, parse_json_object
from utils.time_utils import beijing_now, get_business_date, get_business_window, get_today_noon

admin_bp = Blueprint('admin', __name__)
SCHEDULER_EXPECTED_JOB_IDS = (
    'expire_tickets',
    'clean_sessions',
    'daily_reset',
    'db_keepalive',
    'archive_tickets',
    'archive_uploaded_txt_files',
    'purge_old_auxiliary_records',
)

RESULT_UPLOAD_KIND_HINTS = {
    'predicted': ('\u9884\u6d4b', 'predicted'),
    'final': ('\u6700\u7ec8', 'final'),
}
RESULT_UPLOAD_LOTTERY_TYPES = (
    '\u80dc\u5e73\u8d1f',
    '\u80dc\u8d1f',
    '\u6bd4\u5206',
    '\u4e0a\u4e0b\u76d8',
    '\u603b\u8fdb\u7403',
    '\u534a\u5168\u573a',
)


def _winning_terminal_at(ticket: LotteryTicket):
    return ticket.completed_at or ticket.deadline_time or ticket.assigned_at or ticket.admin_upload_time


def _winning_status_label(status: str) -> str:
    if status == 'expired':
        return '\u5df2\u8fc7\u671f\u672a\u51fa\u7968'
    if status == 'completed':
        return '\u5df2\u51fa\u7968'
    if status == 'revoked':
        return '\u5df2\u64a4\u9500'
    return status or ''


def _get_winning_ticket_or_error(ticket_id_value):
    parsed_ticket_id = _parse_int_arg(ticket_id_value, minimum=1)
    if parsed_ticket_id is None:
        return None, (jsonify({'success': False, 'error': '\u7968ID\u5fc5\u987b\u662f\u5927\u4e8e 0 \u7684\u6574\u6570'}), 400)

    ticket = db.session.get(LotteryTicket, parsed_ticket_id)
    if not ticket:
        return None, (jsonify({'success': False, 'error': '\u7968\u636e\u4e0d\u5b58\u5728'}), 404)
    if not ticket.is_winning:
        return None, (jsonify({'success': False, 'error': '\u672a\u88ab\u7cfb\u7edf\u5224\u5b9a\u4e3a\u4e2d\u5956'}), 400)
    return ticket, None


def _winning_key_matches_ticket(ticket_id: int, oss_key: str) -> bool:
    normalized_key = (oss_key or '').strip()
    if not normalized_key:
        return False

    pattern = rf"^winning(?:[/_]\d{{4}})(?:[/_]\d{{2}})(?:[/_]\d{{2}})[/_]{int(ticket_id)}\.(jpg|jpeg|png|gif|webp)$"
    return re.fullmatch(pattern, normalized_key, flags=re.IGNORECASE) is not None


def _parse_int_arg(value, minimum=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    if minimum is not None and parsed < minimum:
        return None
    return parsed


def _parse_client_mode(value):
    if value not in {'mode_a', 'mode_b'}:
        return None
    return value


def _parse_bool_flag(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {'true', '1'}:
            return True
        if normalized in {'false', '0'}:
            return False
    if isinstance(value, int) and value in (0, 1):
        return bool(value)
    return None


def _safe_uploaded_filename(filename: str) -> str:
    normalized = (filename or '').strip().replace('\\', '/')
    return normalized.rsplit('/', 1)[-1]


def _safe_result_period_folder(detail_period: str) -> str:
    safe_period = re.sub(r'[^0-9A-Za-z_-]+', '_', (detail_period or '').strip()).strip('_')
    return safe_period or 'unknown_period'


def _build_result_upload_relative_path(filename: str, detail_period: str, upload_kind: str) -> str:
    safe_name = _safe_uploaded_filename(filename)
    timestamp = beijing_now().strftime('%Y%m%d%H%M%S')
    unique_id = uuid.uuid4().hex[:8]
    period_folder = _safe_result_period_folder(detail_period)
    stored_name = f"{upload_kind}_{timestamp}_{unique_id}_{safe_name}"
    return os.path.join('results', period_folder, stored_name)


def _validate_result_upload_filename(filename: str, detail_period: str, upload_kind: str):
    basename = _safe_uploaded_filename(filename)
    compact = re.sub(r'\s+', '', basename).lower()

    if not re.search(rf'(?<!\d){re.escape(detail_period)}(?!\d)', compact):
        return f'\u6587\u4ef6\u540d\u9700\u5305\u542b\u671f\u53f7 {detail_period}'

    has_predicted = any(token in compact for token in RESULT_UPLOAD_KIND_HINTS['predicted'])
    has_final = any(token in compact for token in RESULT_UPLOAD_KIND_HINTS['final'])
    if has_predicted and has_final:
        return '\u6587\u4ef6\u540d\u4e0d\u80fd\u540c\u65f6\u5305\u542b\u9884\u6d4b\u548c\u6700\u7ec8'

    expected_tokens = RESULT_UPLOAD_KIND_HINTS[upload_kind]
    if not any(token in compact for token in expected_tokens):
        if upload_kind == 'predicted':
            return '\u4e0a\u4f20\u7c7b\u578b\u4e3a\u201c\u9884\u6d4b\u201d\u65f6\uff0c\u6587\u4ef6\u540d\u9700\u5305\u542b\u201c\u9884\u6d4b\u201d'
        return '\u4e0a\u4f20\u7c7b\u578b\u4e3a\u201c\u6700\u7ec8\u201d\u65f6\uff0c\u6587\u4ef6\u540d\u9700\u5305\u542b\u201c\u6700\u7ec8\u201d'

    return None


def _resolve_result_upload_lottery_type(detail_period: str, lottery_type_raw):
    lottery_type = (lottery_type_raw or '').strip()
    if lottery_type:
        if lottery_type not in RESULT_UPLOAD_LOTTERY_TYPES:
            return None, 'invalid lottery_type'
        return lottery_type, None

    distinct_types = [
        row[0]
        for row in db.session.query(LotteryTicket.lottery_type).filter(
            LotteryTicket.detail_period == detail_period,
            LotteryTicket.lottery_type.isnot(None),
        ).distinct().all()
        if row[0]
    ]

    if len(distinct_types) > 1:
        # Keep admin upload flow backward-compatible when form omits lottery_type.
        return None, None
    if len(distinct_types) == 1:
        return distinct_types[0], None
    return None, None


def _decimal_to_float(value):
    return float(value) if value is not None else None


def _winning_change_percent(predicted_amount, final_amount):
    if predicted_amount in (None, 0) or final_amount is None:
        return None
    return round(((final_amount - predicted_amount) / predicted_amount) * 100, 2)


def _resolve_device_display_name(device_id, client_info):
    if isinstance(client_info, dict):
        for key in ('device_name', 'deviceName', 'name', 'hostname', 'client_name'):
            value = client_info.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
    return (device_id or '').strip() or '未知设备'


def _database_display_info():
    db_uri = current_app.config.get('SQLALCHEMY_DATABASE_URI', '')
    if db_uri.startswith('sqlite:///'):
        raw_path = db_uri[len('sqlite:///'):]
        return {
            'engine': 'sqlite',
            'path': raw_path.replace('/', os.sep),
        }

    return {
        'engine': db_uri.split(':', 1)[0] if db_uri else 'unknown',
        'path': unquote(db_uri),
    }


def _build_health_summary(now=None):
    now = now or beijing_now()
    items = []

    overdue_pending_count = LotteryTicket.query.filter(
        LotteryTicket.status == 'pending',
        LotteryTicket.deadline_time.isnot(None),
        LotteryTicket.deadline_time < now,
    ).count()
    if overdue_pending_count:
        items.append({
            'type': 'overdue_pending_tickets',
            'level': 'warning',
            'count': int(overdue_pending_count),
            'message': '\u53d1\u73b0\u8fc7\u671f\u672a\u5904\u7406\u7968\u636e',
            'action': '\u8bf7\u4f18\u5148\u68c0\u67e5\u6587\u4ef6\u5904\u7406\u9875\u5e76\u5904\u7406\u8fc7\u671f\u6b8b\u7559\u3002',
        })

    stale_assigned_cutoff = now - timedelta(hours=2)
    stale_assigned_count = LotteryTicket.query.filter(
        LotteryTicket.status == 'assigned',
        LotteryTicket.assigned_at.isnot(None),
        LotteryTicket.assigned_at < stale_assigned_cutoff,
    ).count()
    if stale_assigned_count:
        items.append({
            'type': 'stale_assigned_tickets',
            'level': 'warning',
            'count': int(stale_assigned_count),
            'message': '\u53d1\u73b0\u957f\u65f6\u95f4\u672a\u5b8c\u6210\u7684\u5904\u7406\u4e2d\u7968\u636e',
            'action': '\u8bf7\u5148\u786e\u8ba4\u5bf9\u5e94\u8bbe\u5907\u662f\u5426\u4ecd\u5728\u7ebf\u5e76\u6e05\u7406\u6b8b\u7559\u5206\u914d\u3002',
        })

    ticket_counter_sq = db.session.query(
        LotteryTicket.source_file_id.label('file_id'),
        func.sum(case((LotteryTicket.status == 'pending', 1), else_=0)).label('pending_actual'),
        func.sum(case((LotteryTicket.status == 'assigned', 1), else_=0)).label('assigned_actual'),
        func.sum(case((LotteryTicket.status == 'completed', 1), else_=0)).label('completed_actual'),
    ).group_by(LotteryTicket.source_file_id).subquery()

    mismatch_count = db.session.query(func.count(UploadedFile.id)).outerjoin(
        ticket_counter_sq, ticket_counter_sq.c.file_id == UploadedFile.id
    ).filter(
        UploadedFile.status != 'revoked',
        or_(
            UploadedFile.pending_count != func.coalesce(ticket_counter_sq.c.pending_actual, 0),
            UploadedFile.assigned_count != func.coalesce(ticket_counter_sq.c.assigned_actual, 0),
            UploadedFile.completed_count != func.coalesce(ticket_counter_sq.c.completed_actual, 0),
        ),
    ).scalar() or 0
    if mismatch_count:
        items.append({
            'type': 'uploaded_file_counter_mismatch',
            'level': 'critical',
            'count': int(mismatch_count),
            'message': '\u53d1\u73b0\u6587\u4ef6\u8ba1\u6570\u4e0e\u7968\u636e\u72b6\u6001\u4e0d\u4e00\u81f4',
            'action': '\u8bf7\u6682\u505c\u76f8\u5173\u6587\u4ef6\u64cd\u4f5c\u5e76\u8054\u7cfb\u6280\u672f\u6392\u67e5\u8ba1\u6570\u4fee\u590d\u903b\u8f91\u3002',
        })

    result_file_error_count = ResultFile.query.filter(ResultFile.status == 'error').count()
    if result_file_error_count:
        items.append({
            'type': 'result_file_parse_error',
            'level': 'critical',
            'count': int(result_file_error_count),
            'message': '\u53d1\u73b0\u8d5b\u679c\u6587\u4ef6\u89e3\u6790\u5931\u8d25',
            'action': '\u8bf7\u6682\u505c\u91cd\u590d\u4e0a\u4f20\u540c\u4e00\u671f\u8d5b\u679c\u5e76\u5148\u6392\u67e5\u5931\u8d25\u6587\u4ef6\u3002',
        })

    match_calc_error_count = MatchResult.query.filter(MatchResult.calc_status == 'error').count()
    if match_calc_error_count:
        items.append({
            'type': 'match_result_calc_error',
            'level': 'critical',
            'count': int(match_calc_error_count),
            'message': '\u53d1\u73b0\u8d5b\u679c\u91cd\u7b97\u5931\u8d25',
            'action': '\u8bf7\u6682\u505c\u8d5b\u679c\u91cd\u7b97\u64cd\u4f5c\u5e76\u8054\u7cfb\u6280\u672f\u652f\u6301\u3002',
        })

    match_processing_stale_count = MatchResult.query.filter(
        MatchResult.calc_status == 'processing',
        MatchResult.calc_started_at.isnot(None),
        MatchResult.calc_started_at < (now - timedelta(minutes=30)),
    ).count()
    if match_processing_stale_count:
        items.append({
            'type': 'match_result_processing_stale',
            'level': 'warning',
            'count': int(match_processing_stale_count),
            'message': '\u5b58\u5728\u957f\u65f6\u95f4\u672a\u7ed3\u675f\u7684\u8d5b\u679c\u91cd\u7b97',
            'action': '\u8bf7\u5728\u8d5b\u679c\u9875\u786e\u8ba4\u91cd\u7b97\u8fdb\u5ea6\uff0c\u5fc5\u8981\u65f6\u89e6\u53d1\u91cd\u7b97\u3002',
        })

    # 涓氬姟鍙ｅ緞锛氫腑濂栫エ涓嶅己鍒朵笂浼犲浘鐗囨垨琛ヤ腑濂栬褰曪紝鍥犳涓嶇撼鍏ュ仴搴锋憳瑕佹彁閱掗」銆?

    from tasks.scheduler import get_scheduler

    scheduler = get_scheduler()
    expected_job_ids = set(SCHEDULER_EXPECTED_JOB_IDS)
    if scheduler is None:
        items.append({
            'type': 'scheduler_not_started',
            'level': 'critical',
            'count': 1,
            'message': '\u8c03\u5ea6\u5668\u672a\u542f\u52a8',
            'action': '\u8bf7\u5148\u786e\u8ba4\u670d\u52a1\u542f\u52a8\u65e5\u5fd7\u4e0e\u5b9a\u65f6\u4efb\u52a1\u521d\u59cb\u5316\u72b6\u6001\u3002',
        })
    else:
        current_job_ids = {job.id for job in scheduler.get_jobs()}
        missing_job_count = len(expected_job_ids - current_job_ids)
        if missing_job_count:
            items.append({
                'type': 'scheduler_job_missing',
                'level': 'critical',
                'count': int(missing_job_count),
                'message': '\u8c03\u5ea6\u5668\u5173\u952e\u4efb\u52a1\u7f3a\u5931',
                'action': '\u8bf7\u5728\u8bbe\u7f6e\u9875\u6838\u5bf9\u8c03\u5ea6\u4efb\u52a1\u5e76\u8054\u7cfb\u6280\u672f\u8865\u9f50\u3002',
            })

    level_order = {'critical': 2, 'warning': 1}
    items.sort(key=lambda item: (level_order.get(item['level'], 0), item['count']), reverse=True)
    items = items[:5]

    if any(item['level'] == 'critical' for item in items):
        status = 'critical'
        summary = '\u53d1\u73b0\u4e25\u91cd\u5f02\u5e38\uff0c\u8bf7\u5148\u6682\u505c\u76f8\u5173\u64cd\u4f5c\u5e76\u8054\u7cfb\u6280\u672f\u652f\u6301'
    elif items:
        status = 'warning'
        summary = f'\u53d1\u73b0 {len(items)} \u6761\u63d0\u9192\u9879\uff0c\u8bf7\u4f18\u5148\u5904\u7406'
    else:
        status = 'normal'
        summary = '\u7cfb\u7edf\u6b63\u5e38\uff0c\u53ef\u7ee7\u7eed\u5904\u7406\u4e1a\u52a1'

    return {
        'status': status,
        'summary': summary,
        'items': items,
        'generated_at': now.isoformat(),
    }


def _build_scheduler_status():
    expected_job_ids = list(SCHEDULER_EXPECTED_JOB_IDS)

    from tasks.scheduler import get_scheduler

    scheduler = get_scheduler()
    if scheduler is None:
        return {
            'scheduler_present': False,
            'scheduler_running': False,
            'job_count': 0,
            'job_ids': [],
            'expected_job_ids': expected_job_ids,
            'missing_job_ids': expected_job_ids,
            'status': 'critical',
            'message': '\u8c03\u5ea6\u5668\u672a\u542f\u52a8\u3002',
            'action': '\u8bf7\u6682\u505c\u4f9d\u8d56\u81ea\u52a8\u8fc7\u671f\u4e0e\u81ea\u52a8\u6e05\u7406\u7684\u64cd\u4f5c\u5e76\u8054\u7cfb\u6280\u672f\u3002',
        }

    scheduler_running = bool(getattr(scheduler, 'running', True))
    try:
        job_ids = sorted(job.id for job in scheduler.get_jobs())
    except Exception:
        return {
            'scheduler_present': True,
            'scheduler_running': scheduler_running,
            'job_count': 0,
            'job_ids': [],
            'expected_job_ids': expected_job_ids,
            'missing_job_ids': expected_job_ids,
            'status': 'warning',
            'message': '\u8c03\u5ea6\u5668\u72b6\u6001\u90e8\u5206\u5b57\u6bb5\u6682\u4e0d\u53ef\u8bfb\u3002',
            'action': '\u8bf7\u7559\u610f\u8fc7\u671f\u7968\u4e0e\u4f1a\u8bdd\u6e05\u7406\u7ed3\u679c\uff0c\u5fc5\u8981\u65f6\u8054\u7cfb\u6280\u672f\u6392\u67e5\u3002',
        }

    missing_job_ids = sorted(set(expected_job_ids) - set(job_ids))
    if not scheduler_running:
        status = 'critical'
        message = '\u8c03\u5ea6\u5668\u672a\u8fd0\u884c\u3002'
        action = '\u8bf7\u6682\u505c\u4f9d\u8d56\u81ea\u52a8\u4efb\u52a1\u7684\u64cd\u4f5c\u5e76\u8054\u7cfb\u6280\u672f\u3002'
    elif missing_job_ids:
        status = 'critical'
        message = '\u8c03\u5ea6\u5668\u5173\u952e\u4efb\u52a1\u7f3a\u5931\u3002'
        action = '\u8bf7\u5728\u8bbe\u7f6e\u9875\u6838\u5bf9\u4efb\u52a1\u72b6\u6001\u5e76\u8054\u7cfb\u6280\u672f\u8865\u9f50\u3002'
    else:
        status = 'normal'
        message = '\u8c03\u5ea6\u5668\u6b63\u5e38\uff0c\u53ef\u4ee5\u7ee7\u7eed\u64cd\u4f5c\u3002'
        action = '\u53ef\u7ee7\u7eed\u65e5\u5e38\u64cd\u4f5c\u3002'

    return {
        'scheduler_present': True,
        'scheduler_running': scheduler_running,
        'job_count': len(job_ids),
        'job_ids': job_ids,
        'expected_job_ids': expected_job_ids,
        'missing_job_ids': missing_job_ids,
        'status': status,
        'message': message,
        'action': action,
    }


@admin_bp.route('/')
@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    return render_template('admin/dashboard.html')


@admin_bp.route('/api/dashboard-data')
@login_required_json
@login_required
@admin_required
def dashboard_data():
    """瀹炴椂 Dashboard 鏁版嵁鎺ュ彛"""
    pool = get_pool_status()

    from sqlalchemy import text
    from datetime import timedelta
    from models.user import UserSession
    cutoff = beijing_now() - timedelta(minutes=2)  # 2鍒嗛挓鍐呮椿璺冭涓哄湪绾?

    # Get online users via ORM (SQLite compatible)
    active_sessions = UserSession.query.filter(UserSession.last_seen > cutoff).all()
    user_ids = list({s.user_id for s in active_sessions})

    from models.user import User as UserModel
    online_users_objs = UserModel.query.filter(
        UserModel.id.in_(user_ids), UserModel.is_admin == False
    ).all() if user_ids else []

    # 璁＄畻浠婃棩涓氬姟鏃堕棿鑼冨洿
    today_start = get_today_noon()
    today_end = today_start + timedelta(days=1)

    # 鍦ㄧ嚎鐢ㄦ埛缁熻
    user_stats = []
    device_speed_stats = []  # 璁惧閫熷害缁熻
    total_speed = 0.0  # 鎬婚€熷害锛堟瘡鍒嗛挓寮犳暟锛屽彧缁熻褰撳墠鍦ㄧ嚎璁惧锛?

    # 璁＄畻鏈€杩?440鍒嗛挓锛?4灏忔椂锛夌殑鏃堕棿鐐癸紙鐢ㄤ簬閫熷害缁熻锛屾椂闂寸獥鍙ｆ洿闀挎洿绋冲畾锛?
    SPEED_WINDOW_MINUTES = 1440
    speed_window_start = beijing_now() - timedelta(minutes=SPEED_WINDOW_MINUTES)

    # 浼樺寲锛氫竴娆℃€ф煡璇㈡墍鏈夊湪绾跨敤鎴风殑鏈€杩戝畬鎴愮エ锛堥伩鍏峃+1鏌ヨ锛?
    online_user_ids = [ou.id for ou in online_users_objs]
    all_recent_tickets = LotteryTicket.query.filter(
        LotteryTicket.assigned_user_id.in_(online_user_ids),
        LotteryTicket.status == 'completed',
        LotteryTicket.completed_at >= speed_window_start
    ).all() if online_user_ids else []

    # 鎸?(user_id, device_id) 鍒嗙粍
    from collections import defaultdict
    tickets_by_device = defaultdict(list)
    for t in all_recent_tickets:
        key = (t.assigned_user_id, t.assigned_device_id)
        tickets_by_device[key].append(t)

    for ou in online_users_objs:
        # 鐢ㄦ暟鎹簱杩囨护浠婃棩瀹屾垚绁?
        today_tickets = LotteryTicket.query.filter(
            LotteryTicket.assigned_user_id == ou.id,
            LotteryTicket.status == 'completed',
            LotteryTicket.completed_at >= today_start,
            LotteryTicket.completed_at < today_end
        ).all()
        active_count = LotteryTicket.query.filter_by(assigned_user_id=ou.id, status='assigned').count()
        device_count = len({s.device_id for s in active_sessions if s.user_id == ou.id})

        user_stats.append({
            'id': ou.id,
            'username': ou.username,
            'client_mode': ou.client_mode,
            'can_receive': ou.can_receive,
            'device_count': device_count,
            'ticket_count': len(today_tickets),
            'total_amount': sum(float(t.ticket_amount or 0) for t in today_tickets),
            'active_count': active_count,
        })

        # 缁熻璇ョ敤鎴锋瘡涓澶囩殑澶勭悊閫熷害
        user_devices = {s.device_id for s in active_sessions if s.user_id == ou.id and s.device_id}
        for device_id in user_devices:
            # 浠庨鍔犺浇鐨勬暟鎹腑鑾峰彇璇ヨ澶囩殑鏈€杩戝畬鎴愮エ
            recent_tickets = tickets_by_device.get((ou.id, device_id), [])

            if recent_tickets and len(recent_tickets) >= 1:  # 鑷冲皯1寮犵エ鎵嶈兘璁＄畻閫熷害
                # 绛涢€夋湁鏁堢エ锛堟湁鍒嗛厤鍜屽畬鎴愭椂闂达級
                valid_tickets = [t for t in recent_tickets if t.assigned_at and t.completed_at]

                if valid_tickets:
                    # 璁＄畻瀹為檯鏃堕棿璺ㄥ害锛氫粠鏈€鏃╁垎閰嶅埌鏈€鏅氬畬鎴?
                    # 瀵逛簬 B 妯″紡锛屽悓涓€鎵规鐨?assigned_at 鏄浉鍚岀殑
                    # 濡傛灉鍙湁涓€鎵规锛岄€熷害浼氭牴鎹崟寮犵エ鐨勬椂闂磋绠楋紙濡傛灉 span 涓?0 浼氭湁淇濇姢鍊硷級
                    sorted_by_assigned = sorted(valid_tickets, key=lambda t: t.assigned_at)
                    sorted_by_completed = sorted(valid_tickets, key=lambda t: t.completed_at)

                    earliest_assigned = sorted_by_assigned[0].assigned_at
                    latest_completed = sorted_by_completed[-1].completed_at

                    time_span_seconds = (latest_completed - earliest_assigned).total_seconds()
                    time_span_minutes = time_span_seconds / 60.0

                    # 缁熶竴淇濇姢鍊硷細濡傛灉鏃堕棿璺ㄥ害澶煭锛?=0 鎴?<0.1鍒嗛挓鍗?绉掞級锛屼娇鐢?.1鍒嗛挓閬垮厤閫熷害杩囬珮
                    if time_span_minutes <= 0 or time_span_minutes < 0.1:
                        time_span_minutes = 0.1

                    # 閫熷害 = 绁ㄦ暟 / 鏃堕棿璺ㄥ害锛堝垎閽燂級
                    speed_per_minute = len(valid_tickets) / time_span_minutes
                    total_speed += speed_per_minute

                    resolved_device_id = valid_tickets[0].assigned_device_id or device_id
                    device_speed_stats.append({
                        'username': ou.username,
                        'device_id': resolved_device_id,
                        'speed_per_minute': round(speed_per_minute, 2),
                        'recent_count': len(valid_tickets),
                        'time_span_minutes': round(time_span_minutes, 1),
                    })

    # 璁＄畻棰勪及瀹屾垚鏃堕棿
    estimated_minutes = None
    estimated_time_str = None
    if total_speed > 0.01 and pool['total_pending'] > 0:  # 鑷冲皯姣忓垎閽?.01寮?
        estimated_minutes = pool['total_pending'] / total_speed
        # 娣诲姞涓婇檺淇濇姢锛堣秴杩?澶╂樉绀烘彁绀猴級
        if estimated_minutes > 10080:  # 7澶?= 10080鍒嗛挓
            estimated_time_str = "\u8d85\u8fc7 7 \u5929"
        else:
            hours = int(estimated_minutes // 60)
            minutes = int(estimated_minutes % 60)
            if hours > 0:
                estimated_time_str = f"{hours}\u5c0f\u65f6{minutes}\u5206\u949f"
            else:
                estimated_time_str = f"{minutes}\u5206\u949f"

    # 浠婃棩鎵€鏈夌敤鎴峰嚭绁ㄧ粺璁★紙鍖呮嫭涓嶅湪绾跨殑锛?
    daily_stats_query = db.session.query(
        LotteryTicket.assigned_username,
        func.count(LotteryTicket.id).label('count'),
        func.sum(LotteryTicket.ticket_amount).label('amount')
    ).filter(
        LotteryTicket.status == 'completed',
        LotteryTicket.completed_at >= today_start,
        LotteryTicket.completed_at < today_end
    ).group_by(LotteryTicket.assigned_username).all()

    daily_all_users = [
        {
            'username': row.assigned_username,
            'count': row.count,
            'amount': float(row.amount or 0)
        }
        for row in daily_stats_query
    ]

    try:
        health_summary = _build_health_summary()
    except Exception:
        current_app.logger.exception("Failed to build health summary")
        health_summary = {
            'status': 'warning',
            'summary': '\u5065\u5eb7\u6458\u8981\u6682\u4e0d\u53ef\u7528\uff0c\u4e0d\u5f71\u54cd\u5f53\u524d\u4e1a\u52a1',
            'items': [{
                'type': 'summary_unavailable',
                'level': 'warning',
                'count': 1,
                'message': '\u5065\u5eb7\u6458\u8981\u68c0\u67e5\u5931\u8d25',
                'action': '\u53ef\u5148\u7ee7\u7eed\u6838\u5fc3\u64cd\u4f5c\uff0c\u7a0d\u540e\u8054\u7cfb\u6280\u672f\u6392\u67e5\u3002',
            }],
            'generated_at': beijing_now().isoformat(),
        }

    return jsonify({
        'pool': pool,
        'online_users': user_stats,
        'daily_all_users': daily_all_users,
        'device_speed_stats': device_speed_stats,
        'total_speed': round(total_speed, 2),
        'estimated_time': estimated_time_str,
        'estimated_minutes': round(estimated_minutes, 1) if estimated_minutes else None,
        'health_summary': health_summary,
    })


# 鈹€鈹€ File management 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€

@admin_bp.route('/files/upload', methods=['POST'])
@login_required_json
@login_required
@admin_required
def upload_files():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'success': False, 'error': '\u8bf7\u9009\u62e9\u6587\u4ef6'}), 400

    results = []
    for f in files:
        if not f.filename:
            results.append({
                'success': False,
                'filename': '',
                'file_id': None,
                'message': '\u6587\u4ef6\u540d\u4e3a\u7a7a',
            })
            continue

        try:
            result = process_uploaded_file(f, current_user.id)
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception('Admin upload failed for file %s', getattr(f, 'filename', ''))
            result = {
                'success': False,
                'filename': getattr(f, 'filename', '') or '',
                'file_id': None,
                'message': f'\u4e0a\u4f20\u5904\u7406\u5931\u8d25: {exc}',
            }

        results.append(result)

    any_success = any(item.get('success') for item in results)

    if any_success:
        try:
            from services.notify_service import notify_pool_update
            notify_pool_update(get_pool_status())
        except Exception as notify_exc:
            current_app.logger.warning('notify_pool_update failed after upload: %s', notify_exc)

    if not any_success:
        return jsonify({'success': False, 'results': results, 'error': '\u672c\u6b21\u4e0a\u4f20\u5168\u90e8\u5931\u8d25'}), 400

    return jsonify({'success': True, 'results': results})


@admin_bp.route('/files')
@login_required
@admin_required
def files_list():
    return render_template('admin/upload.html')


@admin_bp.route('/api/files')
@login_required_json
@login_required
@admin_required
def api_files_list():
    page = _parse_int_arg(request.args.get('page', 1), minimum=1)
    per_page = _parse_int_arg(request.args.get('per_page', 20), minimum=1)
    if page is None or per_page is None:
        return jsonify({'success': False, 'error': '\u5206\u9875\u53c2\u6570\u5fc5\u987b\u662f\u5927\u4e8e 0 \u7684\u6574\u6570'}), 400
    status_filter = request.args.get('status', '')
    date_str = request.args.get('date', '').strip()

    q = UploadedFile.query.order_by(UploadedFile.uploaded_at.desc())
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': '鏃ユ湡鏍煎紡鏃犳晥锛岃浣跨敤 YYYY-MM-DD'}), 400
        start_at, end_at = get_business_window(selected_date)
        q = q.filter(
            UploadedFile.uploaded_at >= start_at,
            UploadedFile.uploaded_at < end_at,
        )

    # 鏃ユ湡閫夐」锛堜笟鍔℃棩锛?
    uploaded_rows = db.session.query(UploadedFile.uploaded_at).filter(
        UploadedFile.uploaded_at.isnot(None)
    ).all()
    date_options = sorted(
        {str(get_business_date(row[0])) for row in uploaded_rows if row[0]},
        reverse=True,
    )

    all_files = q.all()
    if status_filter:
        all_files = [uploaded_file for uploaded_file in all_files if uploaded_file.derived_status() == status_filter]

    total = len(all_files)
    pages = (total + per_page - 1) // per_page if total else 0
    if pages > 0:
        page = min(page, pages)
    else:
        page = 1
    start = (page - 1) * per_page
    items = all_files[start:start + per_page]

    return jsonify({
        'files': [f.to_dict() for f in items],
        'total': total,
        'pages': pages,
        'page': page,
        'date_options': date_options,
        'current_business_date': str(get_business_date()),
    })


@admin_bp.route('/api/files/<int:file_id>/detail')
@login_required_json
@login_required
@admin_required
def file_detail(file_id):
    uploaded_file = db.session.get(UploadedFile, file_id)
    if not uploaded_file:
        return jsonify({'success': False, 'error': '\u6587\u4ef6\u4e0d\u5b58\u5728'}), 404
    page = _parse_int_arg(request.args.get('page', 1), minimum=1)
    per_page = _parse_int_arg(request.args.get('per_page', 50), minimum=1)
    if page is None or per_page is None:
        return jsonify({'success': False, 'error': '\u5206\u9875\u53c2\u6570\u5fc5\u987b\u662f\u5927\u4e8e 0 \u7684\u6574\u6570'}), 400

    tickets = LotteryTicket.query.filter_by(source_file_id=file_id)\
        .order_by(LotteryTicket.line_number)\
        .paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        'file': uploaded_file.to_dict(),
        'tickets': [t.to_dict() for t in tickets.items],
        'total': tickets.total,
        'pages': tickets.pages,
    })


@admin_bp.route('/api/files/<int:file_id>/revoke', methods=['POST'])
@login_required_json
@login_required
@admin_required
def api_revoke_file(file_id):
    result = revoke_file(file_id, current_user.id)
    if result['success']:
        notify_pool_update = get_pool_status
        try:
            from services.notify_service import notify_pool_update as _npu
            _npu(get_pool_status())
        except Exception as notify_exc:
            current_app.logger.warning('notify_pool_update failed after revoke: %s', notify_exc)
    message = result.get('message') or ''
    if '\u4e0d\u5b58\u5728' in message or 'not found' in message:
        result['message'] = '\u6587\u4ef6\u4e0d\u5b58\u5728'
        return jsonify(result), 404
    if not result['success']:
        return jsonify(result), 400
    return jsonify(result)


@admin_bp.route('/api/tickets/export')
@login_required_json
@login_required
@admin_required
def export_tickets():
    """Export completed and expired tickets as CSV."""
    import csv
    import io

    from models.ticket import LotteryTicket
    from models.file import UploadedFile as UF
    cutoff_start = get_today_noon()
    cutoff_end = cutoff_start + timedelta(days=1)

    all_tickets = LotteryTicket.query.filter(
        LotteryTicket.status.in_(['completed', 'expired']),
    ).order_by(LotteryTicket.id).all()
    tickets_q = [
        ticket for ticket in all_tickets
        if (terminal_at := _winning_terminal_at(ticket)) and cutoff_start <= terminal_at < cutoff_end
    ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'ticket_id', 'line', 'raw_content', 'lottery_type', 'multiplier', 'deadline',
        'period', 'amount', 'status', 'username', '\u8bbe\u5907ID', 'assigned_at', 'completed_at', 'source_file'
    ])
    for t in tickets_q:
        f = db.session.get(UF, t.source_file_id)
        writer.writerow([
            t.id, t.line_number, t.raw_content, t.lottery_type, t.multiplier,
            t.deadline_time, t.detail_period, t.ticket_amount, t.status,
            t.assigned_username, t.assigned_device_id,
            t.assigned_at, t.completed_at, f.original_filename if f else '',
        ])

    output.seek(0)
    from flask import Response
    filename = f"tickets_{get_business_date()}.csv"
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# 鈹€鈹€ User management 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€

@admin_bp.route('/api/tickets/export-by-date')
@login_required_json
@login_required
@admin_required
def export_tickets_by_date():
    """鎸変笂浼犳棩鏈熷鍑鸿鏃ユ墍鏈夌エ鏁版嵁 XLSX"""
    import io as _io
    from openpyxl import Workbook
    from urllib.parse import quote

    date_str = request.args.get('date', '').strip()

    q = db.session.query(
        LotteryTicket.line_number.label('line_number'),
        LotteryTicket.raw_content.label('raw_content'),
        LotteryTicket.lottery_type.label('lottery_type'),
        LotteryTicket.multiplier.label('multiplier'),
        LotteryTicket.deadline_time.label('deadline_time'),
        LotteryTicket.detail_period.label('detail_period'),
        LotteryTicket.ticket_amount.label('ticket_amount'),
        LotteryTicket.status.label('status'),
        LotteryTicket.assigned_username.label('assigned_username'),
        LotteryTicket.assigned_device_id.label('assigned_device_id'),
        LotteryTicket.assigned_at.label('assigned_at'),
        LotteryTicket.completed_at.label('completed_at'),
        LotteryTicket.source_file_id.label('source_file_id'),
        UploadedFile.original_filename.label('original_filename'),
    ).outerjoin(UploadedFile, UploadedFile.id == LotteryTicket.source_file_id)
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': '鏃ユ湡鏍煎紡鏃犳晥锛岃浣跨敤 YYYY-MM-DD'}), 400
        start_at, end_at = get_business_window(selected_date)
        file_exists = db.session.query(UploadedFile.id).filter(
            UploadedFile.uploaded_at >= start_at,
            UploadedFile.uploaded_at < end_at,
        ).first()
        if not file_exists:
            wb = Workbook()
            ws = wb.active
            ws.append(['行号', '原始内容', '彩种', '倍投', '截止时间', '期号', '金额', '状态', '用户名', '设备ID', '分配时间', '完成时间', '来源文件名'])
            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            from flask import Response
            empty_filename = f"{date_str}_无数据投注内容详情.xlsx"
            empty_filename_encoded = quote(empty_filename, encoding='utf-8')
            return Response(buf.read(),
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={'Content-Disposition': f"attachment; filename*=UTF-8''{empty_filename_encoded}"})
        q = q.filter(
            UploadedFile.uploaded_at >= start_at,
            UploadedFile.uploaded_at < end_at,
        )

    tickets = q.order_by(LotteryTicket.source_file_id, LotteryTicket.line_number).all()

    wb = Workbook()
    ws = wb.active
    ws.append(['行号', '原始内容', '彩种', '倍投', '截止时间', '期号', '金额', '状态', '用户名', '设备ID', '分配时间', '完成时间', '来源文件名'])
    status_map = {
        'pending': '\u5f85\u5904\u7406',
        'assigned': '\u5904\u7406\u4e2d',
        'completed': '\u5df2\u5b8c\u6210',
        'revoked': '\u5df2\u64a4\u9500',
        'expired': '\u5df2\u8fc7\u671f',
    }
    for t in tickets:
        ws.append([
            t.line_number,
            t.raw_content or '',
            t.lottery_type or '',
            f"{t.multiplier}x" if t.multiplier else '',
            t.deadline_time.strftime('%Y-%m-%d %H:%M') if t.deadline_time else '',
            t.detail_period or '',
            float(t.ticket_amount or 0),
            status_map.get(t.status, t.status),
            t.assigned_username or '',
            t.assigned_device_id or '',
            t.assigned_at.strftime('%Y-%m-%d %H:%M:%S') if t.assigned_at else '',
            t.completed_at.strftime('%Y-%m-%d %H:%M:%S') if t.completed_at else '',
            t.original_filename or '',
        ])

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import Response
    period_str = next((t.detail_period for t in tickets if t.detail_period), 'unknown_period')
    export_date = date_str or str(get_business_date())
    filename = f"{export_date}_{period_str}_tickets.xlsx"
    filename_encoded = quote(filename, encoding='utf-8')
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{filename_encoded}"},
    )


# 鈹€鈹€ User management 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€

@admin_bp.route('/users')
@login_required
@admin_required
def users_page():
    return render_template('admin/users.html')


@admin_bp.route('/api/users')
@login_required_json
@login_required
@admin_required
def api_users_list():
    mode_sort_key = case(
        (User.client_mode == 'mode_a', 0),
        (User.client_mode == 'mode_b', 1),
        else_=2,
    )
    users = User.query.filter_by(is_admin=False).order_by(mode_sort_key, User.created_at, User.id).all()
    return jsonify({'users': [u.to_dict() for u in users]})


@admin_bp.route('/api/lottery-types')
@login_required_json
@login_required
@admin_required
def api_lottery_types():
    """Return supported lottery types."""
    return jsonify({'lottery_types': list(RESULT_UPLOAD_LOTTERY_TYPES)})


@admin_bp.route('/api/users', methods=['POST'])
@login_required_json
@login_required
@admin_required
def api_create_user():
    data, data_error = parse_json_object()
    if data_error:
        return data_error
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    client_mode = _parse_client_mode(data.get('client_mode', 'mode_a'))
    if client_mode is None:
        return jsonify({'success': False, 'error': '\u5ba2\u6237\u7aef\u6a21\u5f0f\u5fc5\u987b\u662f mode_a \u6216 mode_b'}), 400
    max_devices = _parse_int_arg(data.get('max_devices', 1), minimum=1)
    if max_devices is None:
        return jsonify({'success': False, 'error': '\u6700\u5927\u8bbe\u5907\u6570\u5fc5\u987b\u662f\u5927\u4e8e 0 \u7684\u6574\u6570'}), 400
    desktop_only_b_mode = True
    if 'desktop_only_b_mode' in data:
        desktop_only_b_mode = _parse_bool_flag(data.get('desktop_only_b_mode'))
        if desktop_only_b_mode is None:
            return jsonify({'success': False, 'error': 'desktop_only_b_mode \u5fc5\u987b\u662f\u5e03\u5c14\u503c'}), 400

    # 楠岃瘉 max_processing_b_mode
    max_processing_b_mode = data.get('max_processing_b_mode')
    if max_processing_b_mode is not None:
        try:
            max_processing_b_mode = int(max_processing_b_mode) if max_processing_b_mode else None
            if max_processing_b_mode is not None and (max_processing_b_mode < 1 or max_processing_b_mode > 10000):
                return jsonify({'success': False, 'error': 'B妯″紡涓婇檺蹇呴』鍦?-10000涔嬮棿'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'B妯″紡涓婇檺蹇呴』鏄暣鏁?'}), 400

    # 楠岃瘉 daily_ticket_limit
    daily_ticket_limit = data.get('daily_ticket_limit')
    if daily_ticket_limit is not None:
        try:
            daily_ticket_limit = int(daily_ticket_limit) if daily_ticket_limit else None
            if daily_ticket_limit is not None and (daily_ticket_limit < 1 or daily_ticket_limit > 100000):
                return jsonify({'success': False, 'error': '姣忔棩涓婇檺蹇呴』鍦?-100000涔嬮棿'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': '姣忔棩涓婇檺蹇呴』鏄暣鏁?'}), 400

    # 楠岃瘉 blocked_lottery_types
    blocked_lottery_types = data.get('blocked_lottery_types')
    if blocked_lottery_types is not None:
        if not isinstance(blocked_lottery_types, list):
            return jsonify({'success': False, 'error': '绂佹褰╃蹇呴』鏄暟缁?'}), 400
        if not all(isinstance(t, str) for t in blocked_lottery_types):
            return jsonify({'success': False, 'error': '绂佹褰╃鍒楄〃涓殑姣忛」蹇呴』鏄瓧绗︿覆'}), 400

    if not username or not password:
        return jsonify({'success': False, 'error': '鐢ㄦ埛鍚嶅拰瀵嗙爜涓嶈兘涓虹┖'}), 400
    if len(password) < 6:
        return jsonify({'success': False, 'error': '\u5bc6\u7801\u81f3\u5c11\u9700\u8981 6 \u4f4d'}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({'success': False, 'error': '鐢ㄦ埛鍚嶅凡瀛樺湪'}), 409

    user = User(username=username, client_mode=client_mode, max_devices=max_devices,
                max_processing_b_mode=max_processing_b_mode, daily_ticket_limit=daily_ticket_limit,
                desktop_only_b_mode=desktop_only_b_mode)
    user.set_password(password)
    if blocked_lottery_types is not None:
        user.set_blocked_lottery_types(blocked_lottery_types)
    db.session.add(user)
    db.session.commit()
    return jsonify({'success': True, 'user': user.to_dict()})


@admin_bp.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required_json
@login_required
@admin_required
def api_update_user(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'error': '\u7528\u6237\u4e0d\u5b58\u5728'}), 404
    if user.is_admin:
        return jsonify({'success': False, 'error': '\u4e0d\u5141\u8bb8\u5728\u6b64\u63a5\u53e3\u4fee\u6539\u7ba1\u7406\u5458\u8d26\u53f7'}), 403
    data, data_error = parse_json_object()
    if data_error:
        return data_error
    was_active = user.is_active
    should_refresh_pool_views = False

    if 'client_mode' in data:
        parsed_client_mode = _parse_client_mode(data['client_mode'])
        if parsed_client_mode is None:
            return jsonify({'success': False, 'error': '\u5ba2\u6237\u7aef\u6a21\u5f0f\u5fc5\u987b\u662f mode_a \u6216 mode_b'}), 400
        user.client_mode = parsed_client_mode
        should_refresh_pool_views = True
    if 'max_devices' in data:
        parsed_max_devices = _parse_int_arg(data['max_devices'], minimum=1)
        if parsed_max_devices is None:
            return jsonify({'success': False, 'error': '\u6700\u5927\u8bbe\u5907\u6570\u5fc5\u987b\u662f\u5927\u4e8e 0 \u7684\u6574\u6570'}), 400
        user.max_devices = parsed_max_devices
    if 'max_processing_b_mode' in data:
        val = data['max_processing_b_mode']
        try:
            user.max_processing_b_mode = int(val) if val else None
            if user.max_processing_b_mode is not None and (user.max_processing_b_mode < 1 or user.max_processing_b_mode > 10000):
                return jsonify({'success': False, 'error': 'B妯″紡涓婇檺蹇呴』鍦?-10000涔嬮棿'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'B妯″紡涓婇檺蹇呴』鏄暣鏁?'}), 400
    if 'daily_ticket_limit' in data:
        val = data['daily_ticket_limit']
        try:
            user.daily_ticket_limit = int(val) if val else None
            if user.daily_ticket_limit is not None and (user.daily_ticket_limit < 1 or user.daily_ticket_limit > 100000):
                return jsonify({'success': False, 'error': '姣忔棩涓婇檺蹇呴』鍦?-100000涔嬮棿'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': '姣忔棩涓婇檺蹇呴』鏄暣鏁?'}), 400
    if 'is_active' in data:
        parsed_is_active = _parse_bool_flag(data['is_active'])
        if parsed_is_active is None:
            return jsonify({'success': False, 'error': 'is_active 蹇呴』鏄竷灏斿€?'}), 400
        user.is_active = parsed_is_active
    if 'can_receive' in data:
        parsed_can_receive = _parse_bool_flag(data['can_receive'])
        if parsed_can_receive is None:
            return jsonify({'success': False, 'error': 'can_receive 蹇呴』鏄竷灏斿€?'}), 400
        user.can_receive = parsed_can_receive
        should_refresh_pool_views = True
    if 'desktop_only_b_mode' in data:
        parsed_desktop_only = _parse_bool_flag(data['desktop_only_b_mode'])
        if parsed_desktop_only is None:
            return jsonify({'success': False, 'error': 'desktop_only_b_mode \u5fc5\u987b\u662f\u5e03\u5c14\u503c'}), 400
        user.desktop_only_b_mode = parsed_desktop_only
    if 'password' in data and data['password']:
        if len(data['password']) < 6:
            return jsonify({'success': False, 'error': '\u5bc6\u7801\u81f3\u5c11\u9700\u8981 6 \u4f4d'}), 400
        user.set_password(data['password'])
    if 'blocked_lottery_types' in data:
        blocked_types = data['blocked_lottery_types']
        if blocked_types is not None and not isinstance(blocked_types, list):
            return jsonify({'success': False, 'error': '绂佹褰╃蹇呴』鏄暟缁?'}), 400
        if isinstance(blocked_types, list) and not all(isinstance(t, str) for t in blocked_types):
            return jsonify({'success': False, 'error': '绂佹褰╃鍒楄〃涓殑姣忛」蹇呴』鏄瓧绗︿覆'}), 400
        user.set_blocked_lottery_types(blocked_types)
        should_refresh_pool_views = True

    db.session.commit()

    if was_active and not user.is_active:
        force_logout_user(user.id, '\u8d26\u53f7\u5df2\u88ab\u7ba1\u7406\u5458\u7981\u7528')
        AuditLog.log('force_logout', user_id=current_user.id,
                     resource_type='user', resource_id=user_id,
                     details={'reason': '\u8d26\u53f7\u5df2\u88ab\u7ba1\u7406\u5458\u7981\u7528'})
        db.session.commit()
    elif should_refresh_pool_views:
        try:
            from services.notify_service import notify_pool_update

            notify_pool_update(get_pool_status())
        except Exception:
            current_app.logger.warning('notify_pool_update failed after user config update', exc_info=True)

    return jsonify({'success': True, 'user': user.to_dict()})


@admin_bp.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required_json
@login_required
@admin_required
def api_delete_user(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'error': '\u7528\u6237\u4e0d\u5b58\u5728'}), 404
    if user.is_admin:
        return jsonify({'success': False, 'error': '\u4e0d\u5141\u8bb8\u5728\u6b64\u63a5\u53e3\u5220\u9664\u7ba1\u7406\u5458\u8d26\u53f7'}), 403

    # 鑾峰彇寮哄埗鍒犻櫎鍙傛暟
    force = request.args.get('force', 'false').lower() == 'true'

    # 妫€鏌ュ叧鑱旀暟鎹?
    has_ticket_refs = LotteryTicket.query.filter(LotteryTicket.assigned_user_id == user_id).first() is not None
    has_uploaded_file_refs = UploadedFile.query.filter(
        (UploadedFile.uploaded_by == user_id) | (UploadedFile.revoked_by == user_id)
    ).first() is not None
    has_winning_refs = WinningRecord.query.filter(
        (WinningRecord.uploaded_by == user_id) |
        (WinningRecord.verified_by == user_id) |
        (WinningRecord.checked_by == user_id)
    ).first() is not None
    has_result_refs = ResultFile.query.filter(ResultFile.uploaded_by == user_id).first() is not None or MatchResult.query.filter(
        MatchResult.uploaded_by == user_id
    ).first() is not None
    has_audit_refs = AuditLog.query.filter(AuditLog.user_id == user_id).first() is not None

    has_refs = has_ticket_refs or has_uploaded_file_refs or has_winning_refs or has_result_refs or has_audit_refs

    if has_refs and not force:
        return jsonify({
            'success': False,
            'error': '\u8be5\u7528\u6237\u5df2\u6709\u5386\u53f2\u4e1a\u52a1\u6570\u636e\uff0c\u4e0d\u80fd\u76f4\u63a5\u5220\u9664\uff0c\u8bf7\u6539\u4e3a\u7981\u7528\u8d26\u53f7\u3002\u5982\u786e\u9700\u5f3a\u5236\u5220\u9664\uff0c\u8bf7\u643a\u5e26 force=true \u91cd\u8bd5',
            'has_refs': True
        }), 409

    if has_refs and force:
        # 寮哄埗鍒犻櫎锛氬皢鎵€鏈夊叧鑱斿紩鐢ㄧ疆绌?
        # 1. 绁ㄦ嵁
        LotteryTicket.query.filter(LotteryTicket.assigned_user_id == user_id).update({LotteryTicket.assigned_user_id: None})
        # 2. 涓婁紶鏂囦欢
        UploadedFile.query.filter(UploadedFile.uploaded_by == user_id).update({UploadedFile.uploaded_by: None})
        UploadedFile.query.filter(UploadedFile.revoked_by == user_id).update({UploadedFile.revoked_by: None})
        # 3. 涓璁板綍
        WinningRecord.query.filter(WinningRecord.uploaded_by == user_id).update({WinningRecord.uploaded_by: None})
        WinningRecord.query.filter(WinningRecord.verified_by == user_id).update({WinningRecord.verified_by: None})
        WinningRecord.query.filter(WinningRecord.checked_by == user_id).update({WinningRecord.checked_by: None})
        # 4. 璧涙灉
        ResultFile.query.filter(ResultFile.uploaded_by == user_id).update({ResultFile.uploaded_by: None})
        MatchResult.query.filter(MatchResult.uploaded_by == user_id).update({MatchResult.uploaded_by: None})
        # 5. 瀹¤鏃ュ織涓庤缃?
        AuditLog.query.filter(AuditLog.user_id == user_id).update({AuditLog.user_id: None})
        SystemSettings.query.filter(SystemSettings.updated_by == user_id).update({SystemSettings.updated_by: None})

    db.session.delete(user)
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/users/<int:user_id>/force-logout', methods=['POST'])
@login_required_json
@login_required
@admin_required
def api_force_logout(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'error': '\u7528\u6237\u4e0d\u5b58\u5728'}), 404
    if user.is_admin:
        return jsonify({'success': False, 'error': '\u4e0d\u5141\u8bb8\u5728\u6b64\u63a5\u53e3\u5f3a\u5236\u767b\u51fa\u7ba1\u7406\u5458\u8d26\u53f7'}), 403

    count = force_logout_user(user_id, 'Forced logout by admin')
    AuditLog.log('force_logout', user_id=current_user.id, resource_type='user', resource_id=user_id)
    db.session.commit()
    return jsonify({'success': True, 'sessions_cleared': count})


@admin_bp.route('/api/users/<int:user_id>/can-receive', methods=['PUT'])
@login_required_json
@login_required
@admin_required
def api_toggle_can_receive(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'error': '\u7528\u6237\u4e0d\u5b58\u5728'}), 404
    if user.is_admin:
        return jsonify({'success': False, 'error': '\u4e0d\u5141\u8bb8\u5728\u6b64\u63a5\u53e3\u4fee\u6539\u7ba1\u7406\u5458\u8d26\u53f7'}), 403

    data, data_error = parse_json_object()
    if data_error:
        return data_error

    parsed_can_receive = _parse_bool_flag(data.get('can_receive', True))
    if parsed_can_receive is None:
        return jsonify({'success': False, 'error': 'can_receive \u5fc5\u987b\u662f\u5e03\u5c14\u503c'}), 400

    user.can_receive = parsed_can_receive
    db.session.commit()

    try:
        from services.notify_service import notify_pool_update
        notify_pool_update(get_pool_status())
    except Exception:
        current_app.logger.warning('Failed to notify pool update after can_receive change', exc_info=True)

    return jsonify({'success': True, 'can_receive': user.can_receive})


@admin_bp.route('/api/users/export')
@login_required_json
@login_required
@admin_required
def api_export_users():
    """Export all non-admin users as an XLSX file."""
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from urllib.parse import quote

    users = User.query.filter_by(is_admin=False).order_by(User.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '\u7528\u6237\u5217\u8868'

    headers = [
        '\u7528\u6237\u540d',
        '\u5bc6\u7801\u54c8\u5e0c',
        '\u63a5\u5355\u6a21\u5f0f',
        '\u6700\u5927\u8bbe\u5907\u6570',
        'B\u6a21\u5f0f\u5904\u7406\u4e0a\u9650',
        '\u6bcf\u65e5\u5904\u7406\u4e0a\u9650',
        '\u7981\u6b62\u5f69\u79cd',
        '\u8d26\u53f7\u72b6\u6001',
        '\u63a5\u5355\u5f00\u5173',
        'B\u6a21\u5f0f\u4ec5\u684c\u9762\u7aef',
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for user in users:
        blocked_types = ','.join(user.get_blocked_lottery_types()) if user.get_blocked_lottery_types() else ''
        ws.append([
            user.username,
            user.password_hash,
            user.client_mode,
            user.max_devices,
            user.max_processing_b_mode if user.max_processing_b_mode is not None else '',
            user.daily_ticket_limit if user.daily_ticket_limit is not None else '',
            blocked_types,
            '\u662f' if user.is_active else '\u5426',
            '\u5f00' if user.can_receive else '\u5173',
            '\u662f' if user.desktop_only_b_mode else '\u5426',
        ])

    column_widths = [15, 60, 12, 12, 22, 18, 30, 12, 12, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'users_{timestamp}.xlsx'

    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@admin_bp.route('/api/users/import', methods=['POST'])
@login_required_json
@login_required
@admin_required
def api_import_users():
    """Import users from an XLSX file."""
    import tempfile
    from services.user_import_service import import_users

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Missing file'}), 400

    file = request.files['file']
    safe_filename = _safe_uploaded_filename(file.filename)
    if not safe_filename:
        return jsonify({'success': False, 'error': '\u6587\u4ef6\u540d\u4e0d\u80fd\u4e3a\u7a7a'}), 400
    if not file or file.filename == '':
        return jsonify({'success': False, 'error': 'Empty filename'}), 400

    if not safe_filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Only .xlsx files are supported'}), 400

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            tmp_path = tmp_file.name

        result = import_users(tmp_path, current_user.id)
        status_code = 200 if result.get('success') else 400
        return jsonify(result), status_code
    except Exception as e:
        current_app.logger.error(f'Failed to import users: {str(e)}', exc_info=True)
        return jsonify({'success': False, 'error': f'Import failed: {str(e)}'}), 500
    finally:
        if tmp_path:
            try:
                os.remove(tmp_path)
            except Exception as cleanup_exc:
                current_app.logger.warning(f'Failed to remove temp import file: {tmp_path}, err={cleanup_exc}')


@admin_bp.route('/winning')
@login_required
@admin_required
def winning_page():
    return render_template('admin/winning.html')


@admin_bp.route('/api/winning/filter-options')
@login_required_json
@login_required
@admin_required
def api_winning_filter_options():
    winning_tickets = LotteryTicket.query.filter(
        LotteryTicket.is_winning == True,
        LotteryTicket.status.in_(['completed', 'expired']),
    ).all()
    dates = sorted(
        {str(get_business_date(_winning_terminal_at(ticket))) for ticket in winning_tickets if _winning_terminal_at(ticket)},
        reverse=True,
    )
    return jsonify({
        'usernames': sorted({ticket.assigned_username for ticket in winning_tickets if ticket.assigned_username}),
        'dates': dates,
        'lottery_types': sorted({ticket.lottery_type for ticket in winning_tickets if ticket.lottery_type}),
        'current_business_date': str(get_business_date()),
    })


@admin_bp.route('/api/winning')
@login_required_json
@login_required
@admin_required
def api_winning_list():
    """List winning tickets with optional filters."""
    page = _parse_int_arg(request.args.get('page', 1), minimum=1)
    per_page = _parse_int_arg(request.args.get('per_page', 50), minimum=1)
    if page is None or per_page is None:
        return jsonify({'success': False, 'error': '鍒嗛〉鍙傛暟蹇呴』鏄ぇ浜?0 鐨勬暣鏁?'}), 400
    username = request.args.get('username', '').strip()
    date_str = request.args.get('date', '').strip()
    lottery_type = request.args.get('lottery_type', '').strip()
    image_filter = request.args.get('image_filter', '').strip()  # 'uploaded' | 'missing'
    checked_status = request.args.get('checked_status', '').strip()  # 'all' | 'checked' | 'unchecked'

    from sqlalchemy import func

    terminal_expr = func.coalesce(
        LotteryTicket.completed_at,
        LotteryTicket.deadline_time,
        LotteryTicket.assigned_at,
        LotteryTicket.admin_upload_time,
    )
    q = LotteryTicket.query.filter(
        LotteryTicket.is_winning == True,
        LotteryTicket.status.in_(['completed', 'expired']),
    )
    if username:
        q = q.filter(LotteryTicket.assigned_username == username)
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': '鏃ユ湡鏍煎紡鏃犳晥锛岃浣跨敤 YYYY-MM-DD'}), 400
        start_at, end_at = get_business_window(selected_date)
        q = q.filter(
            terminal_expr >= start_at,
            terminal_expr < end_at,
        )
    if lottery_type:
        q = q.filter(LotteryTicket.lottery_type == lottery_type)
    if image_filter == 'uploaded':
        q = q.filter(LotteryTicket.winning_image_url.isnot(None),
                     LotteryTicket.winning_image_url != '')
    elif image_filter == 'missing':
        q = q.filter(
            (LotteryTicket.winning_image_url == None) |
            (LotteryTicket.winning_image_url == '')
        )

    # 瀹℃牳鐘舵€佺瓫閫夛紙浣跨敤 EXISTS 瀛愭煡璇紝閬垮厤涓㈠け鏁版嵁锛?
    if checked_status == 'checked':
        q = q.filter(
            db.exists().where(
                db.and_(
                    WinningRecord.ticket_id == LotteryTicket.id,
                    WinningRecord.is_checked == True
                )
            )
        )
    elif checked_status == 'unchecked':
        q = q.filter(
            ~db.exists().where(
                db.and_(
                    WinningRecord.ticket_id == LotteryTicket.id,
                    WinningRecord.is_checked == True
                )
            )
        )

    q = q.order_by(terminal_expr.desc(), LotteryTicket.id.desc())

    # 姹囨€伙紙鍏ㄩ噺锛屼笉鍒嗛〉锛?
    all_items = q.all()
    summary_amount = sum(float(t.winning_amount or 0) for t in all_items)
    summary_predicted_amount = sum(float(t.predicted_winning_amount or 0) for t in all_items)
    summary_gross  = sum(float(t.winning_gross  or 0) for t in all_items)
    summary_tax    = sum(float(t.winning_tax    or 0) for t in all_items)
    summary_missing = sum(1 for t in all_items if not t.winning_image_url)
    total = len(all_items)

    # 鍒嗛〉鍒囩墖
    start = (page - 1) * per_page
    page_items = all_items[start:start + per_page]
    import math
    pages = math.ceil(total / per_page) if total else 1

    # 鎵归噺鏌ヨ WinningRecord锛岄伩鍏?N+1 鏌ヨ闂
    ticket_ids = [t.id for t in page_items]
    winning_records_map = {
        wr.ticket_id: wr
        for wr in WinningRecord.query.filter(WinningRecord.ticket_id.in_(ticket_ids)).all()
    }

    records = []
    for t in page_items:
        winning_record = winning_records_map.get(t.id)
        terminal_at = _winning_terminal_at(t)
        predicted_amount = _decimal_to_float(t.predicted_winning_amount)
        final_amount = _decimal_to_float(t.winning_amount)
        records.append({
            'ticket_id': t.id,
            'username': t.assigned_username or '-',
            'device_id': t.assigned_device_id or '-',
            'lottery_type': t.lottery_type,
            'detail_period': t.detail_period,
            'predicted_winning_gross': _decimal_to_float(t.predicted_winning_gross),
            'predicted_winning_amount': predicted_amount,
            'predicted_winning_tax': _decimal_to_float(t.predicted_winning_tax),
            'winning_gross': _decimal_to_float(t.winning_gross),
            'winning_amount': final_amount,
            'winning_tax': _decimal_to_float(t.winning_tax),
            'winning_change_percent': _winning_change_percent(predicted_amount, final_amount),
            'winning_image_url': t.winning_image_url or '',
            'raw_content': t.raw_content or '',
            'ticket_amount': float(t.ticket_amount) if t.ticket_amount else 0,
            'status': t.status,
            'status_label': _winning_status_label(t.status),
            'completed_at': (t.completed_at.isoformat() if t.completed_at else None),
            'terminal_at': (terminal_at.isoformat() if terminal_at else None),
            'terminal_label': '杩囨湡鏃堕棿' if t.status == 'expired' else '瀹屾垚鏃堕棿',
            'is_checked': winning_record.is_checked if winning_record else False,
            'checked_at': winning_record.checked_at.isoformat() if (winning_record and winning_record.checked_at) else None,
            'checked_by_username': winning_record.checker.username if (winning_record and winning_record.checker) else None,
            'winning_record_id': winning_record.id if winning_record else None,
        })
    return jsonify({
        'records': records,
        'total': total,
        'pages': pages,
        'summary': {
            'amount': round(summary_amount, 2),
            'predicted_amount': round(summary_predicted_amount, 2),
            'gross':  round(summary_gross,  2),
            'tax':    round(summary_tax,    2),
            'count':  total,
            'missing': summary_missing,
        },
    })


@admin_bp.route('/api/winning/export')
@login_required_json
@login_required
@admin_required
def api_winning_export():
    """Export winning tickets to XLSX with current filters."""
    import io as _io
    from openpyxl import Workbook
    from urllib.parse import quote
    from sqlalchemy import func

    username = request.args.get('username', '').strip()
    date_str = request.args.get('date', '').strip()
    lottery_type = request.args.get('lottery_type', '').strip()
    image_filter = request.args.get('image_filter', '').strip()
    checked_status = request.args.get('checked_status', '').strip()

    terminal_expr = func.coalesce(
        LotteryTicket.completed_at,
        LotteryTicket.deadline_time,
        LotteryTicket.assigned_at,
        LotteryTicket.admin_upload_time,
    )

    q = LotteryTicket.query.filter(
        LotteryTicket.is_winning == True,
        LotteryTicket.status.in_(['completed', 'expired']),
    )

    if username:
        q = q.filter(LotteryTicket.assigned_username == username)

    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': 'invalid date format, expected YYYY-MM-DD'}), 400
        start_at, end_at = get_business_window(selected_date)
        q = q.filter(terminal_expr >= start_at, terminal_expr < end_at)

    if lottery_type:
        q = q.filter(LotteryTicket.lottery_type == lottery_type)

    if image_filter == 'uploaded':
        q = q.filter(LotteryTicket.winning_image_url.isnot(None), LotteryTicket.winning_image_url != '')
    elif image_filter == 'missing':
        q = q.filter((LotteryTicket.winning_image_url == None) | (LotteryTicket.winning_image_url == ''))

    if checked_status == 'checked':
        q = q.filter(
            db.exists().where(
                db.and_(
                    WinningRecord.ticket_id == LotteryTicket.id,
                    WinningRecord.is_checked == True,
                )
            )
        )
    elif checked_status == 'unchecked':
        q = q.filter(
            ~db.exists().where(
                db.and_(
                    WinningRecord.ticket_id == LotteryTicket.id,
                    WinningRecord.is_checked == True,
                )
            )
        )

    items = q.order_by(terminal_expr.desc(), LotteryTicket.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.append([
        '\u7968ID',
        '\u6295\u6ce8\u5185\u5bb9',
        '\u7968\u9762\u91d1\u989d',
        '\u7528\u6237\u540d',
        '\u8bbe\u5907ID',
        '\u5f69\u79cd',
        '\u671f\u53f7',
        '\u72b6\u6001',
        '\u7a0e\u524d\u91d1\u989d',
        '\u7a0e\u540e\u91d1\u989d',
        '\u7a0e\u989d',
        '\u56fe\u7247\u72b6\u6001',
        '\u7ec8\u6001\u65f6\u95f4',
    ])

    for t in items:
        terminal_at = _winning_terminal_at(t)
        ws.append([
            t.id,
            t.raw_content or '',
            float(t.ticket_amount or 0),
            t.assigned_username or '',
            t.assigned_device_id or '',
            t.lottery_type or '',
            t.detail_period or '',
            _winning_status_label(t.status),
            float(t.winning_gross or 0),
            float(t.winning_amount or 0),
            float(t.winning_tax or 0),
            '\u5df2\u4e0a\u4f20' if t.winning_image_url else '\u672a\u4e0a\u4f20',
            terminal_at.strftime('%Y-%m-%d %H:%M:%S') if terminal_at else '',
        ])

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parts = ['winning_export']
    if username:
        parts.append(username)
    if date_str:
        parts.append(date_str)
    if lottery_type:
        parts.append(lottery_type)
    if image_filter == 'uploaded':
        parts.append('with_image')
    elif image_filter == 'missing':
        parts.append('missing_image')
    parts.append(beijing_now().strftime('%Y%m%d_%H%M%S'))
    filename = '_'.join(parts) + '.xlsx'
    filename_encoded = quote(filename, encoding='utf-8')

    from flask import Response
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{filename_encoded}"},
    )


@admin_bp.route('/api/winning/<int:ticket_id>/presign', methods=['POST'])
@login_required_json
@login_required
@admin_required
def admin_winning_presign(ticket_id):
    ticket, error_response = _get_winning_ticket_or_error(ticket_id)
    if error_response:
        return error_response
    record = WinningRecord.query.filter_by(ticket_id=ticket.id).first()
    if record and record.is_checked:
        return jsonify({'success': False, 'error': '\u8be5\u4e2d\u5956\u8bb0\u5f55\u5df2\u68c0\u67e5\uff0c\u65e0\u6cd5\u66f4\u6362\u56fe\u7247'}), 403
    from services.oss_service import generate_presign_url, build_oss_key
    oss_key = build_oss_key(ticket.id)
    url, key = generate_presign_url(oss_key)
    if url.startswith('/api/winning/upload-local?'):
        url = f"{url}&ticket_id={ticket.id}"
    return jsonify({'success': True, 'url': url, 'oss_key': key})


@admin_bp.route('/api/winning/record', methods=['POST'])
@login_required_json
@login_required
@admin_required
def admin_winning_record():
    """绠＄悊鍛樻洿鏂颁腑濂栧浘鐗嘦RL"""
    data, data_error = parse_json_object()
    if data_error:
        return data_error
    ticket_id = data.get('ticket_id')
    oss_key = data.get('oss_key', '')
    if not oss_key:
        return jsonify({'success': False, 'error': '缂哄皯 oss_key'}), 400
    if not ticket_id:
        return jsonify({'success': False, 'error': '缂哄皯ticket_id'}), 400
    ticket, error_response = _get_winning_ticket_or_error(ticket_id)
    if error_response:
        return error_response
    if not _winning_key_matches_ticket(ticket.id, oss_key):
        return jsonify({'success': False, 'error': 'oss_key 涓庣エ鎹笉鍖归厤'}), 400

    from services.oss_service import delete_stored_image, get_public_url
    image_url = get_public_url(oss_key) if oss_key else ''
    record = WinningRecord.query.filter_by(ticket_id=ticket.id).first()
    if record and record.is_checked:
        return jsonify({'success': False, 'error': '\u8be5\u4e2d\u5956\u8bb0\u5f55\u5df2\u68c0\u67e5\uff0c\u65e0\u6cd5\u66f4\u6362\u56fe\u7247'}), 403
    if record:
        if record.image_oss_key != (oss_key or None) or record.winning_image_url != image_url:
            delete_stored_image(record.image_oss_key, record.winning_image_url)
        record.winning_image_url = image_url
        record.image_oss_key = oss_key or None
        record.uploaded_by = current_user.id
        record.uploaded_at = beijing_now()
    else:
        record = WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=image_url,
            image_oss_key=oss_key or None,
            uploaded_by=current_user.id,
        )
        db.session.add(record)
    ticket.winning_image_url = image_url
    ticket.is_winning = True
    db.session.commit()
    return jsonify({'success': True, 'image_url': image_url, 'record': record.to_dict()})


@admin_bp.route('/api/winning/<int:ticket_id>/upload-image', methods=['POST'])
@login_required_json
@login_required
@admin_required
def admin_winning_upload_image(ticket_id):
    """Upload or replace winning image for a winning ticket."""
    ticket, error_response = _get_winning_ticket_or_error(ticket_id)
    if error_response:
        return error_response
    record = WinningRecord.query.filter_by(ticket_id=ticket_id).first()

    if record and record.is_checked:
        return jsonify({'success': False, 'error': '\u8be5\u4e2d\u5956\u8bb0\u5f55\u5df2\u68c0\u67e5\uff0c\u65e0\u6cd5\u66f4\u6362\u56fe\u7247'}), 403

    if 'image' not in request.files:
        return jsonify({'success': False, 'error': '璇烽€夋嫨鍥剧墖鏂囦欢'}), 400

    file = request.files['image']
    if not file.filename:
        return jsonify({'success': False, 'error': '鏂囦欢鍚嶄负绌?'}), 400

    try:
        from utils.image_upload import prepare_uploaded_image

        compressed, save_ext = prepare_uploaded_image(file)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

    from services.oss_service import _oss_configured, build_oss_key, delete_stored_image, get_public_url

    if _oss_configured():
        from services.oss_service import _get_bucket
        oss_key = build_oss_key(ticket_id, save_ext)
        try:
            _get_bucket().put_object(oss_key, compressed.read())
            image_url = get_public_url(oss_key)
        except Exception as e:
            return jsonify({'success': False, 'error': f'OSS涓婁紶澶辫触: {e}'}), 500
    else:
        upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
        images_dir = os.path.join(upload_folder, 'images')
        os.makedirs(images_dir, exist_ok=True)
        filename = f"winning_{ticket_id}_{uuid.uuid4().hex[:8]}.{save_ext}"
        save_path = os.path.join(images_dir, filename)
        with open(save_path, 'wb') as f:
            f.write(compressed.read())
        image_url = f"/uploads/images/{filename}"

    if record:
        if record.image_oss_key != (oss_key if _oss_configured() else None) or record.winning_image_url != image_url:
            delete_stored_image(record.image_oss_key, record.winning_image_url)
        record.winning_image_url = image_url
        record.image_oss_key = oss_key if _oss_configured() else None
        record.uploaded_by = current_user.id
        record.uploaded_at = beijing_now()
    else:
        record = WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=image_url,
            image_oss_key=oss_key if _oss_configured() else None,
            uploaded_by=current_user.id,
        )
        db.session.add(record)
    ticket.winning_image_url = image_url
    ticket.is_winning = True
    db.session.commit()
    return jsonify({'success': True, 'image_url': image_url, 'record': record.to_dict()})


# 鈹€鈹€ Match results 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€

@admin_bp.route('/match-results/upload', methods=['POST'])
@login_required_json
@login_required
@admin_required
def upload_match_result():
    """Upload match result file and trigger async winning calculation."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '璇烽€夋嫨鏂囦欢'}), 400

    file = request.files['file']
    safe_filename = _safe_uploaded_filename(file.filename)
    if not safe_filename:
        return jsonify({'success': False, 'error': '\u6587\u4ef6\u540d\u4e0d\u80fd\u4e3a\u7a7a'}), 400
    detail_period = (request.form.get('detail_period') or '').strip()
    upload_kind = (request.form.get('upload_kind') or 'final').strip().lower()
    if not detail_period:
        return jsonify({'success': False, 'error': '璇疯緭鍏ユ湡鍙?'}), 400

    upload_lottery_type, lottery_type_error = _resolve_result_upload_lottery_type(
        detail_period,
        request.form.get('lottery_type'),
    )

    if upload_kind not in {'predicted', 'final'}:
        return jsonify({'success': False, 'error': '涓婁紶绫诲瀷鏃犳晥'}), 400
    if lottery_type_error:
        return jsonify({'success': False, 'error': lottery_type_error}), 400

    filename_error = _validate_result_upload_filename(safe_filename, detail_period, upload_kind)
    if filename_error:
        return jsonify({'success': False, 'error': filename_error}), 400

    upload_folder = current_app.config['UPLOAD_FOLDER']
    stored = _build_result_upload_relative_path(safe_filename, detail_period, upload_kind)
    file_path = os.path.join(upload_folder, stored)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    file.save(file_path)

    result_file = ResultFile(
        original_filename=safe_filename,
        stored_filename=stored,
        uploaded_by=current_user.id,
        upload_kind=upload_kind,
    )
    db.session.add(result_file)
    db.session.flush()

    from services.result_parser import parse_result_file
    result = parse_result_file(
        file_path,
        detail_period,
        current_user.id,
        result_file.id,
        upload_kind=upload_kind,
        lottery_type=upload_lottery_type,
    )

    if not result['success']:
        result_file.status = 'error'
        result_file.parse_error = result.get('error')
        db.session.commit()
        return jsonify({'success': False, 'error': result.get('error')}), 400

    result_file.periods_count = result['count']
    db.session.commit()

    # Trigger async winning calculation
    match_result_id = result['match_result_id']
    expected_calc_token = result.get('calc_token') or result.get('uploaded_at')
    from tasks.scheduler import get_scheduler
    from services.winning_calc_service import process_match_result
    sched = get_scheduler()
    if sched:
        try:
            sched.add_job(
                func=process_match_result,
                args=[match_result_id, expected_calc_token],
                id=f'winning_calc_{match_result_id}',
                replace_existing=True,
            )
        except Exception:
            current_app.logger.exception("Failed to enqueue winning calc job, fallback to sync execution")
            process_match_result(
                match_result_id,
                expected_calc_token=expected_calc_token,
                app=current_app._get_current_object(),
            )
    else:
        process_match_result(
            match_result_id,
            expected_calc_token=expected_calc_token,
            app=current_app._get_current_object(),
        )

    return jsonify({'success': True, 'match_result_id': match_result_id, 'count': result['count']})


@admin_bp.route('/api/match-results')
@login_required_json
@login_required
@admin_required
def api_match_results():
    date_str = request.args.get('date', '').strip()
    q = MatchResult.query.order_by(MatchResult.uploaded_at.desc())
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': '鏃ユ湡鏍煎紡鏃犳晥锛岃浣跨敤 YYYY-MM-DD'}), 400
        start_at, end_at = get_business_window(selected_date)
        q = q.filter(
            MatchResult.uploaded_at >= start_at,
            MatchResult.uploaded_at < end_at,
        )
    results = q.limit(100).all()
    # 闄勫甫鏃ユ湡鍒楄〃渚涘墠绔瓫閫?
    uploaded_rows = db.session.query(MatchResult.uploaded_at).filter(
        MatchResult.uploaded_at.isnot(None)
    ).all()
    dates = sorted(
        {str(get_business_date(row[0])) for row in uploaded_rows if row[0]},
        reverse=True,
    )
    return jsonify({
        'results': [r.to_dict() for r in results],
        'dates': dates,
    })


@admin_bp.route('/api/match-results/<int:result_id>/detail')
@login_required_json
@login_required
@admin_required
def api_match_result_detail(result_id):
    """Return raw parsed result_data for a match result."""
    mr = db.session.get(MatchResult, result_id)
    if not mr:
        return jsonify({'success': False, 'error': '\u8d5b\u679c\u4e0d\u5b58\u5728'}), 404
    return jsonify({'success': True, 'result_data': mr.result_data, 'detail_period': mr.detail_period})


@admin_bp.route('/api/match-results/<int:result_id>/export-comparison')
@login_required_json
@login_required
@admin_required
def api_match_result_export_comparison(result_id):
    import io as _io
    from decimal import Decimal
    from openpyxl import Workbook

    match_result = db.session.get(MatchResult, result_id)
    if not match_result:
        return jsonify({'success': False, 'error': '\u8d5b\u679c\u4e0d\u5b58\u5728'}), 404

    tickets_query = LotteryTicket.query.filter(
        LotteryTicket.detail_period == match_result.detail_period,
        LotteryTicket.status.in_(['completed', 'expired']),
    )
    if match_result.lottery_type is not None:
        tickets_query = tickets_query.filter(
            or_(
                LotteryTicket.lottery_type == match_result.lottery_type,
                LotteryTicket.lottery_type.is_(None),
            )
        )
    tickets = tickets_query.all()

    def to_decimal(value):
        if value is None:
            return Decimal('0')
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value))
        except Exception:
            return Decimal('0')

    user_ids = {ticket.assigned_user_id for ticket in tickets if ticket.assigned_user_id is not None}
    device_ids = {ticket.assigned_device_id for ticket in tickets if ticket.assigned_device_id}
    registry_name_map = {}
    if user_ids and device_ids:
        registry_rows = DeviceRegistry.query.filter(
            DeviceRegistry.user_id.in_(user_ids),
            DeviceRegistry.device_id.in_(device_ids),
        ).all()
        for row in registry_rows:
            registry_name_map[(row.user_id, row.device_id)] = _resolve_device_display_name(row.device_id, row.client_info)

    customers = {}
    for ticket in tickets:
        if ticket.assigned_username:
            username = ticket.assigned_username.strip() or '未知客户'
        elif ticket.assigned_user_id is not None:
            username = f"用户#{ticket.assigned_user_id}"
        else:
            username = '未知客户'

        customer_bucket = customers.setdefault(
            username,
            {'predicted': Decimal('0'), 'final': Decimal('0'), 'devices': {}},
        )
        predicted_amount = to_decimal(ticket.predicted_winning_amount)
        final_amount = to_decimal(ticket.winning_amount)
        customer_bucket['predicted'] += predicted_amount
        customer_bucket['final'] += final_amount

        device_id = (ticket.assigned_device_id or '').strip()
        if not device_id:
            device_id = 'unknown'
        device_name = registry_name_map.get((ticket.assigned_user_id, device_id), device_id)
        device_bucket = customer_bucket['devices'].setdefault(
            device_id,
            {'device_name': device_name, 'predicted': Decimal('0'), 'final': Decimal('0')},
        )
        device_bucket['predicted'] += predicted_amount
        device_bucket['final'] += final_amount

    def pct(predicted: Decimal, final_amount: Decimal):
        if predicted == 0:
            return None
        return round(float((final_amount - predicted) / predicted * Decimal('100')), 2)

    wb = Workbook()
    ws = wb.active
    ws.title = '涨跌幅对比'
    ws.append([
        '层级',
        '编号',
        '彩种范围',
        '客户',
        '设备名',
        '设备ID',
        '预测奖金',
        '最终奖金',
        '涨跌幅(%)',
    ])

    for username in sorted(customers.keys()):
        item = customers[username]
        ws.append([
            '客户',
            match_result.detail_period or '',
            match_result.lottery_type or '全部彩种',
            username,
            '',
            '',
            float(item['predicted']),
            float(item['final']),
            pct(item['predicted'], item['final']),
        ])
        for device_id, device_item in sorted(item['devices'].items(), key=lambda pair: pair[0]):
            ws.append([
                '设备',
                match_result.detail_period or '',
                match_result.lottery_type or '全部彩种',
                username,
                device_item['device_name'],
                device_id,
                float(device_item['predicted']),
                float(device_item['final']),
                pct(device_item['predicted'], device_item['final']),
            ])

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"结果计算对比_{match_result.detail_period or 'unknown'}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@admin_bp.route('/api/match-results/<int:result_id>/recalc', methods=['POST'])
@login_required_json
@login_required
@admin_required
def api_recalc(result_id):
    from tasks.scheduler import get_scheduler
    from services.winning_calc_service import process_match_result
    match_result = db.session.get(MatchResult, result_id)
    if not match_result:
        return jsonify({'success': False, 'error': '\u8d5b\u679c\u4e0d\u5b58\u5728'}), 404

    same_period_query = MatchResult.query.filter(
        MatchResult.detail_period == match_result.detail_period
    )
    if match_result.lottery_type is None:
        same_period_query = same_period_query.filter(MatchResult.lottery_type.is_(None))
    else:
        same_period_query = same_period_query.filter(MatchResult.lottery_type == match_result.lottery_type)
    same_period_results = same_period_query.all()
    latest_same_period = max(
        same_period_results,
        key=lambda item: (item.uploaded_at or datetime.min, item.id),
    ) if same_period_results else None
    if latest_same_period and latest_same_period.id != match_result.id:
        return jsonify({
            'success': False,
            'error': 'only latest result_id in same period can be recalculated',
            'latest_result_id': latest_same_period.id,
        }), 409

    match_result.calc_status = 'pending'
    match_result.calc_started_at = None
    match_result.calc_finished_at = None
    match_result.tickets_total = 0
    match_result.tickets_winning = 0
    match_result.predicted_total_winning_amount = 0
    match_result.total_winning_amount = 0
    db.session.commit()
    expected_calc_token = (
        f"rf:{match_result.result_file_id}"
        if match_result.result_file_id is not None
        else (f"ts:{match_result.uploaded_at.isoformat()}" if match_result.uploaded_at else None)
    )

    sched = get_scheduler()
    if sched:
        try:
            sched.add_job(
                func=process_match_result,
                args=[result_id, expected_calc_token],
                id=f'winning_recalc_{result_id}',
                replace_existing=True,
            )
        except Exception:
            current_app.logger.exception("Failed to enqueue winning recalc job, fallback to sync execution")
            process_match_result(
                result_id,
                expected_calc_token=expected_calc_token,
                app=current_app._get_current_object(),
            )
    else:
        process_match_result(
            result_id,
            expected_calc_token=expected_calc_token,
            app=current_app._get_current_object(),
        )
    return jsonify({'success': True})


# 鈹€鈹€ Settings 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€

@admin_bp.route('/settings')
@login_required
@admin_required
def settings_page():
    return render_template('admin/settings.html')


@admin_bp.route('/api/settings', methods=['GET'])
@login_required_json
@login_required
@admin_required
def api_get_settings():
    settings = SystemSettings.get()
    payload = settings.to_dict()
    payload['database_info'] = _database_display_info()
    try:
        payload['scheduler_status'] = _build_scheduler_status()
    except Exception:
        current_app.logger.exception("Failed to build scheduler status")
        payload['scheduler_status'] = {
            'scheduler_present': False,
            'scheduler_running': False,
            'job_count': 0,
            'job_ids': [],
            'expected_job_ids': list(SCHEDULER_EXPECTED_JOB_IDS),
            'missing_job_ids': list(SCHEDULER_EXPECTED_JOB_IDS),
            'status': 'warning',
            'message': '\u8c03\u5ea6\u5668\u72b6\u6001\u6682\u4e0d\u53ef\u7528',
            'action': '\u8bf7\u68c0\u67e5\u5b9a\u65f6\u4efb\u52a1\u670d\u52a1\u662f\u5426\u6b63\u5e38',
        }
    return jsonify(payload)


@admin_bp.route('/api/settings', methods=['PUT'])
@login_required_json
@login_required
@admin_required
def api_update_settings():
    data, data_error = parse_json_object()
    if data_error:
        return data_error
    settings = SystemSettings.get()

    for bool_field in ['registration_enabled', 'pool_enabled', 'mode_a_enabled', 'mode_b_enabled', 'announcement_enabled']:
        if bool_field in data:
            parsed_bool = _parse_bool_flag(data.get(bool_field))
            if parsed_bool is None:
                return jsonify({'success': False, 'error': f'{bool_field} \u5fc5\u987b\u662f\u5e03\u5c14\u503c'}), 400
            data[bool_field] = parsed_bool

    if 'session_lifetime_hours' in data:
        parsed_hours = _parse_int_arg(data.get('session_lifetime_hours'), minimum=1)
        if parsed_hours is None or parsed_hours > 24:
            return jsonify({'success': False, 'error': '\u65e0\u6d3b\u52a8\u8d85\u65f6\u5fc5\u987b\u5728 1 \u5230 24 \u5c0f\u65f6\u4e4b\u95f4'}), 400
        data['session_lifetime_hours'] = parsed_hours

    if 'daily_reset_hour' in data:
        parsed_reset_hour = _parse_int_arg(data.get('daily_reset_hour'), minimum=0)
        if parsed_reset_hour is None or parsed_reset_hour > 23:
            return jsonify({'success': False, 'error': '姣忔棩閲嶇疆鏃堕棿蹇呴』鏄?0 鍒?23 涔嬮棿鐨勬暣鏁?'}), 400
        data['daily_reset_hour'] = parsed_reset_hour

    if 'mode_b_options' in data:
        mode_b_options = data.get('mode_b_options')
        if not isinstance(mode_b_options, list) or not mode_b_options:
            return jsonify({'success': False, 'error': 'B妯″紡鎵归噺閫夐」蹇呴』鏄潪绌烘暣鏁版暟缁?'}), 400

        normalized_options = []
        seen = set()
        for value in mode_b_options:
            parsed_value = _parse_int_arg(value, minimum=1)
            if parsed_value is None:
                return jsonify({'success': False, 'error': 'B妯″紡鎵归噺閫夐」蹇呴』鍏ㄩ儴鏄ぇ浜?0 鐨勬暣鏁?'}), 400
            if parsed_value not in seen:
                seen.add(parsed_value)
                normalized_options.append(parsed_value)
        data['mode_b_options'] = normalized_options

    if 'mode_b_pool_reserve' in data:
        parsed_reserve = _parse_int_arg(data.get('mode_b_pool_reserve'), minimum=0)
        if parsed_reserve is None or parsed_reserve > 100000:
            return jsonify({'success': False, 'error': 'B模式保留张数必须在 0 到 100000 之间'}), 400
        data['mode_b_pool_reserve'] = parsed_reserve

    for field in ['registration_enabled', 'pool_enabled', 'mode_a_enabled', 'mode_b_enabled',
                  'mode_b_options', 'announcement', 'announcement_enabled',
                  'session_lifetime_hours', 'daily_reset_hour', 'mode_b_pool_reserve']:
        if field in data:
            setattr(settings, field, data[field])

    settings.updated_by = current_user.id
    db.session.commit()

    if 'daily_reset_hour' in data:
        try:
            from tasks.scheduler import reschedule_daily_reset

            reschedule_daily_reset(current_app._get_current_object(), settings.daily_reset_hour)
        except Exception:
            current_app.logger.warning('reschedule_daily_reset failed after settings update', exc_info=True)

    if data.get('announcement_enabled') and data.get('announcement'):
        try:
            notify_all('announcement', {'content': data['announcement']})
        except Exception:
            current_app.logger.warning('notify_all announcement failed', exc_info=True)

    if 'pool_enabled' in data:
        try:
            if data['pool_enabled']:
                notify_all('pool_enabled', {'message': '\u7968\u6c60\u5df2\u5f00\u542f'})
            else:
                notify_all('pool_disabled', {'message': '\u7968\u6c60\u5df2\u5173\u95ed'})
        except Exception:
            current_app.logger.warning('notify_all pool state change failed', exc_info=True)

    if 'mode_a_enabled' in data or 'mode_b_enabled' in data:
        try:
            from services.notify_service import notify_pool_update

            notify_pool_update(get_pool_status())
        except Exception:
            current_app.logger.warning('notify_pool_update failed after mode switch update', exc_info=True)

    return jsonify({'success': True, 'settings': settings.to_dict()})

