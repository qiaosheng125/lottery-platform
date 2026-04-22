"""
用户导入服务
支持从 XLSX 文件批量导入用户。
"""

from __future__ import annotations

import os
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from extensions import db
from models.user import User


REQUIRED_COLUMNS = ['用户名', '密码', '接单模式', '最大设备数', '账号状态', '接单开关', 'B模式仅桌面端']
HEADER_ALIASES = {
    '密码哈希': '密码',
}


def _normalize_header(header) -> str:
    if header is None:
        return ''
    return HEADER_ALIASES.get(str(header).strip(), str(header).strip())


def _normalize_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def parse_user_xlsx(file_path: str) -> Tuple[List[Dict], List[Dict]]:
    """
    解析用户 XLSX 文件。

    Returns:
        (valid_rows, errors)
    """
    if not os.path.exists(file_path):
        return [], [{'row': 0, 'field': 'file', 'message': '文件不存在'}]

    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
    except Exception as exc:  # pragma: no cover - exercised via integration route
        return [], [{'row': 0, 'field': 'file', 'message': f'无法读取文件: {exc}'}]

    headers = [_normalize_header(cell.value) for cell in ws[1]]
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing_columns:
        wb.close()
        return [], [{'row': 0, 'field': 'headers', 'message': f'缺少必需列: {", ".join(missing_columns)}'}]

    valid_rows: List[Dict] = []
    errors: List[Dict] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        row_data = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row):
                row_data[header] = row[col_idx]

        if not _normalize_text(row_data.get('用户名')):
            continue

        valid_rows.append({'row_number': row_idx, 'data': row_data})

    wb.close()
    return valid_rows, errors


def _parse_bool_value(value, true_values=None, false_values=None):
    """解析布尔值。"""
    if true_values is None:
        true_values = ['是', '启用', '开启', '开', 'true', 'True', 'TRUE', '1', 1, True]
    if false_values is None:
        false_values = ['否', '禁用', '关闭', '关', 'false', 'False', 'FALSE', '0', 0, False]

    if value in true_values:
        return True
    if value in false_values:
        return False
    return None


def validate_user_row(
    row_data: Dict,
    row_number: int,
    existing_usernames: set,
    file_usernames: set,
) -> Tuple[Dict, List[Dict]]:
    """
    验证单行用户数据。

    Returns:
        (validated_data, errors)
    """
    errors: List[Dict] = []
    validated: Dict = {}

    username = _normalize_text(row_data.get('用户名'))
    if not username:
        errors.append({'row': row_number, 'field': '用户名', 'message': '用户名不能为空'})
    elif len(username) > 64:
        errors.append({'row': row_number, 'field': '用户名', 'message': '用户名长度不能超过64个字符'})
    elif username in existing_usernames:
        errors.append({'row': row_number, 'field': '用户名', 'message': '用户名已存在于数据库'})
    elif username in file_usernames:
        errors.append({'row': row_number, 'field': '用户名', 'message': '用户名在文件中重复'})
    else:
        validated['username'] = username
        file_usernames.add(username)

    password = _normalize_text(row_data.get('密码'))
    if not password:
        errors.append({'row': row_number, 'field': '密码', 'message': '密码不能为空'})
    elif password.startswith(('$2b$', '$2a$', '$2y$')):
        validated['password_hash'] = password
        validated['password'] = None
    elif len(password) < 6:
        errors.append({'row': row_number, 'field': '密码', 'message': '密码至少需要6位'})
    else:
        validated['password'] = password
        validated['password_hash'] = None

    client_mode = _normalize_text(row_data.get('接单模式'))
    if client_mode not in ['mode_a', 'mode_b', 'A模式', 'B模式', '逐单', '批量']:
        errors.append({'row': row_number, 'field': '接单模式', 'message': '接单模式必须是 mode_a 或 mode_b'})
    elif client_mode in ['A模式', '逐单']:
        validated['client_mode'] = 'mode_a'
    elif client_mode in ['B模式', '批量']:
        validated['client_mode'] = 'mode_b'
    else:
        validated['client_mode'] = client_mode

    try:
        max_devices = int(row_data.get('最大设备数') or 1)
        if max_devices < 1 or max_devices > 10:
            errors.append({'row': row_number, 'field': '最大设备数', 'message': '最大设备数必须在1-10之间'})
        else:
            validated['max_devices'] = max_devices
    except (ValueError, TypeError):
        errors.append({'row': row_number, 'field': '最大设备数', 'message': '最大设备数必须是整数'})

    max_processing_b_mode = row_data.get('B模式处理上限')
    if max_processing_b_mode is not None and _normalize_text(max_processing_b_mode):
        try:
            max_processing_b_mode = int(max_processing_b_mode)
            if max_processing_b_mode < 1 or max_processing_b_mode > 10000:
                errors.append({'row': row_number, 'field': 'B模式处理上限', 'message': 'B模式处理上限必须在1-10000之间'})
            else:
                validated['max_processing_b_mode'] = max_processing_b_mode
        except (ValueError, TypeError):
            errors.append({'row': row_number, 'field': 'B模式处理上限', 'message': 'B模式处理上限必须是整数'})
    else:
        validated['max_processing_b_mode'] = None

    daily_ticket_limit = row_data.get('每日处理上限')
    if daily_ticket_limit is not None and _normalize_text(daily_ticket_limit):
        try:
            daily_ticket_limit = int(daily_ticket_limit)
            if daily_ticket_limit < 1 or daily_ticket_limit > 100000:
                errors.append({'row': row_number, 'field': '每日处理上限', 'message': '每日处理上限必须在1-100000之间'})
            else:
                validated['daily_ticket_limit'] = daily_ticket_limit
        except (ValueError, TypeError):
            errors.append({'row': row_number, 'field': '每日处理上限', 'message': '每日处理上限必须是整数'})
    else:
        validated['daily_ticket_limit'] = None

    blocked_lottery_types = row_data.get('禁止彩种')
    if blocked_lottery_types and _normalize_text(blocked_lottery_types):
        validated['blocked_lottery_types'] = [
            item.strip() for item in str(blocked_lottery_types).split(',') if item.strip()
        ]
    else:
        validated['blocked_lottery_types'] = []

    is_active = _parse_bool_value(row_data.get('账号状态'))
    if is_active is None:
        errors.append({'row': row_number, 'field': '账号状态', 'message': '账号状态必须是"启用"或"禁用"'})
    else:
        validated['is_active'] = is_active

    can_receive = _parse_bool_value(row_data.get('接单开关'))
    if can_receive is None:
        errors.append({'row': row_number, 'field': '接单开关', 'message': '接单开关必须是"开启"或"关闭"'})
    else:
        validated['can_receive'] = can_receive

    desktop_only_b_mode = _parse_bool_value(row_data.get('B模式仅桌面端'))
    if desktop_only_b_mode is None:
        errors.append({'row': row_number, 'field': 'B模式仅桌面端', 'message': 'B模式仅桌面端必须是"是"或"否"'})
    else:
        validated['desktop_only_b_mode'] = desktop_only_b_mode

    return validated, errors


def import_users(file_path: str, admin_user_id: int) -> Dict:
    """
    批量导入用户。

    Returns:
        {
            'success': bool,
            'total': int,
            'success_count': int,
            'error_count': int,
            'errors': List[Dict],
        }
    """
    _ = admin_user_id  # reserved for future audit trail support

    rows, parse_errors = parse_user_xlsx(file_path)
    if parse_errors:
        return {
            'success': False,
            'total': 0,
            'success_count': 0,
            'error_count': len(parse_errors),
            'errors': parse_errors,
        }

    if not rows:
        return {
            'success': False,
            'total': 0,
            'success_count': 0,
            'error_count': 1,
            'errors': [{'row': 0, 'field': 'file', 'message': '文件中没有有效的用户数据'}],
        }

    existing_usernames = {user.username for user in User.query.all()}
    file_usernames = set()
    validated_users: List[Dict] = []
    all_errors: List[Dict] = []

    for row_info in rows:
        validated, errors = validate_user_row(
            row_info['data'],
            row_info['row_number'],
            existing_usernames,
            file_usernames,
        )
        if errors:
            all_errors.extend(errors)
        else:
            validated_users.append(validated)

    if all_errors:
        return {
            'success': False,
            'total': len(rows),
            'success_count': 0,
            'error_count': len(all_errors),
            'errors': all_errors,
        }

    try:
        created_count = 0
        for user_data in validated_users:
            user = User(
                username=user_data['username'],
                client_mode=user_data['client_mode'],
                max_devices=user_data['max_devices'],
                max_processing_b_mode=user_data.get('max_processing_b_mode'),
                daily_ticket_limit=user_data.get('daily_ticket_limit'),
                is_active=user_data['is_active'],
                can_receive=user_data['can_receive'],
                desktop_only_b_mode=user_data['desktop_only_b_mode'],
            )

            if user_data.get('password_hash'):
                user.password_hash = user_data['password_hash']
            else:
                user.set_password(user_data['password'])

            user.set_blocked_lottery_types(user_data.get('blocked_lottery_types'))
            db.session.add(user)
            created_count += 1

        db.session.commit()
        return {
            'success': True,
            'total': len(rows),
            'success_count': created_count,
            'error_count': 0,
            'errors': [],
        }
    except Exception as exc:  # pragma: no cover - exercised via integration route
        db.session.rollback()
        return {
            'success': False,
            'total': len(rows),
            'success_count': 0,
            'error_count': 1,
            'errors': [{'row': 0, 'field': 'database', 'message': f'数据库错误: {exc}'}],
        }
