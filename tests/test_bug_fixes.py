import os
import sys
import builtins
import json
from pathlib import Path
from datetime import datetime, timedelta
import io

import pytest
from sqlalchemy import inspect, text
from werkzeug.datastructures import FileStorage

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import tasks.scheduler as scheduler_module
from app import create_app, ensure_runtime_columns, should_start_scheduler
from extensions import db
from models.audit import AuditLog
from models.settings import SystemSettings
from models.file import UploadedFile
from models.ticket import LotteryTicket
from models.device import DeviceRegistry
from models.user import User
from models.result import MatchResult, ResultFile
from models.archive import ArchivedLotteryTicket
from models.winning import WinningRecord
from utils.filename_parser import parse_filename
from utils.time_utils import beijing_now
from routes.admin import _database_display_info


@pytest.fixture()
def app(tmp_path, monkeypatch):
    db_path = tmp_path / "test.sqlite"
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    monkeypatch.setenv("FLASK_ENV", "development")
    monkeypatch.setattr(scheduler_module, "start_scheduler", lambda app: None)

    app = create_app()
    app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)

    with app.app_context():
        db.drop_all()
        db.create_all()
        SystemSettings.get()
        yield app
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def client(app):
    return app.test_client()


def create_user(username: str, password: str, client_mode: str = "mode_b") -> User:
    user = User(username=username, client_mode=client_mode, max_devices=5, can_receive=True)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    return user


def login(client, username: str, password: str):
    return client.post(
        "/auth/login",
        json={"username": username, "password": password},
    )


def test_should_start_scheduler_defaults_to_false_in_production(monkeypatch):
    monkeypatch.setenv("FLASK_ENV", "production")
    monkeypatch.delenv("ENABLE_SCHEDULER", raising=False)
    monkeypatch.delenv("DISABLE_SCHEDULER", raising=False)

    assert should_start_scheduler() is False


def test_should_start_scheduler_can_enable_for_dedicated_process(monkeypatch):
    monkeypatch.setenv("FLASK_ENV", "production")
    monkeypatch.setenv("ENABLE_SCHEDULER", "1")
    monkeypatch.delenv("DISABLE_SCHEDULER", raising=False)

    assert should_start_scheduler() is True


def test_should_start_scheduler_disable_flag_takes_priority(monkeypatch):
    monkeypatch.setenv("FLASK_ENV", "development")
    monkeypatch.setenv("ENABLE_SCHEDULER", "1")
    monkeypatch.setenv("DISABLE_SCHEDULER", "1")

    assert should_start_scheduler() is False


def test_should_start_scheduler_uses_config_name_when_flask_env_missing(monkeypatch):
    monkeypatch.delenv("FLASK_ENV", raising=False)
    monkeypatch.delenv("ENABLE_SCHEDULER", raising=False)
    monkeypatch.delenv("DISABLE_SCHEDULER", raising=False)

    assert should_start_scheduler("production") is False
    assert should_start_scheduler("development") is True


def test_ensure_runtime_columns_backfills_legacy_users_table(app):
    with app.app_context():
        db.drop_all()
        with db.engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE users (
                    id INTEGER PRIMARY KEY,
                    username VARCHAR(64) NOT NULL UNIQUE,
                    password_hash VARCHAR(256) NOT NULL,
                    is_admin BOOLEAN NOT NULL DEFAULT 0,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            """))

        ensure_runtime_columns(app)
        ensure_runtime_columns(app)

        inspector = inspect(db.engine)
        columns = {column["name"] for column in inspector.get_columns("users")}
        expected_columns = {
            "client_mode",
            "max_devices",
            "max_processing_b_mode",
            "daily_ticket_limit",
            "blocked_lottery_types",
            "is_active",
            "can_receive",
            "desktop_only_b_mode",
            "updated_at",
        }
        assert expected_columns.issubset(columns)


def test_ensure_runtime_columns_backfills_system_settings_mode_b_pool_reserve(app):
    with app.app_context():
        db.drop_all()
        with db.engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE system_settings (
                    id INTEGER PRIMARY KEY,
                    registration_enabled BOOLEAN NOT NULL DEFAULT 1,
                    pool_enabled BOOLEAN NOT NULL DEFAULT 1,
                    mode_a_enabled BOOLEAN NOT NULL DEFAULT 1,
                    mode_b_enabled BOOLEAN NOT NULL DEFAULT 1
                )
            """))

        ensure_runtime_columns(app)
        ensure_runtime_columns(app)

        inspector = inspect(db.engine)
        columns = {column["name"] for column in inspector.get_columns("system_settings")}
        assert "mode_b_pool_reserve" in columns


def test_ensure_runtime_columns_backfills_ticket_download_filename(app):
    with app.app_context():
        db.drop_all()
        with db.engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE lottery_tickets (
                    id INTEGER PRIMARY KEY,
                    source_file_id INTEGER NOT NULL,
                    line_number INTEGER NOT NULL,
                    raw_content TEXT NOT NULL,
                    status VARCHAR(20) NOT NULL DEFAULT 'pending'
                )
            """))

        ensure_runtime_columns(app)
        ensure_runtime_columns(app)

        inspector = inspect(db.engine)
        columns = {column["name"] for column in inspector.get_columns("lottery_tickets")}
        assert "download_filename" in columns


def test_login_json_post_stays_json_for_authenticated_user(app, client):
    with app.app_context():
        create_user("login_json_user", "secret123", client_mode="mode_a")

    first = login(client, "login_json_user", "secret123")
    assert first.status_code == 200
    assert first.is_json is True

    second = login(client, "login_json_user", "secret123")
    assert second.status_code == 200
    assert second.is_json is True
    data = second.get_json()
    assert data["success"] is True
    assert data["redirect"] == "/api/user/dashboard"
    assert data["client_mode"] == "mode_a"


def test_login_json_returns_client_mode(app, client):
    with app.app_context():
        create_user("login_modeb_user", "secret123", client_mode="mode_b")

    resp = login(client, "login_modeb_user", "secret123")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["client_mode"] == "mode_b"


def test_login_handles_empty_json_body(app, client):
    resp = client.post("/auth/login", data="", content_type="application/json")
    assert resp.status_code == 401
    data = resp.get_json()
    assert data["success"] is False
    assert "\u7528\u6237\u540d\u6216\u5bc6\u7801\u9519\u8bef" in data["error"]


def test_login_rejects_non_string_password_type(app, client):
    with app.app_context():
        create_user("login_type_guard_user", "secret123", client_mode="mode_b")

    resp = client.post("/auth/login", json={"username": "login_type_guard_user", "password": ["x"]})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "invalid password type" in data["error"]


def test_login_rejects_non_string_username_type(app, client):
    with app.app_context():
        create_user("login_username_guard_user", "secret123", client_mode="mode_b")

    resp = client.post("/auth/login", json={"username": True, "password": "secret123"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "invalid username type" in data["error"]


def test_login_does_not_treat_stale_same_device_session_as_active(app, client):
    with app.app_context():
        user = create_user("login_stale_device_user", "secret123", client_mode="mode_b")
        user.max_devices = 1
        db.session.commit()

        from models.user import UserSession

        db.session.add_all([
            UserSession(
                user_id=user.id,
                session_token="active-other-device",
                device_id="device-active",
                last_seen=beijing_now(),
                expires_at=beijing_now() + timedelta(hours=3),
            ),
            UserSession(
                user_id=user.id,
                session_token="stale-same-device",
                device_id="device-stale",
                last_seen=beijing_now() - timedelta(hours=5),
                expires_at=beijing_now() - timedelta(hours=2),
            ),
        ])
        db.session.commit()

    resp = client.post(
        "/auth/login",
        json={"username": "login_stale_device_user", "password": "secret123", "device_id": "device-stale"},
    )
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_login_without_device_id_still_enforces_max_devices(app, client):
    with app.app_context():
        user = create_user("login_limit_without_device_user", "secret123", client_mode="mode_b")
        user.max_devices = 1
        db.session.commit()

        from models.user import UserSession

        db.session.add(
            UserSession(
                user_id=user.id,
                session_token="existing-active-session",
                device_id="device-a",
                last_seen=beijing_now(),
                expires_at=beijing_now() + timedelta(hours=3),
            )
        )
        db.session.commit()

    resp = client.post(
        "/auth/login",
        json={"username": "login_limit_without_device_user", "password": "secret123"},
    )
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "1" in data["error"]


def test_login_trims_device_id_before_max_devices_check(app, client):
    with app.app_context():
        user = create_user("login_trim_device_user", "secret123", client_mode="mode_b")
        user.max_devices = 1
        db.session.commit()

        from models.user import UserSession

        db.session.add(
            UserSession(
                user_id=user.id,
                session_token="existing-trim-device-session",
                device_id="device-a",
                last_seen=beijing_now(),
                expires_at=beijing_now() + timedelta(hours=3),
            )
        )
        db.session.commit()

    resp = client.post(
        "/auth/login",
        json={"username": "login_trim_device_user", "password": "secret123", "device_id": "device-a "},
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True


def test_heartbeat_rejects_invalid_device_id_format(app, client):
    with app.app_context():
        create_user("heartbeat_invalid_device_user", "secret123", client_mode="mode_a")

    login_resp = login(client, "heartbeat_invalid_device_user", "secret123")
    assert login_resp.status_code == 200

    resp = client.post("/auth/heartbeat", json={"device_id": "bad id"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False

    with app.app_context():
        from models.user import UserSession

        assert UserSession.query.filter_by(device_id="bad id").first() is None


def test_mode_a_current_normalizes_device_id(app, client):
    with app.app_context():
        user = create_user("mode_a_trim_lookup_user", "secret123", client_mode="mode_a")

        uploaded = UploadedFile(
            original_filename="mode_a_source.txt",
            stored_filename="mode_a_source.txt",
            status="active",
            total_tickets=1,
            pending_count=0,
            assigned_count=1,
            completed_count=0,
        )
        db.session.add(uploaded)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="1",
            lottery_type="???",
            deadline_time=beijing_now() + timedelta(hours=2),
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-a",
            assigned_at=beijing_now(),
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

    login_resp = login(client, "mode_a_trim_lookup_user", "secret123")
    assert login_resp.status_code == 200

    resp = client.get("/api/mode-a/current?device_id=device-a%20")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["id"] == ticket_id


def test_admin_export_endpoints_return_json_when_unauthenticated(app, client):
    endpoints = [
        "/admin/api/tickets/export",
        "/admin/api/tickets/export-by-date",
        "/admin/api/users/export",
        "/admin/api/winning/export",
    ]
    for endpoint in endpoints:
        resp = client.get(endpoint)
        assert resp.status_code == 401
        assert resp.is_json is True
        data = resp.get_json()
        assert data["success"] is False


def test_user_json_endpoints_reject_non_object_payload(app, client):
    with app.app_context():
        create_user("non_object_user", "secret123", client_mode="mode_b")

    login_resp = login(client, "non_object_user", "secret123")
    assert login_resp.status_code == 200

    for method, endpoint in [
        ("POST", "/auth/heartbeat"),
        ("POST", "/api/device/register"),
        ("POST", "/api/mode-b/download"),
        ("POST", "/api/mode-b/confirm"),
        ("POST", "/api/user/change-password"),
        ("POST", "/api/winning/record"),
    ]:
        if method == "POST":
            resp = client.post(endpoint, json=[1])
        else:
            resp = client.put(endpoint, json=[1])
        assert resp.status_code == 400
        assert resp.is_json is True
        data = resp.get_json()
        assert data["success"] is False


def test_mode_a_json_endpoints_reject_non_object_payload(app, client):
    with app.app_context():
        create_user("non_object_mode_a_user", "secret123", client_mode="mode_a")

    login_resp = login(client, "non_object_mode_a_user", "secret123")
    assert login_resp.status_code == 200

    for endpoint in ["/api/mode-a/next", "/api/mode-a/stop"]:
        resp = client.post(endpoint, json=[1])
        assert resp.status_code == 400
        assert resp.is_json is True
        data = resp.get_json()
        assert data["success"] is False


def test_mode_a_next_rejects_malformed_json_body(app, client):
    with app.app_context():
        create_user("mode_a_malformed_json_user", "secret123", client_mode="mode_a")

    login_resp = login(client, "mode_a_malformed_json_user", "secret123")
    assert login_resp.status_code == 200

    resp = client.post(
        "/api/mode-a/next?device_id=device-a",
        data='{"device_id":',
        content_type="application/json",
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_admin_recycle_page_is_linked_from_navbar():
    base_template = Path(__file__).resolve().parents[1] / "templates" / "base.html"
    content = base_template.read_text(encoding="utf-8")
    assert "url_for('admin.recycle_page')" in content
    assert "回收处理中票" in content


def test_admin_recycle_template_renders_filters_and_actions():
    recycle_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "recycle.html"
    content = recycle_template.read_text(encoding="utf-8")
    assert "用户名" in content
    assert "设备ID" in content
    assert "分配文件名" in content
    assert '<select v-model="filters.username"' in content
    assert '<select v-model="filters.device_id"' in content
    assert '<select v-model="filters.download_filename"' in content
    assert "filterOptions.usernames" in content
    assert "filterOptions.device_ids" in content
    assert "filterOptions.download_filenames" in content
    assert "const { createApp } = Vue;" in content
    assert "回收单张" in content
    assert "回收当前文件名处理中票" in content
    assert "B模式单张回收风险提示" in content
    assert "/admin/api/tickets/recycle-assigned" in content
    assert "async readJsonResponse(res, fallbackMessage)" in content
    assert "服务返回了非 JSON 响应" in content
    assert "只处理“处理中”的票" in content
    assert "待分配 / pending" not in content
    assert "处理中 / assigned" not in content


def test_ticket_recycle_service_uses_distinct_filter_options_and_row_locks():
    service = Path(__file__).resolve().parents[1] / "services" / "ticket_recycle_service.py"
    content = service.read_text(encoding="utf-8")
    assert ".distinct()" in content
    assert ".with_for_update()" in content


def test_admin_recycle_assigned_list_filters_processing_tickets(app, client):
    from decimal import Decimal

    with app.app_context():
        admin = User(username="admin_recycle_list", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        user = create_user("recycle_list_user", "secret123", client_mode="mode_b")
        other = create_user("recycle_list_other", "secret123", client_mode="mode_b")
        uploaded = UploadedFile(
            original_filename="回收测试.txt",
            stored_filename="recycle-list.txt",
            total_tickets=3,
            pending_count=0,
            assigned_count=2,
            completed_count=1,
        )
        db.session.add(uploaded)
        db.session.flush()
        assigned = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="LIST-ASSIGNED",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="dev-a",
            assigned_at=beijing_now(),
            download_filename="分配文件A.txt",
            detail_period="26051",
            lottery_type="胜平负",
            ticket_amount=Decimal("12.5"),
        )
        completed = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=2,
            raw_content="LIST-COMPLETED",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="dev-a",
            assigned_at=beijing_now(),
            completed_at=beijing_now(),
            download_filename="分配文件A.txt",
            ticket_amount=Decimal("20"),
        )
        other_assigned = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=3,
            raw_content="LIST-OTHER",
            status="assigned",
            assigned_user_id=other.id,
            assigned_username=other.username,
            assigned_device_id="dev-b",
            assigned_at=beijing_now(),
            download_filename="分配文件B.txt",
            ticket_amount=Decimal("30"),
        )
        db.session.add_all([assigned, completed, other_assigned])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_recycle_list", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/tickets/recycle-assigned?username=recycle_list_user&device_id=dev-a&download_filename=分配文件A.txt")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["total"] == 1
    assert data["items"][0]["raw_content"] == "LIST-ASSIGNED"
    assert data["items"][0]["status"] == "assigned"
    assert data["items"][0]["status_label"] == "处理中"
    assert data["items"][0]["ticket_amount"] == 12.5
    assert data["items"][0]["download_filename"] == "分配文件A.txt"
    assert data["filter_options"]["usernames"] == ["recycle_list_user"]
    assert data["filter_options"]["device_ids"] == ["dev-a"]
    assert data["filter_options"]["download_filenames"] == [data["items"][0]["download_filename"]]


def test_admin_recycle_single_assigned_ticket_returns_it_to_pending(app, client):
    from decimal import Decimal

    with app.app_context():
        admin = User(username="admin_recycle_single", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        user = create_user("recycle_single_user", "secret123", client_mode="mode_a")
        uploaded = UploadedFile(
            original_filename="单张回收.txt",
            stored_filename="recycle-single.txt",
            total_tickets=1,
            pending_count=0,
            assigned_count=1,
            completed_count=0,
        )
        db.session.add(uploaded)
        db.session.flush()
        ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="RECYCLE-SINGLE",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="dev-single",
            assigned_at=beijing_now(),
            locked_until=beijing_now() + timedelta(minutes=10),
            download_filename="单张分配.txt",
            completed_at=beijing_now(),
            ticket_amount=Decimal("18"),
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id
        file_id = uploaded.id

    resp = client.post("/auth/login", json={"username": "admin_recycle_single", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post("/admin/api/tickets/recycle-assigned", json={"ticket_ids": [ticket_id]})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["recycled_count"] == 1
    assert data["recycled_amount"] == 18.0

    with app.app_context():
        refreshed = db.session.get(LotteryTicket, ticket_id)
        uploaded = db.session.get(UploadedFile, file_id)
        assert refreshed.status == "pending"
        assert refreshed.assigned_user_id is None
        assert refreshed.assigned_username is None
        assert refreshed.assigned_device_id is None
        assert refreshed.assigned_at is None
        assert refreshed.locked_until is None
        assert refreshed.download_filename is None
        assert refreshed.completed_at is None
        assert uploaded.pending_count == 1
        assert uploaded.assigned_count == 0
        assert uploaded.completed_count == 0
        log = AuditLog.query.filter_by(action_type="ticket_recycle").first()
        assert log is not None
        assert "管理员手动回收处理中票" in log.details


def test_admin_recycle_single_allows_mode_b_ticket_ids(app, client):
    from decimal import Decimal

    with app.app_context():
        admin = User(username="admin_recycle_single_b", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        user = create_user("recycle_single_b_user", "secret123", client_mode="mode_b")
        uploaded = UploadedFile(
            original_filename="B单张回收.txt",
            stored_filename="recycle-single-b.txt",
            total_tickets=1,
            pending_count=0,
            assigned_count=1,
            completed_count=0,
        )
        db.session.add(uploaded)
        db.session.flush()
        ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="RECYCLE-SINGLE-B",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="dev-single-b",
            assigned_at=beijing_now(),
            download_filename="B批量文件.txt",
            ticket_amount=Decimal("18"),
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_recycle_single_b", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post("/admin/api/tickets/recycle-assigned", json={"ticket_ids": [ticket_id]})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["recycled_count"] == 1

    with app.app_context():
        refreshed = db.session.get(LotteryTicket, ticket_id)
        uploaded = UploadedFile.query.filter_by(stored_filename="recycle-single-b.txt").first()
        assert refreshed.status == "pending"
        assert refreshed.assigned_user_id is None
        assert refreshed.assigned_device_id is None
        assert refreshed.download_filename is None
        assert uploaded.pending_count == 1
        assert uploaded.assigned_count == 0


def test_admin_recycle_by_filename_only_recycles_matching_processing_tickets(app, client):
    from decimal import Decimal

    with app.app_context():
        admin = User(username="admin_recycle_filename", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        user = create_user("recycle_file_user", "secret123", client_mode="mode_b")
        other = create_user("recycle_file_other", "secret123", client_mode="mode_b")
        uploaded = UploadedFile(
            original_filename="文件名回收.txt",
            stored_filename="recycle-filename.txt",
            total_tickets=4,
            pending_count=0,
            assigned_count=3,
            completed_count=1,
        )
        db.session.add(uploaded)
        db.session.flush()
        matching = []
        for line in (1, 2):
            matching.append(LotteryTicket(
                source_file_id=uploaded.id,
                line_number=line,
                raw_content=f"RECYCLE-FILE-{line}",
                status="assigned",
                assigned_user_id=user.id,
                assigned_username=user.username,
                assigned_device_id="dev-file",
                assigned_at=beijing_now(),
                download_filename="同一个分配文件.txt",
                ticket_amount=Decimal("10"),
            ))
        same_filename_other_device = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=3,
            raw_content="RECYCLE-FILE-OTHER-DEVICE",
            status="assigned",
            assigned_user_id=other.id,
            assigned_username=other.username,
            assigned_device_id="dev-other",
            assigned_at=beijing_now(),
            download_filename="同一个分配文件.txt",
            ticket_amount=Decimal("10"),
        )
        completed = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=4,
            raw_content="RECYCLE-FILE-COMPLETED",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="dev-file",
            assigned_at=beijing_now(),
            completed_at=beijing_now(),
            download_filename="同一个分配文件.txt",
            ticket_amount=Decimal("10"),
        )
        db.session.add_all(matching + [same_filename_other_device, completed])
        db.session.commit()
        matching_ids = [ticket.id for ticket in matching]
        other_id = same_filename_other_device.id
        completed_id = completed.id
        file_id = uploaded.id

    resp = client.post("/auth/login", json={"username": "admin_recycle_filename", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post("/admin/api/tickets/recycle-assigned", json={
        "username": "recycle_file_user",
        "device_id": "dev-file",
        "download_filename": "同一个分配文件.txt",
    })
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["recycled_count"] == 2

    with app.app_context():
        assert [db.session.get(LotteryTicket, ticket_id).status for ticket_id in matching_ids] == ["pending", "pending"]
        assert db.session.get(LotteryTicket, other_id).status == "assigned"
        assert db.session.get(LotteryTicket, completed_id).status == "completed"
        uploaded = db.session.get(UploadedFile, file_id)
        assert uploaded.pending_count == 2
        assert uploaded.assigned_count == 1
        assert uploaded.completed_count == 1


def test_admin_recycle_by_filename_audit_details_are_sampled_for_large_batches(app, client):
    from decimal import Decimal

    with app.app_context():
        admin = User(username="admin_recycle_large_batch", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        user = create_user("recycle_large_user", "secret123", client_mode="mode_b")
        uploaded = UploadedFile(
            original_filename="large-recycle.txt",
            stored_filename="large-recycle.txt",
            total_tickets=205,
            pending_count=0,
            assigned_count=205,
            completed_count=0,
        )
        db.session.add(uploaded)
        db.session.flush()

        tickets = []
        for line in range(1, 206):
            tickets.append(LotteryTicket(
                source_file_id=uploaded.id,
                line_number=line,
                raw_content=f"LARGE-RECYCLE-{line}",
                status="assigned",
                assigned_user_id=user.id,
                assigned_username=user.username,
                assigned_device_id="dev-large",
                assigned_at=beijing_now(),
                download_filename="large-assigned.txt",
                ticket_amount=Decimal("1"),
            ))
        db.session.add_all(tickets)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_recycle_large_batch", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post("/admin/api/tickets/recycle-assigned", json={
        "username": "recycle_large_user",
        "device_id": "dev-large",
        "download_filename": "large-assigned.txt",
    })
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["success"] is True
    assert payload["recycled_count"] == 205

    with app.app_context():
        latest_log = (
            AuditLog.query.filter_by(action_type="ticket_recycle")
            .order_by(AuditLog.id.desc())
            .first()
        )
        assert latest_log is not None
        details = json.loads(latest_log.details)
        assert details["recycled_count"] == 205
        assert details["ticket_ids_sampled_count"] == 200
        assert details["ticket_ids_omitted_count"] == 5
        assert details["original_sampled_count"] == 200
        assert details["original_omitted_count"] == 5


def test_admin_recycle_by_filename_audit_resource_id_fits_column_for_large_ticket_ids(app, client):
    from decimal import Decimal

    with app.app_context():
        admin = User(username="admin_recycle_large_ids", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        user = create_user("recycle_large_ids_user", "secret123", client_mode="mode_b")
        uploaded = UploadedFile(
            original_filename="large-ids-recycle.txt",
            stored_filename="large-ids-recycle.txt",
            total_tickets=20,
            pending_count=0,
            assigned_count=20,
            completed_count=0,
        )
        db.session.add(uploaded)
        db.session.flush()

        tickets = []
        for line in range(1, 21):
            tickets.append(LotteryTicket(
                id=100000000000 + line,
                source_file_id=uploaded.id,
                line_number=line,
                raw_content=f"LARGE-ID-RECYCLE-{line}",
                status="assigned",
                assigned_user_id=user.id,
                assigned_username=user.username,
                assigned_device_id="dev-large-ids",
                assigned_at=beijing_now(),
                download_filename="large-id-assigned.txt",
                ticket_amount=Decimal("1"),
            ))
        db.session.add_all(tickets)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_recycle_large_ids", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post("/admin/api/tickets/recycle-assigned", json={
        "username": "recycle_large_ids_user",
        "device_id": "dev-large-ids",
        "download_filename": "large-id-assigned.txt",
    })
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["success"] is True
    assert payload["recycled_count"] == 20

    with app.app_context():
        latest_log = (
            AuditLog.query.filter_by(action_type="ticket_recycle")
            .order_by(AuditLog.id.desc())
            .first()
        )
        assert latest_log is not None
        assert len(latest_log.resource_id) <= 64
        assert latest_log.resource_id == "100000000001,100000000002,100000000003,100000000004,100000000005"


def test_mode_b_confirm_reports_recycled_batch_message(app, client):
    with app.app_context():
        user = create_user("mode_b_recycled_confirm_user", "secret123", client_mode="mode_b")
        ticket = create_assigned_ticket(user, "dev-recycled", "RECYCLED-CONFIRM", 1)
        ticket.status = "pending"
        ticket.assigned_user_id = None
        ticket.assigned_username = None
        ticket.assigned_device_id = None
        ticket.assigned_at = None
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "mode_b_recycled_confirm_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/confirm", json={
        "ticket_ids": [ticket_id],
        "device_id": "dev-recycled",
    })
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "管理员回收" in data["error"]


def test_admin_json_endpoints_reject_non_object_payload(app, client):
    with app.app_context():
        admin = User(username="admin_non_object_json", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_non_object_json", "password": "secret123"})
    assert resp.status_code == 200

    for method, endpoint in [
        ("POST", "/admin/api/users"),
        ("PUT", "/admin/api/settings"),
        ("POST", "/admin/api/winning/record"),
        ("POST", "/admin/api/tickets/recycle-assigned"),
    ]:
        if method == "PUT":
            response = client.put(endpoint, json=[1])
        else:
            response = client.post(endpoint, json=[1])
        assert response.status_code == 400
        assert response.is_json is True
        data = response.get_json()
        assert data["success"] is False


def test_admin_recycle_assigned_returns_json_when_internal_error(app, client, monkeypatch):
    with app.app_context():
        admin = User(username="admin_recycle_internal_error", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_recycle_internal_error", "password": "secret123"})
    assert resp.status_code == 200

    def _raise(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr("routes.admin.recycle_assigned_tickets", _raise)

    resp = client.post("/admin/api/tickets/recycle-assigned", json={"ticket_ids": [1]})
    assert resp.status_code == 500
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]
    assert "回收失败" in data["error"]


def test_logout_redirects_to_login_and_invalidates_heartbeat(app, client):
    with app.app_context():
        create_user("logout_user", "secret123", client_mode="mode_b")

    login_resp = login(client, "logout_user", "secret123")
    assert login_resp.status_code == 200
    assert login_resp.get_json()["success"] is True

    logout_resp = client.post("/auth/logout", follow_redirects=False)
    assert logout_resp.status_code in (302, 303)
    assert "/auth/login" in logout_resp.headers.get("Location", "")

    heartbeat_resp = client.post("/auth/heartbeat", json={})
    assert heartbeat_resp.status_code == 401
    assert heartbeat_resp.is_json is True
    heartbeat_data = heartbeat_resp.get_json()
    assert heartbeat_data["success"] is False


def test_logout_get_method_not_allowed(app, client):
    with app.app_context():
        create_user("logout_get_guard_user", "secret123", client_mode="mode_b")

    login_resp = login(client, "logout_get_guard_user", "secret123")
    assert login_resp.status_code == 200

    resp = client.get("/auth/logout", follow_redirects=False)
    assert resp.status_code == 405


def test_create_app_bootstraps_empty_sqlite(monkeypatch, tmp_path):
    db_path = tmp_path / "bootstrap.sqlite"
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    monkeypatch.setenv("FLASK_ENV", "development")
    monkeypatch.setattr(scheduler_module, "start_scheduler", lambda app: None)

    app = create_app()
    app.config.update(TESTING=True)

    with app.app_context():
        admin = User.query.filter_by(username="zucaixu", is_admin=True).first(),
        settings = SystemSettings.get()
        tables = set(inspect(db.engine).get_table_names())
        assert admin is not None
        assert settings is not None
        assert {
            "users",
            "user_sessions",
            "device_registry",
            "uploaded_files",
            "lottery_tickets",
            "archived_lottery_tickets",
            "winning_records",
            "result_files",
            "match_results",
            "audit_logs",
            "system_settings",
        }.issubset(tables)


def test_create_app_normalizes_relative_sqlite_path(monkeypatch):
    monkeypatch.setenv("DATABASE_URL", "sqlite:///single.db")
    monkeypatch.setenv("FLASK_ENV", "development")
    monkeypatch.setattr(scheduler_module, "start_scheduler", lambda app: None)

    app = create_app()
    uri = app.config["SQLALCHEMY_DATABASE_URI"]
    assert uri.startswith("sqlite:///")
    assert uri.endswith("/instance/single.db")


def test_parse_filename_accepts_optional_trailing_parameter():
    parsed = parse_filename("\u81ea_P5\u80dc\u5e73\u8d1f3\u500d\u6295_\u91d1\u989d600\u5143_37\u5f20_01.40_26034_\u5c0f.txt")

    assert parsed["identifier"] == "\u81ea"
    assert parsed["internal_code"] == "P5"
    assert parsed["lottery_type"] == "\u80dc\u5e73\u8d1f"
    assert parsed["multiplier"] == 3
    assert parsed["declared_amount"] == 600
    assert parsed["declared_count"] == 37
    assert parsed["deadline_hhmm"] == "01.40"
    assert parsed["detail_period"] == "26034"
    assert parsed["extra_param"] == "\u5c0f"


def test_database_display_info_uses_runtime_sqlite_path(app):
    with app.app_context():
        info = _database_display_info()
        assert info["engine"] == "sqlite"
        assert "test.sqlite" in info["path"]


def test_admin_create_user_rejects_invalid_client_mode(app, client):
    with app.app_context():
        admin = User(username="admin_invalid_create_mode", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_invalid_create_mode", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/users",
        json={"username": "bad_mode_user", "password": "secret123", "client_mode": "desktop"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_admin_create_user_rejects_short_password(app, client):
    with app.app_context():
        admin = User(username="admin_short_create_pwd", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_short_create_pwd", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/users",
        json={"username": "short_pwd_user", "password": "12345", "client_mode": "mode_a"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_admin_create_user_accepts_desktop_only_b_mode_flag(app, client):
    with app.app_context():
        admin = User(username="admin_create_desktop_only_flag", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_create_desktop_only_flag", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/users",
        json={
            "username": "mode_b_web_allowed_user",
            "password": "secret123",
            "client_mode": "mode_b",
            "desktop_only_b_mode": False,
        },
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["user"]["desktop_only_b_mode"] is False

    with app.app_context():
        created = User.query.filter_by(username="mode_b_web_allowed_user").first()
        assert created is not None
        assert created.desktop_only_b_mode is False


def test_admin_user_list_sorts_by_client_mode(app, client):
    with app.app_context():
        admin = User(username="admin_users_mode_sort", is_admin=True)
        admin.set_password("secret123")
        mode_b_user = create_user("users_mode_b_first", "secret123", client_mode="mode_b")
        mode_a_user = create_user("users_mode_a_second", "secret123", client_mode="mode_a")
        mode_b_user.created_at = datetime(2026, 4, 7, 10, 0, 0)
        mode_a_user.created_at = datetime(2026, 4, 7, 11, 0, 0)
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_users_mode_sort", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/users")
    assert resp.status_code == 200
    data = resp.get_json()
    assert [u["username"] for u in data["users"][:2]] == ["users_mode_a_second", "users_mode_b_first"]


def test_admin_create_user_rejects_invalid_desktop_only_b_mode_flag(app, client):
    with app.app_context():
        admin = User(username="admin_create_bad_desktop_only", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_create_bad_desktop_only", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/users",
        json={
            "username": "bad_desktop_only_flag_user",
            "password": "secret123",
            "client_mode": "mode_b",
            "desktop_only_b_mode": "not-bool",
        },
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_admin_update_user_rejects_invalid_client_mode(app, client):
    with app.app_context():
        admin = User(username="admin_invalid_update_mode", is_admin=True)
        admin.set_password("secret123")
        user = create_user("update_bad_mode_user", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()
        user_id = user.id

    resp = client.post("/auth/login", json={"username": "admin_invalid_update_mode", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put(f"/admin/api/users/{user_id}", json={"client_mode": "desktop"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]

    with app.app_context():
        refreshed_user = User.query.get(user_id)
        assert refreshed_user.client_mode == "mode_a"


def test_admin_update_user_rejects_short_password(app, client):
    with app.app_context():
        admin = User(username="admin_short_update_pwd", is_admin=True)
        admin.set_password("secret123")
        user = create_user("short_update_pwd_user", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()
        user_id = user.id

    resp = client.post("/auth/login", json={"username": "admin_short_update_pwd", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put(f"/admin/api/users/{user_id}", json={"password": "12345"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_admin_update_user_parses_string_boolean_flags(app, client):
    with app.app_context():
        admin = User(username="admin_bool_flag_update", is_admin=True)
        admin.set_password("secret123")
        user = create_user("bool_flag_user", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()
        user_id = user.id

    resp = client.post("/auth/login", json={"username": "admin_bool_flag_update", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put(f"/admin/api/users/{user_id}", json={"is_active": "false", "can_receive": "0"})
    assert resp.status_code == 200

    with app.app_context():
        refreshed_user = User.query.get(user_id)
        assert refreshed_user.is_active is False
        assert refreshed_user.can_receive is False


def test_admin_update_user_parses_desktop_only_b_mode_flag(app, client):
    with app.app_context():
        admin = User(username="admin_update_desktop_only_flag", is_admin=True)
        admin.set_password("secret123")
        user = create_user("update_desktop_only_flag_user", "secret123", client_mode="mode_b")
        db.session.add(admin)
        db.session.commit()
        user_id = user.id

    resp = client.post("/auth/login", json={"username": "admin_update_desktop_only_flag", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put(f"/admin/api/users/{user_id}", json={"desktop_only_b_mode": "false"})
    assert resp.status_code == 200

    with app.app_context():
        refreshed_user = User.query.get(user_id)
        assert refreshed_user.desktop_only_b_mode is False


def test_admin_disabling_user_forces_logout_existing_sessions(app, client):
    notifications = []

    with app.app_context():
        admin = User(username="admin_disable_user", is_admin=True)
        admin.set_password("secret123")
        user = create_user("disable_me_user", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()
        user_id = user.id

        from services.session_service import create_session

        create_session(user, device_id="device-a", ip_address="127.0.0.1")

    resp = client.post("/auth/login", json={"username": "admin_disable_user", "password": "secret123"})
    assert resp.status_code == 200

    with pytest.MonkeyPatch.context() as mp:
        mp.setattr("services.notify_service.notify_user", lambda user_id, event, data: notifications.append((user_id, event, data)))
        resp = client.put(f"/admin/api/users/{user_id}", json={"is_active": False})

    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True

    with app.app_context():
        refreshed_user = User.query.get(user_id)
        assert refreshed_user.is_active is False
        from models.user import UserSession
        assert UserSession.query.filter_by(user_id=user_id).count() == 0
        disable_logs = AuditLog.query.filter_by(action_type='force_logout', resource_id=str(user_id)).all()
        assert disable_logs

    assert notifications
    assert notifications[0][0] == user_id
    assert notifications[0][1] == 'force_logout'
    assert '禁用' in notifications[0][2]['reason']


def test_admin_toggle_can_receive_rejects_invalid_boolean(app, client):
    with app.app_context():
        admin = User(username="admin_toggle_invalid_bool", is_admin=True)
        admin.set_password("secret123")
        user = create_user("toggle_invalid_bool_user", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()
        user_id = user.id

    resp = client.post("/auth/login", json={"username": "admin_toggle_invalid_bool", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put(f"/admin/api/users/{user_id}/can-receive", json={"can_receive": "not-bool"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]

    with app.app_context():
        refreshed_user = User.query.get(user_id)
        assert refreshed_user.can_receive is True


def test_admin_toggle_can_receive_rejects_admin_account(app, client):
    with app.app_context():
        admin = User(username="admin_toggle_target_admin", is_admin=True)
        admin.set_password("secret123")
        operator = User(username="admin_toggle_operator", is_admin=True)
        operator.set_password("secret123")
        db.session.add_all([admin, operator])
        db.session.commit()
        admin_id = admin.id

    resp = client.post("/auth/login", json={"username": "admin_toggle_operator", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put(f"/admin/api/users/{admin_id}/can-receive", json={"can_receive": False})
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "\u4e0d\u5141\u8bb8\u5728\u6b64\u63a5\u53e3\u4fee\u6539\u7ba1\u7406\u5458\u8d26\u53f7" in data["error"]


def test_admin_user_management_routes_require_login_json_response(app, client):
    resp = client.put("/admin/api/users/1", json={"can_receive": False})
    assert resp.status_code == 401
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert "\u8bf7\u5148\u767b\u5f55" in data["error"]

    resp = client.delete("/admin/api/users/1")
    assert resp.status_code == 401
    assert resp.is_json is True

    resp = client.post("/admin/api/users/1/force-logout")
    assert resp.status_code == 401
    assert resp.is_json is True

    resp = client.put("/admin/api/users/1/can-receive", json={"can_receive": False})
    assert resp.status_code == 401
    assert resp.is_json is True


def test_admin_api_returns_json_when_db_session_token_missing(app, client):
    with app.app_context():
        admin = User(username="admin_missing_db_session_token", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_missing_db_session_token", "password": "secret123"})
    assert resp.status_code == 200

    with client.session_transaction() as sess:
        sess.pop("session_token", None)

    resp = client.get("/admin/api/users")
    assert resp.status_code == 401
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False


def test_admin_user_management_routes_return_json_404_for_missing_user(app, client):
    with app.app_context():
        admin = User(username="admin_user_api_missing_target", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_user_api_missing_target", "password": "secret123"})
    assert resp.status_code == 200

    for method, path, payload in [
        ("put", "/admin/api/users/999999", {"can_receive": False}),
        ("delete", "/admin/api/users/999999", None),
        ("post", "/admin/api/users/999999/force-logout", None),
        ("put", "/admin/api/users/999999/can-receive", {"can_receive": False}),
    ]:
        kwargs = {"json": payload} if payload is not None else {}
        resp = getattr(client, method)(path, **kwargs)
        assert resp.status_code == 404
        assert resp.is_json is True
        data = resp.get_json()
        assert data["success"] is False
        assert data["error"]


def test_admin_core_json_routes_require_login_json_response(app, client):
    routes = [
        ("get", "/admin/api/dashboard-data"),
        ("get", "/admin/api/users"),
        ("get", "/admin/api/lottery-types"),
        ("post", "/admin/api/users"),
        ("get", "/admin/api/winning/filter-options"),
        ("get", "/admin/api/winning"),
        ("post", "/admin/api/winning/1/presign"),
        ("post", "/admin/api/winning/record"),
        ("post", "/admin/api/winning/1/upload-image"),
        ("get", "/admin/api/match-results"),
        ("get", "/admin/api/match-results/1/detail"),
        ("get", "/admin/api/match-results/1/export-comparison"),
        ("post", "/admin/api/match-results/1/recalc"),
        ("get", "/admin/api/settings"),
        ("put", "/admin/api/settings"),
    ]

    for method, path in routes:
        resp = getattr(client, method)(path)
        assert resp.status_code == 401, path
        assert resp.is_json is True, path
        data = resp.get_json()
        assert data["success"] is False, path
        assert "\u8bf7\u5148\u767b\u5f55" in data["error"], path


def test_admin_match_result_upload_requires_login_json_response(app, client):
    resp = client.post("/admin/match-results/upload")
    assert resp.status_code == 401
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert "\u8bf7\u5148\u767b\u5f55" in data["error"]


def test_admin_match_result_routes_return_json_404_for_missing_result(app, client):
    with app.app_context():
        admin = User(username="admin_match_result_missing", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_match_result_missing", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/match-results/999999/detail")
    assert resp.status_code == 404
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]

    resp = client.get("/admin/api/match-results/999999/export-comparison")
    assert resp.status_code == 404
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]

    resp = client.post("/admin/api/match-results/999999/recalc")
    assert resp.status_code == 404
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_admin_toggle_can_receive_pushes_pool_refresh(app, client, monkeypatch):
    pushed = []

    def fake_notify_pool_update(payload):
        pushed.append(payload)

    monkeypatch.setattr("services.notify_service.notify_pool_update", fake_notify_pool_update)

    with app.app_context():
        admin = User(username="admin_toggle_receive_pool_refresh", is_admin=True)
        admin.set_password("secret123")
        user = create_user("toggle_receive_target", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="TOGGLE-RECEIVE-TICKET-1",
            status="pending",
            lottery_type="???",
            deadline_time=beijing_now() + timedelta(hours=1),
        ))
        db.session.commit()
        user_id = user.id

    resp = client.post("/auth/login", json={"username": "admin_toggle_receive_pool_refresh", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put(f"/admin/api/users/{user_id}/can-receive", json={"can_receive": False})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["can_receive"] is False
    assert len(pushed) == 1
    assert "total_pending" in pushed[0]


def test_admin_update_settings_parses_boolean_flags(app, client):
    with app.app_context():
        admin = User(username="admin_settings_bool_flags", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_bool_flags", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={
        "pool_enabled": "false",
        "mode_a_enabled": "0",
        "mode_b_enabled": "true",
        "announcement_enabled": "1",
    })
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    settings = data["settings"]
    assert settings["pool_enabled"] is False
    assert settings["mode_a_enabled"] is False
    assert settings["mode_b_enabled"] is True
    assert settings["announcement_enabled"] is True


def test_admin_update_settings_emits_pool_toggle_events(app, client, monkeypatch):
    emitted = []

    def fake_notify(event, payload):
        emitted.append((event, payload))

    monkeypatch.setattr("routes.admin.notify_all", fake_notify)

    with app.app_context():
        admin = User(username="admin_settings_pool_events", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_pool_events", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"pool_enabled": False})
    assert resp.status_code == 200
    resp = client.put("/admin/api/settings", json={"pool_enabled": True})
    assert resp.status_code == 200

    assert any(event == "pool_disabled" for event, _payload in emitted)
    assert any(event == "pool_enabled" for event, _payload in emitted)


def test_admin_update_settings_notify_failures_do_not_break_success_response(app, client, monkeypatch):
    def failing_notify(*_args, **_kwargs):
        raise RuntimeError("notify unavailable")

    monkeypatch.setattr("routes.admin.notify_all", failing_notify)

    with app.app_context():
        admin = User(username="admin_settings_notify_failure_guard", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_notify_failure_guard", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"pool_enabled": False})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True

    with app.app_context():
        settings = SystemSettings.get()
        assert settings.pool_enabled is False


def test_admin_update_settings_pushes_pool_refresh_when_mode_switches(app, client, monkeypatch):
    pushed = []

    def fake_notify_pool_update(payload):
        pushed.append(payload)

    monkeypatch.setattr("services.notify_service.notify_pool_update", fake_notify_pool_update)

    with app.app_context():
        admin = User(username="admin_settings_mode_refresh", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SETTINGS-MODE-REFRESH-1",
            status="pending",
            lottery_type="???",
            deadline_time=beijing_now() + timedelta(hours=1),
        ))
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_mode_refresh", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"mode_b_enabled": False})
    assert resp.status_code == 200
    resp = client.put("/admin/api/settings", json={"mode_a_enabled": False})
    assert resp.status_code == 200

    assert len(pushed) == 2
    assert all("total_pending" in payload for payload in pushed)


def test_admin_update_user_pushes_pool_refresh_for_receive_and_filter_changes(app, client, monkeypatch):
    pushed = []

    def fake_notify_pool_update(payload):
        pushed.append(payload)

    monkeypatch.setattr("services.notify_service.notify_pool_update", fake_notify_pool_update)

    with app.app_context():
        admin = User(username="admin_update_user_pool_refresh", is_admin=True)
        admin.set_password("secret123")
        user = create_user("pool_refresh_target", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="POOL-REFRESH-TICKET-1",
            status="pending",
            lottery_type="???",
            deadline_time=beijing_now() + timedelta(hours=1),
        ))
        db.session.commit()
        user_id = user.id

    resp = client.post("/auth/login", json={"username": "admin_update_user_pool_refresh", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put(f"/admin/api/users/{user_id}", json={"can_receive": False})
    assert resp.status_code == 200
    resp = client.put(f"/admin/api/users/{user_id}", json={"blocked_lottery_types": ["???"]})
    assert resp.status_code == 200
    resp = client.put(f"/admin/api/users/{user_id}", json={"client_mode": "mode_b"})
    assert resp.status_code == 200

    assert len(pushed) == 3
    assert all("total_pending" in payload for payload in pushed)


def test_admin_update_settings_rejects_invalid_boolean_flag(app, client):
    with app.app_context():
        admin = User(username="admin_settings_invalid_bool", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_invalid_bool", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"pool_enabled": "not-bool"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_process_uploaded_file_returns_filename_on_success_and_failure(app, monkeypatch):
    from io import BytesIO
    from services import file_parser

    with app.app_context():
        user = create_user("upload_filename_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-upload-success.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )
        good_result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_600_47_00.55_26034.txt", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )
        assert good_result["success"] is True
        assert good_result["filename"] == "AA_P7TEST_600_47_00.55_26034.txt"

        bad_result = file_parser.process_uploaded_file(
            FileStorage(stream=BytesIO(b"content\n"), filename="bad-name.txt"),
            uploader_id=user.id,
        )
        assert bad_result["success"] is False
        assert bad_result["filename"] == "bad-name.txt"


def test_process_uploaded_file_marks_overdue_tickets_expired_on_import(app, monkeypatch):
    from services import file_parser

    fixed_now = datetime(2026, 4, 7, 1, 0, 0)
    monkeypatch.setattr(file_parser, "beijing_now", lambda: fixed_now)
    monkeypatch.setattr(
        file_parser,
        "build_uploaded_txt_relative_path",
        lambda filename, upload_dt=None: "txt/2026-04-06/mock-overdue.txt",
    )
    monkeypatch.setattr(
        file_parser,
        "parse_filename",
        lambda filename, upload_dt=None: {
            "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 3,
                "declared_amount": 12.0,
                "declared_count": 2,
                "deadline_hhmm": "00.55",
                "deadline_time": datetime(2026, 4, 7, 0, 55, 0),
                "detail_period": "26034",
            },
    )

    with app.app_context():
        user = create_user("upload_overdue_user", "secret123", client_mode="mode_b")
        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_600_47_00.55_26034.txt", "SPF|1=3|1*1|3\nSPF|1=0|1*1|3\n"),
            uploader_id=user.id,
        )
        assert result["success"] is True
        assert result["ticket_count"] == 2
        assert result["pending_ticket_count"] == 0
        assert result["expired_ticket_count"] == 2

        uploaded = UploadedFile.query.get(result["file_id"])
        tickets = LotteryTicket.query.filter_by(source_file_id=uploaded.id).all()

        assert uploaded.pending_count == 0
        assert uploaded.assigned_count == 0
        assert uploaded.completed_count == 0
        assert {ticket.status for ticket in tickets} == {"expired"}


def test_process_uploaded_file_rejects_invalid_ticket_line_without_partial_import(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_invalid_line_user", "secret123", client_mode="mode_b")
        before_files = UploadedFile.query.count()
        before_tickets = LotteryTicket.query.count()

        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-invalid-line.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 8.0,
                "declared_count": 2,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file(
                "AA_P7TEST_600_47_00.55_26034.txt",
                "SPF|1=3|1*1|2\nbad-line\n",
            ),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert "2" in result["message"]
        assert UploadedFile.query.count() == before_files
        assert LotteryTicket.query.count() == before_tickets


def test_process_uploaded_file_rejects_unknown_text_encoding_cleanly(app, monkeypatch):
    from io import BytesIO
    from werkzeug.datastructures import FileStorage
    from services import file_parser

    with app.app_context():
        user = create_user("upload_bad_encoding_user", "secret123", client_mode="mode_b")
        before_count = UploadedFile.query.count()
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-bad-encoding.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )
        result = file_parser.process_uploaded_file(
            FileStorage(
                stream=BytesIO(b"\x80\x80\x80"),
                filename="AA_P7TEST_600_47_00.55_26034.txt",
            ),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert result["filename"] == "AA_P7TEST_600_47_00.55_26034.txt"
        assert "UTF-8" in result["message"]
        assert "GBK" in result["message"]
        assert UploadedFile.query.count() == before_count


def test_process_uploaded_file_rejects_same_business_day_duplicate_filename(app, monkeypatch):
    from services import file_parser
    from services.file_parser import resolve_uploaded_txt_path
    first_now = datetime(2026, 4, 7, 13, 0, 0)
    second_now = datetime(2026, 4, 7, 13, 5, 0)
    calls = {"count": 0}

    def fake_now():
        calls["count"] += 1
        return first_now if calls["count"] == 1 else second_now

    monkeypatch.setattr(file_parser, "beijing_now", fake_now)
    monkeypatch.setattr(
        file_parser,
        "build_uploaded_txt_relative_path",
        lambda filename, upload_dt=None: "txt/2026-04-07/mock-duplicate.txt",
    )
    monkeypatch.setattr(
        file_parser,
        "parse_filename",
        lambda filename, upload_dt=None: {
            "identifier": "AA",
            "internal_code": "P7",
            "lottery_type": "???",
            "multiplier": 2,
            "declared_amount": 4.0,
            "declared_count": 1,
            "deadline_hhmm": "23.55",
            "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
            "detail_period": "26034",
        },
    )

    with app.app_context():
        user = create_user("upload_duplicate_name_user", "secret123", client_mode="mode_b")
        first_result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_600_47_00.55_26034.txt", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )
        assert first_result["success"] is True

        second_result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_600_47_00.55_26034.txt", "SPF|1=0|1*1|2\n"),
            uploader_id=user.id,
        )
        assert second_result["success"] is False
        assert "当前业务日内已上传同名文件" in second_result["message"]

        uploaded = db.session.get(UploadedFile, first_result["file_id"])
        stored_path = resolve_uploaded_txt_path(uploaded.stored_filename, app.config["UPLOAD_FOLDER"])

        assert UploadedFile.query.count() == 1
        assert os.path.exists(stored_path)
        with open(stored_path, "r", encoding="utf-8") as f:
            assert f.read() == "SPF|1=3|1*1|2\n"


def test_process_uploaded_file_rejects_case_only_duplicate_filename_same_business_day(app, monkeypatch):
    from services import file_parser
    first_now = datetime(2026, 4, 7, 13, 0, 0)
    second_now = datetime(2026, 4, 7, 13, 5, 0)
    calls = {"count": 0}

    def fake_now():
        calls["count"] += 1
        return first_now if calls["count"] == 1 else second_now

    monkeypatch.setattr(file_parser, "beijing_now", fake_now)
    monkeypatch.setattr(
        file_parser,
        "build_uploaded_txt_relative_path",
        lambda filename, upload_dt=None: "txt/2026-04-07/mock-duplicate-case.txt",
    )
    monkeypatch.setattr(
        file_parser,
        "parse_filename",
        lambda filename, upload_dt=None: {
            "identifier": "AA",
            "internal_code": "P7",
            "lottery_type": "???",
            "multiplier": 2,
            "declared_amount": 4.0,
            "declared_count": 1,
            "deadline_hhmm": "23.55",
            "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
            "detail_period": "26034",
        },
    )

    with app.app_context():
        user = create_user("upload_duplicate_case_user", "secret123", client_mode="mode_b")
        first_result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_600_47_00.55_26034.TXT", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )
        assert first_result["success"] is True

        second_result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_600_47_00.55_26034.txt", "SPF|1=0|1*1|2\n"),
            uploader_id=user.id,
        )
        assert second_result["success"] is False
        assert "当前业务日内已上传同名文件" in second_result["message"]

        assert UploadedFile.query.count() == 1


def test_process_uploaded_file_rejects_declared_count_mismatch(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_count_mismatch_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-count-mismatch.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 2,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_4_2_00.55_26034.txt", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert result["message"]
        assert result["message"]


def test_process_uploaded_file_rejects_declared_amount_mismatch(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_amount_mismatch_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-amount-mismatch.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 8.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_8_1_00.55_26034.txt", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert result["message"]
        assert result["message"]


def test_process_uploaded_file_rejects_lottery_type_mismatch(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_type_mismatch_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-type-mismatch.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P11",
                "lottery_type": "\u6bd4\u5206",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P11\u6bd4\u52062\u500d\u6295_\u91d1\u989d4\u5143_1\u5f20_00.55_26034.txt", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert result["message"]


def test_process_uploaded_file_rejects_unsupported_lottery_type(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_unsupported_type_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-unsupported-type.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P99",
                "lottery_type": "????",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P99????鍊嶆姇_閲戦4鍏僟1寮燺00.55_26034.txt", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert "\u6587\u4ef6\u540d\u5f69\u79cd\u4e0d\u652f\u6301" in result["message"]


def test_process_uploaded_file_rejects_multiplier_mismatch(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_multiplier_mismatch_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-multiplier-mismatch.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 3,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7???鍊嶆姇_閲戦4鍏僟1寮燺00.55_26034.txt", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert result["message"]


def test_process_uploaded_file_rejects_malformed_field_segment(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_bad_field_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-bad-field.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7???鍊嶆姇_閲戦4鍏僟1寮燺00.55_26034.txt", "SPF|garbage|1*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert "\u5185\u5bb9\u683c\u5f0f\u65e0\u6548" in result["message"]


def test_process_uploaded_file_rejects_invalid_base_segment(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_bad_base_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-bad-base.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7???鍊嶆姇_閲戦4鍏僟1寮燺00.55_26034.txt", "SPF|1=3|9*2|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert "\u5185\u5bb9\u683c\u5f0f\u65e0\u6548" in result["message"]


def test_process_uploaded_file_rejects_non_positive_final_multiplier(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_non_positive_multiplier_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-non-positive-final-multiplier.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        zero_result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7???鍊嶆姇_閲戦4鍏僟1寮燺00.55_26034.txt", "SPF|1=3|1*1|0\n"),
            uploader_id=user.id,
        )
        negative_result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7???鍊嶆姇_閲戦4鍏僟1寮燺00.55_26034.txt", "SPF|1=3|1*1|-2\n"),
            uploader_id=user.id,
        )

        assert zero_result["success"] is False
        assert negative_result["success"] is False
        assert "\u5185\u5bb9\u683c\u5f0f\u65e0\u6548" in zero_result["message"]
        assert "\u5185\u5bb9\u683c\u5f0f\u65e0\u6548" in negative_result["message"]


def test_process_uploaded_file_rejects_non_numeric_field_number(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_non_numeric_field_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-non-numeric-field.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7???鍊嶆姇_閲戦4鍏僟1寮燺00.55_26034.txt", "SPF|A=3|1*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert "\u5185\u5bb9\u683c\u5f0f\u65e0\u6548" in result["message"]


def test_process_uploaded_file_rejects_duplicate_field_number(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_duplicate_field_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-duplicate-field.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7???鍊嶆姇_閲戦4鍏僟1寮燺00.55_26034.txt", "SPF|1=3,1=0|2*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert "\u5185\u5bb9\u683c\u5f0f\u65e0\u6548" in result["message"]


def test_process_uploaded_file_rejects_duplicate_option_in_field(app, monkeypatch):
    from services import file_parser

    with app.app_context():
        user = create_user("upload_duplicate_option_user", "secret123", client_mode="mode_b")
        monkeypatch.setattr(
            file_parser,
            "build_uploaded_txt_relative_path",
            lambda filename, upload_dt=None: "txt/2026-04-07/mock-duplicate-option.txt",
        )
        monkeypatch.setattr(
            file_parser,
            "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 8.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
        )

        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7???鍊嶆姇_閲戦8鍏僟1寮燺00.55_26034.txt", "SPF|1=3/3,2=0|2*1|2\n"),
            uploader_id=user.id,
        )

        assert result["success"] is False
        assert result["file_id"] is None
        assert "\u5185\u5bb9\u683c\u5f0f\u65e0\u6548" in result["message"]


def test_admin_file_upload_returns_http_400_when_all_files_fail(app, client, monkeypatch):
    import routes.admin as admin_routes
    from services import notify_service

    notified = {"count": 0}

    def fake_process_uploaded_file(file, uploader_id):
        return {
            "success": False,
            "filename": file.filename,
            "message": "当前业务日内已上传同名文件",
        }

    def fake_notify_pool_update(_payload):
        notified["count"] += 1

    monkeypatch.setattr(admin_routes, "process_uploaded_file", fake_process_uploaded_file)
    monkeypatch.setattr(notify_service, "notify_pool_update", fake_notify_pool_update)

    with app.app_context():
        admin = User(username="admin_upload_all_fail", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_upload_all_fail", "password": "secret123"})
    assert resp.status_code == 200

    upload = client.post(
        "/admin/files/upload",
        data={"files": (make_upload_file("mock-upload.txt", "3\n1\n"), "mock-upload.txt")},
        content_type="multipart/form-data",
    )
    assert upload.status_code == 400
    data = upload.get_json()
    assert data["success"] is False
    assert data["error"] == "本次上传全部失败"
    assert data["results"][0]["success"] is False
    assert data["results"][0]["filename"] == "mock-upload.txt"
    assert "同名文件" in data["results"][0]["message"]
    assert notified["count"] == 0


def test_admin_file_upload_keeps_batch_running_when_one_file_raises(app, client, monkeypatch):
    import routes.admin as admin_routes
    from services import notify_service

    notified = {"count": 0}

    def fake_process_uploaded_file(file, uploader_id):
        if file.filename == "boom.txt":
            raise RuntimeError("\u89e3\u6790\u5f02\u5e38")
        return {
            "success": True,
            "filename": file.filename,
            "file_id": 123,
            "message": "ok",
        }

    def fake_notify_pool_update(_payload):
        notified["count"] += 1

    monkeypatch.setattr(admin_routes, "process_uploaded_file", fake_process_uploaded_file)
    monkeypatch.setattr(notify_service, "notify_pool_update", fake_notify_pool_update)

    with app.app_context():
        admin = User(username="admin_upload_partial_exception", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_upload_partial_exception", "password": "secret123"})
    assert resp.status_code == 200

    upload = client.post(
        "/admin/files/upload",
        data={
            "files": [
                (make_upload_file("boom.txt", "1\n"), "boom.txt"),
                (make_upload_file("ok.txt", "2\n"), "ok.txt"),
            ]
        },
        content_type="multipart/form-data",
    )
    assert upload.status_code == 200
    data = upload.get_json()
    assert data["success"] is True
    assert [item["filename"] for item in data["results"]] == ["boom.txt", "ok.txt"]
    assert data["results"][0]["success"] is False
    assert "\u4e0a\u4f20\u5904\u7406\u5931\u8d25" in data["results"][0]["message"]
    assert data["results"][1]["success"] is True
    assert notified["count"] == 1


def test_admin_file_upload_rolls_back_dirty_session_after_real_processing_exception(app, client, monkeypatch):
    from services import file_parser

    original_parse_ticket_line = file_parser.parse_ticket_line

    def flaky_parse_ticket_line(line):
        if line == "BOOM":
            raise RuntimeError("解析炸了")
        return original_parse_ticket_line(line)

    monkeypatch.setattr(file_parser, "parse_ticket_line", flaky_parse_ticket_line)

    with app.app_context():
        admin = User(username="admin_upload_real_exception", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_upload_real_exception", "password": "secret123"})
    assert resp.status_code == 200

    upload = client.post(
        "/admin/files/upload",
        data={
            "files": [
                (make_upload_file("AA_P7胜平负2倍投_金额4元_1张_00.55_26034.txt", "BOOM\n"), "AA_P7胜平负2倍投_金额4元_1张_00.55_26034.txt"),
                (make_upload_file("BB_P7胜平负2倍投_金额4元_1张_00.56_26034.txt", "SPF|1=3|1*1|2\n"), "BB_P7胜平负2倍投_金额4元_1张_00.56_26034.txt"),
            ]
        },
        content_type="multipart/form-data",
    )
    assert upload.status_code == 200
    data = upload.get_json()
    assert data["success"] is True
    assert data["results"][0]["success"] is False
    assert "上传处理失败" in data["results"][0]["message"]
    assert data["results"][1]["success"] is True

    with app.app_context():
        assert UploadedFile.query.filter_by(original_filename="AA_P7胜平负2倍投_金额4元_1张_00.55_26034.txt").count() == 0
        assert UploadedFile.query.filter_by(original_filename="BB_P7胜平负2倍投_金额4元_1张_00.56_26034.txt").count() == 1


def test_admin_file_upload_reports_empty_filename_as_failure(app, client):
    with app.app_context():
        admin = User(username="admin_upload_empty_name", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_upload_empty_name", "password": "secret123"})
    assert resp.status_code == 200

    upload = client.post(
        "/admin/files/upload",
        data={"files": [(io.BytesIO(b"abc"), "")]},
        content_type="multipart/form-data",
    )
    assert upload.status_code == 400
    data = upload.get_json()
    assert data["success"] is False
    assert data["error"] == "\u672c\u6b21\u4e0a\u4f20\u5168\u90e8\u5931\u8d25"
    assert data["results"][0]["success"] is False
    assert data["results"][0]["message"]


def test_admin_delete_user_rejects_user_with_ticket_history(app, client):
    with app.app_context():
        admin = User(username="admin_delete_guard", is_admin=True)
        admin.set_password("secret123")
        user = create_user("delete_guard_user", "secret123", client_mode="mode_a")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="DELETE-GUARD-001",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([admin, ticket])
        db.session.commit()
        user_id = user.id

    resp = client.post("/auth/login", json={"username": "admin_delete_guard", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.delete(f"/admin/api/users/{user_id}")
    assert resp.status_code == 409
    data = resp.get_json()
    assert data["success"] is False
    assert "\u4e0d\u80fd\u76f4\u63a5\u5220\u9664" in data["error"]

    with app.app_context():
        assert User.query.get(user_id) is not None


def create_assigned_ticket(user: User, device_id: str, raw_content: str, line_number: int) -> LotteryTicket:
    ticket = LotteryTicket(
        source_file_id=1,
        line_number=line_number,
        raw_content=raw_content,
        status="assigned",
        assigned_user_id=user.id,
        assigned_username=user.username,
        assigned_device_id=device_id,
        assigned_at=beijing_now(),
    )
    db.session.add(ticket)
    db.session.commit()
    return ticket


def create_pending_ticket(raw_content: str, line_number: int) -> LotteryTicket:
    ticket = LotteryTicket(
        source_file_id=1,
        line_number=line_number,
        raw_content=raw_content,
        status="pending",
        deadline_time=(beijing_now() + timedelta(hours=1)).replace(microsecond=0),
    )
    db.session.add(ticket)
    db.session.commit()
    return ticket


def make_upload_file(filename: str, content: str) -> FileStorage:
    return FileStorage(stream=io.BytesIO(content.encode("utf-8")), filename=filename, content_type="text/plain")


def test_mode_b_processing_requires_device_id(app, client):
    with app.app_context():
        user = create_user("modeb_user", "secret123", client_mode="mode_b")
        create_assigned_ticket(user, "device-a", "A001", 1)
        create_assigned_ticket(user, "device-b", "B001", 2)

    resp = login(client, "modeb_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/processing")
    assert resp.status_code == 400

    data = resp.get_json()
    assert data["success"] is False
    assert "设备ID" in data["error"]


def test_mode_b_processing_with_device_id_filters_batches(app, client):
    with app.app_context():
        user = create_user("modeb_user_filter", "secret123", client_mode="mode_b")
        create_assigned_ticket(user, "device-a", "A001", 1)
        create_assigned_ticket(user, "device-b", "B001", 2)

    resp = login(client, "modeb_user_filter", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/processing?device_id=device-a")
    assert resp.status_code == 200

    data = resp.get_json()
    assert data["success"] is True
    assert len(data["batches"]) == 1
    assert data["batches"][0]["count"] == 1


def test_mode_b_pool_status_reflects_reserve_adjusted_available_counts(app, client):
    with app.app_context():
        user = create_user("modeb_pool_status_user", "secret123", client_mode="mode_b")
        first_deadline = beijing_now() + timedelta(hours=1)
        second_deadline = beijing_now() + timedelta(hours=2)
        tickets = []
        for idx in range(12):
            tickets.append(
                LotteryTicket(
                    source_file_id=1,
                    line_number=idx + 1,
                    raw_content=f"EARLY-{idx}",
                    status="pending",
                    lottery_type="???",
                    deadline_time=first_deadline,
                )
            )
        for idx in range(13):
            tickets.append(
                LotteryTicket(
                    source_file_id=1,
                    line_number=100 + idx,
                    raw_content=f"LATE-{idx}",
                    status="pending",
                    lottery_type="璁╃悆???",
                    deadline_time=second_deadline,
                )
            )
        db.session.add_all(tickets)
        db.session.commit()

    resp = login(client, "modeb_pool_status_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/pool-status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["total_pending"] == 5
    assert len(data["by_type"]) == 1
    assert data["by_type"][0]["lottery_type"]
    assert data["by_type"][0]["count"] == 5


def test_mode_b_pool_status_hides_blocked_lottery_types(app, client):
    with app.app_context():
        user = create_user("modeb_pool_blocked_user", "secret123", client_mode="mode_b")
        user.set_blocked_lottery_types(["???"])
        first_deadline = beijing_now() + timedelta(hours=1)
        second_deadline = beijing_now() + timedelta(hours=2)
        tickets = []
        for idx in range(25):
            tickets.append(
                LotteryTicket(
                    source_file_id=1,
                    line_number=idx + 1,
                    raw_content=f"BLOCKED-{idx}",
                    status="pending",
                    lottery_type="???",
                    deadline_time=first_deadline,
                )
            )
        for idx in range(25):
            tickets.append(
                LotteryTicket(
                    source_file_id=1,
                    line_number=100 + idx,
                    raw_content=f"ALLOWED-{idx}",
                    status="pending",
                    lottery_type="璁╃悆???",
                    deadline_time=second_deadline,
                )
            )
        db.session.add_all(tickets)
        db.session.commit()

    resp = login(client, "modeb_pool_blocked_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/pool-status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["total_pending"] == 5
    assert len(data["by_type"]) == 1
    assert data["by_type"][0]["lottery_type"]
    assert data["by_type"][0]["count"] == 5


def test_mode_a_routes_reject_invalid_device_id_and_name(app, client):
    with app.app_context():
        create_user("mode_a_device_guard_user", "secret123", client_mode="mode_a")

    resp = login(client, "mode_a_device_guard_user", "secret123")
    assert resp.status_code == 200

    invalid_id_resp = client.post("/api/mode-a/next", json={"device_id": "bad id"})
    assert invalid_id_resp.status_code == 400
    assert "\u65e0\u6548\u7684\u8bbe\u5907ID" in invalid_id_resp.get_json()["error"]

    too_long_id_resp = client.post("/api/mode-a/next", json={"device_id": "x" * 65})
    assert too_long_id_resp.status_code == 400
    assert "\u65e0\u6548\u7684\u8bbe\u5907ID" in too_long_id_resp.get_json()["error"]


def test_mode_b_download_rejects_invalid_device_info(app, client):
    with app.app_context():
        create_user("mode_b_device_guard_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_device_guard_user", "secret123")
    assert resp.status_code == 200

    invalid_id_resp = client.post("/api/mode-b/download", json={"count": 1, "device_id": "bad id"})
    assert invalid_id_resp.status_code == 400
    assert "无效的设备ID" in invalid_id_resp.get_json()["error"]

    too_long_id_resp = client.post("/api/mode-b/download", json={"count": 1, "device_id": "x" * 65})
    assert too_long_id_resp.status_code == 400
    assert "无效的设备ID" in too_long_id_resp.get_json()["error"]


def test_mode_b_download_requires_device_id(app, client):
    with app.app_context():
        create_user("mode_b_device_required_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_device_required_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/download", json={"count": 1})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "ID" in data["error"]


def test_mode_b_download_blocks_web_when_desktop_only_enabled(app, client):
    with app.app_context():
        create_user("mode_b_desktop_only_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_desktop_only_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/mode-b/download",
        json={"count": 1, "device_id": "web-1", "client_type": "web"},
    )
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_mode_b_download_allows_web_when_desktop_only_disabled(app, client, monkeypatch):
    called = []

    def fake_download_batch(**kwargs):
        called.append(kwargs)
        return {"success": True, "ticket_ids": [101], "count": 1}

    monkeypatch.setattr("routes.mode_b.download_batch", fake_download_batch)

    with app.app_context():
        user = create_user("mode_b_web_allowed_route_user", "secret123", client_mode="mode_b")
        user.desktop_only_b_mode = False
        db.session.commit()

    resp = login(client, "mode_b_web_allowed_route_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/mode-b/download",
        json={"count": 1, "device_id": "web-1", "client_type": "web"},
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert called
    assert called[0]["device_id"] == "web-1"


def test_mode_b_download_rejects_mismatched_session_device(app, client, monkeypatch):
    called = []

    def fake_download_batch(**kwargs):
        called.append(kwargs)
        return {"success": True, "ticket_ids": [101], "count": 1}

    monkeypatch.setattr("routes.mode_b.download_batch", fake_download_batch)

    with app.app_context():
        create_user("mode_b_session_device_guard_user", "secret123", client_mode="mode_b")

    resp = client.post(
        "/auth/login",
        json={"username": "mode_b_session_device_guard_user", "password": "secret123", "device_id": "device-a"},
    )
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/download", json={"count": 1, "device_id": "device-b"})
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "device_id mismatch" in data["error"]
    assert called == []


def test_mode_b_download_backfills_session_device_then_blocks_switch(app, client, monkeypatch):
    called = []

    def fake_download_batch(**kwargs):
        called.append(kwargs)
        return {"success": True, "ticket_ids": [201], "count": 1}

    monkeypatch.setattr("routes.mode_b.download_batch", fake_download_batch)

    with app.app_context():
        create_user("mode_b_session_device_backfill_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_session_device_backfill_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/download", json={"count": 1, "device_id": "device-a"})
    assert resp.status_code == 200
    assert resp.get_json()["success"] is True

    with client.session_transaction() as sess:
        token = sess.get("session_token")

    with app.app_context():
        from models.user import UserSession

        user_session = UserSession.query.filter_by(session_token=token).first()
        assert user_session is not None
        assert user_session.device_id == "device-a"

    resp = client.post("/api/mode-b/download", json={"count": 1, "device_id": "device-b"})
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "device_id mismatch" in data["error"]
    assert len(called) == 1


def test_mode_b_confirm_rejects_non_integer_ticket_ids(app, client):
    with app.app_context():
        create_user("modeb_confirm_guard_user", "secret123", client_mode="mode_b")

    resp = login(client, "modeb_confirm_guard_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/confirm", json={"ticket_ids": ["abc"], "device_id": "device-a"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "整数" in data["error"]


def test_mode_b_confirm_rejects_boolean_ticket_ids(app, client):
    with app.app_context():
        create_user("modeb_confirm_bool_guard_user", "secret123", client_mode="mode_b")

    resp = login(client, "modeb_confirm_bool_guard_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/confirm", json={"ticket_ids": True, "device_id": "device-a"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_mode_b_confirm_rejects_boolean_elements_in_ticket_ids(app, client):
    with app.app_context():
        create_user("modeb_confirm_bool_item_guard_user", "secret123", client_mode="mode_b")

    resp = login(client, "modeb_confirm_bool_item_guard_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/confirm", json={"ticket_ids": [True], "device_id": "device-a"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_mode_b_confirm_requires_device_id(app, client):
    with app.app_context():
        create_user("modeb_confirm_requires_device_user", "secret123", client_mode="mode_b")

    resp = login(client, "modeb_confirm_requires_device_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/confirm", json={"ticket_ids": [1]})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "ID" in data["error"]


def test_mode_b_confirm_rejects_other_device_tickets(app):
    from services.mode_b_service import confirm_batch

    with app.app_context():
        user = create_user("modeb_confirm_device_scope_user", "secret123", client_mode="mode_b")
        ticket = create_assigned_ticket(user, "device-b", "BATCH-DEVICE-SCOPE-001", 1)

        result = confirm_batch([ticket.id], user_id=user.id, device_id="device-a")
        db.session.expire_all()
        refreshed = db.session.get(LotteryTicket, ticket.id)

    assert result["success"] is False
    assert "\u8bbe\u5907" in result["error"]
    assert refreshed.status == "assigned"


def test_mode_b_processing_keeps_same_minute_batches_separate(app, client):
    with app.app_context():
        user = create_user("modeb_processing_separate_user", "secret123", client_mode="mode_b")
        first_time = datetime(2026, 4, 7, 10, 30, 0, 111111)
        second_time = datetime(2026, 4, 7, 10, 30, 0, 222222)
        deadline = datetime(2026, 4, 7, 18, 0, 0)

        tickets = [
            LotteryTicket(
                source_file_id=1,
                line_number=1,
                raw_content="BATCH-SAME-MIN-001",
                lottery_type="???",
                status="assigned",
                assigned_user_id=user.id,
                assigned_username=user.username,
                assigned_device_id="device-b",
                assigned_at=first_time,
                deadline_time=deadline,
                ticket_amount=2,
            ),
            LotteryTicket(
                source_file_id=1,
                line_number=2,
                raw_content="BATCH-SAME-MIN-002",
                lottery_type="???",
                status="assigned",
                assigned_user_id=user.id,
                assigned_username=user.username,
                assigned_device_id="device-b",
                assigned_at=second_time,
                deadline_time=deadline,
                ticket_amount=2,
            ),
        ]
        db.session.add_all(tickets)
        db.session.commit()

    resp = login(client, "modeb_processing_separate_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/processing?device_id=device-b")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert len(data["batches"]) == 2
    assert all(batch["count"] == 1 for batch in data["batches"])


def test_mode_b_download_uses_unique_assigned_at_per_device_batch(app, monkeypatch):
    from services.mode_b_service import download_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0, 123456)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)

    with app.app_context():
        user = create_user("modeb_unique_assigned_at_user", "secret123", client_mode="mode_b")

        existing_assigned = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="EXISTING-BATCH-001",
            lottery_type="???",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-b",
            assigned_at=fixed_now,
            deadline_time=datetime(2026, 4, 7, 18, 0, 0),
            ticket_amount=2,
        )
        db.session.add(existing_assigned)

        for line_number in range(2, 23):
            db.session.add(LotteryTicket(
                source_file_id=1,
                line_number=line_number,
                raw_content=f"PENDING-BATCH-{line_number:03d}",
                lottery_type="???",
                status="pending",
                deadline_time=datetime(2026, 4, 7, 18, 0, 0),
                ticket_amount=2,
            ))

        db.session.commit()

        result = download_batch(
            user_id=user.id,
            device_id="device-b",
            username=user.username,
            count=1
        )

        assert result["success"] is True
        assigned_ticket_id = result["ticket_ids"][0]
        assigned_ticket = db.session.get(LotteryTicket, assigned_ticket_id)
        assert assigned_ticket.assigned_at > fixed_now
        assert assigned_ticket.assigned_at == fixed_now + timedelta(microseconds=1)


def test_mode_b_download_persists_generated_filename_on_tickets(app, monkeypatch):
    from services.mode_b_service import download_batch

    fixed_now = datetime(2026, 4, 29, 1, 13, 9, 123456)
    monkeypatch.setattr("services.mode_b_service.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)

    with app.app_context():
        user = create_user("modeb_download_filename_user", "secret123", client_mode="mode_b")
        settings = SystemSettings.get()
        settings.mode_b_pool_reserve = 0

        for line_number in range(1, 4):
            db.session.add(LotteryTicket(
                source_file_id=1,
                line_number=line_number,
                raw_content=f"DOWNLOAD-FILENAME-{line_number:03d}",
                lottery_type="比分",
                multiplier=2,
                status="pending",
                deadline_time=datetime(2026, 4, 29, 2, 40, 0),
                ticket_amount=2,
            ))
        db.session.commit()

        result = download_batch(
            user_id=user.id,
            device_id="device-b",
            username=user.username,
            count=3,
        )

        filename = result["files"][0]["filename"]
        assigned_tickets = LotteryTicket.query.filter(
            LotteryTicket.id.in_(result["ticket_ids"])
        ).order_by(LotteryTicket.id).all()

    assert result["success"] is True
    assert filename == "比分_2倍_3张_6元_02.40_2026-0429-011309.txt"
    assert {ticket.download_filename for ticket in assigned_tickets} == {filename}


def test_mode_b_download_returns_no_pool_error_when_below_processing_limit(app):
    from services.mode_b_service import download_batch

    with app.app_context():
        user = create_user("modeb_empty_pool_user", "secret123", client_mode="mode_b")
        user.max_processing_b_mode = 5
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ASSIGNED-ONLY-001",
            lottery_type="???",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-b",
            assigned_at=beijing_now(),
            deadline_time=beijing_now() + timedelta(hours=1),
            ticket_amount=2,
        ))
        db.session.commit()

        result = download_batch(
            user_id=user.id,
            device_id="device-b",
            username=user.username,
            count=1
        )

    assert result["success"] is False
    assert result["error"] == "\u5f53\u524d\u7968\u6c60\u65e0\u53ef\u7528\u7968"


def test_order_tickets_by_id_sequence_preserves_requested_batch_order():
    from services.ticket_pool import _order_tickets_by_id_sequence

    class Ticket:
        def __init__(self, ticket_id):
            self.id = ticket_id

    unordered = [Ticket(3), Ticket(1), Ticket(2)]
    ordered = _order_tickets_by_id_sequence(unordered, [1, 2, 3])

    assert [ticket.id for ticket in ordered] == [1, 2, 3]


def test_mode_b_download_rejects_when_pool_disabled(app):
    from services.mode_b_service import download_batch

    with app.app_context():
        user = create_user("mode_b_pool_disabled_user", "secret123", client_mode="mode_b")
        settings = SystemSettings.get()
        settings.pool_enabled = False
        db.session.commit()

        result = download_batch(
            user_id=user.id,
            device_id="device-b",
            username=user.username,
            count=2,
        )

    assert result["success"] is False
    assert result["error"]


def test_mode_b_preview_returns_zero_when_pool_disabled(app):
    from services.mode_b_service import preview_batch

    with app.app_context():
        settings = SystemSettings.get()
        settings.pool_enabled = False
        db.session.commit()

        result = preview_batch(5)

    assert result == {"available": 0, "requested": 5, "sufficient": False}


def test_mode_b_preview_returns_zero_when_mode_b_disabled(app):
    from services.mode_b_service import preview_batch

    with app.app_context():
        settings = SystemSettings.get()
        settings.mode_b_enabled = False
        db.session.commit()

        result = preview_batch(5)

    assert result == {"available": 0, "requested": 5, "sufficient": False}


def test_mode_b_preview_returns_zero_when_user_cannot_receive(app, client):
    with app.app_context():
        user = create_user("mode_b_preview_paused_user", "secret123", client_mode="mode_b")
        user.can_receive = False
        create_pending_ticket("PREVIEW-PAUSED-001", 1)
        db.session.commit()

    resp = login(client, "mode_b_preview_paused_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/preview?count=1")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["available"] == 0
    assert data["requested"] == 1
    assert data["sufficient"] is False


def test_mode_b_preview_caps_available_by_processing_limit(app, client):
    with app.app_context():
        user = create_user("mode_b_preview_processing_cap_user", "secret123", client_mode="mode_b")
        user.max_processing_b_mode = 3
        db.session.add_all([
            LotteryTicket(
                source_file_id=1,
                line_number=1,
                raw_content="ASSIGNED-CAP-001",
                status="assigned",
                lottery_type="???",
                assigned_user_id=user.id,
                assigned_username=user.username,
                assigned_device_id="device-b",
                assigned_at=beijing_now(),
                deadline_time=beijing_now() + timedelta(hours=2),
            ),
            LotteryTicket(
                source_file_id=1,
                line_number=2,
                raw_content="ASSIGNED-CAP-002",
                status="assigned",
                lottery_type="???",
                assigned_user_id=user.id,
                assigned_username=user.username,
                assigned_device_id="device-c",
                assigned_at=beijing_now(),
                deadline_time=beijing_now() + timedelta(hours=2),
            ),
        ])
        deadline = beijing_now() + timedelta(hours=3)
        for idx in range(1, 31):
            db.session.add(LotteryTicket(
                source_file_id=1,
                line_number=100 + idx,
                raw_content=f"PREVIEW-CAP-{idx}",
                status="pending",
                lottery_type="???",
                deadline_time=deadline,
            ))
        db.session.commit()

    resp = login(client, "mode_b_preview_processing_cap_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/preview?count=10")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["available"] == 1
    assert data["requested"] == 10
    assert data["sufficient"] is False


def test_mode_b_preview_caps_available_by_daily_limit(app, client):
    with app.app_context():
        user = create_user("mode_b_preview_daily_cap_user", "secret123", client_mode="mode_b")
        user.daily_ticket_limit = 2
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="COMPLETED-TODAY-CAP-001",
            status="completed",
            lottery_type="???",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            deadline_time=beijing_now() + timedelta(hours=1),
        ))
        deadline = beijing_now() + timedelta(hours=3)
        for idx in range(1, 31):
            db.session.add(LotteryTicket(
                source_file_id=1,
                line_number=200 + idx,
                raw_content=f"PREVIEW-DAILY-CAP-{idx}",
                status="pending",
                lottery_type="???",
                deadline_time=deadline,
            ))
        db.session.commit()

    resp = login(client, "mode_b_preview_daily_cap_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/preview?count=10")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["available"] == 1
    assert data["requested"] == 10
    assert data["sufficient"] is False


def test_mode_b_preview_excludes_blocked_lottery_types(app, client):
    with app.app_context():
        user = create_user("mode_b_preview_blocked_user", "secret123", client_mode="mode_b")
        user.set_blocked_lottery_types(["???"])
        blocked_deadline = beijing_now() + timedelta(hours=1)
        allowed_deadline = beijing_now() + timedelta(hours=2)
        tickets = []
        for idx in range(30):
            tickets.append(
                LotteryTicket(
                    source_file_id=1,
                    line_number=idx + 1,
                    raw_content=f"PREVIEW-BLOCKED-{idx}",
                    status="pending",
                    lottery_type="???",
                    deadline_time=blocked_deadline,
                )
            )
        for idx in range(25):
            tickets.append(
                LotteryTicket(
                    source_file_id=1,
                    line_number=100 + idx,
                    raw_content=f"PREVIEW-ALLOWED-{idx}",
                    status="pending",
                    lottery_type="璁╃悆???",
                    deadline_time=allowed_deadline,
                )
            )
        db.session.add_all(tickets)
        db.session.commit()

    resp = login(client, "mode_b_preview_blocked_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/preview?count=6")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["available"] == 5
    assert data["requested"] == 6
    assert data["sufficient"] is False


def test_mode_b_preview_matches_single_batch_selection_capacity(app, client):
    with app.app_context():
        user = create_user("mode_b_preview_batch_capacity", "secret123", client_mode="mode_b")
        early_deadline = beijing_now() + timedelta(hours=1)
        later_deadline = beijing_now() + timedelta(hours=2)
        for idx in range(30):
            db.session.add(LotteryTicket(
                source_file_id=1,
                line_number=idx + 1,
                raw_content=f"PREVIEW-EARLY-{idx}",
                status="pending",
                lottery_type="???",
                deadline_time=early_deadline,
            ))
        for idx in range(70):
            db.session.add(LotteryTicket(
                source_file_id=1,
                line_number=100 + idx,
                raw_content=f"PREVIEW-LATER-{idx}",
                status="pending",
                lottery_type="璁╃悆???",
                deadline_time=later_deadline,
            ))
        db.session.commit()

    resp = login(client, "mode_b_preview_batch_capacity", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/preview?count=50")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["available"] == 30
    assert data["requested"] == 50
    assert data["sufficient"] is False


def test_mode_b_download_prefers_daily_limit_error_over_generic_limit_message(app):
    from services.mode_b_service import download_batch

    with app.app_context():
        user = create_user("modeb_daily_limit_user", "secret123", client_mode="mode_b")
        user.max_processing_b_mode = 5
        user.daily_ticket_limit = 1
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="COMPLETED-TODAY-001",
            lottery_type="???",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            deadline_time=beijing_now() + timedelta(hours=1),
            ticket_amount=2,
        ))
        for idx in range(2, 24):
            db.session.add(LotteryTicket(
                source_file_id=1,
                line_number=idx,
                raw_content=f"PENDING-TICKET-{idx:03d}",
                lottery_type="???",
                status="pending",
                deadline_time=beijing_now() + timedelta(hours=1),
                ticket_amount=2,
            ))
        db.session.commit()

        result = download_batch(
            user_id=user.id,
            device_id="device-b",
            username=user.username,
            count=1
        )

    assert result["success"] is False
    assert result["error"]


def test_mode_b_postgres_batch_assignment_uses_consistent_now_guard(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)
    deadline = datetime(2026, 4, 7, 18, 0, 0)
    captured = []

    class FakeResult:
        def __init__(self, *, fetchall_data=None, scalar_value=None, rowcount=1):
            self._fetchall_data = fetchall_data or []
            self._scalar_value = scalar_value
            self.rowcount = rowcount

        def fetchall(self):
            return self._fetchall_data

        def scalar(self):
            return self._scalar_value

    class FakeQuery:
        def filter(self, *args, **kwargs):
            return self

        def all(self):
            return [SimpleNamespace(id=123, assigned_at=fixed_now, deadline_time=deadline)]

    class FakeColumn:
        def in_(self, ids):
            return ids

    class FakeLotteryTicket:
        id = FakeColumn()
        query = FakeQuery()

    def fake_execute(statement, params=None):
        sql = str(statement)
        captured.append((sql, params or {}))
        if "SELECT lottery_type, deadline_time, COUNT(*) as cnt" in sql:
            return FakeResult(fetchall_data=[SimpleNamespace(lottery_type="???", deadline_time=deadline, cnt=25)])
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchall_data=[(123,)])
        if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql:
            return FakeResult(fetchall_data=[(123,)], rowcount=1)
        if "UPDATE uploaded_files" in sql:
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._ensure_unique_batch_assigned_at", lambda user_id, device_id, assigned_at: assigned_at)
    monkeypatch.setattr("services.ticket_pool.LotteryTicket", FakeLotteryTicket)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)

    with app.app_context():
        tickets, adjustment_message = assign_tickets_batch(
            user_id=1,
            device_id="device-b",
            username="tester",
            count=1
        )

    assert adjustment_message is None
    assert len(tickets) == 1

    sql_texts = [sql for sql, _ in captured]
    assert any("deadline_time > :now" in sql for sql in sql_texts if "SELECT lottery_type, deadline_time, COUNT(*) as cnt" in sql)
    assert any("deadline_time > :now" in sql for sql in sql_texts if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql)
    assert any("AND deadline_time > :now" in sql for sql in sql_texts if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql)
    assert all("NOW()" not in sql for sql in sql_texts)


def test_mode_b_postgres_batch_assignment_updates_only_returned_ids(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)
    deadline = datetime(2026, 4, 7, 18, 0, 0)
    captured = []

    class FakeResult:
        def __init__(self, *, fetchall_data=None):
            self._fetchall_data = fetchall_data or []

        def fetchall(self):
            return self._fetchall_data

    class FakeFilterQuery:
        def __init__(self):
            self.ids = []

        def all(self):
            return [SimpleNamespace(id=current_id) for current_id in self.ids]

    class FakeColumn:
        def __init__(self, query):
            self.query = query

        def in_(self, ids):
            self.query.ids = list(ids)
            return ids

    fake_filter_query = FakeFilterQuery()

    class FakeLotteryQuery:
        def filter(self, *args, **kwargs):
            return fake_filter_query

    class FakeLotteryTicket:
        id = FakeColumn(fake_filter_query)
        query = FakeLotteryQuery()

    def fake_execute(statement, params=None):
        sql = str(statement)
        params = params or {}
        captured.append((sql, params))
        if "SELECT lottery_type, deadline_time, COUNT(*) as cnt" in sql:
            return FakeResult(fetchall_data=[SimpleNamespace(lottery_type="???", deadline_time=deadline, cnt=25)])
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchall_data=[(101,), (102,)])
        if "UPDATE lottery_tickets" in sql and "RETURNING id" in sql:
            return FakeResult(fetchall_data=[(102,)])
        if "UPDATE uploaded_files" in sql:
            assert params["ids"] == [102]
            return FakeResult(fetchall_data=[])
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._ensure_unique_batch_assigned_at", lambda user_id, device_id, assigned_at: assigned_at)
    monkeypatch.setattr("services.ticket_pool.LotteryTicket", FakeLotteryTicket)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)

    with app.app_context():
        tickets, adjustment_message = assign_tickets_batch(
            user_id=1,
            device_id="device-b",
            username="tester",
            count=2,
        )

    assert adjustment_message is None
    assert [ticket.id for ticket in tickets] == [102]


def test_mode_a_postgres_assignment_update_uses_deadline_guard(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_ticket_atomic

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)
    captured = []

    class FakeResult:
        def __init__(self, *, fetchone_data=None, rowcount=1):
            self._fetchone_data = fetchone_data
            self.rowcount = rowcount

        def fetchone(self):
            return self._fetchone_data

    class FakeTicketRow:
        def __init__(self):
            self.deadline_time = datetime(2026, 4, 7, 18, 0, 0)
            self.lottery_type = "???"
            self.source_file_id = 1

    def fake_execute(statement, params=None):
        sql = str(statement)
        captured.append((sql, params or {}))
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=(321,))
        if "SELECT * FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=FakeTicketRow())
        if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql:
            assert "deadline_time > :now" in sql
            return FakeResult(rowcount=1)
        if "UPDATE uploaded_files" in sql:
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._get_redis", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.get", lambda model, _id: SimpleNamespace(id=_id, assigned_at=fixed_now))

    with app.app_context():
        result = assign_ticket_atomic(user_id=1, device_id="device-a", username="tester")

    assert result is not None
    assert any("deadline_time > :now" in sql for sql, _ in captured if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql)


def test_mode_a_postgres_assignment_prefers_earliest_deadline(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_ticket_atomic

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)
    select_sql = []

    class FakeResult:
        def __init__(self, *, fetchone_data=None, rowcount=1):
            self._fetchone_data = fetchone_data
            self.rowcount = rowcount

        def fetchone(self):
            return self._fetchone_data

    class FakeTicketRow:
        def __init__(self):
            self.deadline_time = datetime(2026, 4, 7, 19, 55, 0)
            self.lottery_type = "SPF"
            self.source_file_id = 1

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            select_sql.append(sql)
            assert "ORDER BY deadline_time, lottery_type, id" in sql
            return FakeResult(fetchone_data=(321,))
        if "SELECT * FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=FakeTicketRow())
        if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql:
            return FakeResult(rowcount=1)
        if "UPDATE uploaded_files" in sql:
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._get_redis", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.get", lambda model, _id: SimpleNamespace(id=_id, assigned_at=fixed_now))

    with app.app_context():
        result = assign_ticket_atomic(user_id=1, device_id="device-a", username="tester")

    assert result is not None
    assert select_sql


def test_mode_a_postgres_assignment_refreshes_now_between_retries(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_ticket_atomic

    first_now = datetime(2026, 4, 7, 10, 30, 0)
    second_now = datetime(2026, 4, 7, 10, 30, 1)
    now_values = iter([first_now, second_now])
    update_now_values = []
    call_state = {"update_calls": 0}

    class FakeResult:
        def __init__(self, *, fetchone_data=None, rowcount=1):
            self._fetchone_data = fetchone_data
            self.rowcount = rowcount

        def fetchone(self):
            return self._fetchone_data

    class FakeTicketRow:
        def __init__(self):
            self.deadline_time = datetime(2026, 4, 7, 18, 0, 0)
            self.lottery_type = "???"
            self.source_file_id = 1

    def fake_now():
        return next(now_values, second_now)

    def fake_execute(statement, params=None):
        sql = str(statement)
        params = params or {}
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=(321,))
        if "SELECT * FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=FakeTicketRow())
        if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql:
            call_state["update_calls"] += 1
            update_now_values.append(params["now"])
            return FakeResult(rowcount=0 if call_state["update_calls"] == 1 else 1)
        if "UPDATE uploaded_files" in sql:
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool._get_redis", lambda: None)
    monkeypatch.setattr("services.ticket_pool._acquire_postgres_user_assignment_lock", lambda _user_id: None)
    monkeypatch.setattr("services.ticket_pool.beijing_now", fake_now)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.get", lambda model, _id: SimpleNamespace(id=_id, assigned_at=second_now))

    with app.app_context():
        result = assign_ticket_atomic(user_id=1, device_id="device-a", username="tester")

    assert result is not None
    assert result.id == 321
    assert update_now_values == [first_now, second_now]


def test_mode_a_postgres_assignment_clamps_file_pending_count(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_ticket_atomic

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeResult:
        def __init__(self, *, fetchone_data=None, rowcount=1):
            self._fetchone_data = fetchone_data
            self.rowcount = rowcount

        def fetchone(self):
            return self._fetchone_data

    class FakeTicketRow:
        def __init__(self):
            self.deadline_time = datetime(2026, 4, 7, 18, 0, 0)
            self.lottery_type = "\u9473\u6ec3\u94a9\u7490?"
            self.source_file_id = 1

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT pg_advisory_xact_lock" in sql:
            return FakeResult(rowcount=1)
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=(321,))
        if "SELECT * FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=FakeTicketRow())
        if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql:
            return FakeResult(rowcount=1)
        if "UPDATE uploaded_files" in sql and "assigned_count = assigned_count + 1" in sql:
            assert "GREATEST(pending_count - 1, 0)" in sql
            assert "AND pending_count > 0" not in sql
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._get_redis", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.get", lambda model, _id: SimpleNamespace(id=_id, assigned_at=fixed_now))

    with app.app_context():
        result = assign_ticket_atomic(user_id=1, device_id="device-a", username="tester")

    assert result is not None


def test_mode_a_postgres_assignment_falls_back_when_redis_returns_stale_id(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_ticket_atomic

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeRedis:
        def __init__(self):
            self.calls = 0

        def lpop(self, key):
            self.calls += 1
            return "999" if self.calls == 1 else None

    class FakeResult:
        def __init__(self, *, fetchone_data=None, rowcount=1):
            self._fetchone_data = fetchone_data
            self.rowcount = rowcount

        def fetchone(self):
            return self._fetchone_data

    class FakeTicketRow:
        def __init__(self):
            self.deadline_time = datetime(2026, 4, 7, 18, 0, 0)
            self.lottery_type = "???"
            self.source_file_id = 1

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=(321,))
        if "SELECT * FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            if params["id"] == 999:
                return FakeResult(fetchone_data=None)
            assert params["id"] == 321
            return FakeResult(fetchone_data=FakeTicketRow())
        if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql:
            assert params["id"] == 321
            return FakeResult(rowcount=1)
        if "UPDATE uploaded_files" in sql:
            assert params["id"] == 321
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    fake_redis = FakeRedis()
    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._get_redis", lambda: fake_redis)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.get", lambda model, _id: SimpleNamespace(id=_id, assigned_at=fixed_now))

    with app.app_context():
        result = assign_ticket_atomic(user_id=1, device_id="device-a", username="tester")

    assert result is not None
    assert result.id == 321


def test_mode_a_postgres_assignment_expires_redis_ticket_at_exact_deadline(app, monkeypatch):
    from services.ticket_pool import assign_ticket_atomic

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)
    calls = []

    class FakeRedis:
        def lpop(self, key):
            return "555"

    class FakeResult:
        def __init__(self, *, fetchone_data=None, rowcount=1):
            self._fetchone_data = fetchone_data
            self.rowcount = rowcount

        def fetchone(self):
            return self._fetchone_data

    class FakeTicketRow:
        def __init__(self):
            self.deadline_time = fixed_now
            self.lottery_type = "???"
            self.source_file_id = 9

    def fake_execute(statement, params=None):
        sql = str(statement)
        calls.append((sql, params or {}))
        if "SELECT * FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=FakeTicketRow())
        if "UPDATE lottery_tickets SET status='expired'" in sql:
            assert params["id"] == 555
            assert "version = version + 1" in sql
            return FakeResult(rowcount=1)
        if "UPDATE uploaded_files" in sql and "pending_count = CASE" in sql:
            assert params["file_id"] == 9
            return FakeResult(rowcount=1)
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=None)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._get_redis", lambda: FakeRedis())
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)

    with app.app_context():
        result = assign_ticket_atomic(user_id=1, device_id="device-a", username="tester")

    assert result is None
    assert not any("SET status = 'assigned'" in sql for sql, _ in calls)


def test_mode_a_postgres_blocked_fallback_ticket_does_not_duplicate_redis_queue(app, monkeypatch):
    from services.ticket_pool import assign_ticket_atomic

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeRedis:
        def __init__(self):
            self.requeued = []

        def lpop(self, key):
            return None

        def rpush(self, key, value):
            self.requeued.append((key, value))

    class FakeResult:
        def __init__(self, *, fetchone_data=None, rowcount=1):
            self._fetchone_data = fetchone_data
            self.rowcount = rowcount

        def fetchone(self):
            return self._fetchone_data

    class FakeTicketRow:
        def __init__(self):
            self.deadline_time = datetime(2026, 4, 7, 18, 0, 0)
            self.lottery_type = "???"
            self.source_file_id = 1

    fake_redis = FakeRedis()
    db_select_calls = {"count": 0}

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            db_select_calls["count"] += 1
            if db_select_calls["count"] == 1:
                return FakeResult(fetchone_data=(321,))
            return FakeResult(fetchone_data=None)
        if "SELECT * FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchone_data=FakeTicketRow())
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._get_redis", lambda: fake_redis)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)

    with app.app_context():
        result = assign_ticket_atomic(
            user_id=1,
            device_id="device-a",
            username="tester",
            blocked_lottery_types=["??"],
        )

    assert result is None
    assert fake_redis.requeued == []


def test_mode_a_sqlite_assignment_retries_after_guarded_update_miss(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_ticket_atomic

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)
    selected_ids = iter([101, 102])
    update_attempts = []
    rollbacks = []

    class FakeResult:
        def __init__(self, *, fetchone_data=None, rowcount=1):
            self._fetchone_data = fetchone_data
            self.rowcount = rowcount

        def fetchone(self):
            return self._fetchone_data

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT id FROM lottery_tickets" in sql and "LIMIT 1" in sql:
            try:
                return FakeResult(fetchone_data=(next(selected_ids),))
            except StopIteration:
                return FakeResult(fetchone_data=None)
        if "UPDATE lottery_tickets" in sql and "SET status = 'assigned'" in sql:
            update_attempts.append(params["id"])
            return FakeResult(rowcount=0 if params["id"] == 101 else 1)
        if "UPDATE uploaded_files" in sql:
            assert params["id"] == 102
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: False)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._get_redis", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: rollbacks.append(True))
    monkeypatch.setattr(
        "services.ticket_pool.LotteryTicket",
        SimpleNamespace(query=SimpleNamespace(get=lambda ticket_id: SimpleNamespace(id=ticket_id, assigned_at=fixed_now))),
    )

    with app.app_context():
        result = assign_ticket_atomic(user_id=1, device_id="device-a", username="tester")

    assert result is not None
    assert result.id == 102
    assert update_attempts == [101, 102]
    assert len(rollbacks) == 1


def test_mode_b_sqlite_batch_assignment_returns_only_freshly_assigned_tickets(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeRow:
        def __init__(self, value):
            self.value = value

        def __getitem__(self, idx):
            return self.value

    class FakeTicketRecord:
        def __init__(self, ticket_id):
            self.id = ticket_id

    class FakeColumn:
        def __init__(self, name):
            self.name = name

        def __eq__(self, other):
            return (self.name, other)

    class FakeAssignedQuery:
        def all(self):
            return [FakeTicketRecord(202)]

    class FakeLotteryQuery:
        def filter(self, *args, **kwargs):
            assert ("assigned_device_id", "device-b") in args
            return FakeAssignedQuery()

    class FakeLotteryTicket:
        query = FakeLotteryQuery()
        id = SimpleNamespace(in_=lambda ids: ids)
        assigned_user_id = FakeColumn("assigned_user_id")
        assigned_device_id = FakeColumn("assigned_device_id")
        status = FakeColumn("status")
        assigned_at = FakeColumn("assigned_at")

    call_state = {"count": 0}

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT COUNT(*) FROM lottery_tickets WHERE status='pending'" in sql:
            call_state["count"] += 1
            return SimpleNamespace(scalar=lambda: 25)
        if "SELECT lottery_type, deadline_time, COUNT(*) as cnt" in sql:
            return SimpleNamespace(fetchall=lambda: [SimpleNamespace(lottery_type="???", deadline_time=datetime(2026, 4, 7, 18, 0, 0), cnt=25)])
        if "SELECT id FROM lottery_tickets" in sql and "LIMIT :count" in sql:
            return SimpleNamespace(fetchall=lambda: [FakeRow(201), FakeRow(202)])
        if "UPDATE lottery_tickets" in sql and "WHERE id IN" in sql:
            assert "deadline_time > :now" in sql
            return SimpleNamespace(rowcount=1)
        if "UPDATE uploaded_files" in sql:
            assert params["id"] == 202
            return SimpleNamespace(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: False)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._ensure_unique_batch_assigned_at", lambda user_id, device_id, assigned_at: assigned_at)
    monkeypatch.setattr("services.ticket_pool.LotteryTicket", FakeLotteryTicket)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.db.session.rollback", lambda: None)

    with app.app_context():
        tickets, adjustment_message = assign_tickets_batch(
            user_id=1,
            device_id="device-b",
            username="tester",
            count=2,
        )

    assert adjustment_message is None
    assert [ticket.id for ticket in tickets] == [202]


def test_mode_b_postgres_batch_assignment_enforces_max_processing_limit(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)
    deadline = datetime(2026, 4, 7, 18, 0, 0)

    class FakeResult:
        def __init__(self, *, fetchall_data=None, scalar_data=None, rowcount=1):
            self._fetchall_data = fetchall_data or []
            self._scalar_data = scalar_data
            self.rowcount = rowcount

        def fetchall(self):
            return self._fetchall_data

        def scalar(self):
            return self._scalar_data

    class FakeTicketRecord:
        def __init__(self, ticket_id):
            self.id = ticket_id

    class FakeAssignedQuery:
        def all(self):
            return [FakeTicketRecord(123)]

    class FakeLotteryQuery:
        def filter(self, *args, **kwargs):
            return FakeAssignedQuery()

    class FakeLotteryTicket:
        query = FakeLotteryQuery()
        id = SimpleNamespace(in_=lambda ids: ids)

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT COUNT(*) FROM lottery_tickets WHERE assigned_user_id = :user_id AND status = 'assigned'" in sql:
            return FakeResult(scalar_data=4)
        if "SELECT lottery_type, deadline_time, COUNT(*) as cnt" in sql:
            return FakeResult(fetchall_data=[SimpleNamespace(lottery_type="胜平负", deadline_time=deadline, cnt=25)])
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            assert params["count"] == 1
            return FakeResult(fetchall_data=[(123,)])
        if "UPDATE lottery_tickets" in sql and "RETURNING id" in sql:
            return FakeResult(fetchall_data=[(123,)], rowcount=1)
        if "UPDATE uploaded_files f" in sql:
            assert params["ids"] == [123]
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._ensure_unique_batch_assigned_at", lambda user_id, device_id, assigned_at: assigned_at)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.LotteryTicket", FakeLotteryTicket)

    with app.app_context():
        tickets, adjustment_message = assign_tickets_batch(
            user_id=1,
            device_id="device-b",
            username="tester",
            count=3,
            max_processing=5,
        )

    assert [ticket.id for ticket in tickets] == [123]
    assert adjustment_message == "已自动调整为1张（当前处理中4张，上限5张）"


def test_mode_b_postgres_batch_assignment_clamps_file_pending_count(app, monkeypatch):
    from types import SimpleNamespace
    from services.ticket_pool import assign_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)
    deadline = datetime(2026, 4, 7, 18, 0, 0)

    class FakeResult:
        def __init__(self, *, fetchall_data=None, scalar_data=None, rowcount=1):
            self._fetchall_data = fetchall_data or []
            self._scalar_data = scalar_data
            self.rowcount = rowcount

        def fetchall(self):
            return self._fetchall_data

        def scalar(self):
            return self._scalar_data

    class FakeTicketRecord:
        def __init__(self, ticket_id):
            self.id = ticket_id

    class FakeAssignedQuery:
        def all(self):
            return [FakeTicketRecord(123), FakeTicketRecord(124)]

    class FakeLotteryQuery:
        def filter(self, *args, **kwargs):
            return FakeAssignedQuery()

    class FakeLotteryTicket:
        query = FakeLotteryQuery()
        id = SimpleNamespace(in_=lambda ids: ids)

    update_called = {"value": False}

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT lottery_type, deadline_time, COUNT(*) as cnt" in sql:
            return FakeResult(fetchall_data=[SimpleNamespace(lottery_type="??ID??", deadline_time=deadline, cnt=25)])
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            return FakeResult(fetchall_data=[(123,), (124,)])
        if "UPDATE lottery_tickets" in sql and "RETURNING id" in sql:
            return FakeResult(fetchall_data=[(123,), (124,)], rowcount=2)
        if "UPDATE uploaded_files f" in sql and "assigned_count = assigned_count + sub.cnt" in sql:
            update_called["value"] = True
            assert "GREATEST(pending_count - sub.cnt, 0)" in sql
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool._ensure_unique_batch_assigned_at", lambda user_id, device_id, assigned_at: assigned_at)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)
    monkeypatch.setattr("services.ticket_pool.LotteryTicket", FakeLotteryTicket)

    with app.app_context():
        tickets, adjustment_message = assign_tickets_batch(
            user_id=1,
            device_id="device-b",
            username="tester",
            count=2,
        )

    assert adjustment_message is None
    assert [ticket.id for ticket in tickets] == [123, 124]
    assert update_called["value"] is True


def test_mode_b_postgres_complete_updates_file_counts_only_for_completed_ids(app, monkeypatch):
    from services.ticket_pool import complete_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeResult:
        def __init__(self, *, fetchall_data=None):
            self._fetchall_data = fetchall_data or []

        def fetchall(self):
            return self._fetchall_data

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "UPDATE lottery_tickets" in sql and "SET status = 'completed'" in sql:
            assert "RETURNING id" in sql
            return FakeResult(fetchall_data=[(201,)])
        if "UPDATE uploaded_files f" in sql:
            assert params["ids"] == [201]
            return FakeResult(fetchall_data=[])
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)

    with app.app_context():
        result = complete_tickets_batch([201, 202], user_id=1)

    assert result == 1


def test_mode_b_postgres_complete_clamps_assigned_count_when_updating_files(app, monkeypatch):
    from services.ticket_pool import complete_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeResult:
        def __init__(self, *, fetchall_data=None):
            self._fetchall_data = fetchall_data or []

        def fetchall(self):
            return self._fetchall_data

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "UPDATE lottery_tickets" in sql and "SET status = 'completed'" in sql:
            return FakeResult(fetchall_data=[(201,)])
        if "UPDATE uploaded_files f" in sql and "completed_count = completed_count + sub.cnt" in sql:
            assert "GREATEST(assigned_count - sub.cnt, 0)" in sql
            return FakeResult(fetchall_data=[])
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)

    with app.app_context():
        assert complete_tickets_batch([201], user_id=1) == 1


def test_mode_a_postgres_complete_ticket_uses_updated_rowcount(app, monkeypatch):
    from services.ticket_pool import complete_ticket

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeResult:
        def __init__(self, *, rowcount=0):
            self.rowcount = rowcount

    calls = []

    def fake_execute(statement, params=None):
        sql = str(statement)
        calls.append((sql, params or {}))
        if "UPDATE lottery_tickets" in sql and "SET status = 'completed'" in sql:
            return FakeResult(rowcount=1)
        if "UPDATE uploaded_files" in sql and "completed_count = completed_count + 1" in sql:
            assert params["id"] == 501
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)

    with app.app_context():
        assert complete_ticket(501, user_id=7) is True

    assert any("UPDATE lottery_tickets" in sql for sql, _ in calls)
    assert any("UPDATE uploaded_files" in sql for sql, _ in calls)


def test_mode_a_postgres_complete_ticket_clamps_negative_assigned_count(app, monkeypatch):
    from services.ticket_pool import complete_ticket

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeResult:
        def __init__(self, *, rowcount=0):
            self.rowcount = rowcount

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "UPDATE lottery_tickets" in sql and "SET status = 'completed'" in sql:
            return FakeResult(rowcount=1)
        if "UPDATE uploaded_files" in sql and "completed_count = completed_count + 1" in sql:
            assert "WHEN assigned_count > 0 THEN assigned_count - 1" in sql
            return FakeResult(rowcount=1)
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)

    with app.app_context():
        assert complete_ticket(501, user_id=7) is True


def test_mode_a_complete_ticket_repairs_file_completed_count_when_assigned_counter_drifted(app):
    from services.ticket_pool import complete_ticket

    with app.app_context():
        user = create_user("mode_a_counter_repair_user", "secret123", client_mode="mode_a")
        uploaded_file = UploadedFile(
            display_id="2026/04/07-01",
            original_filename="repair_a.txt",
            stored_filename="txt/2026-04-07/repair_a.txt",
            status="active",
            uploaded_by=user.id,
            total_tickets=1,
            pending_count=0,
            assigned_count=0,
            completed_count=0,
        )
        db.session.add(uploaded_file)
        db.session.flush()
        ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="A-REPAIR-001",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_at=beijing_now(),
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id
        file_id = uploaded_file.id

        assert complete_ticket(ticket_id, user.id) is True

        refreshed_ticket = LotteryTicket.query.get(ticket_id)
        refreshed_file = UploadedFile.query.get(file_id)
        assert refreshed_ticket.status == "completed"
        assert refreshed_file.assigned_count == 0
        assert refreshed_file.completed_count == 1


def test_mode_b_finalize_repairs_file_counters_when_assigned_counter_drifted(app):
    from services.ticket_pool import finalize_tickets_batch

    with app.app_context():
        user = create_user("mode_b_counter_repair_user", "secret123", client_mode="mode_b")
        uploaded_file = UploadedFile(
            display_id="2026/04/07-02",
            original_filename="repair_b.txt",
            stored_filename="txt/2026-04-07/repair_b.txt",
            status="active",
            uploaded_by=user.id,
            total_tickets=2,
            pending_count=0,
            assigned_count=0,
            completed_count=0,
        )
        db.session.add(uploaded_file)
        db.session.flush()
        ticket1 = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="B-REPAIR-001",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_at=beijing_now(),
        )
        ticket2 = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=2,
            raw_content="B-REPAIR-002",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_at=beijing_now(),
        )
        db.session.add_all([ticket1, ticket2])
        db.session.commit()

        result = finalize_tickets_batch([ticket1.id, ticket2.id], user.id, completed_count=1)

        refreshed_file = UploadedFile.query.get(uploaded_file.id)
        refreshed_ticket1 = LotteryTicket.query.get(ticket1.id)
        refreshed_ticket2 = LotteryTicket.query.get(ticket2.id)
        assert result == {"completed_count": 1, "expired_count": 1}
        assert refreshed_ticket1.status == "completed"
        assert refreshed_ticket2.status == "expired"
        assert refreshed_file.assigned_count == 0
        assert refreshed_file.completed_count == 1


def test_mode_b_postgres_finalize_updates_counts_only_for_returned_ids(app, monkeypatch):
    from services.ticket_pool import finalize_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeResult:
        def __init__(self, *, fetchall_data=None):
            self._fetchall_data = fetchall_data or []

        def fetchall(self):
            return self._fetchall_data

    update_calls = []

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "UPDATE lottery_tickets" in sql and "SET status = 'completed'" in sql and "RETURNING id" in sql:
            update_calls.append(("completed", params["ids"]))
            return FakeResult(fetchall_data=[(301,)])
        if "UPDATE lottery_tickets" in sql and "SET status = 'expired'" in sql and "RETURNING id" in sql:
            update_calls.append(("expired", params["ids"]))
            return FakeResult(fetchall_data=[(303,)])
        if "UPDATE uploaded_files f" in sql and "completed_count = completed_count + sub.cnt" in sql:
            assert params["ids"] == [301]
            return FakeResult(fetchall_data=[])
        if "UPDATE uploaded_files f" in sql and "SET assigned_count = GREATEST(assigned_count - sub.cnt, 0)" in sql:
            assert params["ids"] == [303]
            return FakeResult(fetchall_data=[])
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)

    with app.app_context():
        result = finalize_tickets_batch([301, 302, 303], user_id=1, completed_count=2)

    assert update_calls == [("completed", [301, 302]), ("expired", [303])]
    assert result == {"completed_count": 1, "expired_count": 1}


def test_mode_b_postgres_finalize_clamps_assigned_count_when_updating_files(app, monkeypatch):
    from services.ticket_pool import finalize_tickets_batch

    fixed_now = datetime(2026, 4, 7, 10, 30, 0)

    class FakeResult:
        def __init__(self, *, fetchall_data=None):
            self._fetchall_data = fetchall_data or []

        def fetchall(self):
            return self._fetchall_data

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "UPDATE lottery_tickets" in sql and "SET status = 'completed'" in sql and "RETURNING id" in sql:
            return FakeResult(fetchall_data=[(301,)])
        if "UPDATE lottery_tickets" in sql and "SET status = 'expired'" in sql and "RETURNING id" in sql:
            return FakeResult(fetchall_data=[(302,)])
        if "UPDATE uploaded_files f" in sql and "completed_count = completed_count + sub.cnt" in sql:
            assert "GREATEST(assigned_count - sub.cnt, 0)" in sql
            return FakeResult(fetchall_data=[])
        if "UPDATE uploaded_files f" in sql and "SET assigned_count = GREATEST(assigned_count - sub.cnt, 0)" in sql:
            return FakeResult(fetchall_data=[])
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: fixed_now)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)
    monkeypatch.setattr("services.ticket_pool.db.session.commit", lambda: None)

    with app.app_context():
        result = finalize_tickets_batch([301, 302], user_id=1, completed_count=1)

    assert result == {"completed_count": 1, "expired_count": 1}


def test_mode_b_endpoints_reject_mode_a_user(app, client):
    with app.app_context():
        create_user("mode_a_blocked_user", "secret123", client_mode="mode_a")

    resp = login(client, "mode_a_blocked_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/pool-status")
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_mode_a_endpoints_reject_mode_b_user(app, client):
    with app.app_context():
        create_user("mode_b_blocked_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_blocked_user", "secret123")
    assert resp.status_code == 200

    responses = [
        client.post("/api/mode-a/next", json={"device_id": "device-a"}),
        client.get("/api/mode-a/current?device_id=device-a"),
        client.post("/api/mode-a/stop", json={"device_id": "device-a"}),
        client.get("/api/mode-a/previous?device_id=device-a"),
    ]

    for resp in responses:
        assert resp.status_code == 403
        data = resp.get_json()
        assert data["success"] is False
        assert data["error"]


def test_mode_a_next_does_not_complete_current_ticket_without_explicit_ticket_id(app, client):
    with app.app_context():
        user = create_user("modea_user", "secret123", client_mode="mode_a")
        current_ticket = create_assigned_ticket(user, "device-a", "CUR001", 1)
        create_pending_ticket("NEXT001", 2)
        current_ticket_id = current_ticket.id

    resp = login(client, "modea_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-a/next", json={"device_id": "device-a"})
    assert resp.status_code == 200

    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["id"] == current_ticket_id

    with app.app_context():
        current = LotteryTicket.query.get(current_ticket_id)
        pending = LotteryTicket.query.filter_by(raw_content="NEXT001").first()
        assert current.status == "assigned"
        assert pending.status == "pending"


def test_mode_a_next_ignores_stale_completion_ticket_id(app, client):
    with app.app_context():
        user = create_user("modea_retry_user", "secret123", client_mode="mode_a")
        first_ticket = create_assigned_ticket(user, "device-a", "CUR001", 1)
        next_ticket = create_pending_ticket("NEXT001", 2)
        first_ticket_id = first_ticket.id
        next_ticket_id = next_ticket.id

    resp = login(client, "modea_retry_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
            "/api/mode-a/next",
            json={
                "device_id": "device-a",
                "complete_current_ticket_id": first_ticket_id,
            },
        )
    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["id"] == next_ticket_id

    resp = client.post(
            "/api/mode-a/next",
            json={
                "device_id": "device-a",
                "complete_current_ticket_id": first_ticket_id,
            },
        )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["id"] == next_ticket_id

    with app.app_context():
        first = LotteryTicket.query.get(first_ticket_id)
        current = LotteryTicket.query.get(next_ticket_id)
        assert first.status == "completed"
        assert current.status == "assigned"


def test_mode_a_next_returns_current_device_today_sequence(app, client):
    with app.app_context():
        user = create_user("modea_sequence_user", "secret123", client_mode="mode_a")
        completed = create_assigned_ticket(user, "device-a", "DONE-001", 1)
        completed.status = "completed"
        completed.completed_at = beijing_now() - timedelta(minutes=5)
        create_pending_ticket("NEXT-SEQUENCE-002", 2)
        db.session.commit()

    resp = login(client, "modea_sequence_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-a/next", json={"device_id": "device-a"})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["device_today_sequence"] == 2


def test_mode_a_next_enforces_server_side_cooldown_before_completing_current(app, client):
    with app.app_context():
        user = create_user("modea_server_cooldown_user", "secret123", client_mode="mode_a")
        first = create_pending_ticket("SERVER-COOLDOWN-1", 1)
        second = create_pending_ticket("SERVER-COOLDOWN-2", 2)
        first_id = first.id
        second_id = second.id

    resp = login(client, "modea_server_cooldown_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-a/next", json={"device_id": "device-a"})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["id"] == first_id

    resp = client.post(
        "/api/mode-a/next",
        json={"device_id": "device-a", "complete_current_ticket_id": first_id},
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["id"] == first_id
    assert data["completed_current"] is False
    assert data["cooldown_remaining"] > 0

    with app.app_context():
        first = LotteryTicket.query.get(first_id)
        second = LotteryTicket.query.get(second_id)
        assert first.status == "assigned"
        assert second.status == "pending"


def test_mode_a_next_can_expire_overdue_current_ticket(app, client):
    with app.app_context():
        user = create_user("modea_expire_user", "secret123", client_mode="mode_a")
        current_ticket = create_assigned_ticket(user, "device-a", "CUR-EXPIRE", 1)
        current_ticket.deadline_time = beijing_now() - timedelta(minutes=1)
        next_ticket = create_pending_ticket("NEXT-AFTER-EXPIRE", 2)
        current_ticket_id = current_ticket.id
        next_ticket_id = next_ticket.id
        db.session.commit()

    resp = login(client, "modea_expire_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/mode-a/next",
        json={
            "device_id": "device-a",
            "complete_current_ticket_id": current_ticket_id,
            "complete_current_ticket_action": "expired",
        },
    )
    assert resp.status_code == 200

    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["id"] == next_ticket_id

    with app.app_context():
        expired_ticket = LotteryTicket.query.get(current_ticket_id)
        assigned_next = LotteryTicket.query.get(next_ticket_id)
        assert expired_ticket.status == "expired"
        assert assigned_next.status == "assigned"


def test_mode_a_next_prefers_earliest_deadline_even_when_later_deadline_has_smaller_id(app):
    from services.mode_a_service import get_next_ticket

    with app.app_context():
        user = create_user("modea_deadline_priority_user", "secret123", client_mode="mode_a")
        later_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="DEADLINE-LATER",
            status="pending",
            deadline_time=(beijing_now() + timedelta(hours=2, minutes=25)).replace(microsecond=0),
            detail_period="2055",
            lottery_type="SPF",
        )
        earlier_ticket = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="DEADLINE-EARLIER",
            status="pending",
            deadline_time=(beijing_now() + timedelta(hours=1, minutes=25)).replace(microsecond=0),
            detail_period="1955",
            lottery_type="SPF",
        )
        db.session.add_all([later_ticket, earlier_ticket])
        db.session.commit()

        result = get_next_ticket(
            user_id=user.id,
            device_id="device-a",
            username=user.username,
        )

        assert result["success"] is True
        assert result["ticket"]["id"] == earlier_ticket.id
        assert result["ticket"]["detail_period"] == "1955"

        refreshed_earlier = LotteryTicket.query.get(earlier_ticket.id)
        refreshed_later = LotteryTicket.query.get(later_ticket.id)
        assert refreshed_earlier.status == "assigned"
        assert refreshed_later.status == "pending"


def test_mode_a_next_stops_after_completing_current_when_pool_disabled(app):
    from services.mode_a_service import get_next_ticket

    with app.app_context():
        user = create_user("modea_pool_disabled_user", "secret123", client_mode="mode_a")
        settings = SystemSettings.get()
        settings.pool_enabled = False
        db.session.commit()

        current_ticket = create_assigned_ticket(user, "device-a", "MODEA-POOL-DISABLED", 1)
        current_ticket_id = current_ticket.id

        result = get_next_ticket(
            user_id=user.id,
            device_id="device-a",
            username=user.username,
            complete_current_ticket_id=current_ticket_id,
            complete_current_ticket_action="completed",
        )

        refreshed = LotteryTicket.query.get(current_ticket_id)
        assert refreshed.status == "completed"
        assert result["success"] is False
        assert result["error"]
        assert result["completed_current"] is True


def test_mode_a_current_prefers_latest_assigned_ticket_for_device(app, client):
    with app.app_context():
        user = create_user("modea_latest_current_user", "secret123", client_mode="mode_a")
        older = create_assigned_ticket(user, "device-a", "CUR-OLDER", 1)
        newer = create_assigned_ticket(user, "device-a", "CUR-NEWER", 2)
        older.assigned_at = datetime(2026, 4, 7, 10, 0, 0)
        newer.assigned_at = datetime(2026, 4, 7, 10, 5, 0)
        db.session.commit()
        newer_id = newer.id

    resp = login(client, "modea_latest_current_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-a/current?device_id=device-a")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket"]["id"] == newer_id


def test_mode_a_previous_rejects_invalid_offset(app, client):
    with app.app_context():
        create_user("modea_previous_guard_user", "secret123", client_mode="mode_a")

    resp = login(client, "modea_previous_guard_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-a/previous?device_id=device-a&offset=bad")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "offset" in data["error"]


def test_mode_a_stop_reports_finalize_race_failure(app, client, monkeypatch):
    with app.app_context():
        user = create_user("modea_stop_race_user", "secret123", client_mode="mode_a")
        create_assigned_ticket(user, "device-stop-race", "STOP-RACE-001", 1)

    resp = login(client, "modea_stop_race_user", "secret123")
    assert resp.status_code == 200

    monkeypatch.setattr("services.mode_a_service.finalize_ticket", lambda *args, **kwargs: False)

    resp = client.post("/api/mode-a/stop", json={"device_id": "device-stop-race"})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is False
    assert "状态已变化" in data["error"]


def test_mode_a_device_daily_records_include_current_device_future_deadline_tickets(app, client):
    with app.app_context():
        user = create_user("modea_daily_list_user", "secret123", client_mode="mode_a")
        other_user = create_user("modea_daily_list_other_user", "secret123", client_mode="mode_a")

        first = create_assigned_ticket(user, "device-a", "DEVICE-A-DONE-1", 1)
        first.status = "completed"
        first.deadline_time = beijing_now() + timedelta(hours=2)
        first.completed_at = beijing_now() - timedelta(minutes=8)
        first.ticket_amount = 12

        assigned = create_assigned_ticket(user, "device-a", "DEVICE-A-CURRENT-2", 2)
        assigned.deadline_time = beijing_now() + timedelta(hours=3)
        assigned.assigned_at = beijing_now() - timedelta(minutes=2)
        assigned.ticket_amount = 34

        other_device = create_assigned_ticket(user, "device-b", "DEVICE-B-DONE", 3)
        other_device.status = "completed"
        other_device.completed_at = beijing_now() - timedelta(minutes=6)

        other_owner = create_assigned_ticket(other_user, "device-a", "OTHER-USER-DONE", 4)
        other_owner.status = "completed"
        other_owner.completed_at = beijing_now() - timedelta(minutes=4)

        db.session.commit()

    resp = login(client, "modea_daily_list_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-a/device-daily?device_id=device-a")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["count"] == 2
    assert [record["raw_content"] for record in data["records"]] == [
        "DEVICE-A-CURRENT-2",
        "DEVICE-A-DONE-1",
    ]
    assert [record["device_today_sequence"] for record in data["records"]] == [2, 1]
    assert [record["ticket_amount"] for record in data["records"]] == [34.0, 12.0]
    assert data["records"][0]["status"] == "assigned"
    assert data["records"][1]["status"] == "completed"


def test_pool_status_returns_empty_when_pool_disabled(app, client):
    with app.app_context():
        user = create_user("pool_user", "secret123", client_mode="mode_b")
        create_pending_ticket("PENDING001", 1)
        settings = SystemSettings.get()
        settings.pool_enabled = False
        db.session.commit()

    resp = login(client, "pool_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/pool/status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["pool_enabled"] is False
    assert data["total_pending"] == 0
    assert data["by_type"] == []


def test_pool_status_hides_blocked_lottery_types_for_user(app, client):
    with app.app_context():
        user = create_user("pool_blocked_user", "secret123", client_mode="mode_a")
        user.set_blocked_lottery_types(["???"])
        blocked_deadline = beijing_now() + timedelta(hours=1)
        allowed_deadline = beijing_now() + timedelta(hours=2)
        db.session.add_all([
            LotteryTicket(source_file_id=1, line_number=1, raw_content="POOL-BLOCKED-1", status="pending", lottery_type="???", deadline_time=blocked_deadline),
            LotteryTicket(source_file_id=1, line_number=2, raw_content="POOL-ALLOWED-1", status="pending", lottery_type="璁╃悆???", deadline_time=allowed_deadline),
        ])
        db.session.commit()

    resp = login(client, "pool_blocked_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/pool/status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["total_pending"] == 1
    assert len(data["by_type"]) == 1
    assert data["by_type"][0]["lottery_type"]


def test_pool_status_uses_mode_b_reserve_rule_for_mode_b_users(app, client):
    with app.app_context():
        user = create_user("pool_mode_b_reserve_user", "secret123", client_mode="mode_b")
        deadline = beijing_now() + timedelta(hours=1)
        tickets = [
            LotteryTicket(
                source_file_id=1,
                line_number=index + 1,
                raw_content=f"POOL-MODEB-{index}",
                status="pending",
                lottery_type="???",
                deadline_time=deadline,
            )
            for index in range(25)
        ]
        db.session.add_all(tickets)
        db.session.commit()

    resp = login(client, "pool_mode_b_reserve_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/pool/status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["total_pending"] == 5
    assert len(data["by_type"]) == 1
    assert data["by_type"][0]["count"] == 5


def test_pool_status_uses_configured_mode_b_pool_reserve(app, client):
    with app.app_context():
        user = create_user("pool_mode_b_custom_reserve_user", "secret123", client_mode="mode_b")
        settings = SystemSettings.get()
        settings.mode_b_pool_reserve = 10
        deadline = beijing_now() + timedelta(hours=1)
        tickets = [
            LotteryTicket(
                source_file_id=1,
                line_number=index + 1,
                raw_content=f"POOL-MODEB-CUSTOM-{index}",
                status="pending",
                lottery_type="胜平负",
                deadline_time=deadline,
            )
            for index in range(25)
        ]
        db.session.add_all(tickets)
        db.session.commit()

    resp = login(client, "pool_mode_b_custom_reserve_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/pool/status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["total_pending"] == 15
    assert len(data["by_type"]) == 1
    assert data["by_type"][0]["count"] == 15


def test_pool_status_returns_zero_for_mode_b_users_when_mode_b_disabled(app, client):
    with app.app_context():
        user = create_user("pool_status_mode_b_disabled", "secret123", client_mode="mode_b")
        deadline = beijing_now() + timedelta(hours=1)
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="POOL-MODEB-DISABLED",
            status="pending",
            lottery_type="???",
            deadline_time=deadline,
        ))
        settings = SystemSettings.get()
        settings.mode_b_enabled = False
        db.session.commit()

    resp = login(client, "pool_status_mode_b_disabled", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/pool/status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["pool_enabled"] is True
    assert data["total_pending"] == 0
    assert data["by_type"] == []


def test_pool_status_requires_login_json_response(app, client):
    resp = client.get("/api/pool/status")
    assert resp.status_code == 401
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert "\u8bf7\u5148\u767b\u5f55" in data["error"]


def test_pool_status_returns_zero_for_mode_a_users_when_mode_a_disabled(app, client):
    with app.app_context():
        user = create_user("pool_status_mode_a_disabled", "secret123", client_mode="mode_a")
        deadline = beijing_now() + timedelta(hours=1)
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="POOL-MODEA-DISABLED",
            status="pending",
            lottery_type="???",
            deadline_time=deadline,
        ))
        settings = SystemSettings.get()
        settings.mode_a_enabled = False
        db.session.commit()

    resp = login(client, "pool_status_mode_a_disabled", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/pool/status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["pool_enabled"] is True
    assert data["total_pending"] == 0
    assert data["by_type"] == []


def test_get_client_ip_ignores_x_forwarded_for_without_trusted_proxy(app):
    from utils.decorators import get_client_ip

    with app.test_request_context(
        '/',
        headers={'X-Forwarded-For': '198.51.100.77'},
        environ_base={'REMOTE_ADDR': '203.0.113.10'},
    ):
        assert get_client_ip() == '203.0.113.10'


def test_get_client_ip_uses_x_forwarded_for_when_remote_addr_trusted(app):
    from utils.decorators import get_client_ip

    app.config['TRUSTED_PROXY_IPS'] = '203.0.113.10'
    with app.test_request_context(
        '/',
        headers={'X-Forwarded-For': '198.51.100.77'},
        environ_base={'REMOTE_ADDR': '203.0.113.10'},
    ):
        assert get_client_ip() == '198.51.100.77'


def test_heartbeat_requires_login_json_response(app, client):
    resp = client.post("/auth/heartbeat")
    assert resp.status_code == 401
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert "\u8bf7\u5148\u767b\u5f55" in data["error"]


def test_mode_b_pool_status_returns_empty_when_pool_disabled(app, client):
    with app.app_context():
        user = create_user("pool_mode_b_user", "secret123", client_mode="mode_b")
        create_pending_ticket("PENDING002", 1)
        settings = SystemSettings.get()
        settings.pool_enabled = False
        db.session.commit()

    resp = login(client, "pool_mode_b_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/pool-status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["total_pending"] == 0
    assert data["by_type"] == []


def test_mode_b_pool_status_returns_empty_when_mode_b_disabled(app, client):
    with app.app_context():
        user = create_user("pool_mode_b_disabled_user", "secret123", client_mode="mode_b")
        create_pending_ticket("PENDING-MODEB-DISABLED-1", 1)
        settings = SystemSettings.get()
        settings.mode_b_enabled = False
        db.session.commit()

    resp = login(client, "pool_mode_b_disabled_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/pool-status")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["total_pending"] == 0
    assert data["by_type"] == []


def test_deleted_user_session_invalidates_api_access(app, client):
    with app.app_context():
        user = create_user("session_user", "secret123", client_mode="mode_a")

    resp = login(client, "session_user", "secret123")
    assert resp.status_code == 200

    with app.app_context():
        from models.user import UserSession

        UserSession.query.delete()
        db.session.commit()

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 401
    data = resp.get_json()
    assert data["success"] is False


def test_daily_stats_returns_announcement_for_receiving_user(app, client):
    with app.app_context():
        user = create_user("daily_stats_announcement_user", "secret123", client_mode="mode_a")
        settings = SystemSettings.get()
        settings.announcement_enabled = True
        settings.announcement = "浠婃櫄 8 鐐圭淮鎶?"
        db.session.commit()

    resp = login(client, "daily_stats_announcement_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["can_receive"] is True
    assert data["announcement"]


def test_daily_stats_hides_announcement_for_non_receiving_user(app, client):
    with app.app_context():
        user = create_user("daily_stats_no_announcement_user", "secret123", client_mode="mode_a")
        user.can_receive = False
        settings = SystemSettings.get()
        settings.announcement_enabled = True
        settings.announcement = "\u8fd9\u6761\u516c\u544a\u4e0d\u5e94\u663e\u793a"
        db.session.commit()

    resp = login(client, "daily_stats_no_announcement_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["can_receive"] is False
    assert data["announcement"] == ""


def test_expired_user_session_invalidates_api_access(app, client):
    with app.app_context():
        user = create_user("expired_session_user", "secret123", client_mode="mode_a")

    resp = login(client, "expired_session_user", "secret123")
    assert resp.status_code == 200

    with app.app_context():
        from models.user import UserSession

        session = UserSession.query.first()
        session.expires_at = beijing_now() - timedelta(minutes=1)
        db.session.commit()

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 401
    data = resp.get_json()
    assert data["success"] is False


def test_device_register_handles_empty_json_body(app, client):
    with app.app_context():
        create_user("device_empty_body_user", "secret123", client_mode="mode_b")

    resp = login(client, "device_empty_body_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/device/register", data="", content_type="application/json")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "\u8bbe\u5907ID" in data["error"]


def test_device_register_rejects_claiming_other_users_device_id(app, client):
    with app.app_context():
        user_a = create_user("device_owner_user", "secret123", client_mode="mode_b")
        create_user("device_other_user", "secret123", client_mode="mode_b")
        from models.device import DeviceRegistry

        db.session.add(DeviceRegistry(device_id="shared-device", user_id=user_a.id))
        db.session.commit()

    resp = login(client, "device_other_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/device/register", json={"device_id": "shared-device"})
    assert resp.status_code == 409
    data = resp.get_json()
    assert data["success"] is False
    assert "\u5176\u4ed6\u7528\u6237" in data["error"]


def test_device_register_returns_409_when_commit_hits_integrity_error(app, client, monkeypatch):
    from sqlalchemy.exc import IntegrityError

    with app.app_context():
        create_user("device_integrity_conflict_user", "secret123", client_mode="mode_b")

    resp = login(client, "device_integrity_conflict_user", "secret123")
    assert resp.status_code == 200

    def failing_commit():
        raise IntegrityError("INSERT INTO device_registry ...", {}, Exception("duplicate key"))

    monkeypatch.setattr("extensions.db.session.commit", failing_commit)

    resp = client.post("/api/device/register", json={"device_id": "device-race-1"})
    assert resp.status_code == 409
    data = resp.get_json()
    assert data["success"] is False


def test_device_register_requires_login_json_response(app, client):
    resp = client.post("/api/device/register", json={"device_id": "device-a"})
    assert resp.status_code == 401
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert "\u8bf7\u5148\u767b\u5f55" in data["error"]


def test_device_register_rejects_invalid_device_id_format(app, client):
    with app.app_context():
        create_user("device_invalid_format_user", "secret123", client_mode="mode_b")

    resp = login(client, "device_invalid_format_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/device/register", json={"device_id": "bad id"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_device_register_returns_device_id_only(app, client):
    with app.app_context():
        create_user("device_default_name_user", "secret123", client_mode="mode_b")

    resp = login(client, "device_default_name_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/device/register", json={"device_id": "device-one", "client_info": {"client_type": "web"}})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["device"]["device_id"] == "device-one"
    assert data["device"]["device_id"] == "device-one"


def test_device_register_rejects_too_long_device_id(app, client):
    with app.app_context():
        user = create_user("device_long_name_user", "secret123", client_mode="mode_b")

    resp = login(client, "device_long_name_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/device/register", json={"device_id": "x" * 51})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "\u957f\u5ea6\u4e0d\u80fd\u8d85\u8fc720" in data["error"]


def test_device_register_returns_json_for_missing_device_id(app, client):
    with app.app_context():
        create_user("device_missing_name_user", "secret123", client_mode="mode_b")

    resp = login(client, "device_missing_name_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/device/register", json={})
    assert resp.status_code == 400
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert "\u8bf7\u8f93\u5165\u8bbe\u5907ID" in data["error"]


def test_device_update_changes_current_session_and_registry(app, client):
    with app.app_context():
        user = create_user("device_update_user", "secret123", client_mode="mode_b")
        user_id = user.id

    resp = client.post(
        "/auth/login",
        json={"username": "device_update_user", "password": "secret123", "device_id": "device-old"},
    )
    assert resp.status_code == 200

    resp = client.post("/api/device/register", json={"device_id": "device-old", "client_info": {"client_type": "web"}})
    assert resp.status_code == 200

    resp = client.post(
        "/api/device/update",
        json={
            "current_device_id": "device-old",
            "new_device_id": "device-new",
            "client_info": {"client_type": "web"},
        },
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["device"]["device_id"] == "device-new"

    with app.app_context():
        from models.user import UserSession

        session_record = UserSession.query.filter_by(user_id=user_id).first()
        assert session_record.device_id == "device-new"
        assert DeviceRegistry.query.filter_by(device_id="device-old").first() is None
        device = DeviceRegistry.query.filter_by(device_id="device-new").first()
        assert device is not None
        assert device.user_id == user_id


def test_device_update_rejects_claiming_other_users_device_id(app, client):
    with app.app_context():
        owner = create_user("device_update_owner", "secret123", client_mode="mode_b")
        create_user("device_update_actor", "secret123", client_mode="mode_b")
        db.session.add(DeviceRegistry(device_id="taken-device", user_id=owner.id))
        db.session.commit()

    resp = client.post(
        "/auth/login",
        json={"username": "device_update_actor", "password": "secret123", "device_id": "actor-device"},
    )
    assert resp.status_code == 200

    resp = client.post(
        "/api/device/update",
        json={"current_device_id": "actor-device", "new_device_id": "taken-device"},
    )
    assert resp.status_code == 409
    data = resp.get_json()
    assert data["success"] is False
    assert "\u5176\u4ed6\u7528\u6237" in data["error"]


def test_device_update_rejects_when_old_device_has_assigned_tickets(app, client):
    with app.app_context():
        user = create_user("device_update_assigned_user", "secret123", client_mode="mode_b")
        file = UploadedFile(
            original_filename="assigned.txt",
            stored_filename="assigned.txt",
            total_tickets=1,
            assigned_count=1,
        )
        db.session.add(file)
        db.session.flush()
        db.session.add(
            LotteryTicket(
                source_file_id=file.id,
                line_number=1,
                raw_content="ticket",
                status="assigned",
                assigned_user_id=user.id,
                assigned_username=user.username,
                assigned_device_id="busy-device",
                assigned_at=beijing_now(),
                deadline_time=beijing_now() + timedelta(hours=1),
            )
        )
        db.session.commit()

    resp = client.post(
        "/auth/login",
        json={"username": "device_update_assigned_user", "password": "secret123", "device_id": "busy-device"},
    )
    assert resp.status_code == 200

    resp = client.post(
        "/api/device/update",
        json={"current_device_id": "busy-device", "new_device_id": "free-device"},
    )
    assert resp.status_code == 409
    data = resp.get_json()
    assert data["success"] is False
    assert "\u5904\u7406\u4e2d" in data["error"]


def test_change_password_handles_empty_json_body(app, client):
    with app.app_context():
        create_user("change_password_empty_body_user", "secret123", client_mode="mode_a")

    resp = login(client, "change_password_empty_body_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/user/change-password", data="", content_type="application/json")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_mode_b_confirm_handles_empty_json_body(app, client):
    with app.app_context():
        create_user("modeb_empty_confirm_user", "secret123", client_mode="mode_b")

    resp = login(client, "modeb_empty_confirm_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/confirm", data="", content_type="application/json")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "缺少设备ID" in data["error"]


def test_heartbeat_can_backfill_session_device_id(app, client):
    with app.app_context():
        user = create_user("heartbeat_device_user", "secret123", client_mode="mode_b")
        user_id = user.id
        settings = SystemSettings.get()
        settings.session_lifetime_hours = 6
        db.session.commit()

    resp = login(client, "heartbeat_device_user", "secret123")
    assert resp.status_code == 200

    with app.app_context():
        from models.user import UserSession

        session = UserSession.query.filter_by(user_id=user_id).first()
        session.expires_at = beijing_now() + timedelta(minutes=1)
        db.session.commit()

    resp = client.post("/auth/heartbeat", json={"device_id": "device-01"})
    assert resp.status_code == 200
    assert resp.get_json()["success"] is True

    with app.app_context():
        from models.user import UserSession

        session = UserSession.query.filter_by(user_id=user_id).first()
        assert session is not None
        assert session.device_id == "device-01"
        remaining_hours = (session.expires_at - beijing_now()).total_seconds() / 3600
        assert remaining_hours > 5.8


def test_create_session_uses_db_session_lifetime_setting(app):
    from services.session_service import create_session

    with app.app_context():
        user = create_user("session_lifetime_user", "secret123", client_mode="mode_b")
        settings = SystemSettings.get()
        settings.session_lifetime_hours = 6
        db.session.commit()

        before = beijing_now()
        session = create_session(user, device_id="device-session")
        delta_hours = (session.expires_at - before).total_seconds() / 3600

    assert 5.9 <= delta_hours <= 6.1


def test_clean_inactive_sessions_cleans_expired_sessions_even_if_last_seen_recent(app):
    from services.session_service import clean_inactive_sessions
    from models.user import UserSession

    with app.app_context():
        user = create_user("cleanup_expired_user", "secret123", client_mode="mode_b")
        now = beijing_now()
        session = UserSession(
            user_id=user.id,
            session_token="cleanup-expired-token",
            device_id="cleanup-device",
            last_seen=now,
            expires_at=now - timedelta(minutes=1),
        )
        db.session.add(session)
        db.session.commit()

        cleaned = clean_inactive_sessions(hours=3)

        assert cleaned == 1
        assert UserSession.query.filter_by(session_token="cleanup-expired-token").first() is None


def test_clean_inactive_sessions_keeps_recent_non_expired_session(app):
    from services.session_service import clean_inactive_sessions
    from models.user import UserSession

    with app.app_context():
        user = create_user("cleanup_keep_user", "secret123", client_mode="mode_b")
        now = beijing_now()
        session = UserSession(
            user_id=user.id,
            session_token="cleanup-keep-token",
            device_id="cleanup-keep-device",
            last_seen=now,
            expires_at=now + timedelta(hours=2),
        )
        db.session.add(session)
        db.session.commit()

        cleaned = clean_inactive_sessions(hours=3)

        assert cleaned == 0
        assert UserSession.query.filter_by(session_token="cleanup-keep-token").first() is not None


def test_authenticated_request_extends_session_expiry(app, client):
    with app.app_context():
        user = create_user("session_extend_user", "secret123", client_mode="mode_a")
        settings = SystemSettings.get()
        settings.session_lifetime_hours = 6
        db.session.commit()

    resp = login(client, "session_extend_user", "secret123")
    assert resp.status_code == 200

    with app.app_context():
        from models.user import UserSession

        session = UserSession.query.first()
        session.expires_at = beijing_now() + timedelta(minutes=1)
        db.session.commit()

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200

    with app.app_context():
        from models.user import UserSession

        session = UserSession.query.first()
        remaining_hours = (session.expires_at - beijing_now()).total_seconds() / 3600

    assert remaining_hours > 5.8


def test_time_utils_use_configured_daily_reset_hour(app):
    from utils.time_utils import get_business_date, get_business_window, resolve_deadline_datetime

    with app.app_context():
        settings = SystemSettings.get()
        settings.daily_reset_hour = 6
        db.session.commit()

        assert str(get_business_date(datetime(2026, 4, 7, 5, 59, 0))) == "2026-04-06"
        assert str(get_business_date(datetime(2026, 4, 7, 6, 0, 0))) == "2026-04-07"

        window_start, window_end = get_business_window(datetime(2026, 4, 7).date())
        assert window_start == datetime(2026, 4, 7, 6, 0, 0)
        assert window_end == datetime(2026, 4, 8, 6, 0, 0)

        deadline = resolve_deadline_datetime("05.30", datetime(2026, 4, 7, 7, 0, 0))
        assert deadline == datetime(2026, 4, 8, 5, 30, 0)


def test_my_winning_uses_configured_business_reset_hour(app, client, monkeypatch):
    monkeypatch.setattr("utils.time_utils.beijing_now", lambda: datetime(2026, 4, 7, 7, 0, 0))

    with app.app_context():
        user = create_user("winning_reset_hour_user", "secret123", client_mode="mode_a")
        settings = SystemSettings.get()
        settings.daily_reset_hour = 6

        inside_window = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-RESET-INSIDE",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 7, 5, 30, 0),
            is_winning=True,
        )
        outside_window = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="WIN-RESET-OUTSIDE",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 7, 6, 30, 0),
            is_winning=True,
        )
        db.session.add_all([inside_window, outside_window])
        db.session.commit()

    resp = login(client, "winning_reset_hour_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my?date=2026-04-06")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    records = data["grouped"]["2026-04-06"]
    raw_contents = [record["raw_content"] for record in records]
    assert "WIN-RESET-INSIDE" in raw_contents
    assert "WIN-RESET-OUTSIDE" not in raw_contents


def test_user_daily_stats_uses_current_business_window_before_noon(app, client, monkeypatch):
    business_start = datetime(2026, 4, 6, 12, 0, 0)
    business_end = datetime(2026, 4, 7, 12, 0, 0)

    with app.app_context():
        user = create_user("daily_stats_user", "secret123", client_mode="mode_b")

        in_window = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="IN-WINDOW",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-a",
            ticket_amount=2,
            completed_at=business_start + timedelta(hours=1),
        )
        out_of_window = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="OUT-OF-WINDOW",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-a",
            ticket_amount=3,
            completed_at=business_start - timedelta(hours=1),
        )
        db.session.add_all([in_window, out_of_window])
        db.session.commit()

    monkeypatch.setattr("routes.user.get_today_noon", lambda: business_start)
    monkeypatch.setattr("routes.user.get_business_date", lambda dt=None: business_start.date())

    resp = login(client, "daily_stats_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["ticket_count"] == 1
    assert data["total_amount"] == 2.0
    assert data["today"] == "2026-04-06"
    assert data["device_stats"][0]["count"] == 1
    assert data["device_stats"][0]["amount"] == 2.0


def test_user_daily_stats_pool_total_pending_excludes_blocked_lottery_types(app, client):
    with app.app_context():
        user = create_user("daily_stats_blocked_user", "secret123", client_mode="mode_a")
        user.set_blocked_lottery_types(["???"])
        blocked_deadline = beijing_now() + timedelta(hours=1)
        allowed_deadline = beijing_now() + timedelta(hours=2)
        db.session.add_all([
            LotteryTicket(source_file_id=1, line_number=1, raw_content="STATS-BLOCKED-1", status="pending", lottery_type="???", deadline_time=blocked_deadline),
            LotteryTicket(source_file_id=1, line_number=2, raw_content="STATS-ALLOWED-1", status="pending", lottery_type="璁╃悆???", deadline_time=allowed_deadline),
        ])
        db.session.commit()

    resp = login(client, "daily_stats_blocked_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["pool_total_pending"] == 1


def test_user_daily_stats_pool_total_pending_uses_mode_b_reserve_rule(app, client):
    with app.app_context():
        user = create_user("daily_stats_mode_b_reserve_user", "secret123", client_mode="mode_b")
        deadline = beijing_now() + timedelta(hours=1)
        tickets = [
            LotteryTicket(
                source_file_id=1,
                line_number=index + 1,
                raw_content=f"MODEB-POOL-{index}",
                status="pending",
                lottery_type="???",
                deadline_time=deadline,
            )
            for index in range(25)
        ]
        db.session.add_all(tickets)
        db.session.commit()

    resp = login(client, "daily_stats_mode_b_reserve_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["pool_total_pending"] == 5


def test_user_daily_stats_pool_total_pending_is_zero_when_pool_disabled(app, client):
    with app.app_context():
        create_user("daily_stats_pool_disabled", "secret123", client_mode="mode_b")
        settings = SystemSettings.get()
        settings.pool_enabled = False
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="POOL-DISABLED-1",
            status="pending",
            lottery_type="璁╃悆???",
            deadline_time=beijing_now() + timedelta(hours=1),
        ))
        db.session.commit()

    resp = login(client, "daily_stats_pool_disabled", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["pool_total_pending"] == 0


def test_user_daily_stats_pool_total_pending_is_zero_when_mode_b_disabled(app, client):
    with app.app_context():
        create_user("daily_stats_mode_b_disabled", "secret123", client_mode="mode_b")
        settings = SystemSettings.get()
        settings.mode_b_enabled = False
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="MODE-B-DISABLED-1",
            status="pending",
            lottery_type="璁╃悆???",
            deadline_time=beijing_now() + timedelta(hours=1),
        ))
        db.session.commit()

    resp = login(client, "daily_stats_mode_b_disabled", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["pool_total_pending"] == 0


def test_user_daily_stats_pool_total_pending_is_zero_when_mode_a_disabled(app, client):
    with app.app_context():
        create_user("daily_stats_mode_a_disabled", "secret123", client_mode="mode_a")
        settings = SystemSettings.get()
        settings.mode_a_enabled = False
        db.session.add(LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="MODE-A-DISABLED-1",
            status="pending",
            lottery_type="???",
            deadline_time=beijing_now() + timedelta(hours=1),
        ))
        db.session.commit()

    resp = login(client, "daily_stats_mode_a_disabled", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/user/daily-stats")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["pool_total_pending"] == 0


def test_file_display_id_uses_business_date_before_noon(app, monkeypatch):
    fixed_now = datetime(2026, 4, 7, 11, 0, 0)

    with app.app_context():
        monkeypatch.setattr("services.file_parser.beijing_now", lambda: fixed_now)

        from services.file_parser import _generate_display_id

        display_id = _generate_display_id()
        assert display_id == "2026/04/06-01"


def test_upload_winning_image_creates_winning_record(app, client):
    from PIL import Image

    with app.app_context():
        user = create_user("winning_user", "secret123", client_mode="mode_a")
        ticket = create_assigned_ticket(user, "device-a", "WIN001", 1)
        ticket.is_winning = True
        db.session.commit()
        user_id = user.id
        ticket_id = ticket.id

    resp = login(client, "winning_user", "secret123")
    assert resp.status_code == 200

    image_bytes = io.BytesIO()
    Image.new("RGB", (20, 20), color="red").save(image_bytes, format="PNG")
    image_bytes.seek(0)

    resp = client.post(
        f"/api/winning/upload-image/{ticket_id}",
        data={"image": (image_bytes, "winning.png")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True

    with app.app_context():
        from models.winning import WinningRecord

        ticket = LotteryTicket.query.get(ticket_id)
        record = WinningRecord.query.filter_by(ticket_id=ticket_id).first()
        assert ticket.winning_image_url
        assert ticket.is_winning is True
        assert record is not None
        assert record.winning_image_url == ticket.winning_image_url
        assert record.uploaded_by == user_id


def test_upload_winning_image_works_without_pillow(app, client, monkeypatch):
    original_import = builtins.__import__

    def fake_import(name, globals=None, locals=None, fromlist=(), level=0):
        if name == "PIL":
            raise ModuleNotFoundError("No module named 'PIL'")
        return original_import(name, globals, locals, fromlist, level)

    monkeypatch.setattr(builtins, "__import__", fake_import)

    with app.app_context():
        user = create_user("winning_user_no_pillow", "secret123", client_mode="mode_a")
        ticket = create_assigned_ticket(user, "device-b", "WIN002", 2)
        ticket.is_winning = True
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_user_no_pillow", "secret123")
    assert resp.status_code == 200

    raw_image = io.BytesIO(b"raw-image-without-pillow")
    resp = client.post(
        f"/api/winning/upload-image/{ticket_id}",
        data={"image": (raw_image, "winning.png")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["image_url"].startswith("/uploads/images/winning_")


def test_upload_winning_image_rejects_empty_filename(app, client):
    with app.app_context():
        user = create_user("winning_empty_filename_user", "secret123", client_mode="mode_a")
        ticket = create_assigned_ticket(user, "device-empty-image", "WIN-EMPTY-NAME", 1)
        ticket.is_winning = True
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_empty_filename_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        f"/api/winning/upload-image/{ticket_id}",
        data={"image": (io.BytesIO(b"fake-image"), "")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "文件名不能为空" in data["error"]


def test_upload_winning_image_replaces_old_local_file(app, client):
    from PIL import Image

    with app.app_context():
        user = create_user("winning_replace_user", "secret123", client_mode="mode_a")
        ticket = create_assigned_ticket(user, "device-c", "WIN003", 3)
        ticket.is_winning = True
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_replace_user", "secret123")
    assert resp.status_code == 200

    first_image = io.BytesIO()
    Image.new("RGB", (20, 20), color="red").save(first_image, format="PNG")
    first_image.seek(0)
    first_resp = client.post(
        f"/api/winning/upload-image/{ticket_id}",
        data={"image": (first_image, "winning1.png")},
        content_type="multipart/form-data",
    )
    assert first_resp.status_code == 200
    first_url = first_resp.get_json()["image_url"]

    with app.app_context():
        first_record = WinningRecord.query.filter_by(ticket_id=ticket_id).first()
        first_path = Path(app.config["UPLOAD_FOLDER"]) / "images" / first_url.rsplit("/", 1)[-1]
        assert first_record is not None
        assert first_path.exists()

    second_image = io.BytesIO()
    Image.new("RGB", (20, 20), color="blue").save(second_image, format="PNG")
    second_image.seek(0)
    second_resp = client.post(
        f"/api/winning/upload-image/{ticket_id}",
        data={"image": (second_image, "winning2.png")},
        content_type="multipart/form-data",
    )
    assert second_resp.status_code == 200
    second_url = second_resp.get_json()["image_url"]

    with app.app_context():
        refreshed_record = WinningRecord.query.filter_by(ticket_id=ticket_id).first()
        second_path = Path(app.config["UPLOAD_FOLDER"]) / "images" / second_url.rsplit("/", 1)[-1]

    assert first_url != second_url
    assert not first_path.exists()
    assert second_path.exists()
    assert refreshed_record.winning_image_url == second_url


def test_expire_overdue_tickets_updates_file_counters(app):
    from tasks.expire_tickets import expire_overdue_tickets

    with app.app_context():
        user = create_user("expire_user", "secret123", client_mode="mode_a")
        uploaded_file = UploadedFile(
            display_id="2026/03/30-01",
            original_filename="expired.txt",
            stored_filename="expired.txt",
            uploaded_by=user.id,
            total_tickets=2,
            pending_count=1,
            assigned_count=1,
            completed_count=0,
        )
        db.session.add(uploaded_file)
        db.session.commit()

        db.session.add(LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="PENDING-EXPIRED",
            status="pending",
            deadline_time=beijing_now() - timedelta(minutes=1),
        ))
        db.session.add(LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=2,
            raw_content="ASSIGNED-EXPIRED",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-a",
            deadline_time=beijing_now() - timedelta(minutes=1),
        ))
        db.session.commit()

        expire_overdue_tickets()

        refreshed = db.session.get(UploadedFile, uploaded_file.id)
        statuses = {t.status for t in LotteryTicket.query.filter_by(source_file_id=uploaded_file.id).all()}
        assert statuses == {"expired", "assigned"}
        assert refreshed.pending_count == 0
        assert refreshed.assigned_count == 1


def test_expire_overdue_tickets_removes_pending_ids_from_redis(app, monkeypatch):
    from tasks.expire_tickets import expire_overdue_tickets

    removed = []

    class FakePipe:
        def lrem(self, key, count, value):
            removed.append((key, count, value))
            return self

        def execute(self):
            return True

    class FakeRedis:
        def pipeline(self):
            return FakePipe()

    monkeypatch.setattr("extensions.redis_client", FakeRedis())

    with app.app_context():
        user = create_user("expire_redis_user", "secret123", client_mode="mode_a")
        uploaded_file = UploadedFile(
            display_id="2026/03/30-02",
            original_filename="expired-redis.txt",
            stored_filename="expired-redis.txt",
            uploaded_by=user.id,
            total_tickets=2,
            pending_count=1,
            assigned_count=1,
            completed_count=0,
        )
        db.session.add(uploaded_file)
        db.session.commit()

        pending_ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="PENDING-REDIS-EXPIRED",
            status="pending",
            deadline_time=beijing_now() - timedelta(minutes=1),
        )
        assigned_ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=2,
            raw_content="ASSIGNED-REDIS-EXPIRED",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-a",
            deadline_time=beijing_now() - timedelta(minutes=1),
        )
        db.session.add_all([pending_ticket, assigned_ticket])
        db.session.commit()
        pending_ticket_id = pending_ticket.id
        assigned_ticket_id = assigned_ticket.id

        expire_overdue_tickets()

    assert ("pool:pending", 0, str(pending_ticket_id)) in removed
    assert ("pool:pending", 0, str(assigned_ticket_id)) not in removed


def test_expire_overdue_tickets_pushes_pool_update(app, monkeypatch):
    from tasks.expire_tickets import expire_overdue_tickets

    pushed = []

    monkeypatch.setattr("services.notify_service.notify_pool_update", lambda payload: pushed.append(payload))
    monkeypatch.setattr("services.ticket_pool.get_pool_status", lambda: {"total_pending": 0, "by_type": [], "assigned": 0, "completed_today": 0})

    with app.app_context():
        user = create_user("expire_notify_user", "secret123", client_mode="mode_a")
        uploaded_file = UploadedFile(
            display_id="2026/03/30-03",
            original_filename="expired-notify.txt",
            stored_filename="expired-notify.txt",
            uploaded_by=user.id,
            total_tickets=1,
            pending_count=1,
            assigned_count=0,
            completed_count=0,
        )
        db.session.add(uploaded_file)
        db.session.flush()
        db.session.add(LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="PENDING-NOTIFY-EXPIRED",
            status="pending",
            deadline_time=beijing_now() - timedelta(minutes=1),
        ))
        db.session.commit()

        expire_overdue_tickets()

    assert pushed == [{"total_pending": 0, "by_type": [], "assigned": 0, "completed_today": 0}]


def test_expire_overdue_tickets_expires_exact_deadline(app, monkeypatch):
    from tasks.expire_tickets import expire_overdue_tickets

    fixed_now = datetime(2026, 4, 7, 12, 0, 0)
    monkeypatch.setattr("tasks.expire_tickets.beijing_now", lambda: fixed_now)

    with app.app_context():
        user = create_user("expire_exact_deadline_user", "secret123", client_mode="mode_a")
        uploaded_file = UploadedFile(
            display_id="2026/04/07-90",
            original_filename="expired-exact-deadline.txt",
            stored_filename="expired-exact-deadline.txt",
            uploaded_by=user.id,
            total_tickets=1,
            pending_count=1,
            assigned_count=0,
            completed_count=0,
        )
        db.session.add(uploaded_file)
        db.session.flush()
        ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="PENDING-EXACT-DEADLINE",
            status="pending",
            deadline_time=fixed_now,
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id
        file_id = uploaded_file.id

        expire_overdue_tickets()

        refreshed_ticket = db.session.get(LotteryTicket, ticket_id)
        refreshed_file = db.session.get(UploadedFile, file_id)
        assert refreshed_ticket.status == "expired"
        assert refreshed_file.pending_count == 0


def test_expire_overdue_tickets_increments_ticket_version(app):
    from tasks.expire_tickets import expire_overdue_tickets

    with app.app_context():
        uploaded_file = UploadedFile(
            display_id="2026/04/07-91",
            original_filename="expired-version.txt",
            stored_filename="expired-version.txt",
            uploaded_by=1,
            total_tickets=1,
            pending_count=1,
            assigned_count=0,
            completed_count=0,
        )
        db.session.add(uploaded_file)
        db.session.flush()
        ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="PENDING-VERSION-001",
            status="pending",
            deadline_time=beijing_now() - timedelta(minutes=1),
            version=3,
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

        expire_overdue_tickets()

        refreshed_ticket = db.session.get(LotteryTicket, ticket_id)
        assert refreshed_ticket.status == "expired"
        assert refreshed_ticket.version == 4


def test_archive_old_tickets_deletes_completed_ticket_after_retention(app):
    from tasks.archive import archive_old_tickets

    with app.app_context():
        user = create_user("archive_completed_user", "secret123", client_mode="mode_b")
        uploaded_file = UploadedFile(
            display_id="2026/03/01-01",
            original_filename="archive_completed.txt",
            stored_filename="archive_completed.txt",
            uploaded_by=user.id,
            total_tickets=1,
            completed_count=1,
        )
        db.session.add(uploaded_file)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="ARCHIVE-COMPLETED",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-a",
            admin_upload_time=beijing_now() - timedelta(days=40),
            assigned_at=beijing_now() - timedelta(days=40),
            completed_at=beijing_now() - timedelta(days=35),
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

        deleted_count = archive_old_tickets(days_ago=30)
        archived = ArchivedLotteryTicket.query.filter_by(original_ticket_id=ticket_id).first()
        remaining = LotteryTicket.query.get(ticket_id)

    assert deleted_count == 1
    assert archived is None
    assert remaining is None


def test_archive_old_tickets_deletes_stale_winning_image_files(app):
    from tasks.archive import archive_old_tickets

    with app.app_context():
        user = create_user("archive_winning_image_user", "secret123", client_mode="mode_b")
        uploaded_file = UploadedFile(
            display_id="2026/03/01-03",
            original_filename="archive_winning_image.txt",
            stored_filename="archive_winning_image.txt",
            uploaded_by=user.id,
            total_tickets=1,
            completed_count=1,
        )
        db.session.add(uploaded_file)
        db.session.commit()

        image_dir = Path(app.config["UPLOAD_FOLDER"]) / "images"
        image_dir.mkdir(parents=True, exist_ok=True)
        image_path = image_dir / "archive-winning-stale.png"
        image_path.write_bytes(b"stale-image")

        ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="ARCHIVE-WIN-IMAGE",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now() - timedelta(days=35),
            is_winning=True,
            winning_image_url="/uploads/images/archive-winning-stale.png",
        )
        db.session.add(ticket)
        db.session.commit()
        db.session.add(WinningRecord(
            ticket_id=ticket.id,
            source_file_id=uploaded_file.id,
            winning_image_url=ticket.winning_image_url,
            uploaded_by=user.id,
        ))
        db.session.commit()

        deleted_count = archive_old_tickets(days_ago=30)

    assert deleted_count == 1
    assert not image_path.exists()


def test_archive_old_tickets_uses_fallback_terminal_time_for_expired_and_revoked(app):
    from tasks.archive import archive_old_tickets

    with app.app_context():
        user = create_user("archive_terminal_user", "secret123", client_mode="mode_a")
        uploaded_file = UploadedFile(
            display_id="2026/03/01-02",
            original_filename="archive_terminal.txt",
            stored_filename="archive_terminal.txt",
            uploaded_by=user.id,
            total_tickets=2,
        )
        db.session.add(uploaded_file)
        db.session.commit()

        expired_ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="ARCHIVE-EXPIRED",
            status="expired",
            deadline_time=beijing_now() - timedelta(days=31),
            admin_upload_time=beijing_now() - timedelta(days=32),
        )
        revoked_ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=2,
            raw_content="ARCHIVE-REVOKED",
            status="revoked",
            admin_upload_time=beijing_now() - timedelta(days=33),
        )
        db.session.add_all([expired_ticket, revoked_ticket])
        db.session.commit()
        expired_id = expired_ticket.id
        revoked_id = revoked_ticket.id

        deleted_count = archive_old_tickets(days_ago=30)
        remaining_ids = {
            row.id
            for row in LotteryTicket.query.filter(LotteryTicket.id.in_([expired_id, revoked_id])).all()
        }

    assert deleted_count == 2
    assert remaining_ids == set()


def test_process_uploaded_file_stores_txt_under_business_date_folder(app, monkeypatch):
    from services import file_parser

    monkeypatch.setattr(
        file_parser,
        "build_uploaded_txt_relative_path",
        lambda filename, upload_dt=None: "txt/2026-04-07/mock-folder-check.txt",
    )
    monkeypatch.setattr(
        file_parser,
        "parse_filename",
            lambda filename, upload_dt=None: {
                "identifier": "AA",
                "internal_code": "P7",
                "lottery_type": "???",
                "multiplier": 2,
                "declared_amount": 4.0,
                "declared_count": 1,
                "deadline_hhmm": "23.55",
                "deadline_time": datetime(2026, 4, 7, 23, 55, 0),
                "detail_period": "26034",
            },
    )

    with app.app_context():
        user = create_user("upload_folder_user", "secret123", client_mode="mode_b")
        result = file_parser.process_uploaded_file(
            make_upload_file("AA_P7TEST_600_47_00.55_26034.txt", "SPF|1=3|1*1|2\n"),
            uploader_id=user.id,
        )
        assert result["success"] is True

        uploaded = UploadedFile.query.get(result["file_id"])
        assert uploaded is not None
        normalized = uploaded.stored_filename.replace('\\', '/')
        assert normalized.startswith('txt/2026-04-07/')


def test_archive_old_uploaded_txt_files_deletes_closed_txt_after_retention(app):
    from services.file_parser import process_uploaded_file, resolve_uploaded_txt_path
    from tasks.archive import archive_old_tickets, archive_old_uploaded_txt_files

    with app.app_context():
        user = create_user("archive_txt_user", "secret123", client_mode="mode_b")
        result = process_uploaded_file(
            make_upload_file("AA_P7胜平负3倍投_金额6元_1张_00.55_26034.txt", "SPF|1=3|1*1|3\n"),
            uploader_id=user.id,
        )
        uploaded = UploadedFile.query.get(result["file_id"])
        old_path = resolve_uploaded_txt_path(uploaded.stored_filename, app.config["UPLOAD_FOLDER"])
        assert os.path.exists(old_path)

        uploaded.uploaded_at = beijing_now() - timedelta(days=31)
        uploaded.pending_count = 0
        uploaded.assigned_count = 0
        uploaded.completed_count = uploaded.total_tickets

        for ticket in LotteryTicket.query.filter_by(source_file_id=uploaded.id).all():
            ticket.status = "completed"
            ticket.completed_at = beijing_now() - timedelta(days=31)
        db.session.commit()

        archive_old_tickets(days_ago=30)
        moved = archive_old_uploaded_txt_files(days_ago=30)
        remaining = UploadedFile.query.get(result["file_id"])

    assert moved == 1
    assert not os.path.exists(old_path)
    assert remaining is None


def test_archive_old_uploaded_txt_files_keeps_file_with_recent_ticket_history(app):
    from services.file_parser import process_uploaded_file, resolve_uploaded_txt_path
    from tasks.archive import archive_old_uploaded_txt_files

    with app.app_context():
        user = create_user("archive_txt_recent_ticket_user", "secret123", client_mode="mode_b")
        result = process_uploaded_file(
            make_upload_file("AA_P7胜平负3倍投_金额6元_1张_00.55_26034.txt", "SPF|1=3|1*1|3\n"),
            uploader_id=user.id,
        )
        uploaded = UploadedFile.query.get(result["file_id"])
        old_path = resolve_uploaded_txt_path(uploaded.stored_filename, app.config["UPLOAD_FOLDER"])
        assert os.path.exists(old_path)

        uploaded.uploaded_at = beijing_now() - timedelta(days=31)
        uploaded.pending_count = 0
        uploaded.assigned_count = 0
        uploaded.completed_count = uploaded.total_tickets

        ticket = LotteryTicket.query.filter_by(source_file_id=uploaded.id).first()
        ticket.status = "completed"
        ticket.completed_at = beijing_now() - timedelta(days=5)
        db.session.commit()

        moved = archive_old_uploaded_txt_files(days_ago=30)
        remaining = UploadedFile.query.get(result["file_id"])

    assert moved == 0
    assert os.path.exists(old_path)
    assert remaining is not None


def test_purge_old_auxiliary_records_deletes_old_match_results_before_result_files(app):
    from tasks.archive import purge_old_auxiliary_records
    from models.result import MatchResult, ResultFile

    with app.app_context():
        user = create_user("purge_aux_user", "secret123", client_mode="mode_b")
        result_file = ResultFile(
            original_filename="results.txt",
            stored_filename="results_old.txt",
            uploaded_by=user.id,
            uploaded_at=beijing_now() - timedelta(days=31),
            periods_count=1,
        )
        db.session.add(result_file)
        db.session.commit()

        stored_path = Path(app.config["UPLOAD_FOLDER"]) / result_file.stored_filename
        stored_path.write_text("old result payload", encoding="utf-8")

        match_result = MatchResult(
            detail_period="26034",
            lottery_type="P7",
            result_data={"61": {"SPF": {"result": "3", "sp": 1.85}}},
            result_file_id=result_file.id,
            uploaded_by=user.id,
            uploaded_at=beijing_now() - timedelta(days=31),
        )
        db.session.add(match_result)
        db.session.commit()
        result_file_id = result_file.id
        match_result_id = match_result.id

        purge_old_auxiliary_records(days_ago=30)

        remaining_result_file = ResultFile.query.get(result_file_id)
        remaining_match_result = MatchResult.query.get(match_result_id)

    assert remaining_match_result is None
    assert remaining_result_file is None
    assert not stored_path.exists()


def test_purge_old_auxiliary_records_does_not_delete_outside_upload_folder(app, tmp_path):
    from tasks.archive import purge_old_auxiliary_records
    from models.result import MatchResult, ResultFile

    with app.app_context():
        upload_root = tmp_path / "uploads"
        upload_root.mkdir(parents=True, exist_ok=True)
        app.config["UPLOAD_FOLDER"] = str(upload_root)

        user = create_user("purge_aux_path_guard_user", "secret123", client_mode="mode_b")
        result_file = ResultFile(
            original_filename="malicious.txt",
            stored_filename="..\\outside_result_payload.txt",
            uploaded_by=user.id,
            uploaded_at=beijing_now() - timedelta(days=31),
            periods_count=1,
        )
        db.session.add(result_file)
        db.session.commit()

        outside_path = tmp_path / "outside_result_payload.txt"
        outside_path.write_text("outside", encoding="utf-8")

        match_result = MatchResult(
            detail_period="26035",
            lottery_type="P7",
            result_data={"61": {"SPF": {"result": "3", "sp": 1.85}}},
            result_file_id=result_file.id,
            uploaded_by=user.id,
            uploaded_at=beijing_now() - timedelta(days=31),
        )
        db.session.add(match_result)
        db.session.commit()
        result_file_id = result_file.id
        match_result_id = match_result.id

        purge_old_auxiliary_records(days_ago=30)

        remaining_result_file = ResultFile.query.get(result_file_id)
        remaining_match_result = MatchResult.query.get(match_result_id)

    assert outside_path.exists() is True
    assert remaining_match_result is None
    assert remaining_result_file is None


def test_resolve_uploaded_txt_path_rejects_paths_outside_upload_folder(tmp_path):
    from services.file_parser import resolve_uploaded_txt_path

    upload_root = tmp_path / "uploads"
    upload_root.mkdir(parents=True, exist_ok=True)

    outside_resolved = resolve_uploaded_txt_path("..\\outside_payload.txt", str(upload_root))
    assert outside_resolved == ""

    absolute_outside = resolve_uploaded_txt_path(str(tmp_path / "outside_absolute.txt"), str(upload_root))
    assert absolute_outside == ""


def test_resolve_uploaded_txt_path_does_not_fallback_to_basename_file(tmp_path):
    from services.file_parser import resolve_uploaded_txt_path

    upload_root = tmp_path / "uploads"
    upload_root.mkdir(parents=True, exist_ok=True)
    (upload_root / "existing.txt").write_text("safe", encoding="utf-8")

    resolved = resolve_uploaded_txt_path("txt/2026-04-07/existing.txt", str(upload_root))
    expected = os.path.abspath(os.path.join(str(upload_root), "txt/2026-04-07/existing.txt"))
    assert resolved == expected
    assert os.path.exists(resolved) is False


def test_delete_uploaded_txt_file_does_not_delete_outside_upload_folder(app, tmp_path):
    from services.file_parser import delete_uploaded_txt_file

    with app.app_context():
        upload_root = tmp_path / "uploads"
        upload_root.mkdir(parents=True, exist_ok=True)
        app.config["UPLOAD_FOLDER"] = str(upload_root)

        outside_path = tmp_path / "outside_uploaded_txt.txt"
        outside_path.write_text("outside", encoding="utf-8")

        uploaded = UploadedFile(
            original_filename="outside_uploaded_txt.txt",
            stored_filename="..\\outside_uploaded_txt.txt",
        )
        deleted = delete_uploaded_txt_file(uploaded, app.config["UPLOAD_FOLDER"])

    assert deleted is False
    assert outside_path.exists() is True


def test_public_register_page_redirects_to_login(client):
    resp = client.get("/auth/register", follow_redirects=False)
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/auth/login")


def test_public_register_json_returns_403(client):
    resp = client.post(
        "/auth/register",
        json={"username": "new_form_user", "password": "secret123"},
    )
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False


def test_dashboard_uses_correct_change_password_endpoint():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "fetch('/api/user/change-password'" in content
    assert "fetch('/user/change-password'" not in content


def test_login_page_submits_device_id():
    login_template = Path(__file__).resolve().parents[1] / "templates" / "login.html"
    content = login_template.read_text(encoding="utf-8")
    assert "getOrCreateDeviceId" in content
    assert "device_id: deviceId || undefined" in content


def test_login_page_uses_readable_chinese_labels():
    login_template = Path(__file__).resolve().parents[1] / "templates" / "login.html"
    content = login_template.read_text(encoding="utf-8")
    assert "登录 - 数据文件管理平台" in content
    assert "📁 数据文件管理平台" in content
    assert "用户名" in content
    assert "请输入用户名" in content
    assert "密码" in content
    assert "请输入密码" in content
    assert "鐧诲綍" not in content


def test_admin_dashboard_template_uses_readable_chinese_labels():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "管理后台" in content
    # 数据库信息已迁移到设置页，不再在 dashboard 中
    assert "总速度(张/分)" in content
    assert "暂无在线用户" in content
    assert "showToast(e.message || '操作失败', 'danger');" in content


def test_admin_settings_template_handles_save_failures():
    settings_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "settings.html"
    content = settings_template.read_text(encoding="utf-8")
    assert "v-if=\"error\"" in content
    assert "error: ''" in content
    assert "mode_b_pool_reserve" in content
    assert "if (!res.ok || data.success === false) {" in content
    assert "throw new Error(data.error || '\u4fdd\u5b58\u5931\u8d25');" in content
    assert "this.error = e.message || '\u4fdd\u5b58\u5931\u8d25';" in content
    assert "this.error = e.message || '\u52a0\u8f7d\u8bbe\u7f6e\u5931\u8d25';" in content


def test_admin_users_template_uses_readable_chinese_labels():
    users_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "users.html"
    content = users_template.read_text(encoding="utf-8")
    assert "用户管理" in content
    assert "最大设备数" in content
    assert "接单开关" in content
    assert "确认踢出用户" in content
    assert "请输入用户名和密码" in content


def test_admin_users_template_handles_update_failures():
    users_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "users.html"
    content = users_template.read_text(encoding="utf-8")
    assert "const original = {" in content
    assert "try {" in content
    assert "if (!res.ok || data.success === false) {" in content
    assert "throw new Error(data.error || '\u66f4\u65b0\u5931\u8d25');" in content
    assert "Object.assign(u, original);" in content
    assert "showToast(e.message || '\u66f4\u65b0\u5931\u8d25', 'danger');" in content
    assert "await this.loadUsers();" in content


def test_admin_users_template_checks_http_status_for_mutations():
    users_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "users.html"
    content = users_template.read_text(encoding="utf-8")
    assert content.count("if (!res.ok || data.success === false) {") >= 5
    assert "throw new Error(data.error || '创建失败');" in content
    assert "throw new Error(data.error || '更新失败');" in content
    assert "throw new Error(data.error || '操作失败');" in content
    assert "throw new Error(data.error || '删除失败');" in content
    assert "showToast('已强制下线', 'success');" in content
    assert "showToast('密码已更新', 'success');" in content


def test_winning_presign_uses_local_upload_api_when_oss_disabled(app):
    with app.app_context():
        from services.oss_service import generate_presign_url

        url, oss_key = generate_presign_url("winning/2026/04/07/123.jpg")

    assert url == "/api/winning/upload-local?key=winning_2026_04_07_123.jpg"
    assert oss_key == "winning_2026_04_07_123.jpg"


def test_winning_upload_local_saves_image_under_uploads_images(app, client):
    from PIL import Image

    with app.app_context():
        user = create_user("winning_local_upload_user", "secret123", client_mode="mode_a")
        upload_folder = Path(app.config["UPLOAD_FOLDER"])
        ticket = create_assigned_ticket(user, "device-local", "WIN-LOCAL-001", 1)
        ticket.is_winning = True
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_local_upload_user", "secret123")
    assert resp.status_code == 200

    presign_resp = client.get(f"/api/winning/presign?ticket_id={ticket_id}")
    assert presign_resp.status_code == 200
    presign_data = presign_resp.get_json()
    assert presign_data["success"] is True
    assert presign_data["url"].startswith("/api/winning/upload-local?key=")

    image_bytes = io.BytesIO()
    Image.new("RGB", (20, 20), color="blue").save(image_bytes, format="PNG")
    image_bytes.seek(0)

    upload_resp = client.post(
        presign_data["url"],
        data={"file": (image_bytes, "winning.png")},
        content_type="multipart/form-data",
    )
    assert upload_resp.status_code == 200
    upload_data = upload_resp.get_json()
    assert upload_data["success"] is True
    assert upload_data["image_url"].startswith("/uploads/images/winning_")
    assert upload_data["oss_key"].startswith("winning_")

    saved_path = upload_folder / "images" / upload_data["oss_key"]
    assert saved_path.exists()


def test_winning_upload_local_requires_matching_ticket_key(app, client):
    from PIL import Image

    with app.app_context():
        user = create_user("winning_local_guard_user", "secret123", client_mode="mode_a")
        ticket = create_assigned_ticket(user, "device-local", "WIN-LOCAL-GUARD", 1)
        ticket.is_winning = True
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_local_guard_user", "secret123")
    assert resp.status_code == 200

    image_bytes = io.BytesIO()
    Image.new("RGB", (20, 20), color="green").save(image_bytes, format="PNG")
    image_bytes.seek(0)

    upload_resp = client.post(
        f"/api/winning/upload-local?ticket_id={ticket_id}&key=winning_2026_04_07_999.jpg",
        data={"file": (image_bytes, "winning.png")},
        content_type="multipart/form-data",
    )
    assert upload_resp.status_code == 400
    data = upload_resp.get_json()
    assert data["success"] is False
    assert "key" in data["error"]


def test_winning_upload_local_accepts_ticket_bound_key_even_when_date_differs(app, client):
    from PIL import Image

    with app.app_context():
        user = create_user("winning_local_date_tolerant_user", "secret123", client_mode="mode_a")
        ticket = create_assigned_ticket(user, "device-local", "WIN-LOCAL-DATE-TOLERANT", 1)
        ticket.is_winning = True
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_local_date_tolerant_user", "secret123")
    assert resp.status_code == 200

    image_bytes = io.BytesIO()
    Image.new("RGB", (20, 20), color="purple").save(image_bytes, format="PNG")
    image_bytes.seek(0)

    upload_resp = client.post(
        f"/api/winning/upload-local?ticket_id={ticket_id}&key=winning_2000_01_01_{ticket_id}.jpg",
        data={"file": (image_bytes, "winning.png")},
        content_type="multipart/form-data",
    )
    assert upload_resp.status_code == 200
    data = upload_resp.get_json()
    assert data["success"] is True
    assert data["oss_key"].startswith(f"winning_2000_01_01_{ticket_id}")


def test_admin_export_tickets_csv_uses_business_window_without_name_error(app, client, monkeypatch):
    business_start = datetime(2026, 4, 6, 12, 0, 0)

    with app.app_context():
        admin = User(username="admin_csv_export_user", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        in_window = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="CSV-IN-WINDOW",
            status="completed",
            assigned_username="user-a",
            completed_at=business_start + timedelta(hours=1),
        )
        expired_in_window = LotteryTicket(
            source_file_id=1,
            line_number=3,
            raw_content="CSV-EXPIRED-IN-WINDOW",
            status="expired",
            assigned_username="user-a",
            deadline_time=business_start + timedelta(hours=2),
        )
        out_of_window = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="CSV-OUT-OF-WINDOW",
            status="completed",
            assigned_username="user-a",
            completed_at=business_start - timedelta(hours=1),
        )
        db.session.add_all([in_window, expired_in_window, out_of_window])
        db.session.commit()

    monkeypatch.setattr("routes.admin.get_today_noon", lambda: business_start)
    monkeypatch.setattr("routes.admin.get_business_date", lambda dt=None: business_start.date())

    resp = client.post("/auth/login", json={"username": "admin_csv_export_user", "password": "secret123"})
    assert resp.status_code == 200

    export_resp = client.get("/admin/api/tickets/export")
    assert export_resp.status_code == 200
    csv_text = export_resp.data.decode("utf-8-sig")
    assert "\u8bbe\u5907ID" in csv_text
    assert "CSV-IN-WINDOW" in csv_text
    assert "CSV-EXPIRED-IN-WINDOW" in csv_text
    assert "CSV-OUT-OF-WINDOW" not in csv_text


def test_user_export_daily_uses_business_date_filename(app, client, monkeypatch):
    from decimal import Decimal

    business_start = datetime(2026, 4, 6, 12, 0, 0)

    with app.app_context():
        user = create_user("user_export_daily_name", "secret123", client_mode="mode_a")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="USER-EXPORT-DAILY",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="\u8bbe\u5907A",
            deadline_time=business_start + timedelta(hours=1),
            completed_at=business_start + timedelta(hours=2),
            detail_period="26034",
            ticket_amount=Decimal("4"),
        )
        db.session.add(ticket)
        db.session.commit()

    monkeypatch.setattr("routes.user.get_today_noon", lambda: business_start)
    monkeypatch.setattr("routes.user.get_business_date", lambda dt=None: business_start.date())
    monkeypatch.setattr("routes.user.beijing_now", lambda: business_start + timedelta(hours=3))

    resp = login(client, "user_export_daily_name", "secret123")
    assert resp.status_code == 200

    export_resp = client.get("/api/user/export-daily")
    assert export_resp.status_code == 200
    assert "attachment;" in export_resp.headers["Content-Disposition"]
    assert "2026-04-06" in export_resp.headers["Content-Disposition"]


def test_admin_export_tickets_by_date_includes_device_id_column():
    admin_route = Path(__file__).resolve().parents[1] / "routes" / "admin.py"
    content = admin_route.read_text(encoding="utf-8")
    assert "['行号', '原始内容', '彩种', '倍投', '截止时间', '期号', '金额', '状态', '用户名', '设备ID', '分配时间', '分配文件名', '完成时间', '来源文件名']" in content
    assert "t.assigned_device_id or ''" in content


def test_admin_export_tickets_by_date_empty_uses_selected_business_date_filename():
    admin_route = Path(__file__).resolve().parents[1] / "routes" / "admin.py"
    content = admin_route.read_text(encoding="utf-8")
    assert 'empty_filename = f"{date_str}_无数据投注内容详情.xlsx"' in content
    assert "filename*=UTF-8''{empty_filename_encoded}" in content


def test_admin_export_tickets_by_date_default_filename_uses_business_date(app, client, monkeypatch):
    with app.app_context():
        admin = User(username="admin_export_default_date_name", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    monkeypatch.setattr("routes.admin.get_business_date", lambda dt=None: datetime(2026, 4, 9).date())

    resp = client.post("/auth/login", json={"username": "admin_export_default_date_name", "password": "secret123"})
    assert resp.status_code == 200

    export_resp = client.get("/admin/api/tickets/export-by-date")
    assert export_resp.status_code == 200
    assert "attachment;" in export_resp.headers["Content-Disposition"]
    assert "2026-04-09" in export_resp.headers["Content-Disposition"]


def test_admin_export_tickets_by_date_localizes_status_labels_to_chinese(app, client):
    from decimal import Decimal
    from openpyxl import load_workbook

    with app.app_context():
        admin = User(username="admin_export_status_cn", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        uploaded = UploadedFile(
            original_filename="status_source.txt",
            stored_filename="status_source.txt",
            status="active",
            uploaded_at=datetime(2026, 4, 8, 13, 0, 0),
            total_tickets=1,
            pending_count=0,
            assigned_count=1,
            completed_count=0,
        )
        db.session.add(uploaded)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="STATUS-LOCALIZED",
            lottery_type="胜平负",
            multiplier=1,
            detail_period="26040",
            ticket_amount=Decimal("2"),
            status="assigned",
            assigned_username="tester",
            assigned_device_id="dev-status",
            assigned_at=datetime(2026, 4, 8, 13, 5, 0),
        )
        db.session.add(ticket)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_export_status_cn", "password": "secret123"})
    assert resp.status_code == 200

    export_resp = client.get("/admin/api/tickets/export-by-date?date=2026-04-08")
    assert export_resp.status_code == 200

    workbook = load_workbook(io.BytesIO(export_resp.data))
    worksheet = workbook.active
    assert worksheet.cell(row=2, column=8).value == "处理中"


def test_admin_export_tickets_by_date_includes_download_filename(app, client):
    from decimal import Decimal
    from openpyxl import load_workbook

    download_filename = "比分_2倍_53张_1112元_02.40_2026-0429-011309.txt"

    with app.app_context():
        admin = User(username="admin_export_download_filename", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        uploaded = UploadedFile(
            original_filename="source_file.txt",
            stored_filename="source_file.txt",
            status="active",
            uploaded_at=datetime(2026, 4, 8, 13, 0, 0),
            total_tickets=1,
            pending_count=0,
            assigned_count=0,
            completed_count=1,
        )
        db.session.add(uploaded)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="EXPORT-DOWNLOAD-FILENAME",
            lottery_type="比分",
            multiplier=2,
            detail_period="26040",
            ticket_amount=Decimal("4"),
            status="completed",
            assigned_username="tester",
            assigned_device_id="dev-export",
            assigned_at=datetime(2026, 4, 8, 13, 5, 0),
            completed_at=datetime(2026, 4, 8, 13, 10, 0),
            download_filename=download_filename,
        )
        db.session.add(ticket)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_export_download_filename", "password": "secret123"})
    assert resp.status_code == 200

    export_resp = client.get("/admin/api/tickets/export-by-date?date=2026-04-08")
    assert export_resp.status_code == 200

    workbook = load_workbook(io.BytesIO(export_resp.data))
    worksheet = workbook.active
    header = [cell for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = [cell for cell in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    assert "分配文件名" in header
    assert row[header.index("分配文件名")] == download_filename
    assert row[header.index("来源文件名")] == "source_file.txt"


def test_admin_dashboard_eta_uses_chinese_labels():
    admin_route = Path(__file__).resolve().parents[1] / "routes" / "admin.py"
    content = admin_route.read_text(encoding="utf-8")
    assert '"\\u8d85\\u8fc7 7 \\u5929"' in content
    assert "\\u5c0f\\u65f6" in content
    assert "\\u5206\\u949f" in content
    assert '"over 7 days"' not in content


def test_admin_routes_avoid_legacy_query_get_for_file_reads():
    admin_route = Path(__file__).resolve().parents[1] / "routes" / "admin.py"
    content = admin_route.read_text(encoding="utf-8")
    assert "db.session.get(UploadedFile, file_id)" in content
    assert "db.session.get(UF, t.source_file_id)" in content


def test_core_services_avoid_legacy_query_get_usage():
    root = Path(__file__).resolve().parents[1]
    targets = [
        root / "models" / "user.py",
        root / "routes" / "winning.py",
        root / "services" / "mode_a_service.py",
        root / "services" / "mode_b_service.py",
        root / "services" / "ticket_pool.py",
        root / "services" / "winning_calc_service.py",
    ]
    for path in targets:
        content = path.read_text(encoding="utf-8")
        assert ".query.get(" not in content
        assert "Query.get(" not in content


def test_admin_files_list_rejects_invalid_page_params(app, client):
    with app.app_context():
        admin = User(username="admin_invalid_page_user", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_invalid_page_user", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/files?page=bad")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "\u5206\u9875\u53c2\u6570" in data["error"]


def test_uploaded_file_to_dict_uses_derived_status(app):
    with app.app_context():
        now = beijing_now()
        completed_file = UploadedFile(
            display_id="2026/04/07-01",
            original_filename="done.txt",
            stored_filename="txt/2026-04-07/done.txt",
            uploaded_by=1,
            total_tickets=5,
            completed_count=5,
            pending_count=0,
            assigned_count=0,
            deadline_time=now + timedelta(hours=1),
        )
        expired_file = UploadedFile(
            display_id="2026/04/07-02",
            original_filename="expired.txt",
            stored_filename="txt/2026-04-07/expired.txt",
            uploaded_by=1,
            total_tickets=5,
            completed_count=2,
            pending_count=0,
            assigned_count=0,
            deadline_time=now - timedelta(hours=1),
        )
        exact_deadline_file = UploadedFile(
            display_id="2026/04/07-03",
            original_filename="expired-exact.txt",
            stored_filename="txt/2026-04-07/expired-exact.txt",
            uploaded_by=1,
            total_tickets=5,
            completed_count=2,
            pending_count=0,
            assigned_count=0,
            deadline_time=now,
        )
        db.session.add_all([completed_file, expired_file, exact_deadline_file])
        db.session.commit()

        assert completed_file.to_dict()["status"] == "exhausted"
        assert expired_file.to_dict()["status"] == "expired"
        assert exact_deadline_file.to_dict()["status"] == "expired"


def test_revoke_file_succeeds_even_when_realtime_notify_fails(app, monkeypatch):
    from decimal import Decimal
    from services import file_parser

    with app.app_context():
        admin = User(username="revoke_notify_admin", is_admin=True)
        admin.set_password("secret123")
        user = create_user("revoke_notify_user", "secret123", client_mode="mode_b")
        db.session.add(admin)
        db.session.commit()

        uploaded = UploadedFile(
            display_id="2026/04/07-03",
            original_filename="revoke-test.txt",
            stored_filename="txt/2026-04-07/revoke-test.txt",
            uploaded_by=admin.id,
            total_tickets=2,
            pending_count=1,
            assigned_count=1,
            completed_count=0,
            deadline_time=beijing_now() + timedelta(hours=1),
        )
        db.session.add(uploaded)
        db.session.flush()

        pending_ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="SPF|1=3|1*1|2",
            lottery_type="???",
            multiplier=1,
            detail_period="26034",
            ticket_amount=Decimal("4"),
            deadline_time=uploaded.deadline_time,
            status="pending",
            admin_upload_time=beijing_now(),
        )
        assigned_ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=2,
            raw_content="SPF|1=0|1*1|2",
            lottery_type="???",
            multiplier=1,
            detail_period="26034",
            ticket_amount=Decimal("4"),
            deadline_time=uploaded.deadline_time,
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-rv",
            assigned_at=beijing_now(),
            admin_upload_time=beijing_now(),
        )
        db.session.add_all([pending_ticket, assigned_ticket])
        db.session.commit()

        monkeypatch.setattr(
            "services.notify_service.notify_all",
            lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("socket down")),
        )

        result = file_parser.revoke_file(uploaded.id, admin.id)

        assert result["success"] is True

        refreshed_file = db.session.get(UploadedFile, uploaded.id)
        statuses = {
            ticket.status
            for ticket in LotteryTicket.query.filter_by(source_file_id=uploaded.id).all()
        }
        assert refreshed_file.status == "revoked"
        assert refreshed_file.pending_count == 0
        assert refreshed_file.assigned_count == 0
        assert statuses == {"revoked"}


def test_revoke_file_increments_ticket_versions(app):
    from decimal import Decimal
    from services import file_parser

    with app.app_context():
        admin = User(username="revoke_version_admin", is_admin=True)
        admin.set_password("secret123")
        user = create_user("revoke_version_user", "secret123", client_mode="mode_b")
        db.session.add(admin)
        db.session.commit()

        uploaded = UploadedFile(
            display_id="2026/04/07-04",
            original_filename="revoke-version.txt",
            stored_filename="txt/2026-04-07/revoke-version.txt",
            uploaded_by=admin.id,
            total_tickets=2,
            pending_count=1,
            assigned_count=1,
            completed_count=0,
            deadline_time=beijing_now() + timedelta(hours=1),
        )
        db.session.add(uploaded)
        db.session.flush()

        pending_ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="SPF|2=3|1*1|2",
            lottery_type="???",
            multiplier=1,
            detail_period="26034",
            ticket_amount=Decimal("4"),
            deadline_time=uploaded.deadline_time,
            status="pending",
            version=2,
            admin_upload_time=beijing_now(),
        )
        assigned_ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=2,
            raw_content="SPF|2=0|1*1|2",
            lottery_type="???",
            multiplier=1,
            detail_period="26034",
            ticket_amount=Decimal("4"),
            deadline_time=uploaded.deadline_time,
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-rv",
            assigned_at=beijing_now(),
            version=5,
            admin_upload_time=beijing_now(),
        )
        db.session.add_all([pending_ticket, assigned_ticket])
        db.session.commit()

        result = file_parser.revoke_file(uploaded.id, admin.id)
        db.session.expire_all()
        refreshed_pending = db.session.get(LotteryTicket, pending_ticket.id)
        refreshed_assigned = db.session.get(LotteryTicket, assigned_ticket.id)

    assert result["success"] is True
    assert refreshed_pending.status == "revoked"
    assert refreshed_assigned.status == "revoked"
    assert refreshed_pending.version == 3
    assert refreshed_assigned.version == 6


def test_admin_revoke_rejects_non_active_files(app, client):
    with app.app_context():
        admin = User(username="admin_revoke_non_active", is_admin=True)
        admin.set_password("secret123")
        exhausted_file = UploadedFile(
            display_id="2026/04/07-98",
            original_filename="exhausted-no-revoke.txt",
            stored_filename="txt/2026-04-07/exhausted-no-revoke.txt",
            uploaded_by=admin.id,
            total_tickets=2,
            pending_count=0,
            assigned_count=0,
            completed_count=2,
            deadline_time=beijing_now() + timedelta(hours=1),
            status="active",
        )
        db.session.add(admin)
        db.session.flush()
        exhausted_file.uploaded_by = admin.id
        db.session.add(exhausted_file)
        db.session.commit()
        file_id = exhausted_file.id

    resp = login(client, "admin_revoke_non_active", "secret123")
    assert resp.status_code == 200

    revoke_resp = client.post(f"/admin/api/files/{file_id}/revoke")
    assert revoke_resp.status_code == 400
    data = revoke_resp.get_json()
    assert data["success"] is False
    assert "\u4e0d\u80fd\u64a4\u56de" in data["message"]

    with app.app_context():
        refreshed = db.session.get(UploadedFile, file_id)
        assert refreshed.status == "active"


def test_admin_files_list_filters_by_derived_status(app, client):
    with app.app_context():
        admin = User(username="admin_file_status_filter", is_admin=True)
        admin.set_password("secret123")
        exhausted_file = UploadedFile(
            display_id="2026/04/07-03",
            original_filename="exhausted.txt",
            stored_filename="txt/2026-04-07/exhausted.txt",
            uploaded_by=1,
            total_tickets=3,
            completed_count=3,
            pending_count=0,
            assigned_count=0,
            deadline_time=beijing_now() + timedelta(hours=1),
        )
        active_file = UploadedFile(
            display_id="2026/04/07-04",
            original_filename="active.txt",
            stored_filename="txt/2026-04-07/active.txt",
            uploaded_by=1,
            total_tickets=3,
            completed_count=1,
            pending_count=2,
            assigned_count=0,
            deadline_time=beijing_now() + timedelta(hours=1),
        )
        db.session.add_all([admin, exhausted_file, active_file])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_file_status_filter", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/files?status=exhausted")
    assert resp.status_code == 200
    data = resp.get_json()
    names = [item["original_filename"] for item in data["files"]]
    assert "exhausted.txt" in names
    assert "active.txt" not in names


def test_admin_file_management_endpoints_require_login_json_response(app, client):
    list_resp = client.get("/admin/api/files")
    assert list_resp.status_code == 401
    assert list_resp.is_json is True
    assert list_resp.get_json()["success"] is False

    upload_resp = client.post("/admin/files/upload")
    assert upload_resp.status_code == 401
    assert upload_resp.is_json is True
    assert upload_resp.get_json()["success"] is False


def test_admin_file_detail_returns_json_for_missing_file(app, client):
    with app.app_context():
        admin = User(username="admin_missing_file_detail", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_missing_file_detail", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/files/999999/detail")
    assert resp.status_code == 404
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_admin_revoke_missing_file_returns_404_json(app, client):
    with app.app_context():
        admin = User(username="admin_missing_revoke_file", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_missing_revoke_file", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post("/admin/api/files/999999/revoke")
    assert resp.status_code == 404
    assert resp.is_json is True
    data = resp.get_json()
    assert data["success"] is False
    assert data["message"]


def test_admin_create_user_rejects_invalid_max_devices(app, client):
    with app.app_context():
        admin = User(username="admin_create_user_guard", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_create_user_guard", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/users",
        json={"username": "bad_max_devices_user", "password": "secret123", "max_devices": "bad"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_admin_update_settings_handles_empty_json_body(app, client):
    with app.app_context():
        admin = User(username="admin_settings_empty_body", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_empty_body", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", data="", content_type="application/json")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert "settings" in data


def test_admin_update_settings_rejects_invalid_session_hours(app, client):
    with app.app_context():
        admin = User(username="admin_settings_hours_guard", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_hours_guard", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"session_lifetime_hours": 0})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_admin_update_settings_reschedules_daily_reset_job(app, client, monkeypatch):
    rescheduled = []

    def fake_reschedule(app_obj, hour):
        rescheduled.append(hour)

    monkeypatch.setattr("tasks.scheduler.reschedule_daily_reset", fake_reschedule)

    with app.app_context():
        admin = User(username="admin_settings_reschedule", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_reschedule", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"daily_reset_hour": 8})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["settings"]["daily_reset_hour"] == 8
    assert rescheduled == [8]


def test_admin_update_settings_normalizes_mode_b_options(app, client):
    with app.app_context():
        admin = User(username="admin_settings_modeb_guard", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_modeb_guard", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"mode_b_options": [200, "50", 200, "100"]})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["settings"]["mode_b_options"] == [200, 50, 100]


def test_admin_update_settings_accepts_mode_b_pool_reserve(app, client):
    with app.app_context():
        admin = User(username="admin_settings_modeb_reserve", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_modeb_reserve", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"mode_b_pool_reserve": 35})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["settings"]["mode_b_pool_reserve"] == 35

    with app.app_context():
        assert SystemSettings.get().mode_b_pool_reserve == 35


def test_admin_update_settings_rejects_negative_mode_b_pool_reserve(app, client):
    with app.app_context():
        admin = User(username="admin_settings_modeb_reserve_guard", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_modeb_reserve_guard", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.put("/admin/api/settings", json={"mode_b_pool_reserve": -1})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_admin_dashboard_data_includes_normal_health_summary(app, client, monkeypatch):
    expected_job_ids = {
        "expire_tickets",
        "clean_sessions",
        "daily_reset",
        "db_keepalive",
        "archive_tickets",
        "archive_uploaded_txt_files",
        "purge_old_auxiliary_records",
    }

    class FakeJob:
        def __init__(self, job_id):
            self.id = job_id

    class FakeScheduler:
        def get_jobs(self):
            return [FakeJob(job_id) for job_id in sorted(expected_job_ids)]

    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: FakeScheduler())

    with app.app_context():
        admin = User(username="admin_dashboard_health_normal", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_dashboard_health_normal", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/dashboard-data")
    assert resp.status_code == 200
    data = resp.get_json()
    health = data["health_summary"]
    assert health["status"] == "normal"
    assert health["items"] == []
    assert "系统正常" in health["summary"]
    assert health["generated_at"]


def test_admin_dashboard_data_marks_overdue_pending_as_warning(app, client, monkeypatch):
    expected_job_ids = {
        "expire_tickets",
        "clean_sessions",
        "daily_reset",
        "db_keepalive",
        "archive_tickets",
        "archive_uploaded_txt_files",
        "purge_old_auxiliary_records",
    }

    class FakeJob:
        def __init__(self, job_id):
            self.id = job_id

    class FakeScheduler:
        def get_jobs(self):
            return [FakeJob(job_id) for job_id in sorted(expected_job_ids)]

    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: FakeScheduler())

    with app.app_context():
        admin = User(username="admin_dashboard_health_warning", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        uploaded_file = UploadedFile(
            original_filename="warning.txt",
            stored_filename="warning.txt",
            uploaded_by=admin.id,
            status="active",
            total_tickets=1,
            pending_count=1,
            assigned_count=0,
            completed_count=0,
            deadline_time=beijing_now() - timedelta(hours=1),
        )
        db.session.add(uploaded_file)
        db.session.flush()

        overdue_ticket = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="1(3.00)",
            status="pending",
            deadline_time=beijing_now() - timedelta(minutes=30),
        )
        db.session.add(overdue_ticket)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_dashboard_health_warning", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/dashboard-data")
    assert resp.status_code == 200
    data = resp.get_json()
    health = data["health_summary"]
    assert health["status"] == "warning"
    overdue_item = next(item for item in health["items"] if item["type"] == "overdue_pending_tickets")
    assert overdue_item["count"] >= 1


def test_admin_dashboard_data_marks_result_parse_error_as_critical(app, client, monkeypatch):
    expected_job_ids = {
        "expire_tickets",
        "clean_sessions",
        "daily_reset",
        "db_keepalive",
        "archive_tickets",
        "archive_uploaded_txt_files",
        "purge_old_auxiliary_records",
    }

    class FakeJob:
        def __init__(self, job_id):
            self.id = job_id

    class FakeScheduler:
        def get_jobs(self):
            return [FakeJob(job_id) for job_id in sorted(expected_job_ids)]

    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: FakeScheduler())

    with app.app_context():
        admin = User(username="admin_dashboard_health_critical", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        db.session.add(
            ResultFile(
                original_filename="bad_result.txt",
                stored_filename="bad_result.txt",
                uploaded_by=admin.id,
                status="error",
                parse_error="bad format",
            )
        )
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_dashboard_health_critical", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/dashboard-data")
    assert resp.status_code == 200
    data = resp.get_json()
    health = data["health_summary"]
    assert health["status"] == "critical"
    parse_error_item = next(item for item in health["items"] if item["type"] == "result_file_parse_error")
    assert parse_error_item["count"] >= 1


def test_admin_dashboard_data_does_not_warn_for_missing_winning_record_or_image(app, client, monkeypatch):
    expected_job_ids = {
        "expire_tickets",
        "clean_sessions",
        "daily_reset",
        "db_keepalive",
        "archive_tickets",
        "archive_uploaded_txt_files",
        "purge_old_auxiliary_records",
    }

    class FakeJob:
        def __init__(self, job_id):
            self.id = job_id

    class FakeScheduler:
        def get_jobs(self):
            return [FakeJob(job_id) for job_id in sorted(expected_job_ids)]

    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: FakeScheduler())

    with app.app_context():
        admin = User(username="admin_dashboard_health_winning_optional", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        uploaded_file = UploadedFile(
            original_filename="winning_optional.txt",
            stored_filename="winning_optional.txt",
            uploaded_by=admin.id,
            status="active",
            total_tickets=2,
            pending_count=0,
            assigned_count=0,
            completed_count=2,
            deadline_time=beijing_now() + timedelta(hours=1),
        )
        db.session.add(uploaded_file)
        db.session.flush()

        ticket_with_record = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=1,
            raw_content="SPF|61=3|1*1|1",
            status="completed",
            is_winning=True,
            winning_amount=10,
            winning_gross=12,
            winning_tax=2,
            deadline_time=beijing_now() + timedelta(hours=1),
            assigned_at=beijing_now(),
            completed_at=beijing_now(),
        )
        ticket_without_record = LotteryTicket(
            source_file_id=uploaded_file.id,
            line_number=2,
            raw_content="SPF|62=3|1*1|1",
            status="completed",
            is_winning=True,
            winning_amount=8,
            winning_gross=10,
            winning_tax=2,
            deadline_time=beijing_now() + timedelta(hours=1),
            assigned_at=beijing_now(),
            completed_at=beijing_now(),
        )
        db.session.add_all([ticket_with_record, ticket_without_record])
        db.session.flush()

        db.session.add(
            WinningRecord(
                ticket_id=ticket_with_record.id,
                source_file_id=uploaded_file.id,
                detail_period="26034",
                lottery_type="???",
                winning_amount=10,
                winning_image_url=None,
                uploaded_by=admin.id,
                is_checked=False,
            )
        )
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_dashboard_health_winning_optional", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/dashboard-data")
    assert resp.status_code == 200
    data = resp.get_json()
    health = data["health_summary"]
    item_types = {item["type"] for item in health["items"]}
    assert "winning_ticket_missing_record" not in item_types
    assert "winning_record_missing_image" not in item_types
    assert health["status"] == "normal"


def test_admin_dashboard_data_health_summary_fallback_does_not_break_core_payload(app, client, monkeypatch):
    def _raise_summary_error(*_args, **_kwargs):
        raise RuntimeError("summary unavailable")

    monkeypatch.setattr("routes.admin._build_health_summary", _raise_summary_error)

    with app.app_context():
        admin = User(username="admin_dashboard_health_fallback", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_dashboard_health_fallback", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/dashboard-data")
    assert resp.status_code == 200
    data = resp.get_json()
    assert "pool" in data
    assert "online_users" in data
    health = data["health_summary"]
    assert health["status"] == "warning"
    assert health["items"][0]["type"] == "summary_unavailable"


def test_admin_settings_includes_scheduler_status_normal(app, client, monkeypatch):
    expected_job_ids = [
        "expire_tickets",
        "clean_sessions",
        "daily_reset",
        "db_keepalive",
        "archive_tickets",
        "archive_uploaded_txt_files",
        "purge_old_auxiliary_records",
    ]

    class FakeJob:
        def __init__(self, job_id):
            self.id = job_id

    class FakeScheduler:
        running = True

        def get_jobs(self):
            return [FakeJob(job_id) for job_id in expected_job_ids]

    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: FakeScheduler())

    with app.app_context():
        admin = User(username="admin_settings_scheduler_normal", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_scheduler_normal", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/settings")
    assert resp.status_code == 200
    data = resp.get_json()
    scheduler_status = data["scheduler_status"]
    assert scheduler_status["status"] == "normal"
    assert scheduler_status["scheduler_present"] is True
    assert scheduler_status["scheduler_running"] is True
    assert scheduler_status["job_count"] == len(expected_job_ids)
    assert scheduler_status["missing_job_ids"] == []


def test_admin_settings_marks_scheduler_none_as_critical(app, client, monkeypatch):
    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: None)

    with app.app_context():
        admin = User(username="admin_settings_scheduler_none", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_scheduler_none", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/settings")
    assert resp.status_code == 200
    data = resp.get_json()
    scheduler_status = data["scheduler_status"]
    assert scheduler_status["status"] == "critical"
    assert scheduler_status["scheduler_present"] is False
    assert scheduler_status["scheduler_running"] is False
    assert "daily_reset" in scheduler_status["missing_job_ids"]


def test_admin_settings_lists_missing_scheduler_jobs(app, client, monkeypatch):
    class FakeJob:
        def __init__(self, job_id):
            self.id = job_id

    class FakeScheduler:
        running = True

        def get_jobs(self):
            return [FakeJob("expire_tickets"), FakeJob("clean_sessions")]

    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: FakeScheduler())

    with app.app_context():
        admin = User(username="admin_settings_scheduler_missing", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_scheduler_missing", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/settings")
    assert resp.status_code == 200
    data = resp.get_json()
    scheduler_status = data["scheduler_status"]
    assert scheduler_status["status"] == "critical"
    assert "daily_reset" in scheduler_status["missing_job_ids"]
    assert scheduler_status["job_count"] == 2


def test_admin_settings_scheduler_status_fallback_when_status_builder_errors(app, client, monkeypatch):
    def _raise_scheduler_error(*_args, **_kwargs):
        raise RuntimeError("scheduler status unavailable")

    monkeypatch.setattr("routes.admin._build_scheduler_status", _raise_scheduler_error)

    with app.app_context():
        admin = User(username="admin_settings_scheduler_fallback", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_settings_scheduler_fallback", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/settings")
    assert resp.status_code == 200
    data = resp.get_json()
    scheduler_status = data["scheduler_status"]
    assert scheduler_status["status"] == "warning"
    assert "暂不可用" in scheduler_status["message"]


def test_admin_winning_template_uses_readable_chinese_labels():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "结果管理" in content
    assert "全部图片状态" in content
    assert "税后合计：" in content
    assert "确认将这条中奖记录标记为已检查吗？" in content
    assert "已解析 ${data.count} 条赛果，中奖计算已加入队列" in content
    assert "showToast('\u5df2\u63d0\u4ea4\u91cd\u7b97', 'success');" in content
    assert "showToast(e.message || '\u63d0\u4ea4\u5931\u8d25', 'danger');" in content


def test_admin_winning_template_declares_match_results_and_uploading_state():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "matchResults: [], mrFilterDate: '', mrDateOptions: []" in content
    assert "uploadingImageId: null" in content


def test_admin_winning_template_shows_uploaded_profit_summary():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "今日上传金额：" in content
    assert "盈利金额：" in content
    assert "盈利百分比：" in content
    assert "summaryUploadedAmount" in content
    assert "summaryProfitAmount" in content
    assert "summaryProfitPercent" in content


def test_client_dashboard_template_uses_readable_chinese_labels():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "我的主页" in content
    assert "今日处理张数" in content
    assert "今日各设备出票统计" in content
    assert "确认停止接单？当前票将标记为已完成。" in content


def test_client_dashboard_preserves_winning_group_mode_after_filter():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "this.applyWinningGrouping();" in content
    assert "const key = r.business_date || 'unknown';" in content
    assert "this.winningGroupBy === 'date'" in content
    assert "winningGroupBy === 'device'" in content
    assert "const key = r.assigned_device_id || '未知设备';" in content
    assert "按设备ID" in content


def test_client_dashboard_clears_password_success_timer_before_reopen():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "pwdSuccessTimer: null" in content
    assert "clearTimeout(this.pwdSuccessTimer);" in content
    assert "this.pwdSuccessTimer = setTimeout(() => {" in content


def test_client_dashboard_only_shows_no_ticket_toast_once():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert content.count("showToast(data.error || '\u6682\u65e0\u53ef\u7528\u7968\u636e', 'warning');") == 1


def test_client_dashboard_handles_mode_a_stop_failures_and_localizes_next_ticket_messages():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "async playNoTicketWarningAlert()" in content
    assert "this.playNoTicketWarningAlert();" in content
    assert "this.startCountdown(data.ticket.deadline_time);\n          // Refresh daily device totals immediately after successful Mode A progression.\n          this.loadStats();" in content
    assert "showToast('请等待 ' + remaining + ' 秒后再获取下一张', 'warning');" in content
    assert "showToast(data.error || '\u6682\u65e0\u53ef\u7528\u7968\u636e', 'warning');" in content
    assert "showToast('获取下一张失败，请稍后重试', 'danger');" in content
    assert "if (!res.ok || data.success === false) {" in content
    assert "showToast(e.message || '停止接单失败', 'danger');" in content
    assert "showToast(data.message || '已停止接单', 'success');" in content
    assert "this.countdown = '已截止';" in content


def test_admin_winning_template_uses_api_mark_checked_endpoint():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "fetch(`/api/winning/admin/mark-checked/${record.winning_record_id}`" in content
    assert "fetch(`/winning/admin/mark-checked/${record.winning_record_id}`" not in content


def test_admin_winning_template_tracks_summary_tax_and_missing():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "summaryTax" in content
    assert "summaryMissing" in content
    assert "data.summary.tax" in content
    assert "data.summary.missing" in content


def test_admin_winning_export_preserves_checked_filter():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "params.set('checked_status', this.filterChecked)" in content


def test_admin_winning_export_honors_checked_status_filter(app, client):
    from io import BytesIO
    from openpyxl import load_workbook

    with app.app_context():
        admin = User(username="admin_export_user", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)

        ticket_checked = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="CHK-001",
            status="completed",
            assigned_username="user-a",
            completed_at=beijing_now(),
            is_winning=True,
        )
        ticket_unchecked = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="UNCHK-001",
            status="completed",
            assigned_username="user-a",
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add_all([ticket_checked, ticket_unchecked])
        db.session.commit()

        db.session.add(WinningRecord(ticket_id=ticket_checked.id, is_checked=True, uploaded_by=admin.id))
        db.session.add(WinningRecord(ticket_id=ticket_unchecked.id, is_checked=False, uploaded_by=admin.id))
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_export_user", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/winning/export?checked_status=checked")
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.data))
    ws = wb.active
    exported_values = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "CHK-001" in exported_values
    assert "UNCHK-001" not in exported_values


def test_admin_users_export_contains_desktop_only_column_and_value(app, client):
    from io import BytesIO
    from openpyxl import load_workbook

    with app.app_context():
        admin = User(username="admin_users_export", is_admin=True)
        admin.set_password("secret123")
        user = create_user("users_export_target", "secret123", client_mode="mode_b")
        user.desktop_only_b_mode = False
        user.can_receive = False
        user.set_blocked_lottery_types(["胜平负", "比分"])
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_users_export", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/users/export")
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.data))
    ws = wb.active
    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    target_row = next(row for row in rows if row[0] == "users_export_target")

    assert "密码" in header
    assert "B模式仅桌面端" in header
    assert target_row[1].startswith("$2")
    assert target_row[9] == "否"
    assert target_row[8] == "关闭"
    assert target_row[6] == "胜平负,比分"


def test_admin_users_import_accepts_hashed_password_and_desktop_only_flag(app, client):
    from io import BytesIO
    from openpyxl import Workbook

    with app.app_context():
        admin = User(username="admin_users_import", is_admin=True)
        admin.set_password("secret123")
        hashed_user = User(username="hash_seed_user")
        hashed_user.set_password("import-secret")
        password_hash = hashed_user.password_hash
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_users_import", "password": "secret123"})
    assert resp.status_code == 200

    wb = Workbook()
    ws = wb.active
    ws.append(["用户名", "密码哈希", "接单模式", "最大设备数", "B模式处理上限", "每日处理上限", "禁止彩种", "账号状态", "接单开关", "B模式仅桌面端"])
    ws.append(["imported_hash_user", password_hash, "mode_b", 2, 88, 99, "胜平负,比分", "是", "关", "否"])
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    resp = client.post(
        "/admin/api/users/import",
        data={"file": (payload, "users-import.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["success_count"] == 1

    with app.app_context():
        imported = User.query.filter_by(username="imported_hash_user").first()
        assert imported is not None
        assert imported.client_mode == "mode_b"
        assert imported.is_active is True
        assert imported.can_receive is False
        assert imported.desktop_only_b_mode is False
        assert imported.max_processing_b_mode == 88
        assert imported.daily_ticket_limit == 99
        assert imported.get_blocked_lottery_types() == ["胜平负", "比分"]
        assert imported.check_password("import-secret") is True


def test_admin_users_export_file_can_be_used_as_import_template(app, client):
    from io import BytesIO
    from openpyxl import load_workbook

    with app.app_context():
        admin = User(username="admin_users_roundtrip", is_admin=True)
        admin.set_password("secret123")
        user = create_user("users_export_source", "export-secret", client_mode="mode_b")
        user.max_devices = 3
        user.max_processing_b_mode = 66
        user.daily_ticket_limit = 77
        user.can_receive = False
        user.desktop_only_b_mode = False
        user.set_blocked_lottery_types(["胜平负", "比分"])
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_users_roundtrip", "password": "secret123"})
    assert resp.status_code == 200

    export_resp = client.get("/admin/api/users/export")
    assert export_resp.status_code == 200

    wb = load_workbook(BytesIO(export_resp.data))
    ws = wb.active
    ws["A2"] = "users_export_copy"
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    import_resp = client.post(
        "/admin/api/users/import",
        data={"file": (payload, "users-roundtrip.xlsx")},
        content_type="multipart/form-data",
    )
    assert import_resp.status_code == 200
    data = import_resp.get_json()
    assert data["success"] is True
    assert data["success_count"] == 1

    with app.app_context():
        imported = User.query.filter_by(username="users_export_copy").first()
        assert imported is not None
        assert imported.client_mode == "mode_b"
        assert imported.max_devices == 3
        assert imported.max_processing_b_mode == 66
        assert imported.daily_ticket_limit == 77
        assert imported.can_receive is False
        assert imported.desktop_only_b_mode is False
        assert imported.get_blocked_lottery_types() == ["胜平负", "比分"]
        assert imported.check_password("export-secret") is True


def test_admin_users_import_accepts_uppercase_xlsx_extension(app, client, monkeypatch):
    from io import BytesIO

    with app.app_context():
        admin = User(username="admin_users_import_upper_ext", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_users_import_upper_ext", "password": "secret123"})
    assert resp.status_code == 200

    monkeypatch.setattr(
        "services.user_import_service.import_users",
        lambda *_args, **_kwargs: {"success": True, "success_count": 1, "failure_count": 0},
    )

    resp = client.post(
        "/admin/api/users/import",
        data={"file": (BytesIO(b"dummy"), "users-import.XLSX")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True


def test_admin_checked_winning_record_cannot_replace_image(app, client):
    from io import BytesIO

    with app.app_context():
        admin = User(username="admin_checked_image", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)

        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="CHK-IMG-001",
            status="completed",
            assigned_username="user-a",
            completed_at=beijing_now(),
            is_winning=True,
            winning_image_url="/uploads/images/original.jpg",
        )
        db.session.add(ticket)
        db.session.commit()

        db.session.add(WinningRecord(ticket_id=ticket.id, is_checked=True, uploaded_by=admin.id))
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_checked_image", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        f"/admin/api/winning/{ticket_id}/upload-image",
        data={"image": (BytesIO(b"fake-image"), "winning.png")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "已检查" in data["error"]


def test_admin_winning_record_creates_winning_record(app, client):
    with app.app_context():
        admin = User(username="admin_record_creator", is_admin=True)
        admin.set_password("secret123")
        user = create_user("admin_record_target", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-OSS-001",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_record_creator", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/winning/record",
        json={"ticket_id": ticket_id, "oss_key": f"winning/2026/04/07/{ticket_id}.jpg"},
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["record"]["ticket_id"] == ticket_id

    with app.app_context():
        ticket = LotteryTicket.query.get(ticket_id)
        record = WinningRecord.query.filter_by(ticket_id=ticket_id).first()
        assert ticket.is_winning is True
        assert ticket.winning_image_url
        assert record is not None
        assert record.image_oss_key == f"winning/2026/04/07/{ticket_id}.jpg"


def test_record_winning_does_not_delete_reuploaded_same_oss_key(app, client, monkeypatch):
    deleted = []
    monkeypatch.setattr("services.oss_service.delete_stored_image", lambda key=None, url=None: deleted.append((key, url)))
    monkeypatch.setattr("services.oss_service.get_public_url", lambda oss_key: f"https://oss.example.com/{oss_key}")

    with app.app_context():
        user = create_user("winning_same_key_user", "secret123", client_mode="mode_a")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-SAME-KEY",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_image_url="https://oss.example.com/winning/2026/04/07/1.jpg",
        )
        db.session.add(ticket)
        db.session.commit()
        db.session.add(WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=ticket.winning_image_url,
            image_oss_key="winning/2026/04/07/1.jpg",
            uploaded_by=user.id,
        ))
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_same_key_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/winning/record",
        json={"ticket_id": ticket_id, "oss_key": "winning/2026/04/07/1.jpg"},
    )
    assert resp.status_code == 200
    assert deleted == []


def test_record_winning_rejects_checked_record_replacement(app, client):
    with app.app_context():
        user = create_user("winning_checked_guard_user", "secret123", client_mode="mode_a")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-CHECKED-GUARD",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_image_url="https://oss.example.com/winning/original.jpg",
        )
        db.session.add(ticket)
        db.session.commit()
        db.session.add(WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=ticket.winning_image_url,
            image_oss_key="winning/original.jpg",
            uploaded_by=user.id,
            is_checked=True,
        ))
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_checked_guard_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/winning/record",
        json={"ticket_id": ticket_id, "oss_key": "winning/replaced.jpg", "winning_amount": 100},
    )
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "已检查" in data["error"]


def test_winning_presign_rejects_checked_record(app, client):
    with app.app_context():
        user = create_user("winning_presign_checked_user", "secret123", client_mode="mode_a")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-PRESIGN-CHECKED",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add(ticket)
        db.session.commit()
        db.session.add(WinningRecord(ticket_id=ticket.id, uploaded_by=user.id, is_checked=True))
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_presign_checked_user", "secret123")
    assert resp.status_code == 200

    resp = client.get(f"/api/winning/presign?ticket_id={ticket_id}")
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "已检查" in data["error"]


def test_record_winning_handles_empty_json_body(app, client):
    with app.app_context():
        user = create_user("winning_record_empty_body_user", "secret123", client_mode="mode_a")
        create_assigned_ticket(user, "device-win-empty", "WIN-EMPTY", 1)

    resp = login(client, "winning_record_empty_body_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/winning/record", data="", content_type="application/json")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "参数不完整" in data["error"]


@pytest.mark.parametrize("bad_amount", [True, "abc", {"x": 1}])
def test_record_winning_rejects_invalid_winning_amount_values(app, client, bad_amount):
    with app.app_context():
        user = create_user("winning_amount_invalid_user", "secret123", client_mode="mode_a")
        ticket = create_assigned_ticket(user, "device-win-amount", "WIN-AMOUNT-INVALID", 1)
        ticket.is_winning = True
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_amount_invalid_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/winning/record",
        json={
            "ticket_id": ticket_id,
            "oss_key": f"winning/2026/04/07/{ticket_id}.jpg",
            "winning_amount": bad_amount,
        },
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_record_winning_rejects_negative_winning_amount(app, client):
    with app.app_context():
        user = create_user("winning_amount_negative_user", "secret123", client_mode="mode_a")
        ticket = create_assigned_ticket(user, "device-win-amount-negative", "WIN-AMOUNT-NEG", 1)
        ticket.is_winning = True
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_amount_negative_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/winning/record",
        json={
            "ticket_id": ticket_id,
            "oss_key": f"winning/2026/04/07/{ticket_id}.jpg",
            "winning_amount": -1,
        },
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_admin_winning_record_does_not_delete_reuploaded_same_oss_key(app, client, monkeypatch):
    deleted = []
    monkeypatch.setattr("services.oss_service.delete_stored_image", lambda key=None, url=None: deleted.append((key, url)))
    monkeypatch.setattr("services.oss_service.get_public_url", lambda oss_key: f"https://oss.example.com/{oss_key}")

    with app.app_context():
        admin = User(username="admin_same_key", is_admin=True)
        admin.set_password("secret123")
        user = create_user("admin_same_key_target", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-SAME-KEY",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_image_url="",
        )
        db.session.add(ticket)
        db.session.commit()
        current_key = f"winning/2026/04/07/{ticket.id}.jpg"
        ticket.winning_image_url = f"https://oss.example.com/{current_key}"
        db.session.add(WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=ticket.winning_image_url,
            image_oss_key=current_key,
            uploaded_by=admin.id,
        ))
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_same_key", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/winning/record",
        json={"ticket_id": ticket_id, "oss_key": f"winning/2026/04/07/{ticket_id}.jpg"},
    )
    assert resp.status_code == 200
    assert deleted == []


def test_admin_winning_record_rejects_empty_oss_key(app, client):
    with app.app_context():
        admin = User(username="admin_empty_oss_key", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-EMPTY-KEY",
            status="completed",
            completed_at=beijing_now(),
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_empty_oss_key", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post("/admin/api/winning/record", json={"ticket_id": ticket_id, "oss_key": ""})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "oss_key" in data["error"]


def test_admin_winning_record_handles_empty_json_body(app, client):
    with app.app_context():
        admin = User(username="admin_winning_empty_body", is_admin=True)
        admin.set_password("secret123")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-WIN-EMPTY",
            status="completed",
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add_all([admin, ticket])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_winning_empty_body", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post("/admin/api/winning/record", data="", content_type="application/json")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "oss_key" in data["error"]


def test_admin_winning_record_checked_error_uses_readable_chinese(app, client):
    with app.app_context():
        admin = User(username="admin_checked_record_text", is_admin=True)
        admin.set_password("secret123")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-CHECKED-TEXT",
            status="completed",
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add_all([admin, ticket])
        db.session.commit()
        db.session.add(WinningRecord(ticket_id=ticket.id, is_checked=True, uploaded_by=admin.id))
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_checked_record_text", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/winning/record",
        json={"ticket_id": ticket_id, "oss_key": f"winning/2026/04/07/{ticket_id}.jpg"},
    )
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "已检查" in data["error"]
    assert "无法更换图片" in data["error"]


def test_admin_winning_presign_rejects_checked_record(app, client):
    with app.app_context():
        admin = User(username="admin_presign_checked_record", is_admin=True)
        admin.set_password("secret123")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-PRESIGN-CHECKED",
            status="completed",
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add_all([admin, ticket])
        db.session.commit()
        db.session.add(WinningRecord(ticket_id=ticket.id, uploaded_by=admin.id, is_checked=True))
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_presign_checked_record", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(f"/admin/api/winning/{ticket_id}/presign")
    assert resp.status_code == 403
    data = resp.get_json()
    assert data["success"] is False
    assert "已检查" in data["error"]


def test_winning_presign_rejects_invalid_ticket_id(app, client):
    with app.app_context():
        user = create_user("winning_invalid_ticket_id_user", "secret123", client_mode="mode_a")

    resp = login(client, "winning_invalid_ticket_id_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/presign?ticket_id=abc")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "票ID必须是整数" in data["error"]


def test_winning_record_rejects_non_winning_ticket(app, client):
    with app.app_context():
        user = create_user("winning_non_winning_user", "secret123", client_mode="mode_a")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="NOT-WIN-001",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=False,
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_non_winning_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/winning/record",
        json={"ticket_id": ticket_id, "oss_key": "winning/test/not-win.jpg"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "未被系统判定为中奖" in data["error"]

    with app.app_context():
        ticket = LotteryTicket.query.get(ticket_id)
        assert ticket.is_winning is False
        assert WinningRecord.query.filter_by(ticket_id=ticket_id).first() is None


def test_admin_winning_record_rejects_invalid_ticket_id(app, client):
    with app.app_context():
        admin = User(username="admin_invalid_winning_ticket_id", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_invalid_winning_ticket_id", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/winning/record",
        json={"ticket_id": "bad-id", "oss_key": "winning/test/admin-invalid.jpg"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "票ID必须是大于 0 的整数" in data["error"]


def test_admin_winning_presign_rejects_non_winning_ticket(app, client):
    with app.app_context():
        admin = User(username="admin_presign_non_winning", is_admin=True)
        admin.set_password("secret123")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-NOT-WIN-001",
            status="completed",
            completed_at=beijing_now(),
            is_winning=False,
        )
        db.session.add_all([admin, ticket])
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_presign_non_winning", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(f"/admin/api/winning/{ticket_id}/presign")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "未被系统判定为中奖" in data["error"]


def test_record_winning_rejects_oss_key_for_other_ticket(app, client):
    with app.app_context():
        user = create_user("winning_bad_key_user", "secret123", client_mode="mode_a")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-BAD-KEY-001",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

    resp = login(client, "winning_bad_key_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/winning/record",
        json={"ticket_id": ticket_id, "oss_key": "winning/test/not-this-ticket.jpg"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "oss_key" in data["error"]


def test_admin_winning_record_rejects_oss_key_for_other_ticket(app, client):
    with app.app_context():
        admin = User(username="admin_bad_winning_key", is_admin=True)
        admin.set_password("secret123")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-BAD-KEY-001",
            status="completed",
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add_all([admin, ticket])
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_bad_winning_key", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/api/winning/record",
        json={"ticket_id": ticket_id, "oss_key": "winning/test/not-this-ticket.jpg"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "oss_key" in data["error"]


def test_admin_upload_winning_image_creates_winning_record(app, client):
    from io import BytesIO
    from PIL import Image

    with app.app_context():
        admin = User(username="admin_upload_creator", is_admin=True)
        admin.set_password("secret123")
        user = create_user("admin_upload_target", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="ADMIN-UPLOAD-001",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add(ticket)
        db.session.commit()
        ticket_id = ticket.id

    resp = client.post("/auth/login", json={"username": "admin_upload_creator", "password": "secret123"})
    assert resp.status_code == 200

    image_bytes = BytesIO()
    Image.new("RGB", (20, 20), color="blue").save(image_bytes, format="PNG")
    image_bytes.seek(0)

    resp = client.post(
        f"/admin/api/winning/{ticket_id}/upload-image",
        data={"image": (image_bytes, "winning.png")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["record"]["ticket_id"] == ticket_id

    with app.app_context():
        ticket = LotteryTicket.query.get(ticket_id)
        record = WinningRecord.query.filter_by(ticket_id=ticket_id).first()
        assert ticket.is_winning is True
        assert ticket.winning_image_url
        assert record is not None
        assert record.winning_image_url == ticket.winning_image_url


def test_my_winning_returns_business_date(app, client):
    with app.app_context():
        user = create_user("winning_business_date", "secret123", client_mode="mode_a")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-BIZ-001",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-a",
            completed_at=beijing_now(),
            is_winning=True,
        )
        db.session.add(ticket)
        db.session.commit()

    resp = login(client, "winning_business_date", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my")
    assert resp.status_code == 200
    data = resp.get_json()
    records = [item for items in data["grouped"].values() for item in items]
    assert records
    assert records[0]["business_date"]


def test_my_winning_returns_download_filename(app, client):
    with app.app_context():
        user = create_user("winning_download_filename", "secret123", client_mode="mode_b")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-FILENAME-001",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-a",
            completed_at=beijing_now(),
            is_winning=True,
            download_filename="比分_2倍_53张_1112元_02.40_2026-0429-011309.txt",
        )
        db.session.add(ticket)
        db.session.commit()

    resp = login(client, "winning_download_filename", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my")
    assert resp.status_code == 200
    data = resp.get_json()
    record = [item for items in data["grouped"].values() for item in items][0]
    assert record["download_filename"] == "比分_2倍_53张_1112元_02.40_2026-0429-011309.txt"


def test_my_winning_returns_original_ticket_amount(app, client):
    with app.app_context():
        user = create_user("winning_ticket_amount", "secret123", client_mode="mode_b")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-AMOUNT-001",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            ticket_amount=12.5,
            winning_amount=88,
        )
        db.session.add(ticket)
        db.session.commit()

    resp = login(client, "winning_ticket_amount", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my")
    assert resp.status_code == 200
    data = resp.get_json()
    record = [item for items in data["grouped"].values() for item in items][0]
    assert record["ticket_amount"] == 12.5


def test_my_winning_default_date_uses_previous_business_day(app, client, monkeypatch):
    business_today = datetime(2026, 4, 8, 13, 0, 0)

    with app.app_context():
        user = create_user("winning_default_prev_day", "secret123", client_mode="mode_b")
        prev_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-DEFAULT-PREV",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 7, 13, 0, 0),
            is_winning=True,
            winning_amount=10,
        )
        current_ticket = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="WIN-DEFAULT-CURRENT",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 8, 13, 0, 0),
            is_winning=True,
            winning_amount=20,
        )
        db.session.add_all([prev_ticket, current_ticket])
        db.session.commit()

    monkeypatch.setattr("routes.winning.get_business_date", lambda dt=None: (dt or business_today).date())

    resp = login(client, "winning_default_prev_day", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my?default_date=1")
    assert resp.status_code == 200
    data = resp.get_json()
    records = [item["raw_content"] for items in data["grouped"].values() for item in items]
    assert records == ["WIN-DEFAULT-PREV"]
    assert data["current_business_date"] == "2026-04-08"
    assert data["default_business_date"] == "2026-04-07"
    assert "2026-04-07" in data["filter_options"]["dates"]


def test_my_winning_returns_assigned_time_and_filtered_summary(app, client):
    assigned_at = beijing_now() - timedelta(hours=2)

    with app.app_context():
        user = create_user("winning_assigned_summary", "secret123", client_mode="mode_b")
        final_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-SUMMARY-FINAL",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_at=assigned_at,
            completed_at=beijing_now(),
            is_winning=True,
            lottery_type="胜平负",
            winning_amount=100.12,
        )
        predicted_ticket = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="WIN-SUMMARY-PREDICTED",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_at=assigned_at + timedelta(minutes=5),
            completed_at=beijing_now(),
            is_winning=True,
            lottery_type="胜平负",
            predicted_winning_amount=20.50,
        )
        filtered_out_ticket = LotteryTicket(
            source_file_id=1,
            line_number=3,
            raw_content="WIN-SUMMARY-OTHER",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_at=assigned_at,
            completed_at=beijing_now(),
            is_winning=True,
            lottery_type="比分",
            winning_amount=999,
        )
        db.session.add_all([final_ticket, predicted_ticket, filtered_out_ticket])
        db.session.commit()

    resp = login(client, "winning_assigned_summary", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my?lottery_type=胜平负")
    assert resp.status_code == 200
    data = resp.get_json()
    records = [item for items in data["grouped"].values() for item in items]
    records_by_raw = {record["raw_content"]: record for record in records}
    assert set(records_by_raw) == {"WIN-SUMMARY-FINAL", "WIN-SUMMARY-PREDICTED"}
    assert records_by_raw["WIN-SUMMARY-FINAL"]["assigned_at"] == assigned_at.isoformat()
    assert data["summary"]["record_count"] == 2
    assert data["summary"]["total_display_winning_amount"] == 120.62


def test_my_winning_keeps_recent_four_business_days(app, client, monkeypatch):
    business_today = datetime(2026, 4, 7, 13, 0, 0)

    with app.app_context():
        user = create_user("winning_recent_four", "secret123", client_mode="mode_a")
        keep_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-KEEP-4D",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 4, 13, 0, 0),
            is_winning=True,
        )
        drop_ticket = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="WIN-DROP-5D",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 3, 11, 0, 0),
            is_winning=True,
        )
        db.session.add_all([keep_ticket, drop_ticket])
        db.session.commit()

    monkeypatch.setattr("routes.winning.get_business_date", lambda dt=None: (dt or business_today).date() if dt else business_today.date())

    resp = login(client, "winning_recent_four", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my")
    assert resp.status_code == 200
    data = resp.get_json()
    records = [item["raw_content"] for items in data["grouped"].values() for item in items]
    assert "WIN-KEEP-4D" in records
    assert "WIN-DROP-5D" not in records


def test_winning_calc_includes_expired_but_excludes_revoked(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    def fake_calculate_winning(raw_content, result_data, multiplier):
        if raw_content in {"WIN-COMPLETED", "WIN-EXPIRED"}:
            return True, 100, 100, 0
        return False, 0, 0, 0

    monkeypatch.setattr("services.winning_calc_service.calculate_winning", fake_calculate_winning)

    with app.app_context():
        user = create_user("winning_calc_status_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26066",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.23}}},
            uploaded_by=user.id,
        )
        db.session.add(match_result)
        db.session.commit()

        completed_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-COMPLETED",
            status="completed",
            detail_period="26066",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        expired_ticket = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="WIN-EXPIRED",
            status="expired",
            detail_period="26066",
            assigned_user_id=user.id,
            assigned_username=user.username,
            deadline_time=beijing_now(),
        )
        revoked_ticket = LotteryTicket(
            source_file_id=1,
            line_number=3,
            raw_content="WIN-REVOKED",
            status="revoked",
            detail_period="26066",
            assigned_user_id=user.id,
            assigned_username=user.username,
            admin_upload_time=beijing_now(),
        )
        db.session.add_all([completed_ticket, expired_ticket, revoked_ticket])
        db.session.commit()

        process_match_result(match_result.id, app=app)
        db.session.expire_all()

        refreshed_completed = LotteryTicket.query.get(completed_ticket.id)
        refreshed_expired = LotteryTicket.query.get(expired_ticket.id)
        refreshed_revoked = LotteryTicket.query.get(revoked_ticket.id)
        refreshed_result = MatchResult.query.get(match_result.id)

    assert refreshed_completed.is_winning is True
    assert refreshed_expired.is_winning is True
    assert refreshed_revoked.is_winning is None
    assert refreshed_result.tickets_total == 2
    assert refreshed_result.tickets_winning == 2


def test_winning_calculator_matches_hyphenated_score_results():
    from decimal import Decimal

    from utils.winning_calculator import calculate_winning

    result_data = {
        "1": {
            "CBF": {"result": "1-1", "sp": 10.674},
            "BQC": {"result": "1-1", "sp": 8.591},
        },
        "2": {
            "CBF": {"result": "0-0", "sp": 13.707},
            "BQC": {"result": "0-0", "sp": 5.799},
        },
    }

    cbf_win, cbf_gross, cbf_net, cbf_tax = calculate_winning(
        "CBF|1=11,2=00|2*1|2",
        result_data,
        2,
    )
    bqc_win, bqc_gross, bqc_net, bqc_tax = calculate_winning(
        "BQC|1=11,2=00|2*1|2",
        result_data,
        2,
    )

    assert cbf_win is True
    assert cbf_gross == Decimal("380.40")
    assert cbf_net == Decimal("380.40")
    assert cbf_tax == Decimal("0")
    assert bqc_win is True
    assert bqc_gross == Decimal("129.53")
    assert bqc_net == Decimal("129.53")
    assert bqc_tax == Decimal("0")


def test_winning_calc_clears_stale_amounts_when_ticket_calc_errors(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    def fake_calculate_winning(raw_content, result_data, multiplier):
        raise RuntimeError("boom")

    monkeypatch.setattr("services.winning_calc_service.calculate_winning", fake_calculate_winning)

    with app.app_context():
        user = create_user("winning_calc_error_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26067",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.23}}},
            uploaded_by=user.id,
        )
        db.session.add(match_result)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-ERROR",
            status="completed",
            detail_period="26067",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_gross=123,
            winning_amount=100,
            winning_tax=23,
        )
        db.session.add(ticket)
        db.session.commit()

        process_match_result(match_result.id, app=app)
        db.session.expire_all()

        refreshed_ticket = LotteryTicket.query.get(ticket.id)
        refreshed_result = MatchResult.query.get(match_result.id)

    assert refreshed_ticket.is_winning is False
    assert refreshed_ticket.winning_gross is None
    assert refreshed_ticket.winning_amount is None
    assert refreshed_ticket.winning_tax is None
    assert refreshed_result.tickets_winning == 0
    assert float(refreshed_result.total_winning_amount or 0) == 0.0


def test_winning_calc_removes_stale_winning_record_when_ticket_is_no_longer_winning(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    def fake_calculate_winning(raw_content, result_data, multiplier):
        return False, 0, 0, 0

    monkeypatch.setattr("services.winning_calc_service.calculate_winning", fake_calculate_winning)

    with app.app_context():
        user = create_user("winning_calc_stale_record_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26068",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.23}}},
            uploaded_by=user.id,
        )
        db.session.add(match_result)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-STALE-RECORD",
            status="completed",
            detail_period="26068",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_image_url="/uploads/images/stale.png",
        )
        db.session.add(ticket)
        db.session.commit()
        db.session.add(WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=ticket.winning_image_url,
            uploaded_by=user.id,
        ))
        db.session.commit()

        process_match_result(match_result.id, app=app)
        db.session.expire_all()

        refreshed_ticket = LotteryTicket.query.get(ticket.id)
        refreshed_record = WinningRecord.query.filter_by(ticket_id=ticket.id).first()

    assert refreshed_ticket.is_winning is False
    assert refreshed_ticket.winning_image_url is None
    assert refreshed_record is None


def test_winning_calc_keeps_checked_winning_record_when_ticket_is_no_longer_winning(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    def fake_calculate_winning(raw_content, result_data, multiplier):
        return False, 0, 0, 0

    monkeypatch.setattr("services.winning_calc_service.calculate_winning", fake_calculate_winning)

    with app.app_context():
        user = create_user("winning_calc_checked_record_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="260681",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.23}}},
            uploaded_by=user.id,
        )
        db.session.add(match_result)
        db.session.commit()

        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-CHECKED-RECORD",
            status="completed",
            detail_period="260681",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_image_url="/uploads/images/checked-stale.png",
        )
        db.session.add(ticket)
        db.session.commit()
        db.session.add(WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=ticket.winning_image_url,
            uploaded_by=user.id,
            is_checked=True,
        ))
        db.session.commit()

        process_match_result(match_result.id, app=app)
        db.session.expire_all()

        refreshed_ticket = LotteryTicket.query.get(ticket.id)
        refreshed_record = WinningRecord.query.filter_by(ticket_id=ticket.id).first()

    assert refreshed_ticket.is_winning is False
    assert refreshed_ticket.winning_image_url == "/uploads/images/checked-stale.png"
    assert refreshed_record is not None
    assert refreshed_record.is_checked is True


def test_winning_calc_removes_stale_local_image_when_ticket_is_no_longer_winning(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    def fake_calculate_winning(raw_content, result_data, multiplier):
        return False, 0, 0, 0

    monkeypatch.setattr("services.winning_calc_service.calculate_winning", fake_calculate_winning)

    with app.app_context():
        user = create_user("winning_calc_stale_image_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26069",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.23}}},
            uploaded_by=user.id,
        )
        db.session.add(match_result)
        db.session.commit()

        images_dir = Path(app.config["UPLOAD_FOLDER"]) / "images"
        images_dir.mkdir(parents=True, exist_ok=True)
        stale_path = images_dir / "stale-winning.png"
        stale_path.write_bytes(b"stale-image")

        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-STALE-IMAGE",
            status="completed",
            detail_period="26069",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_image_url="/uploads/images/stale-winning.png",
        )
        db.session.add(ticket)
        db.session.commit()
        db.session.add(WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=ticket.winning_image_url,
            uploaded_by=user.id,
        ))
        db.session.commit()

        process_match_result(match_result.id, app=app)

    assert not stale_path.exists()


def test_admin_winning_lists_expired_ticket_with_special_status(app, client):
    from io import BytesIO
    from openpyxl import load_workbook

    with app.app_context():
        admin = User(username="admin_winning_expired", is_admin=True)
        admin.set_password("secret123")
        user = create_user("expired_winning_user", "secret123", client_mode="mode_a")
        db.session.add(admin)
        db.session.commit()

        expired_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-EXPIRED-LIST",
            status="expired",
            assigned_user_id=user.id,
            assigned_username=user.username,
            deadline_time=datetime(2026, 4, 7, 11, 0, 0),
            is_winning=True,
            winning_amount=88,
        )
        db.session.add(expired_ticket)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_winning_expired", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/winning?date=2026-04-06")
    assert resp.status_code == 200
    data = resp.get_json()
    assert [r["raw_content"] for r in data["records"]] == ["WIN-EXPIRED-LIST"]
    assert data["records"][0]["status"] == "expired"
    assert data["records"][0]["status_label"] == "已过期未出票"
    assert data["records"][0]["terminal_at"].startswith("2026-04-07T11:00:00")

    export_resp = client.get("/admin/api/winning/export?date=2026-04-06")
    assert export_resp.status_code == 200
    wb = load_workbook(BytesIO(export_resp.data))
    ws = wb.active
    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    row = [cell for cell in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    assert "状态" in header
    assert "终态时间" in header
    assert "已过期未出票" in row
    assert "2026-04-07 11:00:00" in row


def test_my_winning_still_hides_expired_tickets(app, client):
    with app.app_context():
        user = create_user("winning_hide_expired", "secret123", client_mode="mode_a")
        completed_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-COMPLETED-VIEW",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
        )
        expired_ticket = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="WIN-EXPIRED-HIDDEN",
            status="expired",
            assigned_user_id=user.id,
            assigned_username=user.username,
            deadline_time=beijing_now(),
            is_winning=True,
        )
        db.session.add_all([completed_ticket, expired_ticket])
        db.session.commit()

    resp = login(client, "winning_hide_expired", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my")
    assert resp.status_code == 200
    data = resp.get_json()
    records = [item["raw_content"] for items in data["grouped"].values() for item in items]
    assert "WIN-COMPLETED-VIEW" in records
    assert "WIN-EXPIRED-HIDDEN" not in records


def test_admin_file_list_uses_business_date_for_date_filter(app, client):
    with app.app_context():
        admin = User(username="admin_file_date", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        file_prev_business = UploadedFile(
            display_id="2026/04/06-01",
            original_filename="before_noon.txt",
            stored_filename="before_noon.txt",
            uploaded_by=admin.id,
            total_tickets=1,
            pending_count=1,
            uploaded_at=datetime(2026, 4, 7, 11, 0, 0),
        )
        file_current_business = UploadedFile(
            display_id="2026/04/07-01",
            original_filename="after_noon.txt",
            stored_filename="after_noon.txt",
            uploaded_by=admin.id,
            total_tickets=1,
            pending_count=1,
            uploaded_at=datetime(2026, 4, 7, 13, 0, 0),
        )
        db.session.add_all([file_prev_business, file_current_business])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_file_date", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/files?date=2026-04-06&include_date_options=1")
    assert resp.status_code == 200
    data = resp.get_json()
    assert [f["original_filename"] for f in data["files"]] == ["before_noon.txt"]
    assert "2026-04-06" in data["date_options"]
    assert "2026-04-07" in data["date_options"]


def test_admin_files_list_returns_current_business_date(app, client, monkeypatch):
    monkeypatch.setattr("routes.admin.get_business_date", lambda dt=None: datetime(2026, 4, 7).date())
    with app.app_context():
        admin = User(username="admin_file_current_business_date", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_file_current_business_date", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/files")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["current_business_date"] == "2026-04-07"


def test_admin_files_page_embeds_current_business_date(app, client, monkeypatch):
    monkeypatch.setattr("routes.admin.get_business_date", lambda dt=None: datetime(2026, 4, 7).date())
    with app.app_context():
        admin = User(username="admin_file_page_default_date", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_file_page_default_date", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/files")
    assert resp.status_code == 200
    content = resp.get_data(as_text=True)
    assert "window.DEFAULT_FILE_FILTER_DATE = \"2026-04-07\";" in content


def test_admin_winning_uses_business_date_for_date_filter(app, client):
    with app.app_context():
        admin = User(username="admin_winning_date", is_admin=True)
        admin.set_password("secret123")
        user = create_user("winning_date_user", "secret123", client_mode="mode_a")

        db.session.add(admin)
        db.session.commit()

        ticket_prev_business = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="WIN-PREV-BIZ",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 7, 11, 0, 0),
            is_winning=True,
        )
        ticket_current_business = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="WIN-CUR-BIZ",
            status="completed",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 7, 13, 0, 0),
            is_winning=True,
        )
        db.session.add_all([ticket_prev_business, ticket_current_business])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_winning_date", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/winning/filter-options")
    assert resp.status_code == 200
    options = resp.get_json()
    assert "2026-04-06" in options["dates"]
    assert "2026-04-07" in options["dates"]

    resp = client.get("/admin/api/winning?date=2026-04-06")
    assert resp.status_code == 200
    data = resp.get_json()
    assert [r["raw_content"] for r in data["records"]] == ["WIN-PREV-BIZ"]

    export_resp = client.get("/admin/api/winning/export?date=2026-04-06")
    assert export_resp.status_code == 200
    assert export_resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_admin_match_results_use_business_date_for_date_filter(app, client):
    with app.app_context():
        admin = User(username="admin_match_result_date", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        prev_business = MatchResult(
            detail_period="26034",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.23}}},
            uploaded_by=admin.id,
            uploaded_at=datetime(2026, 4, 7, 11, 0, 0),
        )
        current_business = MatchResult(
            detail_period="26035",
            result_data={"1": {"SPF": {"result": "1", "sp": 2.34}}},
            uploaded_by=admin.id,
            uploaded_at=datetime(2026, 4, 7, 13, 0, 0),
        )
        db.session.add_all([prev_business, current_business])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_match_result_date", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/match-results?date=2026-04-06")
    assert resp.status_code == 200
    data = resp.get_json()
    assert [r["detail_period"] for r in data["results"]] == ["26034"]
    assert "2026-04-06" in data["dates"]
    assert "2026-04-07" in data["dates"]


def test_upload_match_result_falls_back_to_sync_calc_when_scheduler_missing(app, client, monkeypatch):
    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: None)

    with app.app_context():
        admin = User(username="admin_result_sync", is_admin=True)
        admin.set_password("secret123")
        user = create_user("result_sync_user", "secret123", client_mode="mode_b")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26080",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([admin, ticket])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_result_sync", "password": "secret123"})
    assert resp.status_code == 200

    payload = "序号\t让球胜平负彩果\t让球胜平负SP值\n1\t3\t1.85\n".encode("utf-8")
    resp = client.post(
        "/admin/match-results/upload",
        data={
            "detail_period": "26080",
            "file": (io.BytesIO(payload), "26080期彩果-最终.txt"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200

    with app.app_context():
        match_result = MatchResult.query.filter_by(detail_period="26080").first()
        ticket = LotteryTicket.query.filter_by(detail_period="26080").first()

    assert match_result is not None
    assert match_result.calc_status == "done"
    assert ticket.is_winning is True


def test_upload_match_result_rejects_empty_filename(app, client):
    with app.app_context():
        admin = User(username="admin_result_empty_name", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_result_empty_name", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(
        "/admin/match-results/upload",
        data={
            "detail_period": "26100",
            "file": (io.BytesIO(b"payload"), ""),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_upload_match_result_rejects_filename_period_mismatch(app, client):
    with app.app_context():
        admin = User(username="admin_result_bad_period_name", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_result_bad_period_name", "password": "secret123"})
    assert resp.status_code == 200

    payload = "\u5e8f\u53f7\t\u8ba9\u7403\u80dc\u5e73\u8d1f\u5f69\u679c\\t\u8ba9\u7403\u80dc\u5e73\u8d1fSP\u503c\\n1\t3\t1.85\n".encode("utf-8")
    resp = client.post(
        "/admin/match-results/upload",
        data={
            "detail_period": "26080",
            "upload_kind": "final",
            "file": (io.BytesIO(payload), "26081鏈熷僵鏋?鏈€缁?txt"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "\u6587\u4ef6\u540d\u9700\u5305\u542b\u671f\u53f7 26080" in data["error"]


def test_upload_match_result_rejects_filename_kind_mismatch(app, client):
    with app.app_context():
        admin = User(username="admin_result_bad_kind_name", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_result_bad_kind_name", "password": "secret123"})
    assert resp.status_code == 200

    payload = "\u5e8f\u53f7\t\u8ba9\u7403\u80dc\u5e73\u8d1f\u5f69\u679c\\t\u8ba9\u7403\u80dc\u5e73\u8d1fSP\u503c\\n1\t3\t1.85\n".encode("utf-8")
    resp = client.post(
        "/admin/match-results/upload",
        data={
            "detail_period": "26080",
            "upload_kind": "final",
            "file": (io.BytesIO(payload), "26080鏈熷僵鏋?棰勬祴.txt"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert data["error"]


def test_parse_result_file_updates_latest_duplicate_match_result(app, tmp_path):
    from services.result_parser import parse_result_file

    with app.app_context():
        admin = User(username="result_duplicate_admin", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        admin_id = admin.id

        older = MatchResult(
            detail_period="26099",
            result_data={"61": {"SPF": {"result": "0", "sp": 1.1}}},
            uploaded_by=admin_id,
            uploaded_at=beijing_now() - timedelta(days=2),
        )
        newer = MatchResult(
            detail_period="26099",
            result_data={"61": {"SPF": {"result": "1", "sp": 1.2}}},
            uploaded_by=admin_id,
            uploaded_at=beijing_now() - timedelta(days=1),
        )
        db.session.add_all([older, newer])
        db.session.commit()
        older_id = older.id
        newer_id = newer.id

    result_file = tmp_path / "result_duplicate.txt"
    result_file.write_text(
        "序号\t让球胜平负彩果\t让球胜平负SP值\n61\t3\t1.88\n",
        encoding="utf-8",
    )

    with app.app_context():
        result = parse_result_file(str(result_file), "26099", admin_id)
        assert result["success"] is True
        assert result["match_result_id"] == newer_id

        refreshed_older = MatchResult.query.get(older_id)
        refreshed_newer = MatchResult.query.get(newer_id)

    assert refreshed_older.result_data == {"61": {"SPF": {"result": "0", "sp": 1.1}}}
    assert refreshed_newer.result_data["61"]["SPF"]["result"] == "3"


def test_parse_result_file_rejects_incomplete_same_kind_replace(app, tmp_path):
    from services.result_parser import parse_result_file

    with app.app_context():
        admin = User(username="result_incomplete_replace_admin", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        admin_id = admin.id

        existing = MatchResult(
            detail_period="26110",
            result_data={
                "1": {"SPF": {"result": "3", "sp": 1.80}},
                "2": {"SPF": {"result": "0", "sp": 2.10}},
            },
            uploaded_by=admin_id,
        )
        db.session.add(existing)
        db.session.commit()
        existing_id = existing.id

    incomplete_file = tmp_path / "result_incomplete_replace.txt"
    incomplete_file.write_text("\u5e8f\u53f7\tA\tB\n1\t3\t1.95\n", encoding="utf-8")

    with app.app_context():
        result = parse_result_file(str(incomplete_file), "26110", admin_id, upload_kind="final")
        refreshed = MatchResult.query.get(existing_id)

    assert result["success"] is False
    assert "incomplete final upload" in result["error"]
    assert refreshed.result_data["1"]["SPF"]["sp"] == 1.80
    assert refreshed.result_data["2"]["SPF"]["sp"] == 2.10


def test_recalc_rejects_non_latest_match_result_id(app, client):
    with app.app_context():
        admin = User(username="admin_recalc_latest_guard", is_admin=True)
        admin.set_password("secret123")
        older = MatchResult(
            detail_period="26081",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.11}}},
            uploaded_by=1,
            calc_status="done",
            tickets_total=9,
        )
        newer = MatchResult(
            detail_period="26081",
            result_data={"1": {"SPF": {"result": "1", "sp": 1.22}}},
            uploaded_by=1,
            calc_status="done",
            tickets_total=5,
        )
        db.session.add_all([admin, older, newer])
        db.session.commit()
        older_id = older.id
        newer_id = newer.id

    resp = client.post("/auth/login", json={"username": "admin_recalc_latest_guard", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(f"/admin/api/match-results/{older_id}/recalc")
    assert resp.status_code == 409
    data = resp.get_json()
    assert data["success"] is False
    assert data["latest_result_id"] == newer_id

    with app.app_context():
        refreshed_older = MatchResult.query.get(older_id)

    assert refreshed_older.calc_status == "done"
    assert refreshed_older.tickets_total == 9


def test_recalc_allows_latest_match_result_id(app, client, monkeypatch):
    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: None)
    monkeypatch.setattr(
        "services.winning_calc_service.process_match_result",
        lambda result_id, expected_calc_token=None, expected_uploaded_at=None, app=None: None,
    )

    with app.app_context():
        admin = User(username="admin_recalc_latest_allowed", is_admin=True)
        admin.set_password("secret123")
        older = MatchResult(
            detail_period="26082",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.11}}},
            uploaded_by=1,
            calc_status="done",
            tickets_total=9,
        )
        newer = MatchResult(
            detail_period="26082",
            result_data={"1": {"SPF": {"result": "1", "sp": 1.22}}},
            uploaded_by=1,
            calc_status="done",
            tickets_total=5,
        )
        db.session.add_all([admin, older, newer])
        db.session.commit()
        newer_id = newer.id

    resp = client.post("/auth/login", json={"username": "admin_recalc_latest_allowed", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(f"/admin/api/match-results/{newer_id}/recalc")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True

    with app.app_context():
        refreshed_newer = MatchResult.query.get(newer_id)

    assert refreshed_newer.calc_status == "pending"
    assert refreshed_newer.tickets_total == 0


def test_recalc_resets_stale_summary_when_scheduler_missing(app, client, monkeypatch):
    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: None)
    monkeypatch.setattr(
        "services.winning_calc_service.process_match_result",
        lambda result_id, expected_calc_token=None, expected_uploaded_at=None, app=None: None,
    )

    with app.app_context():
        admin = User(username="admin_recalc_reset", is_admin=True)
        admin.set_password("secret123")
        match_result = MatchResult(
            detail_period="26081",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.23}}},
            uploaded_by=1,
            calc_status="done",
            calc_started_at=beijing_now(),
            calc_finished_at=beijing_now(),
            tickets_total=12,
            tickets_winning=5,
            total_winning_amount=666,
        )
        db.session.add_all([admin, match_result])
        db.session.commit()
        result_id = match_result.id

    resp = client.post("/auth/login", json={"username": "admin_recalc_reset", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(f"/admin/api/match-results/{result_id}/recalc")
    assert resp.status_code == 200

    with app.app_context():
        refreshed = MatchResult.query.get(result_id)

    assert refreshed.calc_status == "pending"
    assert refreshed.calc_started_at is None
    assert refreshed.calc_finished_at is None
    assert refreshed.tickets_total == 0


def test_match_result_export_comparison_includes_user_and_device_rows(app, client):
    from openpyxl import load_workbook

    with app.app_context():
        admin = User(username="admin_match_export", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        user_a = create_user("match_export_user_a", "secret123", client_mode="mode_b")
        user_b = create_user("match_export_user_b", "secret123", client_mode="mode_b")

        match_result = MatchResult(
            detail_period="26999",
            lottery_type=None,
            result_data={"1": {"SPF": {"result": "3", "predicted_sp": 2.0, "sp": 2.2}}},
            uploaded_by=admin.id,
            calc_status="done",
        )
        db.session.add(match_result)
        db.session.flush()

        db.session.add_all([
            DeviceRegistry(device_id="dev-export-a", user_id=user_a.id, client_info={"device_name": "设备A"}),
            DeviceRegistry(device_id="dev-export-b", user_id=user_a.id, client_info={"device_name": "设备B"}),
            DeviceRegistry(device_id="dev-export-c", user_id=user_b.id, client_info={"device_name": "设备C"}),
        ])
        db.session.add_all([
            LotteryTicket(
                source_file_id=1,
                line_number=1,
                raw_content="EXPORT-A-1",
                status="completed",
                detail_period="26999",
                assigned_user_id=user_a.id,
                assigned_username=user_a.username,
                assigned_device_id="dev-export-a",
                predicted_winning_amount=100,
                winning_amount=120,
            ),
            LotteryTicket(
                source_file_id=1,
                line_number=2,
                raw_content="EXPORT-A-2",
                status="expired",
                detail_period="26999",
                assigned_user_id=user_a.id,
                assigned_username=user_a.username,
                assigned_device_id="dev-export-b",
                predicted_winning_amount=50,
                winning_amount=25,
            ),
            LotteryTicket(
                source_file_id=1,
                line_number=3,
                raw_content="EXPORT-B-1",
                status="completed",
                detail_period="26999",
                assigned_user_id=user_b.id,
                assigned_username=user_b.username,
                assigned_device_id="dev-export-c",
                predicted_winning_amount=0,
                winning_amount=30,
            ),
        ])
        db.session.commit()
        result_id = match_result.id

    resp = client.post("/auth/login", json={"username": "admin_match_export", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get(f"/admin/api/match-results/{result_id}/export-comparison")
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers["Content-Type"]

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == (
        "层级", "编号", "彩种范围", "客户", "设备名",
        "设备ID", "预测奖金", "最终奖金", "涨跌幅(%)"
    )

    data_rows = rows[1:]
    assert any(r[0] == "客户" and r[3] == "match_export_user_a" and r[6] == pytest.approx(150.0, rel=1e-3) and r[7] == pytest.approx(145.0, rel=1e-3) and r[8] == pytest.approx(-3.33, rel=1e-3) for r in data_rows)
    assert any(r[0] == "设备" and r[3] == "match_export_user_a" and r[4] == "设备A" and r[5] == "dev-export-a" and r[6] == pytest.approx(100.0, rel=1e-3) and r[7] == pytest.approx(120.0, rel=1e-3) and r[8] == pytest.approx(20.0, rel=1e-3) for r in data_rows)
    assert any(r[0] == "设备" and r[3] == "match_export_user_a" and r[4] == "设备B" and r[5] == "dev-export-b" and r[6] == pytest.approx(50.0, rel=1e-3) and r[7] == pytest.approx(25.0, rel=1e-3) and r[8] == pytest.approx(-50.0, rel=1e-3) for r in data_rows)
    assert any(r[0] == "客户" and r[3] == "match_export_user_b" and r[6] == pytest.approx(0.0, rel=1e-3) and r[7] == pytest.approx(30.0, rel=1e-3) and r[8] is None for r in data_rows)


def test_parse_result_file_keeps_predicted_and_final_sp_separate(app, tmp_path):
    from services.result_parser import parse_result_file

    with app.app_context():
        admin = User(username="predicted_final_parser_admin", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        admin_id = admin.id

    predicted_file = tmp_path / "predicted_result.txt"
    predicted_file.write_text("\u5e8f\u53f7\tA\tB\n1\t0\t1.85\n", encoding="utf-8")
    final_file = tmp_path / "final_result.txt"
    final_file.write_text("\u5e8f\u53f7\tA\tB\n1\t3\t2.05\n", encoding="utf-8")
    predicted_file_2 = tmp_path / "predicted_result_2.txt"
    predicted_file_2.write_text("\u5e8f\u53f7\tA\tB\n1\t1\t1.95\n", encoding="utf-8")

    with app.app_context():
        first = parse_result_file(str(predicted_file), "26188", admin_id, upload_kind="predicted")
        second = parse_result_file(str(final_file), "26188", admin_id, upload_kind="final")
        third = parse_result_file(str(predicted_file_2), "26188", admin_id, upload_kind="predicted")
        match_result = MatchResult.query.filter_by(detail_period="26188").first()

    assert first["success"] is True
    assert second["success"] is True
    assert third["success"] is True
    assert first["calc_token"].startswith("ts:")
    assert second["calc_token"].startswith("ts:")
    assert third["calc_token"].startswith("ts:")
    assert match_result.result_data["1"]["SPF"]["result"] == "3"
    assert match_result.result_data["1"]["SPF"]["predicted_result"] == "1"
    assert match_result.result_data["1"]["SPF"]["predicted_sp"] == 1.95
    assert match_result.result_data["1"]["SPF"]["sp"] == 2.05


def test_parse_result_file_returns_error_for_invalid_sp_value(app, tmp_path):
    from services.result_parser import parse_result_file

    with app.app_context():
        admin = User(username="invalid_sp_parser_admin", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        admin_id = admin.id

    bad_sp_file = tmp_path / "bad_sp_result.txt"
    bad_sp_file.write_text(
        "序号\t让球胜平负彩果\t让球胜平负SP值\n1\t3\tabc\n",
        encoding="utf-8",
    )

    with app.app_context():
        result = parse_result_file(str(bad_sp_file), "26901", admin_id, upload_kind="final")

    assert result["success"] is False
    assert result["count"] == 0
    assert "invalid sp value" in result["error"]


def test_parse_result_file_accepts_dash_sp_as_missing_value(app, tmp_path):
    from services.result_parser import parse_result_file

    with app.app_context():
        admin = User(username="dash_sp_parser_admin", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        admin_id = admin.id

    result_file = tmp_path / "dash_sp_result.txt"
    result_file.write_text("序号\tA\tB\n1\t3\t-\n", encoding="utf-8")

    with app.app_context():
        result = parse_result_file(str(result_file), "26903", admin_id, upload_kind="final")
        match_result = MatchResult.query.filter_by(detail_period="26903").first()

    assert result["success"] is True
    assert match_result.result_data["1"]["SPF"]["result"] == "3"
    assert match_result.result_data["1"]["SPF"]["sp"] is None


def test_parse_result_file_dash_sp_replaces_existing_sp(app, tmp_path):
    from services.result_parser import parse_result_file

    with app.app_context():
        admin = User(username="dash_sp_replace_admin", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        admin_id = admin.id

    full_file = tmp_path / "full_sp_result.txt"
    full_file.write_text("序号\tA\tB\n1\t3\t2.05\n", encoding="utf-8")
    dash_file = tmp_path / "dash_sp_replace_result.txt"
    dash_file.write_text("序号\tA\tB\n1\t3\t-\n", encoding="utf-8")

    with app.app_context():
        first = parse_result_file(str(full_file), "26904", admin_id, upload_kind="final")
        second = parse_result_file(str(dash_file), "26904", admin_id, upload_kind="final")
        match_result = MatchResult.query.filter_by(detail_period="26904").first()

    assert first["success"] is True
    assert second["success"] is True
    assert match_result.result_data["1"]["SPF"]["result"] == "3"
    assert match_result.result_data["1"]["SPF"]["sp"] is None


def test_admin_match_result_upload_returns_400_for_invalid_sp_value(app, client):
    with app.app_context():
        admin = User(username="admin_invalid_sp_upload", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_invalid_sp_upload", "password": "secret123"})
    assert resp.status_code == 200

    payload = "序号\t让球胜平负彩果\t让球胜平负SP值\n1\t3\tabc\n".encode("utf-8")
    resp = client.post(
        "/admin/match-results/upload",
        data={
            "detail_period": "26902",
            "upload_kind": "final",
            "file": (io.BytesIO(payload), "26902最终.txt"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "invalid sp value" in data["error"]


def test_winning_calc_stores_predicted_and_final_amounts_separately(app):
    from services.winning_calc_service import process_match_result

    with app.app_context():
        user = create_user("predicted_final_calc_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26189",
            result_data={"1": {"SPF": {"result": "3", "predicted_sp": 5, "sp": 10}}},
            uploaded_by=user.id,
        )
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26189",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([match_result, ticket])
        db.session.commit()
        match_result_id = match_result.id
        ticket_id = ticket.id

    process_match_result(match_result_id, app=app)

    with app.app_context():
        refreshed_match_result = MatchResult.query.get(match_result_id)
        refreshed_ticket = LotteryTicket.query.get(ticket_id)

    assert refreshed_ticket.is_winning is True
    assert float(refreshed_ticket.predicted_winning_amount) == 6.5
    assert float(refreshed_ticket.winning_amount) == 13.0
    assert float(refreshed_match_result.predicted_total_winning_amount) == 6.5
    assert float(refreshed_match_result.total_winning_amount) == 13.0


def test_winning_calc_final_result_wins_after_new_predicted_upload(app):
    from services.winning_calc_service import process_match_result

    with app.app_context():
        user = create_user("predicted_after_final_calc_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="261890",
            result_data={
                "1": {
                    "SPF": {
                        "predicted_result": "0",
                        "predicted_sp": 5,
                        "result": "3",
                        "sp": 10,
                    }
                }
            },
            uploaded_by=user.id,
        )
        final_win_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="261890",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        predicted_win_ticket = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="SPF|1=0|1*1|1",
            status="completed",
            detail_period="261890",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([match_result, final_win_ticket, predicted_win_ticket])
        db.session.commit()
        match_result_id = match_result.id
        final_win_ticket_id = final_win_ticket.id
        predicted_win_ticket_id = predicted_win_ticket.id

    process_match_result(match_result_id, app=app)

    with app.app_context():
        refreshed_match_result = db.session.get(MatchResult, match_result_id)
        refreshed_final_ticket = db.session.get(LotteryTicket, final_win_ticket_id)
        refreshed_predicted_ticket = db.session.get(LotteryTicket, predicted_win_ticket_id)

    assert refreshed_final_ticket.is_winning is True
    assert float(refreshed_final_ticket.winning_amount) == 13.0
    assert refreshed_final_ticket.predicted_winning_amount is None
    assert refreshed_predicted_ticket.is_winning is False
    assert refreshed_predicted_ticket.winning_amount is None
    assert float(refreshed_predicted_ticket.predicted_winning_amount) == 6.5
    assert refreshed_match_result.tickets_winning == 1
    assert float(refreshed_match_result.predicted_total_winning_amount) == 6.5
    assert float(refreshed_match_result.total_winning_amount) == 13.0


def test_winning_calc_notify_failure_does_not_flip_done_status(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    def fake_notify_admins(*args, **kwargs):
        raise RuntimeError("notify down")

    monkeypatch.setattr("services.notify_service.notify_admins", fake_notify_admins)

    with app.app_context():
        user = create_user("winning_notify_guard_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="261891",
            result_data={"1": {"SPF": {"result": "3", "sp": 10}}},
            uploaded_by=user.id,
            calc_status="pending",
        )
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="261891",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([match_result, ticket])
        db.session.commit()
        match_result_id = match_result.id
        ticket_id = ticket.id

    process_match_result(match_result_id, app=app)

    with app.app_context():
        refreshed_match_result = db.session.get(MatchResult, match_result_id)
        refreshed_ticket = db.session.get(LotteryTicket, ticket_id)

    assert refreshed_match_result.calc_status == "done"
    assert refreshed_match_result.calc_finished_at is not None
    assert refreshed_ticket.is_winning is True
    assert float(refreshed_ticket.winning_amount) == 13.0


def test_winning_calc_skips_stale_upload_token(app):
    from services.winning_calc_service import process_match_result

    with app.app_context():
        user = create_user("stale_result_token_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26190",
            result_data={"1": {"SPF": {"result": "3", "sp": 2.0}}},
            uploaded_by=user.id,
            calc_status="pending",
        )
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26190",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([match_result, ticket])
        db.session.commit()
        match_result_id = match_result.id
        ticket_id = ticket.id
        stale_token = match_result.uploaded_at.isoformat()

        match_result.uploaded_at = beijing_now() + timedelta(seconds=3)
        db.session.commit()

    process_match_result(match_result_id, expected_uploaded_at=stale_token, app=app)

    with app.app_context():
        refreshed_match_result = db.session.get(MatchResult, match_result_id)
        refreshed_ticket = db.session.get(LotteryTicket, ticket_id)

    assert refreshed_match_result.calc_status == "pending"
    assert refreshed_match_result.calc_started_at is None
    assert refreshed_match_result.calc_finished_at is None
    assert refreshed_ticket.winning_amount is None
    assert refreshed_ticket.predicted_winning_amount is None


def test_winning_calc_falls_back_to_predicted_when_final_is_incomplete(app):
    from services.winning_calc_service import process_match_result

    with app.app_context():
        user = create_user("final_incomplete_fallback_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26192",
            result_data={
                "1": {"SPF": {"result": "3", "predicted_sp": 5.0, "sp": None}},
                "2": {"SPF": {"result": "3", "sp": 1.5}},
            },
            uploaded_by=user.id,
        )
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26192",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([match_result, ticket])
        db.session.commit()
        match_result_id = match_result.id
        ticket_id = ticket.id

    process_match_result(match_result_id, app=app)

    with app.app_context():
        refreshed_match_result = db.session.get(MatchResult, match_result_id)
        refreshed_ticket = db.session.get(LotteryTicket, ticket_id)

    assert refreshed_ticket.is_winning is True
    assert float(refreshed_ticket.predicted_winning_amount) == 6.5
    assert refreshed_ticket.winning_amount is None
    assert refreshed_match_result.tickets_winning == 1
    assert float(refreshed_match_result.predicted_total_winning_amount) == 6.5
    assert float(refreshed_match_result.total_winning_amount) == 6.5


def test_winning_calc_uses_win_result_when_completeness_helper_disagrees(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    def fake_calculate_winning(raw_content, result_data, multiplier, sp_field='sp'):
        if sp_field == 'sp':
            return True, 13, 13, 0
        return False, 0, 0, 0

    monkeypatch.setattr("services.winning_calc_service.calculate_winning", fake_calculate_winning)
    monkeypatch.setattr("services.winning_calc_service.has_complete_result_data", lambda *args, **kwargs: False)

    with app.app_context():
        user = create_user("calc_vs_complete_guard_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26194",
            result_data={"1": {"SPF": {"result": "3", "sp": 10.0}}},
            uploaded_by=user.id,
        )
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26194",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([match_result, ticket])
        db.session.commit()
        match_result_id = match_result.id
        ticket_id = ticket.id

    process_match_result(match_result_id, app=app)

    with app.app_context():
        refreshed_match_result = db.session.get(MatchResult, match_result_id)
        refreshed_ticket = db.session.get(LotteryTicket, ticket_id)

    assert refreshed_ticket.is_winning is True
    assert float(refreshed_ticket.winning_amount) == 13.0
    assert refreshed_match_result.tickets_winning == 1
    assert float(refreshed_match_result.total_winning_amount) == 13.0


def test_winning_calc_stale_commit_does_not_delete_winning_image(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    calls = {'token_checks': 0, 'deleted_images': 0}

    def fake_token_matches(_token_meta, _result_file_id, _uploaded_at):
        calls['token_checks'] += 1
        return calls['token_checks'] == 1

    def fake_delete_stored_image(_oss_key, _image_url):
        calls['deleted_images'] += 1

    monkeypatch.setattr("services.winning_calc_service._token_matches", fake_token_matches)
    monkeypatch.setattr("services.winning_calc_service.delete_stored_image", fake_delete_stored_image)

    with app.app_context():
        user = create_user("stale_commit_image_guard_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="26193",
            result_data={"1": {"SPF": {"result": "0", "sp": 1.1}}},
            uploaded_by=user.id,
            calc_status="pending",
        )
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26193",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_amount=100,
            winning_image_url="/uploads/images/stale-guard.png",
        )
        db.session.add_all([match_result, ticket])
        db.session.commit()
        record = WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=ticket.winning_image_url,
            image_oss_key="oss/stale-guard.png",
            uploaded_by=user.id,
        )
        db.session.add(record)
        db.session.commit()
        match_result_id = match_result.id
        ticket_id = ticket.id
        record_id = record.id

    process_match_result(match_result_id, expected_calc_token="rf:999999", app=app)

    with app.app_context():
        refreshed_ticket = db.session.get(LotteryTicket, ticket_id)
        refreshed_record = db.session.get(WinningRecord, record_id)
        refreshed_result = db.session.get(MatchResult, match_result_id)

    assert calls['token_checks'] >= 2
    assert calls['deleted_images'] == 0
    assert refreshed_result.calc_status in {"pending", "processing"}
    assert refreshed_ticket.winning_image_url == "/uploads/images/stale-guard.png"
    assert refreshed_record is not None


def test_my_winning_falls_back_to_predicted_amount_when_final_missing(app, client):
    with app.app_context():
        user = create_user("predicted_only_client_user", "secret123", client_mode="mode_b")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26190",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            predicted_winning_gross=120,
            predicted_winning_amount=120,
            predicted_winning_tax=0,
        )
        db.session.add(ticket)
        db.session.commit()

    resp = login(client, "predicted_only_client_user", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/winning/my")
    assert resp.status_code == 200
    data = resp.get_json()
    record = [item for items in data["grouped"].values() for item in items][0]

    assert record["winning_amount"] is None
    assert record["display_winning_amount"] == 120.0
    assert record["display_winning_tax"] == 0.0
    assert record["is_predicted_display"] is True


def test_admin_winning_api_returns_predicted_amount_and_change_percent(app, client):
    with app.app_context():
        admin = User(username="predicted_final_admin", is_admin=True)
        admin.set_password("secret123")
        user = create_user("predicted_final_member", "secret123", client_mode="mode_b")
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26191",
            lottery_type="SPF",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            predicted_winning_amount=100,
            predicted_winning_tax=0,
            winning_amount=120,
            winning_tax=0,
        )
        db.session.add_all([admin, ticket])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "predicted_final_admin", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/winning")
    assert resp.status_code == 200
    data = resp.get_json()
    record = data["records"][0]

    assert data["summary"]["predicted_amount"] == 100.0
    assert data["summary"]["amount"] == 120.0
    assert record["predicted_winning_amount"] == 100.0
    assert record["winning_amount"] == 120.0
    assert record["winning_change_percent"] == 20.0


def test_admin_winning_api_returns_uploaded_profit_summary(app, client):
    with app.app_context():
        admin = User(username="winning_profit_summary_admin", is_admin=True)
        admin.set_password("secret123")
        user = create_user("winning_profit_summary_user", "secret123", client_mode="mode_b")
        uploaded = UploadedFile(
            original_filename="profit_source.txt",
            stored_filename="profit_source.txt",
            uploaded_at=datetime(2026, 4, 8, 13, 0, 0),
            status="active",
            total_tickets=1,
            actual_total_amount=1000,
        )
        revoked_uploaded = UploadedFile(
            original_filename="profit_revoked.txt",
            stored_filename="profit_revoked.txt",
            uploaded_at=datetime(2026, 4, 8, 14, 0, 0),
            status="revoked",
            total_tickets=1,
            actual_total_amount=500,
        )
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="26192",
            lottery_type="SPF",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=datetime(2026, 4, 8, 13, 30, 0),
            is_winning=True,
            winning_amount=300,
            winning_tax=0,
        )
        db.session.add_all([admin, uploaded, revoked_uploaded, ticket])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "winning_profit_summary_admin", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.get("/admin/api/winning?date=2026-04-08")
    assert resp.status_code == 200
    summary = resp.get_json()["summary"]

    assert summary["amount"] == 300.0
    assert summary["uploaded_amount"] == 1000.0
    assert summary["profit_amount"] == -700.0
    assert summary["profit_percent"] == -70.0


def test_admin_user_management_endpoints_reject_admin_targets(app, client):
    with app.app_context():
        super_admin = User(username="admin_guard_actor", is_admin=True)
        super_admin.set_password("secret123")
        protected_admin = User(username="admin_guard_target", is_admin=True)
        protected_admin.set_password("secret123")
        db.session.add_all([super_admin, protected_admin])
        db.session.commit()
        protected_admin_id = protected_admin.id

    resp = client.post("/auth/login", json={"username": "admin_guard_actor", "password": "secret123"})
    assert resp.status_code == 200

    update_resp = client.put(f"/admin/api/users/{protected_admin_id}", json={"max_devices": 9})
    assert update_resp.status_code == 403
    assert "管理员账号" in update_resp.get_json()["error"]

    delete_resp = client.delete(f"/admin/api/users/{protected_admin_id}")
    assert delete_resp.status_code == 403
    assert "管理员账号" in delete_resp.get_json()["error"]

    logout_resp = client.post(f"/admin/api/users/{protected_admin_id}/force-logout")
    assert logout_resp.status_code == 403
    assert "管理员账号" in logout_resp.get_json()["error"]


def test_mode_b_confirm_returns_error_when_nothing_completed(app):
    from services.mode_b_service import confirm_batch

    with app.app_context():
        result = confirm_batch([], user_id=1)
    assert result["success"] is False
    assert "\u7968\u636e" in result["error"]


def test_mode_b_confirm_can_complete_prefix_and_expire_rest(app):
    from services.mode_b_service import confirm_batch

    with app.app_context():
        user = create_user("modeb_partial_user", "secret123", client_mode="mode_b")
        first = create_assigned_ticket(user, "device-b", "BATCH-001", 1)
        second = create_assigned_ticket(user, "device-b", "BATCH-002", 2)
        third = create_assigned_ticket(user, "device-b", "BATCH-003", 3)
        for ticket in (first, second, third):
            ticket.deadline_time = beijing_now() - timedelta(minutes=1)
        db.session.commit()

        result = confirm_batch([first.id, second.id, third.id], user_id=user.id, completed_count=2)
        db.session.expire_all()
        refreshed = [db.session.get(LotteryTicket, ticket_id) for ticket_id in (first.id, second.id, third.id)]

    assert result["success"] is True
    assert result["completed_count"] == 2
    assert result["expired_count"] == 1
    assert [ticket.status for ticket in refreshed] == ["completed", "completed", "expired"]


def test_mode_b_confirm_validates_completed_count_after_deduping_ticket_ids(app):
    from services.mode_b_service import confirm_batch

    with app.app_context():
        user = create_user("modeb_confirm_dedup_user", "secret123", client_mode="mode_b")
        first = create_assigned_ticket(user, "device-b", "BATCH-DEDUP-001", 1)
        second = create_assigned_ticket(user, "device-b", "BATCH-DEDUP-002", 2)

        result = confirm_batch([first.id, first.id, second.id], user_id=user.id, completed_count=3)

    assert result["success"] is False
    assert "\u8303\u56f4" in result["error"]


def test_mode_b_finalize_ignores_duplicate_ticket_ids(app):
    from services.ticket_pool import finalize_tickets_batch

    with app.app_context():
        user = create_user("modeb_duplicate_finalize_user", "secret123", client_mode="mode_b")
        uploaded = UploadedFile(
            display_id="2026/04/07-99",
            original_filename="duplicate-finalize.txt",
            stored_filename="txt/2026-04-07/duplicate-finalize.txt",
            uploaded_by=user.id,
            total_tickets=1,
            pending_count=0,
            assigned_count=1,
            completed_count=0,
        )
        db.session.add(uploaded)
        db.session.flush()

        ticket = LotteryTicket(
            source_file_id=uploaded.id,
            line_number=1,
            raw_content="DUP-FINALIZE-001",
            status="assigned",
            assigned_user_id=user.id,
            assigned_username=user.username,
            assigned_device_id="device-b",
            assigned_at=beijing_now(),
            deadline_time=beijing_now() - timedelta(minutes=1),
        )
        db.session.add(ticket)
        db.session.commit()

        result = finalize_tickets_batch([ticket.id, ticket.id], user.id, completed_count=2)
        db.session.expire_all()
        refreshed_ticket = db.session.get(LotteryTicket, ticket.id)
        refreshed_file = db.session.get(UploadedFile, uploaded.id)

    assert result == {"completed_count": 1, "expired_count": 0}
    assert refreshed_ticket.status == "completed"
    assert refreshed_file.assigned_count == 0
    assert refreshed_file.completed_count == 1


def test_mode_b_preview_rejects_invalid_count(app, client):
    with app.app_context():
        user = create_user("mode_b_preview_invalid", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_preview_invalid", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/preview?count=abc")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert '整数' in data["error"]


def test_mode_b_download_rejects_invalid_count(app, client):
    with app.app_context():
        user = create_user("mode_b_download_invalid", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_download_invalid", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/download", json={"count": 0, "device_id": "dev-1"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert '整数' in data["error"]


def test_mode_b_preview_rejects_excessive_count(app, client):
    with app.app_context():
        create_user("mode_b_preview_excessive", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_preview_excessive", "secret123")
    assert resp.status_code == 200

    resp = client.get("/api/mode-b/preview?count=1001")
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "1000" in data["error"]


def test_mode_b_download_rejects_excessive_count_before_assignment(app, client, monkeypatch):
    called = []

    def fake_download_batch(**kwargs):
        called.append(kwargs)
        return {"success": True, "ticket_ids": [1], "actual_count": 1}

    monkeypatch.setattr("routes.mode_b.download_batch", fake_download_batch)

    with app.app_context():
        create_user("mode_b_download_excessive", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_download_excessive", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/download", json={"count": 1001, "device_id": "dev-1"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "1000" in data["error"]
    assert called == []


def test_client_dashboard_handles_mode_b_confirm_failure():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "throw new Error(data.error ||" in content
    assert "showToast(e.message ||" in content
    assert "body: JSON.stringify({ ticket_ids: batch.ticket_ids, completed_count: completedCount, device_id: currentDeviceId() })," in content
    assert "showToast(message, 'success');" in content
    assert "bDownloadCooldownUntil > Date.now()" in content


def test_client_dashboard_mode_b_overdue_prompts_completed_count_directly():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    start = content.index("resolveModeBCompletedCount(batch) {")
    end = content.index("async doStop()", start)
    snippet = content[start:end]
    assert "window.prompt(" in snippet
    assert "window.confirm(" not in snippet
    assert "请输入已完成的张数" in snippet
    assert "return batch.count;" not in snippet


def test_client_dashboard_replaces_processing_batches_from_server():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "this.bPendingBatches = data.batches || [];" in content


def test_client_dashboard_reloads_processing_batches_after_confirm():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "await this.loadProcessingBatches();" in content
    assert "this.bPendingBatches.splice(index, 1);" not in content


def test_client_dashboard_reloads_processing_batches_after_download():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert content.count("await this.loadProcessingBatches();") >= 2
    assert "this.bPendingBatches.push({" not in content


def test_client_dashboard_defaults_mode_b_count_to_smallest_option():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "bCount: 50," in content
    assert "bCountInitialized: false," in content
    assert "const minOption = Math.min(...this.bOptions);" in content
    assert "if (!this.bCountInitialized) {\n            this.bCount = minOption;\n            this.bCountInitialized = true;" in content
    assert "else if (!this.bOptions.includes(this.bCount)) {\n            this.bCount = minOption;" in content


def test_client_dashboard_shows_winning_download_filename():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "r.download_filename" in content
    assert "文件：" in content


def test_client_dashboard_shows_winning_assigned_time_and_summary():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "winningSummary: { record_count: 0, total_display_winning_amount: 0 }" in content
    assert "data.summary || { record_count: 0, total_display_winning_amount: 0 }" in content
    assert "winningSummary.total_display_winning_amount" in content
    assert "获取时间：" in content
    assert "r.assigned_at" in content
    assert ".winning-card-times {\n  display: flex; align-items: center; gap: .75rem;" in content


def test_client_dashboard_shows_winning_original_ticket_amount():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "票面金额：" in content
    assert "r.ticket_amount" in content
    assert ".winning-card-meta {" in content


def test_client_dashboard_defaults_winning_records_to_previous_business_day():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "winningDefaultBusinessDate: ''" in content
    assert "new URLSearchParams({ default_date: '1' })" in content
    assert "this.winningDefaultBusinessDate = data.default_business_date || '';" in content
    assert "this.winningFilterDate = this.winningDefaultBusinessDate;" in content
    assert "this.winningFilterDate = this.winningDefaultBusinessDate || '';" in content


def test_client_dashboard_auto_applies_winning_filters_on_select_change():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert 'v-model="winningFilterDate" style="min-width:120px" @change="applyWinningFilter"' in content
    assert 'v-model="winningFilterType" style="min-width:120px" @change="applyWinningFilter"' in content


def test_client_dashboard_emphasizes_winning_raw_content():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "font-size: .9rem; font-weight: 700; color: #495057;" in content


def test_client_dashboard_resets_matching_state_on_load_failures():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "} catch(e) {" in content
    assert "if (requestSeq !== dailyStatsRequestSeq) return;" in content
    assert "this.stats = { ticket_count: 0, total_amount: 0, pool_total_pending: 0, active_count: 0, device_stats: [] };" in content
    assert "throw new Error(data.error || '加载处理中批次失败');" in content
    assert "if (requestSeq !== modeBProcessingRequestSeq) return;" in content
    assert "this.bPendingBatches = [];" in content
    assert "throw new Error(data.error || '加载票池状态失败');" in content


def test_client_dashboard_only_calls_mode_b_endpoints_for_mode_b_users():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "isModeB: {{ (current_user.client_mode == 'mode_b') | tojson }}," in content
    assert "if (!this.isModeB) {\n      this.loadCurrentModeATicket();\n      this.loadModeAPoolStatus();\n    }" in content
    assert "if (this.isModeB) {\n      this.loadPoolStatus();\n      this.loadProcessingBatches();\n    }" in content
    assert "if (this.isModeB) {\n      setInterval(this.loadPoolStatus, 15000);\n    } else {\n      setInterval(this.loadModeAPoolStatus, 15000);\n    }" in content
    assert "if (this.isModeB) {\n        this.loadProcessingBatches();\n        this.loadPoolStatus();\n      }\n      this.loadStats();" in content


def test_client_dashboard_restores_mode_a_current_ticket_on_mount():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "async loadCurrentModeATicket()" in content
    assert "fetch(`/api/mode-a/current?device_id=${encodeURIComponent(currentDeviceId())}`)" in content
    assert "document.body.classList.add('mode-a-active');" in content
    assert "this.ticketHistory = [data.ticket];" in content
    assert "this.modeAActive = true;" in content


def test_client_dashboard_uses_dynamic_device_identifier_helpers():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "const currentDeviceId = () => getOrCreateDeviceId();" in content
    assert "device_id: currentDeviceId()," in content
    assert "client_type: 'web'," in content
    assert "id=\"current-device-id-display\"" in content
    assert "showDeviceIdPrompt('', 'edit')" in content


def test_web_device_registration_uses_single_identifier_flow():
    app_js = Path(__file__).resolve().parents[1] / "static" / "js" / "app.js"
    content = app_js.read_text(encoding="utf-8")
    assert "client_info: { client_type: 'web' }," in content
    assert "lottery_device_id" in content
    assert "fetch('/api/device/update'" in content
    assert "current_device_id: oldId || undefined" in content
    assert "new_device_id: id" in content
    assert "refreshDeviceIdDisplays()" in content


def test_mode_a_active_hides_navbar_logout_button():
    base_template = Path(__file__).resolve().parents[1] / "templates" / "base.html"
    style_path = Path(__file__).resolve().parents[1] / "static" / "css" / "style.css"

    base_content = base_template.read_text(encoding="utf-8")
    style_content = style_path.read_text(encoding="utf-8")

    assert 'class="d-inline mb-0 app-navbar-logout"' in base_content
    assert "body.mode-a-active .app-navbar-logout {" in style_content
    assert "display: none !important;" in style_content


def test_mode_a_mobile_layout_keeps_stop_action_visible():
    style_path = Path(__file__).resolve().parents[1] / "static" / "css" / "style.css"
    content = style_path.read_text(encoding="utf-8")
    assert "body.mode-a-active {" in content
    assert "overflow: hidden;" in content
    assert "body.mode-a-active .client-device-stats," in content
    assert "body.mode-a-active .client-quick-links," in content
    assert "body.mode-a-active .app-navbar-logout {" in content
    assert "body.mode-a-active #mode-a {" in content
    assert "position: fixed;" in content
    assert "body.mode-a-active #mode-a .card-body {" in content
    assert ".mode-a-bottom-bar {" in content
    assert "env(safe-area-inset-bottom)" in content
    assert "body.mode-a-active .mode-a-fullscreen-container {" in content
    assert "flex: 1;" in content


def test_client_dashboard_fully_resets_mode_a_state_when_current_ticket_missing():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "resetModeAState()" in content
    assert "this.showStopConfirm = false;" in content
    assert "this.nextCooldownUntil = 0;" in content
    assert "document.body.classList.remove('mode-a-active');" in content
    assert "if (!res.ok || data.success === false || !data.ticket) {\n          if (requestSeq !== modeACurrentTicketRequestSeq) return;\n          this.resetModeAState();" in content


def test_client_dashboard_stop_reanchors_from_history_to_latest_ticket():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "if (this.historyOffset > 0 && this.ticketHistory.length > 0) {" in content
    assert "this.currentTicket = this.ticketHistory[0];" in content
    assert "this.startCountdown(this.currentTicket.deadline_time);" in content


def test_client_dashboard_clears_current_ticket_when_next_returns_empty():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    next_ticket_start = content.index("async nextTicket() {")
    next_ticket_error_toast = content.index("showToast(data.error || '暂无可用票据', 'warning');")
    next_ticket_segment = content[next_ticket_start:next_ticket_error_toast]
    assert "this.historyOffset = 0;" in next_ticket_segment
    assert "this.currentTicket = null;" in next_ticket_segment
    assert "this.currentTicket = this.ticketHistory[0];" not in next_ticket_segment


def test_client_dashboard_resets_full_mode_a_state_after_stop_success():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "async doStop()" in content
    assert "this.resetModeAState();" in content
    assert "this.loadStats();" in content


def test_client_dashboard_cancel_stop_restores_mode_a_active_state():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "@click=\"cancelStopConfirm\"" in content
    assert "cancelStopConfirm()" in content
    assert "this.modeAActive = true;" in content
    assert "document.body.classList.add('mode-a-active');" in content


def test_client_dashboard_stop_without_current_ticket_resets_mode_a_state():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "if (this.currentTicket) {" in content
    assert "} else {\n        this.resetModeAState();" in content


def test_client_dashboard_deduplicates_mode_a_history_when_server_returns_same_ticket():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "this.ticketHistory = this.ticketHistory.filter(t => t && t.id !== data.ticket.id);" in content
    assert "this.ticketHistory.unshift(data.ticket);" in content


def test_client_dashboard_skips_mode_a_cooldown_when_server_returns_same_ticket():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "const previousCurrentTicketId = this.currentTicket ? this.currentTicket.id : null;" in content
    assert "if (!(data.completed_current === false && previousCurrentTicketId === data.ticket.id)) {" in content
    assert "this.startNextCooldown(data.cooldown_seconds || 3);" in content


def test_client_dashboard_mode_a_uses_three_second_cooldown_and_server_remaining():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "startNextCooldown(seconds = 3)" in content
    assert "const cooldownSeconds = Math.max(1, Number(seconds) || 3);" in content
    assert "data.cooldown_remaining" in content
    assert "服务端冷却中，请等待 " in content


def test_client_dashboard_mode_a_device_daily_records_ui():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "当前设备今日处理清单" in content
    assert "openModeADailyRecords" in content
    assert "fetch(`/api/mode-a/device-daily?device_id=${encodeURIComponent(currentDeviceId())}`)" in content
    assert "第 {{ currentTicket.device_today_sequence }} 张" in content
    assert "本设备当日 ¥{{ currentDeviceTodayAmount.toFixed(2) }}" in content
    assert "currentDeviceTodayAmount()" in content
    assert "票面金额：¥{{ Number(r.ticket_amount || 0).toFixed(2) }}" in content


def test_client_dashboard_mode_a_incoming_alert_only_on_zero_to_positive_pool():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "lastModeAPoolTotalPending: 0" in content
    assert "let modeAPoolStatusRequestSeq = 0;" in content
    assert "previousModeAPoolTotal === 0" in content
    assert "if (previousModeAPoolTotal === 0 && currentModeAPoolTotal > 0) {" in content
    assert "if (this.modeAActive && previousModeAPoolTotal === 0 && currentModeAPoolTotal > 0) {" not in content
    assert "this.loadModeAPoolStatus();" in content
    assert "if (requestSeq !== modeAPoolStatusRequestSeq) return;" in content


def test_client_dashboard_mode_b_incoming_alert_only_on_zero_to_positive_pool():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "lastModeBPoolTotalPending: 0" in content
    assert "const previousModeBPoolTotal = Number(this.lastModeBPoolTotalPending || 0);" in content
    assert "this.lastModeBPoolTotalPending = currentModeBPoolTotal;" in content
    assert "if (previousModeBPoolTotal === 0 && currentModeBPoolTotal > 0) {" in content
    assert "this.playIncomingTicketAlert();" in content
    assert "showToast('B模式票池有新票，可以下载了', 'info');" in content


def test_client_dashboard_blocks_mode_a_next_when_current_ticket_is_overdue():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert ":disabled=\"loadingNext || (historyOffset === 0 && nextCooldownSec > 0) || (currentTicket && isDeadlinePassed(currentTicket.deadline_time))\"" in content
    assert "currentTicket && isDeadlinePassed(currentTicket.deadline_time) ? '当前票已截止，请点击停止接单'" in content
    assert "if (this.currentTicket && this.isDeadlinePassed(this.currentTicket.deadline_time)) {" in content
    assert "showToast('当前票已截止，请先点击停止接单', 'warning');" in content


def test_client_dashboard_listens_for_realtime_revoke_and_announcement_events():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "window.addEventListener('pool_updated', this._onPoolUpdated);" in content
    assert "window.addEventListener('announcement', this._onAnnouncement);" in content
    assert "window.addEventListener('pool_disabled', this._onPoolDisabled);" in content
    assert "window.addEventListener('pool_enabled', this._onPoolEnabled);" in content
    assert "window.addEventListener('file_revoked', this._onFileRevoked);" in content
    assert "window.addEventListener('keydown', this._onModeAArrowKey);" in content
    assert "window.removeEventListener('keydown', this._onModeAArrowKey);" in content
    assert "if (event.key === 'ArrowLeft') {" in content
    assert "if (event.key === 'ArrowRight') {" in content
    assert "this._onPoolUpdated = (event) => {" in content
    assert "this.loadProcessingBatches();" in content
    assert "this.currentTicket = null;" in content
    assert "this._onPoolUpdated = (event) => {\n      if (this.isModeB) {\n        this.loadPoolStatus();\n        this.loadProcessingBatches();\n      } else {\n        this.handleModeAPoolUpdate(event);\n        this.loadModeAPoolStatus();\n      }\n      this.loadStats();\n    };" in content
    assert "this._onPoolDisabled = () => {\n      if (this.isModeB) {\n        this.loadPoolStatus();\n        this.loadProcessingBatches();\n      } else {\n        this.loadCurrentModeATicket();\n        this.loadModeAPoolStatus();\n      }\n      this.loadStats();\n    };" in content
    assert "this._onPoolEnabled = () => {\n      if (this.isModeB) {\n        this.loadPoolStatus();\n        this.loadProcessingBatches();\n      } else {\n        this.loadCurrentModeATicket();\n        this.loadModeAPoolStatus();\n      }\n      this.loadStats();\n    };" in content


def test_client_dashboard_mode_a_flip_buttons_show_text_hints():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "mode-a-side-hint" in content
    assert "mode-a-side-hint" in content


def test_client_dashboard_mode_a_history_indicator_uses_two_fixed_slots():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "v-for=\"slot in 2\"" in content
    assert ":class=\"historyDotClass(slot - 1)\"" in content
    assert "if (this.ticketHistory.length > 2) this.ticketHistory.pop();" in content
    assert "const start = totalSlots - count;" in content
    assert "const historyIndex = count - 1 - fromLeft;" in content


def test_client_dashboard_handles_mode_b_preview_failure():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "this.bPreview = null;" in content
    assert "showToast(e.message || '\u83b7\u53d6\u9884\u89c8\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content


def test_client_dashboard_handles_export_daily_network_failure():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "showToast('导出失败，请稍后重试', 'danger');" in content
    assert "showToast(data.error || data.message || '暂无可导出记录', 'warning');" in content


def test_user_export_daily_requires_login_json_response(app, client):
    resp = client.get("/api/user/export-daily")
    assert resp.status_code == 401
    assert resp.is_json
    data = resp.get_json()
    assert data["success"] is False


def test_client_dashboard_merges_returned_winning_record_after_upload():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "Object.assign(record, data.record || {}, { winning_image_url: data.image_url });" in content


def test_admin_winning_template_merges_returned_winning_record_after_upload():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "Object.assign(record, data.record || {}, { winning_image_url: data.image_url });" in content


def test_socket_client_dispatches_realtime_custom_events():
    socket_client = Path(__file__).resolve().parents[1] / "static" / "js" / "socket_client.js"
    content = socket_client.read_text(encoding="utf-8")
    assert "window.dispatchEvent(new CustomEvent('announcement', { detail: data }));" in content
    assert "window.dispatchEvent(new CustomEvent('pool_disabled', { detail: data }));" in content
    assert "window.dispatchEvent(new CustomEvent('pool_enabled', { detail: data }));" in content


def test_socket_request_pool_status_trims_mode_b_counts(app, monkeypatch):
    from types import SimpleNamespace
    from sockets.pool_events import on_request_pool_status

    emitted = []

    monkeypatch.setattr(
        "sockets.pool_events.current_user",
        SimpleNamespace(
            is_authenticated=True,
            client_mode="mode_b",
            get_blocked_lottery_types=lambda: ["???"],
        ),
    )
    monkeypatch.setattr(
        "sockets.pool_events.emit",
        lambda event, payload: emitted.append((event, payload)),
    )
    monkeypatch.setattr(
        "services.ticket_pool.get_pool_status",
        lambda blocked_types=None: {
            "total_pending": 25,
            "by_type": [
                {"lottery_type": "璁╃悆???", "deadline_time": "2026-04-08T10:00:00", "count": 25}
            ],
            "assigned": 0,
            "completed_today": 0,
        },
    )

    with app.app_context():
        on_request_pool_status()

    assert emitted == [
        (
            "pool_updated",
            {
                "total_pending": 5,
                "by_type": [
                    {"lottery_type": "璁╃悆???", "deadline_time": "2026-04-08T10:00:00", "count": 5}
                ],
                "assigned": 0,
                "completed_today": 0,
            },
        )
    ]


def test_admin_upload_template_uses_xlsx_export_label():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "\u5bfc\u51faXLSX" in content
    assert "\u5bfc\u51faCSV" not in content


def test_admin_upload_template_uses_server_derived_file_status():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "if (f.status === 'exhausted') return 'bg-success';" in content
    assert "if (f.status === 'expired') return 'bg-secondary';" in content
    assert "if (f.status === 'exhausted') return '已完成';" in content
    assert "if (f.status === 'expired') return '已过期';" in content
    assert "new Date(f.deadline_time) < new Date()" not in content


def test_admin_upload_template_loads_all_detail_pages():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "let page = 1;" in content
    assert "let totalPages = 1;" in content
    assert "detail?page=${page}&per_page=100" in content
    assert "if (!res.ok || data.success === false) {" in content
    assert "detailTickets.push(...(data.tickets || []));" in content
    assert "} while (page <= totalPages);" in content
    assert "showToast(e.message || '\u52a0\u8f7d\u8be6\u60c5\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content
    assert "showToast(e.message || '\u64a4\u56de\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content


def test_admin_upload_template_shows_assigned_count_column():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "<th>处理中</th>" in content
    assert "{{ f.assigned_count }}" in content


def test_admin_upload_template_distinguishes_pending_files_from_in_progress_files():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "if (f.status === 'active' && (f.assigned_count || 0) === 0 && (f.pending_count || 0) > 0) return 'bg-primary';" in content
    assert "if (f.status === 'active' && (f.assigned_count || 0) === 0 && (f.pending_count || 0) > 0) return '待处理';" in content
    assert "return '处理中';" in content


def test_admin_upload_template_shows_assigned_device_id_in_detail():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "<th>\u8bbe\u5907ID</th>" in content
    assert "{{ t.assigned_device_id || '-' }}" in content


def test_admin_upload_template_accepts_uppercase_txt_files():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert 'accept=".txt,.TXT"' in content
    assert "f.name.toLowerCase().endsWith('.txt')" in content
    assert "Array.from(e.target.files)\n        .filter(f => f.name.toLowerCase().endsWith('.txt'))" in content


def test_admin_users_template_handles_initial_load_failures():
    users_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "users.html"
    content = users_template.read_text(encoding="utf-8")
    assert "v-if=\"loadError\"" in content
    assert "loadError: ''" in content
    assert "showToast(e.message || '\u52a0\u8f7d\u5f69\u79cd\u5217\u8868\u5931\u8d25', 'danger');" in content
    assert "this.loadError = e.message || '\u52a0\u8f7d\u7528\u6237\u5217\u8868\u5931\u8d25';" in content
    assert "showToast(this.loadError, 'danger');" in content
    assert "finally {" in content
    assert "this.loading = false;" in content


def test_admin_users_template_handles_action_network_failures():
    users_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "users.html"
    content = users_template.read_text(encoding="utf-8")
    assert "this.createError = e.message || '\u7f51\u7edc\u5f02\u5e38\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5';" in content
    assert "showToast(e.message || '\u66f4\u65b0\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content
    assert "showToast(e.message || '\u64cd\u4f5c\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content
    assert "showToast(e.message || '\u5220\u9664\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content


def test_admin_upload_template_handles_file_list_failures():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "v-if=\"listError\"" in content
    assert "listError: ''" in content
    assert "this.listError = '';" in content
    assert "throw new Error(data.error || '\u52a0\u8f7d\u6587\u4ef6\u5217\u8868\u5931\u8d25');" in content
    assert "this.listError = e.message || '\u52a0\u8f7d\u6587\u4ef6\u5217\u8868\u5931\u8d25';" in content
    assert "this.page = 1;" in content
    assert "this.dateOptions = [];" in content
    assert "showToast(this.listError, 'danger');" in content
    assert "finally {" in content
    assert "this.loading = false;" in content


def test_admin_upload_template_handles_http_upload_failures():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "if (!res.ok || data.success === false) {" in content
    assert "throw new Error(data.error || '\u4e0a\u4f20\u5931\u8d25');" in content
    assert "i.message=e.message || '\u4e0a\u4f20\u5931\u8d25'" in content
    assert "showToast(e.message || '\u4e0a\u4f20\u5931\u8d25', 'danger');" in content
    assert "throw new Error(data.error || data.message || '\u64a4\u56de\u5931\u8d25');" in content


def test_admin_upload_template_maps_per_file_results_before_batch_failure_throw():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert content.index("if (data.results) {") < content.index("throw new Error(data.error || '\u4e0a\u4f20\u5931\u8d25');")


def test_admin_upload_template_retries_non_done_items():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "const pendingItems = this.uploadQueue.filter(i => i.status !== 'done');" in content
    assert "i.status = 'pending';" in content
    assert "i.message = '';" in content


def test_admin_upload_template_maps_results_by_queue_position_before_fallback_name_match():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "const item = pendingItems[i] || this.uploadQueue.find(q => q.name === (r.filename || pendingItems[i]?.name));" in content


def test_admin_upload_template_skips_empty_mutation_batches_before_network_call():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "if (pendingItems.length === 0) {" in content
    assert "showToast('\u6ca1\u6709\u53ef\u4e0a\u4f20\u7684\u6587\u4ef6', 'warning');" in content
    assert content.index("if (pendingItems.length === 0) {") < content.index("this.uploading = true;")


def test_admin_upload_template_clears_stale_detail_rows_before_and_after_failed_load():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "async viewDetail(fileId) {" in content
    assert content.count("this.detailTickets = [];") >= 2
    assert "this.detailTickets = detailTickets;" in content


def test_admin_upload_template_ignores_stale_detail_responses():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "detailRequestSeq: 0," in content
    assert "const requestSeq = ++this.detailRequestSeq;" in content
    assert content.count("if (requestSeq !== this.detailRequestSeq) return;") >= 3


def test_admin_upload_template_syncs_page_with_server_response():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "this.page = data.page || 1;" in content
    assert "this.totalPages = data.pages || 1;" in content


def test_admin_upload_template_clears_stale_date_filter_and_reloads():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "if (this.filterDate && this.dateOptions.length > 0 && !this.dateOptions.includes(this.filterDate)) {" in content
    assert "this.filterDate = '';" in content
    assert "await this.loadFiles();" in content
    assert "return;" in content


def test_admin_upload_template_resets_to_first_page_after_successful_mutations():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert content.count("this.page = 1;") >= 2
    assert "if (data.success) {\n          this.page = 1;\n          this.loadFiles();\n        }" in content


def test_admin_upload_template_defaults_date_filter_to_current_business_date():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "filterDate: window.DEFAULT_FILE_FILTER_DATE || ''," in content
    assert "currentBusinessDate: ''," in content
    assert "defaultDateInitialized" not in content
    assert "this.currentBusinessDate = data.current_business_date || '';" in content
    assert "include_date_options" in content


def test_admin_upload_template_checks_export_http_errors_before_download():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "async exportByDate() {" in content
    assert "const res = await fetch(`/admin/api/tickets/export-by-date?${params}`);" in content
    assert "if (!res.ok) {" in content
    assert "throw new Error(data.error || data.message || '\u5bfc\u51fa\u5931\u8d25');" in content
    assert "const blob = await res.blob();" in content
    assert "showToast(e.message || '\u5bfc\u51fa\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content


def test_admin_winning_template_checks_export_http_errors_before_download():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "async exportWinning() {" in content
    assert "const res = await fetch(`/admin/api/winning/export?${params}`);" in content
    assert "async exportMatchResultComparison(row) {" in content
    assert "const res = await fetch(`/admin/api/match-results/${row.id}/export-comparison`);" in content
    assert "if (!res.ok) {" in content
    assert "const blob = await res.blob();" in content
    assert "showToast(e.message || '\u5bfc\u51fa\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content


def test_admin_upload_template_listens_for_realtime_file_events():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "this._reloadFileList = () => {" in content
    assert "window.addEventListener('file_uploaded', this._reloadFileList);" in content
    assert "window.addEventListener('file_revoked', this._reloadFileList);" in content
    assert "window.removeEventListener('file_uploaded', this._reloadFileList);" in content
    assert "window.removeEventListener('file_revoked', this._reloadFileList);" in content
    assert "window.addEventListener('pool_updated', this._reloadFileList);" not in content
    assert "window.removeEventListener('pool_updated', this._reloadFileList);" not in content


def test_uploaded_files_model_indexes_uploaded_at_for_file_list_performance():
    model = Path(__file__).resolve().parents[1] / "models" / "file.py"
    content = model.read_text(encoding="utf-8")
    assert "idx_uploaded_files_uploaded_at" in content


def test_uploaded_files_uploaded_at_index_migration_exists():
    root = Path(__file__).resolve().parents[1]
    up = root / "migrations" / "add_uploaded_files_uploaded_at_index.up.sql"
    down = root / "migrations" / "add_uploaded_files_uploaded_at_index.down.sql"
    assert up.exists()
    assert down.exists()
    assert "idx_uploaded_files_uploaded_at" in up.read_text(encoding="utf-8")
    assert "idx_uploaded_files_uploaded_at" in down.read_text(encoding="utf-8")


def test_admin_upload_template_ignores_stale_file_list_responses():
    upload_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "upload.html"
    content = upload_template.read_text(encoding="utf-8")
    assert "listRequestSeq: 0," in content
    assert "const requestSeq = ++this.listRequestSeq;" in content
    assert content.count("if (requestSeq !== this.listRequestSeq) return;") >= 3


def test_socket_client_dispatches_file_uploaded_custom_event():
    socket_client = Path(__file__).resolve().parents[1] / "static" / "js" / "socket_client.js"
    content = socket_client.read_text(encoding="utf-8")
    assert "socket.on('file_uploaded', (data) => {" in content
    assert "window.dispatchEvent(new CustomEvent('file_uploaded', { detail: data }));" in content


def test_admin_files_list_clamps_page_after_result_set_shrinks(app, client):
    with app.app_context():
        admin = User(username="admin_file_page_clamp", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        for i in range(3):
            db.session.add(
                UploadedFile(
                    original_filename=f"page-clamp-{i}.txt",
                    stored_filename=f"page-clamp-{i}.txt",
                    uploaded_by=admin.id,
                    total_tickets=1,
                    pending_count=1,
                    assigned_count=0,
                    completed_count=0,
                    uploaded_at=beijing_now(),
                )
            )
        db.session.commit()

    resp = login(client, "admin_file_page_clamp", "secret123")
    assert resp.status_code == 200

    resp = client.get("/admin/api/files?page=3&per_page=2")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["page"] == 2
    assert len(data["files"]) == 1


def test_admin_files_list_uses_database_pagination_before_serialization(app, client, monkeypatch):
    serialized_ids = []
    original_to_dict = UploadedFile.to_dict

    def tracking_to_dict(self):
        serialized_ids.append(self.id)
        return original_to_dict(self)

    monkeypatch.setattr(UploadedFile, "to_dict", tracking_to_dict)

    with app.app_context():
        admin = User(username="admin_file_db_pagination", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.flush()
        for i in range(5):
            db.session.add(
                UploadedFile(
                    original_filename=f"db-page-{i}.txt",
                    stored_filename=f"db-page-{i}.txt",
                    uploaded_by=admin.id,
                    total_tickets=1,
                    pending_count=1,
                    assigned_count=0,
                    completed_count=0,
                    uploaded_at=beijing_now() + timedelta(minutes=i),
                )
            )
        db.session.commit()

    resp = login(client, "admin_file_db_pagination", "secret123")
    assert resp.status_code == 200

    resp = client.get("/admin/api/files?page=2&per_page=2")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["total"] == 5
    assert data["pages"] == 3
    assert data["page"] == 2
    assert len(data["files"]) == 2
    assert len(serialized_ids) == 2


def test_admin_winning_template_handles_list_and_detail_load_failures():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "throw new Error(data.error || '\u52a0\u8f7d\u4e2d\u5956\u8bb0\u5f55\u5931\u8d25');" in content
    assert "showToast(e.message || '\u52a0\u8f7d\u4e2d\u5956\u8bb0\u5f55\u5931\u8d25', 'danger');" in content
    assert "throw new Error(data.error || '\u52a0\u8f7d\u8d5b\u679c\u5217\u8868\u5931\u8d25');" in content
    assert "showToast(e.message || '\u52a0\u8f7d\u8d5b\u679c\u5217\u8868\u5931\u8d25', 'danger');" in content
    assert "throw new Error(data.error || '\u52a0\u8f7d\u8d5b\u679c\u8be6\u60c5\u5931\u8d25');" in content
    assert "showToast(e.message || '\u52a0\u8f7d\u8d5b\u679c\u8be6\u60c5\u5931\u8d25', 'danger');" in content


def test_admin_winning_template_formats_sp_display_to_two_decimals():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "formatSpDisplay(value) {" in content
    assert "return num.toFixed(2);" in content
    assert "const predicted = this.formatSpDisplay(play.predicted_sp);" in content
    assert "const finalSp = this.formatSpDisplay(play.sp);" in content


def test_admin_dashboard_and_winning_templates_check_http_status_on_actions():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    dashboard_content = dashboard_template.read_text(encoding="utf-8")
    assert "if (!res.ok || data.success === false) {" in dashboard_content
    assert "throw new Error(data.error || '\u64cd\u4f5c\u5931\u8d25');" in dashboard_content
    assert "showToast(e.message || '\u64cd\u4f5c\u5931\u8d25', 'danger');" in dashboard_content

    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    winning_content = winning_template.read_text(encoding="utf-8")
    assert winning_content.count("if (!res.ok || data.success === false) {") >= 4
    assert "throw new Error(data.error || '\u6807\u8bb0\u5931\u8d25');" in winning_content
    assert "throw new Error(data.error || '\u4e0a\u4f20\u5931\u8d25');" in winning_content
    assert "throw new Error(data.error || '\u63d0\u4ea4\u5931\u8d25');" in winning_content
    assert "showToast(e.message || '\u6807\u8bb0\u5931\u8d25', 'danger');" in winning_content
    assert "this.uploadMsg = e.message || '\u7f51\u7edc\u5f02\u5e38\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5';" in winning_content


def test_admin_winning_defaults_date_filter_to_current_business_day():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "currentBusinessDate: ''" in content
    assert "this.currentBusinessDate = data.current_business_date || '';" in content
    assert "if (!this.filterDate && this.currentBusinessDate) {" in content
    assert "this.filterDate = this.currentBusinessDate;" in content
    assert "resetWinningFilters()" in content

def test_client_dashboard_handles_mode_b_network_failures():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "throw new Error(data.error || '\u83b7\u53d6\u9884\u89c8\u5931\u8d25');" in content
    assert "showToast(e.message || '\u83b7\u53d6\u9884\u89c8\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content
    assert "throw new Error(data.error || '\u786e\u8ba4\u5b8c\u6210\u5931\u8d25');" in content
    assert "showToast(e.message || '\u786e\u8ba4\u5b8c\u6210\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content


def test_client_dashboard_renders_fixed_announcement_card_and_hides_global_bar():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "当前公告" in content
    assert "announcementDisplay" in content
    assert "if (!this.stats.can_receive) {" in content
    assert "return '无公告';" in content
    assert "const announcementBar = document.getElementById('announcement-bar');" in content
    assert "announcementBar.classList.add('d-none');" in content
    assert ".client-announcement-card {" in content


def test_client_dashboard_plays_distinct_announcement_alert_sound_for_all_modes():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "announcementAudioContext: null" in content
    assert "announcementAlertAt: 0" in content
    assert "ensureAnnouncementAudioReady()" in content
    assert "playAnnouncementAlert()" in content
    assert "[523.25, 659.25, 783.99].forEach" in content
    assert "gain.gain.exponentialRampToValueAtTime(2, startAt + 0.02);" in content
    assert "oscillator.frequency.setValueAtTime(frequency, startAt);" in content
    assert "this.playAnnouncementAlert();" in content
    assert "if (!this.stats.can_receive) {" in content


def test_client_dashboard_handles_download_and_open_winning_failures():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "throw new Error(data.error || '\u4e0b\u8f7d\u5931\u8d25');" in content
    assert "showToast(e.message || '\u4e0b\u8f7d\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content
    assert "throw new Error(data.error || '\u52a0\u8f7d\u4e2d\u5956\u8bb0\u5f55\u5931\u8d25');" in content
    assert "throw new Error(data.error || '\u7b5b\u9009\u5931\u8d25');" in content
    assert "showToast(e.message || '\u52a0\u8f7d\u4e2d\u5956\u8bb0\u5f55\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5, 'danger');" not in content
    assert "showToast(e.message || '\u52a0\u8f7d\u4e2d\u5956\u8bb0\u5f55\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content
    assert "showToast(e.message || '\u7b5b\u9009\u5931\u8d25', 'danger');" in content


def test_admin_dashboard_marks_refresh_failure_in_indicator():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "const onlineIndicator = document.getElementById('online-indicator');" in content
    assert "throw new Error(data.error ||" in content
    assert "onlineIndicator.textContent = '\u8fde\u63a5\u5f02\u5e38';" in content
    assert "onlineIndicator.className = 'badge bg-danger';" in content


def test_admin_dashboard_ignores_stale_refresh_responses():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "let dashboardRequestSeq = 0;" in content
    assert "const requestSeq = ++dashboardRequestSeq;" in content
    assert content.count("if (requestSeq !== dashboardRequestSeq) return;") >= 2


def test_admin_dashboard_ignores_stale_announcement_responses():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "let announcementRequestSeq = 0;" in content
    assert content.count("const requestSeq = ++announcementRequestSeq;") >= 2
    assert content.count("if (requestSeq !== announcementRequestSeq) return;") >= 5


def test_admin_dashboard_contains_announcement_panel():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "发送公告" in content
    assert "id=\"announcement-input\"" in content
    assert "async function loadAnnouncementSettings()" in content
    assert "async function saveAnnouncement()" in content
    assert "loadAnnouncementSettings();" in content


def test_admin_dashboard_renders_health_summary_panel():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "id=\"health-summary-badge\"" in content
    assert "id=\"health-summary-items\"" in content
    assert "function renderHealthSummary(summary)" in content
    assert "renderHealthSummary(data.health_summary);" in content


def test_admin_dashboard_renders_device_speed_detail_table_and_speed_overview():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "\u8bbe\u5907\u5904\u7406\u901f\u5ea6\u7edf\u8ba1" in content
    assert "id=\"device-speed-tbody\"" in content
    assert "data.device_speed_stats" in content
    assert "id=\"stat-speed\"" in content
    assert "id=\"stat-eta\"" in content


def test_admin_settings_no_longer_renders_announcement_editor():
    settings_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "settings.html"
    content = settings_template.read_text(encoding="utf-8")
    assert "\u516c\u544a\u5185\u5bb9" not in content
    assert "\u663e\u793a\u516c\u544a" not in content
def test_client_dashboard_validates_winning_image_type_and_handles_password_http_errors():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "if (!res.ok) {" in content
    assert "showToast(data.error || '\u83b7\u53d6\u4e0b\u4e00\u5f20\u5931\u8d25\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5', 'danger');" in content
    assert "if (!file.type.startsWith('image/')) {" in content
    assert "showToast('\u8bf7\u4e0a\u4f20\u56fe\u7247\u6587\u4ef6', 'warning');" in content
    assert "if (!res.ok || data.success === false) {" in content
    assert "throw new Error(data.error || '\u5bc6\u7801\u4fee\u6539\u5931\u8d25');" in content
    assert "this.pwdError = e.message || '\u7f51\u7edc\u5f02\u5e38\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5';" in content
    assert "showToast(data.error || '\u4e0a\u4f20\u5931\u8d25', 'danger');" in content
    assert "showToast('\u4e0a\u4f20\u5931\u8d25', 'danger');" in content


def test_client_dashboard_resets_stats_when_load_fails():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "throw new Error(data.error || '\u52a0\u8f7d\u7edf\u8ba1\u5931\u8d25');" in content
    assert "this.stats = { ticket_count: 0, total_amount: 0, pool_total_pending: 0, active_count: 0, device_stats: [] };" in content
def test_admin_settings_template_checks_http_status_on_load():
    settings_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "settings.html"
    content = settings_template.read_text(encoding="utf-8")
    assert "if (!res.ok || data.success === false) {" in content
    assert "throw new Error(data.error ||" in content
    assert "this.saved = true;" in content


def test_admin_settings_template_renders_scheduler_status_panel():
    settings_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "settings.html"
    content = settings_template.read_text(encoding="utf-8")
    assert "调度器状态" in content
    assert "scheduler_status" in content
    assert "schedulerStatusLabel(status)" in content
    assert "schedulerStatusBadgeClass(status)" in content


def test_admin_settings_template_ignores_stale_load_and_save_responses():
    settings_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "settings.html"
    content = settings_template.read_text(encoding="utf-8")
    assert "let settingsRequestSeq = 0;" in content
    assert "const requestSeq = ++settingsRequestSeq;" in content
    assert "if (requestSeq !== settingsRequestSeq) return;" in content
    assert "setTimeout(() => this.saved = false, 3000);" in content


def test_admin_users_template_ignores_stale_list_and_lottery_type_responses():
    users_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "users.html"
    content = users_template.read_text(encoding="utf-8")
    assert "let usersListRequestSeq = 0;" in content
    assert "let lotteryTypesRequestSeq = 0;" in content
    assert "const requestSeq = ++lotteryTypesRequestSeq;" in content
    assert "const requestSeq = ++usersListRequestSeq;" in content
    assert "if (requestSeq !== lotteryTypesRequestSeq) return;" in content
    assert "if (requestSeq !== usersListRequestSeq) return;" in content


def test_admin_winning_template_ignores_stale_filter_record_and_match_result_responses():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "let winningFilterOptionsRequestSeq = 0;" in content
    assert "let winningRecordsRequestSeq = 0;" in content
    assert "let matchResultsRequestSeq = 0;" in content
    assert "const requestSeq = ++winningFilterOptionsRequestSeq;" in content
    assert "const requestSeq = ++winningRecordsRequestSeq;" in content
    assert "const requestSeq = ++matchResultsRequestSeq;" in content
    assert "if (requestSeq !== winningFilterOptionsRequestSeq) return;" in content
    assert "if (requestSeq !== winningRecordsRequestSeq) return;" in content
    assert "if (requestSeq !== matchResultsRequestSeq) return;" in content


def test_admin_winning_template_ignores_stale_match_result_detail_responses():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "let matchResultDetailRequestSeq = 0;" in content
    assert "const requestSeq = ++matchResultDetailRequestSeq;" in content
    assert "if (requestSeq !== matchResultDetailRequestSeq) return;" in content
    assert "this.detailLoading = true;" in content


def test_client_dashboard_ignores_stale_mode_a_and_mode_b_load_responses():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "let modeACurrentTicketRequestSeq = 0;" in content
    assert "let dailyStatsRequestSeq = 0;" in content
    assert "let modeBProcessingRequestSeq = 0;" in content
    assert "let modeBPoolStatusRequestSeq = 0;" in content
    assert "const requestSeq = ++modeACurrentTicketRequestSeq;" in content
    assert "const requestSeq = ++dailyStatsRequestSeq;" in content
    assert "const requestSeq = ++modeBProcessingRequestSeq;" in content
    assert "const requestSeq = ++modeBPoolStatusRequestSeq;" in content
    assert "if (requestSeq !== modeACurrentTicketRequestSeq) return;" in content
    assert "if (requestSeq !== dailyStatsRequestSeq) return;" in content
    assert "if (requestSeq !== modeBProcessingRequestSeq) return;" in content
    assert "if (requestSeq !== modeBPoolStatusRequestSeq) return;" in content


def test_client_dashboard_ignores_stale_winning_record_responses():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "let winningRecordsRequestSeq = 0;" in content
    assert "const requestSeq = ++winningRecordsRequestSeq;" in content
    assert "if (requestSeq !== winningRecordsRequestSeq) return;" in content
    assert "this.winningVisible = true;" in content


def test_database_info_moves_to_settings_page():
    settings_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "settings.html"
    settings_content = settings_template.read_text(encoding="utf-8")
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "dashboard.html"
    dashboard_content = dashboard_template.read_text(encoding="utf-8")
    assert "数据库信息" in settings_content
    assert "databaseInfo.engine" in settings_content
    assert "databaseInfo.path" in settings_content
    assert "数据库信息" not in dashboard_content


def test_admin_settings_api_includes_database_info():
    admin_route = Path(__file__).resolve().parents[1] / "routes" / "admin.py"
    content = admin_route.read_text(encoding="utf-8")
    assert "payload['database_info'] = _database_display_info()" in content
    assert "return jsonify(payload)" in content
    assert "return render_template('admin/dashboard.html')" in content
    assert "return render_template('admin/dashboard.html', database_info=_database_display_info())" not in content


def test_admin_winning_template_handles_filter_option_failures_and_localizes_mark_checked():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    content = winning_template.read_text(encoding="utf-8")
    assert "throw new Error(data.error ||" in content
    assert "showToast(e.message ||" in content
    assert "showToast(" in content and "success" in content
def test_client_dashboard_clears_stale_state_when_background_loads_fail():
    dashboard_template = Path(__file__).resolve().parents[1] / "templates" / "client" / "dashboard.html"
    content = dashboard_template.read_text(encoding="utf-8")
    assert "this.stats = { ticket_count: 0, total_amount: 0, pool_total_pending: 0, active_count: 0, device_stats: [] };" in content
    assert "this.bPendingBatches = [];" in content
    assert "this.poolStatus = { total_pending: 0, by_type: [] };" in content
def test_templates_use_chinese_for_recent_admin_prompt_fixes():
    winning_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "winning.html"
    winning_content = winning_template.read_text(encoding="utf-8")
    settings_template = Path(__file__).resolve().parents[1] / "templates" / "admin" / "settings.html"
    settings_content = settings_template.read_text(encoding="utf-8")
    login_template = Path(__file__).resolve().parents[1] / "templates" / "login.html"
    login_content = login_template.read_text(encoding="utf-8")
    assert "throw new Error(data.error ||" in winning_content
    assert "showToast(" in winning_content and "'success'" in winning_content
    assert "加载设置失败" in settings_content
    assert "登录失败（HTTP ${res.status}）" in login_content
    assert "this.error = data.error || '登录失败';" in login_content
    assert "this.error = '网络异常，请稍后重试';" in login_content
def test_mode_b_download_rejects_float_count_payload(app, client):
    with app.app_context():
        create_user("mode_b_download_float_count_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_download_float_count_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/download", json={"count": 1.9, "device_id": "dev-float"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_mode_b_download_rejects_bool_count_payload(app, client):
    with app.app_context():
        create_user("mode_b_download_bool_count_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_download_bool_count_user", "secret123")
    assert resp.status_code == 200

    resp = client.post("/api/mode-b/download", json={"count": True, "device_id": "dev-bool"})
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_mode_b_confirm_rejects_float_ticket_ids(app, client):
    with app.app_context():
        create_user("mode_b_confirm_float_ticket_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_confirm_float_ticket_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/mode-b/confirm",
        json={"ticket_ids": [1.9], "completed_count": 1, "device_id": "dev-confirm"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "integer" in data["error"]


def test_mode_b_confirm_rejects_non_integer_completed_count_payload(app, client):
    with app.app_context():
        create_user("mode_b_confirm_float_completed_user", "secret123", client_mode="mode_b")

    resp = login(client, "mode_b_confirm_float_completed_user", "secret123")
    assert resp.status_code == 200

    resp = client.post(
        "/api/mode-b/confirm",
        json={"ticket_ids": [1], "completed_count": 1.2, "device_id": "dev-confirm"},
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False
    assert "completed_count" in data["error"]


def test_mode_b_confirm_batch_rejects_non_integer_completed_count(app):
    from services.mode_b_service import confirm_batch

    with app.app_context():
        result = confirm_batch([1, 2], user_id=1, completed_count=True)

    assert result["success"] is False
    assert "completed_count" in result["error"]


def test_mode_b_batch_now_is_computed_after_postgres_user_lock(app, monkeypatch):
    from services.ticket_pool import assign_tickets_batch

    order = []

    def fake_acquire(_user_id):
        order.append("lock")

    def fake_ensure(_user_id, _device_id, assigned_at):
        order.append("ensure")
        return assigned_at

    class _Result:
        def fetchall(self):
            return []

    def fake_execute(*_args, **_kwargs):
        return _Result()

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool._acquire_postgres_user_assignment_lock", fake_acquire)
    monkeypatch.setattr("services.ticket_pool._ensure_unique_batch_assigned_at", fake_ensure)
    monkeypatch.setattr(db.session, "execute", fake_execute)

    with app.app_context():
        tickets, message = assign_tickets_batch(
            user_id=1,
            device_id="dev-order",
            username="tester",
            count=5,
            max_processing=None,
            daily_limit=None,
            blocked_lottery_types=[],
        )

    assert tickets == []
    assert message is None
    assert order[:2] == ["lock", "ensure"]


def test_recalc_falls_back_to_sync_when_scheduler_enqueue_fails(app, client, monkeypatch):
    calls = {}

    class FailingScheduler:
        def add_job(self, *args, **kwargs):
            raise RuntimeError("scheduler down")

    def fake_process(result_id, expected_calc_token=None, expected_uploaded_at=None, app=None):
        calls["result_id"] = result_id
        calls["token"] = expected_calc_token

    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: FailingScheduler())
    monkeypatch.setattr("services.winning_calc_service.process_match_result", fake_process)

    with app.app_context():
        admin = User(username="admin_recalc_sync_fallback", is_admin=True)
        admin.set_password("secret123")
        match_result = MatchResult(
            detail_period="27101",
            result_data={"1": {"SPF": {"result": "3", "sp": 1.23}}},
            uploaded_by=1,
            calc_status="done",
            tickets_total=7,
        )
        db.session.add_all([admin, match_result])
        db.session.commit()
        result_id = match_result.id

    resp = client.post("/auth/login", json={"username": "admin_recalc_sync_fallback", "password": "secret123"})
    assert resp.status_code == 200

    resp = client.post(f"/admin/api/match-results/{result_id}/recalc")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert calls["result_id"] == result_id
    assert isinstance(calls["token"], str)


def test_match_result_upload_falls_back_to_sync_when_scheduler_enqueue_fails(app, client, monkeypatch):
    calls = {}

    class FailingScheduler:
        def add_job(self, *args, **kwargs):
            raise RuntimeError("scheduler down")

    def fake_parse_result_file(*_args, **_kwargs):
        return {
            "success": True,
            "count": 1,
            "match_result_id": 99991,
            "calc_token": "rf:99991",
        }

    def fake_process(result_id, expected_calc_token=None, expected_uploaded_at=None, app=None):
        calls["result_id"] = result_id
        calls["token"] = expected_calc_token

    monkeypatch.setattr("tasks.scheduler.get_scheduler", lambda: FailingScheduler())
    monkeypatch.setattr("services.result_parser.parse_result_file", fake_parse_result_file)
    monkeypatch.setattr("services.winning_calc_service.process_match_result", fake_process)

    with app.app_context():
        admin = User(username="admin_upload_sync_fallback", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_upload_sync_fallback", "password": "secret123"})
    assert resp.status_code == 200

    payload = "header\n1\t3\t1.80\n".encode("utf-8")
    resp = client.post(
        "/admin/match-results/upload",
        data={
            "detail_period": "27102",
            "upload_kind": "final",
            "file": (io.BytesIO(payload), "27102_final.txt"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert calls["result_id"] == 99991
    assert calls["token"] == "rf:99991"


def test_winning_calc_cleanup_failure_does_not_flip_done_status(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    def fake_calculate_winning(raw_content, result_data, multiplier, sp_field='sp'):
        return False, 0, 0, 0

    def fake_delete_stored_image(_oss_key, _image_url):
        raise RuntimeError("oss unavailable")

    monkeypatch.setattr("services.winning_calc_service.calculate_winning", fake_calculate_winning)
    monkeypatch.setattr("services.winning_calc_service.delete_stored_image", fake_delete_stored_image)

    with app.app_context():
        user = create_user("winning_cleanup_fallback_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="27103",
            result_data={"1": {"SPF": {"result": "0", "sp": 1.2}}},
            uploaded_by=user.id,
            calc_status="pending",
        )
        ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="27103",
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
            is_winning=True,
            winning_image_url="/uploads/images/cleanup-failed.png",
        )
        db.session.add_all([match_result, ticket])
        db.session.commit()
        record = WinningRecord(
            ticket_id=ticket.id,
            source_file_id=ticket.source_file_id,
            detail_period=ticket.detail_period,
            lottery_type=ticket.lottery_type,
            winning_image_url=ticket.winning_image_url,
            image_oss_key="oss/cleanup-failed.png",
            uploaded_by=user.id,
        )
        db.session.add(record)
        db.session.commit()
        match_result_id = match_result.id
        ticket_id = ticket.id

    process_match_result(match_result_id, app=app)

    with app.app_context():
        refreshed_result = db.session.get(MatchResult, match_result_id)
        refreshed_ticket = db.session.get(LotteryTicket, ticket_id)
        refreshed_record = WinningRecord.query.filter_by(ticket_id=ticket_id).first()

    assert refreshed_result.calc_status == "done"
    assert refreshed_ticket.is_winning is False
    assert refreshed_record is None


def test_parse_result_file_updates_latest_result_with_same_period_and_lottery_type(app, tmp_path):
    from services.result_parser import parse_result_file

    with app.app_context():
        admin = User(username="result_type_isolation_admin", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        admin_id = admin.id

        older_spf = MatchResult(
            detail_period="27210",
            lottery_type="???",
            result_data={"61": {"SPF": {"result": "0", "sp": 1.10}}},
            uploaded_by=admin_id,
            uploaded_at=beijing_now() - timedelta(days=2),
        )
        newer_spf = MatchResult(
            detail_period="27210",
            lottery_type="???",
            result_data={"61": {"SPF": {"result": "1", "sp": 1.20}}},
            uploaded_by=admin_id,
            uploaded_at=beijing_now() - timedelta(days=1),
        )
        other_type = MatchResult(
            detail_period="27210",
            lottery_type="\u6bd4\u5206",
            result_data={"61": {"CBF": {"result": "1:0", "sp": 7.80}}},
            uploaded_by=admin_id,
            uploaded_at=beijing_now() - timedelta(hours=12),
        )
        db.session.add_all([older_spf, newer_spf, other_type])
        db.session.commit()
        older_id = older_spf.id
        newer_id = newer_spf.id
        other_id = other_type.id

    result_file = tmp_path / "result_type_isolation.txt"
    result_file.write_text("\u5e8f\u53f7\tA\tB\n61\t3\t1.88\n", encoding="utf-8")

    with app.app_context():
        result = parse_result_file(
            str(result_file),
            "27210",
            admin_id,
            upload_kind="final",
            lottery_type="???",
        )
        assert result["success"] is True
        assert result["match_result_id"] == newer_id

        refreshed_older = db.session.get(MatchResult, older_id)
        refreshed_newer = db.session.get(MatchResult, newer_id)
        refreshed_other = db.session.get(MatchResult, other_id)

    assert refreshed_older.result_data == {"61": {"SPF": {"result": "0", "sp": 1.10}}}
    assert refreshed_newer.result_data["61"]["SPF"]["result"] == "3"
    assert refreshed_other.result_data == {"61": {"CBF": {"result": "1:0", "sp": 7.80}}}


def test_upload_match_result_allows_ambiguous_period_without_lottery_type(app, client):
    with app.app_context():
        admin = User(username="admin_result_ambiguous_type", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

        db.session.add_all([
            LotteryTicket(
                source_file_id=1,
                line_number=1,
                raw_content="SPF|1=3|1*1|1",
                detail_period="27211",
                lottery_type="???",
                status="completed",
                assigned_user_id=admin.id,
                assigned_username=admin.username,
                completed_at=beijing_now(),
            ),
            LotteryTicket(
                source_file_id=1,
                line_number=2,
                raw_content="CBF|1=1:0|1*1|1",
                detail_period="27211",
                lottery_type="\u6bd4\u5206",
                status="completed",
                assigned_user_id=admin.id,
                assigned_username=admin.username,
                completed_at=beijing_now(),
            ),
        ])
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_result_ambiguous_type", "password": "secret123"})
    assert resp.status_code == 200

    payload = "\u5e8f\u53f7\tA\tB\n61\t3\t1.88\n".encode("utf-8")
    resp = client.post(
        "/admin/match-results/upload",
        data={
            "detail_period": "27211",
            "upload_kind": "final",
            "file": (io.BytesIO(payload), "27211_final.txt"),
        },
        content_type="multipart/form-data",
    )

    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True
    assert data["match_result_id"]

    with app.app_context():
        match_result = db.session.get(MatchResult, data["match_result_id"])
        assert match_result is not None
        assert match_result.detail_period == "27211"
        assert match_result.lottery_type is None


def test_winning_calc_processes_only_matching_lottery_type_tickets(app, monkeypatch):
    from services.winning_calc_service import process_match_result

    monkeypatch.setattr(
        "services.winning_calc_service.calculate_winning",
        lambda raw_content, result_data, multiplier, sp_field='sp': (True, 10, 10, 0),
    )

    with app.app_context():
        user = create_user("winning_calc_type_scope_user", "secret123", client_mode="mode_b")
        match_result = MatchResult(
            detail_period="27212",
            lottery_type="???",
            result_data={"1": {"SPF": {"result": "3", "sp": 2.0}}},
            uploaded_by=user.id,
            calc_status="pending",
        )
        spf_ticket = LotteryTicket(
            source_file_id=1,
            line_number=1,
            raw_content="SPF|1=3|1*1|1",
            status="completed",
            detail_period="27212",
            lottery_type="???",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        cbf_ticket = LotteryTicket(
            source_file_id=1,
            line_number=2,
            raw_content="CBF|1=1:0|1*1|1",
            status="completed",
            detail_period="27212",
            lottery_type="\u6bd4\u5206",
            multiplier=1,
            assigned_user_id=user.id,
            assigned_username=user.username,
            completed_at=beijing_now(),
        )
        db.session.add_all([match_result, spf_ticket, cbf_ticket])
        db.session.commit()
        match_result_id = match_result.id
        spf_ticket_id = spf_ticket.id
        cbf_ticket_id = cbf_ticket.id

    process_match_result(match_result_id, app=app)

    with app.app_context():
        refreshed_result = db.session.get(MatchResult, match_result_id)
        refreshed_spf = db.session.get(LotteryTicket, spf_ticket_id)
        refreshed_cbf = db.session.get(LotteryTicket, cbf_ticket_id)

    assert refreshed_result.tickets_total == 1
    assert refreshed_result.tickets_winning == 1
    assert refreshed_spf.is_winning is True
    assert refreshed_cbf.is_winning is None


def test_mode_b_batch_acquires_reserve_lock_before_type_stats_query(app, monkeypatch):
    from services.ticket_pool import assign_tickets_batch

    events = []

    class FakeResult:
        def __init__(self, fetchall_data=None):
            self._fetchall_data = fetchall_data or []

        def fetchall(self):
            return self._fetchall_data

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT lottery_type, deadline_time, COUNT(*) as cnt" in sql:
            events.append("type_stats")
            return FakeResult(fetchall_data=[])
        raise AssertionError(f"Unexpected SQL: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool._acquire_postgres_user_assignment_lock", lambda _user_id: events.append("user_lock"))
    monkeypatch.setattr("services.ticket_pool._acquire_postgres_mode_b_reserve_lock", lambda: events.append("reserve_lock"))
    monkeypatch.setattr("services.ticket_pool._ensure_unique_batch_assigned_at", lambda _user_id, _device_id, assigned_at: assigned_at)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: datetime(2026, 4, 7, 10, 30, 0))
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)

    with app.app_context():
        tickets, message = assign_tickets_batch(
            user_id=1,
            device_id="device-b",
            username="tester",
            count=1,
        )

    assert tickets == []
    assert message is None
    assert events[:3] == ["user_lock", "reserve_lock", "type_stats"]


def test_upload_match_result_sanitizes_nested_filename_paths(app, client):
    with app.app_context():
        admin = User(username="admin_result_nested_filename", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_result_nested_filename", "password": "secret123"})
    assert resp.status_code == 200

    payload = "\u5e8f\u53f7\tA\tB\n1\t3\t1.85\n".encode("utf-8")
    resp = client.post(
        "/admin/match-results/upload",
        data={
            "detail_period": "27213",
            "upload_kind": "final",
            "file": (io.BytesIO(payload), "nested/27213_final.txt"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["success"] is True

    with app.app_context():
        rf = ResultFile.query.order_by(ResultFile.id.desc()).first()

    assert rf is not None
    normalized_stored = rf.stored_filename.replace("\\", "/")
    assert normalized_stored.startswith("results/27213/")
    assert ".." not in normalized_stored
    assert "/" not in rf.original_filename
    assert "\\" not in rf.original_filename


def test_parse_result_file_duplicate_seq_keeps_unique_count(app, tmp_path):
    from services.result_parser import parse_result_file

    with app.app_context():
        admin = User(username="result_duplicate_seq_count_admin", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()
        admin_id = admin.id

    result_file = tmp_path / "duplicate_seq_result.txt"
    result_file.write_text("\u5e8f\u53f7\tA\tB\n1\t3\t1.80\n1\t1\t2.20\n", encoding="utf-8")

    with app.app_context():
        result = parse_result_file(str(result_file), "27214", admin_id, upload_kind="final")
        assert result["success"] is True
        assert result["count"] == 1
        mr = db.session.get(MatchResult, result["match_result_id"])

    assert mr is not None
    assert sorted(mr.result_data.keys()) == ["1"]
    assert mr.result_data["1"]["SPF"]["result"] == "1"
    assert mr.result_data["1"]["SPF"]["sp"] == 2.20


def test_mode_b_batch_sqlite_assigns_null_lottery_type_tickets(app):
    from services.ticket_pool import assign_tickets_batch

    with app.app_context():
        user = create_user("mode_b_null_type_sqlite_user", "secret123", client_mode="mode_b")
        deadline = beijing_now() + timedelta(hours=2)
        db.session.add_all([
            LotteryTicket(
                source_file_id=1,
                line_number=i,
                raw_content=f"NULL-TYPE-{i}",
                lottery_type=None,
                deadline_time=deadline,
                status="pending",
            )
            for i in range(1, 26)
        ])
        db.session.commit()

        tickets, adjustment_message = assign_tickets_batch(
            user_id=user.id,
            device_id="device-null-type",
            username=user.username,
            count=3,
            max_processing=None,
            daily_limit=None,
            blocked_lottery_types=[],
        )

        assigned_count = LotteryTicket.query.filter_by(
            assigned_user_id=user.id,
            assigned_device_id="device-null-type",
            status="assigned",
        ).count()
        assigned_types = [ticket.lottery_type for ticket in tickets]

    assert adjustment_message is None
    assert len(tickets) == 3
    assert all(lottery_type is None for lottery_type in assigned_types)
    assert assigned_count == 3


def test_mode_b_batch_postgres_uses_is_null_clause_for_null_lottery_type(app, monkeypatch):
    from services.ticket_pool import assign_tickets_batch

    class FakeResult:
        def __init__(self, fetchall_data=None):
            self._fetchall_data = fetchall_data or []

        def fetchall(self):
            return self._fetchall_data

    class TypeStat:
        def __init__(self, lottery_type, deadline_time, cnt):
            self.lottery_type = lottery_type
            self.deadline_time = deadline_time
            self.cnt = cnt

    now = datetime(2026, 4, 8, 9, 0, 0)
    selected_deadline = datetime(2026, 4, 8, 12, 0, 0)
    observed_sql = {"value": None}

    def fake_execute(statement, params=None):
        sql = str(statement)
        if "SELECT lottery_type, deadline_time, COUNT(*) as cnt" in sql:
            return FakeResult(fetchall_data=[TypeStat(None, selected_deadline, 30)])
        if "SELECT id FROM lottery_tickets" in sql and "FOR UPDATE SKIP LOCKED" in sql:
            observed_sql["value"] = sql
            return FakeResult(fetchall_data=[])
        raise AssertionError(f"Unexpected SQL executed: {sql}")

    monkeypatch.setattr("services.ticket_pool._is_postgres", lambda: True)
    monkeypatch.setattr("services.ticket_pool._acquire_postgres_user_assignment_lock", lambda _user_id: None)
    monkeypatch.setattr("services.ticket_pool._acquire_postgres_mode_b_reserve_lock", lambda: None)
    monkeypatch.setattr("services.ticket_pool._ensure_unique_batch_assigned_at", lambda _user_id, _device_id, assigned_at: assigned_at)
    monkeypatch.setattr("services.ticket_pool.beijing_now", lambda: now)
    monkeypatch.setattr("services.ticket_pool.db.session.execute", fake_execute)

    with app.app_context():
        tickets, adjustment_message = assign_tickets_batch(
            user_id=1,
            device_id="device-null-type-pg",
            username="tester",
            count=3,
            max_processing=None,
            daily_limit=None,
            blocked_lottery_types=[],
        )

    assert tickets == []
    assert adjustment_message is None
    assert observed_sql["value"] is not None
    assert "lottery_type IS NULL" in observed_sql["value"]


def test_admin_users_import_returns_400_when_service_reports_failure(app, client, monkeypatch):
    with app.app_context():
        admin = User(username="admin_import_failure_status", is_admin=True)
        admin.set_password("secret123")
        db.session.add(admin)
        db.session.commit()

    resp = client.post("/auth/login", json={"username": "admin_import_failure_status", "password": "secret123"})
    assert resp.status_code == 200

    monkeypatch.setattr(
        "services.user_import_service.import_users",
        lambda *_args, **_kwargs: {"success": False, "error": "invalid workbook"},
    )

    resp = client.post(
        "/admin/api/users/import",
        data={"file": (io.BytesIO(b"dummy"), "users-import.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    data = resp.get_json()
    assert data["success"] is False


def test_match_result_flags_ignore_malformed_payload_shapes():
    record = MatchResult(detail_period="shape-guard", result_data=["bad", "payload"])
    assert record.has_predicted_results() is False
    assert record.has_final_results() is False

    record.result_data = {"1": {"SPF": "not-an-object"}}
    assert record.has_predicted_results() is False
    assert record.has_final_results() is False

    record.result_data = {"1": {"SPF": {"predicted_sp": 1.55}}}
    assert record.has_predicted_results() is True
    assert record.has_final_results() is False

    record.result_data = {"1": {"SPF": {"sp": 2.05}}}
    assert record.has_predicted_results() is False
    assert record.has_final_results() is True


def test_can_receive_required_returns_json_401_for_anonymous_user(app, client):
    from flask import jsonify
    from utils.decorators import can_receive_required

    endpoint = "probe_can_receive_guard"
    if endpoint not in app.view_functions:
        @app.route("/__probe-can-receive-guard", endpoint=endpoint)
        @can_receive_required
        def _probe_can_receive_guard():
            return jsonify({"success": True})

    resp = client.get("/__probe-can-receive-guard")
    assert resp.status_code == 401
    data = resp.get_json()
    assert data["success"] is False


def test_delete_object_local_mode_flattens_slash_oss_key_to_local_filename(app, tmp_path):
    from services.oss_service import delete_object

    with app.app_context():
        app.config["UPLOAD_FOLDER"] = str(tmp_path)
        images_dir = tmp_path / "images"
        images_dir.mkdir(parents=True, exist_ok=True)
        local_file = images_dir / "winning_2026_04_19_123.jpg"
        local_file.write_bytes(b"img")

        ok = delete_object("winning/2026/04/19/123.jpg")

    assert ok is True
    assert local_file.exists() is False


def test_delete_stored_image_falls_back_to_image_url_when_oss_key_delete_fails(app, tmp_path):
    from services.oss_service import delete_stored_image

    with app.app_context():
        app.config["UPLOAD_FOLDER"] = str(tmp_path)
        images_dir = tmp_path / "images"
        images_dir.mkdir(parents=True, exist_ok=True)
        fallback_file = images_dir / "fallback-cleanup.jpg"
        fallback_file.write_bytes(b"img")

        ok = delete_stored_image(
            image_oss_key="winning/2026/04/19/not-existing.jpg",
            image_url="/uploads/images/fallback-cleanup.jpg",
        )

    assert ok is True
    assert fallback_file.exists() is False


def test_delete_object_local_mode_rejects_backslash_path_traversal_key(app, tmp_path):
    from services.oss_service import delete_object

    with app.app_context():
        app.config["UPLOAD_FOLDER"] = str(tmp_path)
        images_dir = tmp_path / "images"
        images_dir.mkdir(parents=True, exist_ok=True)
        outside_file = tmp_path / "outside.txt"
        outside_file.write_bytes(b"outside")

        ok = delete_object("..\\outside.txt")

    assert ok is False
    assert outside_file.exists() is True


def test_delete_stored_image_rejects_backslash_path_traversal_in_image_url(app, tmp_path):
    from services.oss_service import delete_stored_image

    with app.app_context():
        app.config["UPLOAD_FOLDER"] = str(tmp_path)
        images_dir = tmp_path / "images"
        images_dir.mkdir(parents=True, exist_ok=True)
        outside_file = tmp_path / "outside-image.txt"
        outside_file.write_bytes(b"outside")

        ok = delete_stored_image(image_url="/uploads/images/..\\outside-image.txt")

    assert ok is False
    assert outside_file.exists() is True

